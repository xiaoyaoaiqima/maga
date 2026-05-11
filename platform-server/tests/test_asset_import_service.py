"""Tests for MAGA marketing asset registry and Excel import service."""

from pathlib import Path

import pytest
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.models.base import Base
from app.models.maga_assets import AssetImportRun, AssetRegistry
from app.services.asset_import_service import import_yuanyue_training_rules


@pytest.mark.asyncio
async def test_import_yuanyue_training_rules_creates_versioned_assets(tmp_path: Path):
    from openpyxl import Workbook

    workbook_path = tmp_path / "源悦种草活动-ai训练规则.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "品牌资料整理"
    ws["B3"] = "内容focus"
    ws["C3"] = "好消化易吸收，对应便便好，不上火"
    ws["B4"] = "内容风格"
    ws["C4"] = "高质量真实用户ugc"
    ws["B8"] = "分级"
    ws["C8"] = "卖点"
    ws["D8"] = "成分"
    ws["E8"] = "源悦优势"
    ws["F8"] = "生动理解"
    ws["B9"] = "核心卖点"
    ws["C9"] = "好消化易吸收"
    ws["D9"] = "软分子蛋白"
    ws["E9"] = "形成结构松散的软凝乳"
    ws["F9"] = "软软的米糊"

    ws2 = wb.create_sheet("内容模型")
    ws2.append(["序号", "宝宝阶段", "核心痛点", "具体表现", "痛点描述", "对应卖点"])
    ws2.append([None, None, "便便不规律", "羊屎蛋/干硬", "便便又干又硬", "好消化易吸收"])

    ws3 = wb.create_sheet("ugc卖点表述")
    ws3.append(["序号", "对应卖点", "卖点描述", "负责人"])
    ws3.append([None, "便便不规律", "便便基本一天一次，拉起来也不费劲", "东昕"])

    ws4 = wb.create_sheet("审核规则")
    ws4.append(["序号", "审核内容", "分类", "审核维度（问题分类）", "审核意见（返回给用户的）"])
    ws4.append([1, "文案审核", "草稿审核", "夸大产品效果或虚构使用经历", "文本不符合活动要求，请修改后重新提交"])

    ws5 = wb.create_sheet("例文收集")
    ws5.append(["序号", "来源", "方向", "ID昵称", "粉丝量", "发布链接", "发布形式", "笔记标题", "笔记正文", "发布时间", "统计时间", "痛点"])
    ws5.append([None, "品牌", "吃", "大小周妈妈", "0.87", "http://xhslink.com/a", "用后分享", "源悦你给我出来 我是真的会谢！", "新手妈妈别急着焦虑，先看宝宝喝奶和便便状态。", None, None, "肠胃弱/奶量上不去"])
    ws5.append([None, "品牌", "吃", "六六", "1", "http://xhslink.com/b", "用后分享", "真实经验！转奶终于不踩坑！😭", "我会看拉臭费不费劲，也会看肚肚舒不舒服。", None, None, "转奶/消化吸收"])
    wb.save(workbook_path)

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all, tables=[AssetRegistry.__table__, AssetImportRun.__table__])
    session_factory = async_sessionmaker(engine, expire_on_commit=False)

    async with session_factory() as session:
        result = await import_yuanyue_training_rules(session, workbook_path, source_name="源悦种草活动-ai训练规则.xlsx")
        await session.commit()

    assert result.imported_assets >= 4
    assert result.import_run_id is not None

    async with session_factory() as session:
        assets = (await session.execute(AssetRegistry.__table__.select())).mappings().all()

    by_key = {(row["asset_type"], row["asset_key"]): row for row in assets}
    assert ("brand_profile", "yuanyue") in by_key
    assert ("product_selling_points", "yuanyue") in by_key
    assert ("painpoint_model", "yuanyue") in by_key
    assert ("ugc_expression_corpus", "yuanyue") in by_key
    assert ("compliance_rules", "yuanyue") in by_key
    assert ("reference_examples", "yuanyue") in by_key

    painpoints = by_key[("painpoint_model", "yuanyue")]["content_json"]["items"]
    assert painpoints == [
        {
            "baby_stage": None,
            "painpoint": "便便不规律",
            "symptom": "羊屎蛋/干硬",
            "description": "便便又干又硬",
            "selling_point": "好消化易吸收",
            "extra_descriptions": [],
        }
    ]

    expressions = by_key[("ugc_expression_corpus", "yuanyue")]["content_json"]["items"]
    assert expressions[0]["painpoint_or_selling_point"] == "便便不规律"
    assert expressions[0]["expression"] == "便便基本一天一次，拉起来也不费劲"

    examples = by_key[("reference_examples", "yuanyue")]["content_json"]["items"]
    assert examples == [
        {
            "example_id": "yuanyue_ref_001",
            "source": "品牌",
            "direction": "吃",
            "author_name": "大小周妈妈",
            "follower_count_w": "0.87",
            "post_url": "http://xhslink.com/a",
            "post_format": "用后分享",
            "title": "源悦你给我出来 我是真的会谢！",
            "body": "新手妈妈别急着焦虑，先看宝宝喝奶和便便状态。",
            "painpoint": "肠胃弱/奶量上不去",
            "reference_type": "用后分享",
            "owner": "大小周妈妈",
            "style_tags": ["用后分享", "吃"],
            "structure_tags": [],
        },
        {
            "example_id": "yuanyue_ref_002",
            "source": "品牌",
            "direction": "吃",
            "author_name": "六六",
            "follower_count_w": "1",
            "post_url": "http://xhslink.com/b",
            "post_format": "用后分享",
            "title": "真实经验！转奶终于不踩坑！😭",
            "body": "我会看拉臭费不费劲，也会看肚肚舒不舒服。",
            "painpoint": "转奶/消化吸收",
            "reference_type": "用后分享",
            "owner": "六六",
            "style_tags": ["用后分享", "吃"],
            "structure_tags": [],
        },
    ]
