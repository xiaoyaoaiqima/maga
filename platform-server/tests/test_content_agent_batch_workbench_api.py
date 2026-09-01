"""API tests for the operator-facing content-agent workbench batch flow."""

import csv
import json
import re
from collections import Counter
from io import BytesIO, StringIO
from types import SimpleNamespace

from openpyxl import load_workbook
from pydantic import ValidationError
import pytest
import pytest_asyncio
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.api.v1.endpoints.content_agent import router
from app.core.database import get_db
from app.models.base import Base
from app.models.content_agent import (
    ContentBatchJob,
    ContentAgentEvent,
    ContentAgentStageCall,
    ContentBatchItem,
    ContentBatchItemVersion,
    ContentFeedback,
    ExecutorRegistry,
)
from app.models.llm_provider_config import LLMProviderConfig
from app.models.maga_assets import AssetChangeRequest, AssetRegistry
from app.models.maga_core import MAGA_CORE_TABLE_NAMES
from app.schemas.content_batch_report import (
    ContentBatchReportResponse,
    ContentBatchReportSummary,
    ContentBatchReportItem,
    ContentBatchStartRequest,
    ContentCommentBatchStartRequest,
)
from app.services.activity_quality_guard_service import (
    ActivityQualityGuardService,
    build_article_pool_context_list,
    resolve_quality_guard_profile,
)
from app.services.comment_batch_variation_review_service import CommentBatchVariationReviewService
from app.services.comment_batch_delivery_selection_service import CommentBatchDeliverySelectionService
from app.services.content_comment_batch_service import (
    ContentCommentBatchService,
)
from app.services.content_batch_report_service import (
    ContentBatchReportService,
    _article_pool_csv_filename,
    _article_pool_export_items,
    _build_article_pool_csv,
)
from app.services.comment_realness_review_service import (
    CommentRealnessReviewService,
    STATIC_COMMENT_REALNESS_REPLACEMENTS,
    _remove_or_replace_realness_terms,
    _rewrite_input_payload,
    find_comment_realness_hits,
)


@pytest_asyncio.fixture
async def content_agent_workbench_client():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        tables = [Base.metadata.tables[name] for name in MAGA_CORE_TABLE_NAMES]
        await conn.run_sync(lambda sync_conn: Base.metadata.create_all(sync_conn, tables=tables))
    session_factory = async_sessionmaker(engine, expire_on_commit=False)

    async with session_factory() as session:
        session.add(
            ExecutorRegistry(
                executor_code="maga_direct_llm_executor",
                executor_type="direct_llm",
                profile_name=None,
                display_name="MAGA direct LLM executor",
                invoke_url="mock://direct-llm/content",
                supported_capabilities_json=[
                    {"capability": "asset.import", "schema_version": "1"},
                    {"capability": "content.generate", "schema_version": "1"},
                    {"capability": "content.rewrite", "schema_version": "1"},
                ],
                enabled=1,
            )
        )
        session.add_all(_yuanyue_assets())
        await session.commit()

    app = FastAPI()
    app.include_router(router, prefix="/api/v1/content-agent")

    async def override_get_db():
        async with session_factory() as session:
            try:
                yield session
                await session.commit()
            except Exception:
                await session.rollback()
                raise

    app.dependency_overrides[get_db] = override_get_db
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as client:
        yield client, session_factory

    await engine.dispose()


def test_article_batch_start_request_allows_one_thousand_items():
    assert ContentBatchStartRequest(count=1000).count == 1000
    with pytest.raises(ValidationError):
        ContentBatchStartRequest(count=1001)


def test_comment_batch_start_request_allows_three_hundred_items():
    assert ContentCommentBatchStartRequest(count=300).count == 300
    assert ContentCommentBatchStartRequest().concurrency == 5
    assert ContentCommentBatchStartRequest(concurrency=10).concurrency == 10
    with pytest.raises(ValidationError):
        ContentCommentBatchStartRequest(count=301)
    with pytest.raises(ValidationError):
        ContentCommentBatchStartRequest(concurrency=51)


def test_comment_batch_start_request_accepts_multi_rule_prompt_slot_probe():
    request = ContentCommentBatchStartRequest(
        rule_ids=["a2_direct_28", "a2_direct_29"],
        comment_prompt_slots={"开头方式": ["不要用固定开头，直接接话。"]},
        comment_batch_variation_review={"enabled": True, "affects_hard_pass": False},
        comment_delivery_selection={"enabled": True, "target_count": 105},
        comment_post_context="你正在回复一篇消费者吐槽价格的帖子。",
    )

    assert request.rule_ids == ["a2_direct_28", "a2_direct_29"]
    assert request.comment_prompt_slots == {"开头方式": ["不要用固定开头，直接接话。"]}
    assert request.comment_batch_variation_review == {"enabled": True, "affects_hard_pass": False}
    assert request.comment_delivery_selection == {"enabled": True, "target_count": 105}
    assert request.comment_post_context == "你正在回复一篇消费者吐槽价格的帖子。"


@pytest.mark.asyncio
async def test_ppl_profile_list_exposes_brand_generation_profiles(content_agent_workbench_client):
    client, _session_factory = content_agent_workbench_client

    response = await client.get("/api/v1/content-agent/ppl-runs/profiles")

    assert response.status_code == 200
    profiles = response.json()["data"]["items"]
    profile_codes = {profile["profile_code"] for profile in profiles}
    assert {
        "royal_friso_ugc_article",
        "wangyue_0705_article",
        "wangyue_v3_0705_article",
        "a2_sentiment_post_article",
        "a2_sentiment_comment",
    }.issubset(profile_codes)
    royal = next(profile for profile in profiles if profile["profile_code"] == "royal_friso_ugc_article")
    assert royal["content_type"] == "article"
    assert royal["asset_key"] == "royal_friso_ugc_post_rules_v1"
    assert royal["keyword_asset_key"] == "royal_friso_ugc_post_keywords_v1"
    assert royal["prompt_mode"] == "royal_compact"
    wangyue_v3 = next(profile for profile in profiles if profile["profile_code"] == "wangyue_v3_0705_article")
    assert wangyue_v3["asset_key"] == "wangyue_v3_core_storyline_article_rules"
    assert wangyue_v3["keyword_asset_key"] is None
    assert wangyue_v3["prompt_mode"] == "layered_article"
    assert wangyue_v3["default_count"] == 20
    assert wangyue_v3["default_articles_per_prompt"] == 1


def test_comment_rule_selection_balances_angles_when_rule_pool_is_large():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rules = [
        {
            "business_rule": angle,
            "corpus": f"{angle} corpus {index}",
            "source_row_no": index,
        }
        for angle, start in [("便便问题", 1), ("奶量补充", 11), ("生长发育", 21)]
        for index in range(start, start + 10)
    ]

    selected = service._select_rules(rules, 6)

    assert len(selected) == 6
    assert {rule["business_rule"] for rule in selected} == {"便便问题", "奶量补充", "生长发育"}
    assert [rule["source_row_no"] for rule in selected] != list(range(1, 7))


def test_a2_profile_rule_selection_balances_required_keywords():
    profile = resolve_quality_guard_profile("a2_sentiment_comment_202606")
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rules = [
        {"business_rule": "A2舆情改善评论", "corpus": f"有货+批批检规则-{index}\n有货补货物流码报告0.03"}
        for index in range(6)
    ] + [
        {"business_rule": "A2舆情改善评论", "corpus": f"批批检+转奶规则-{index}\n转奶肚肚物流码报告0.03"}
        for index in range(6)
    ] + [
        {"business_rule": "A2舆情改善评论", "corpus": f"有货+转奶规则-{index}\n有货转奶物流码报告0.03"}
        for index in range(6)
    ]

    selected, selection_mode = service._select_rules_for_batch(
        rules,
        12,
        focus_business_rule=None,
        profile=profile,
    )

    assert selection_mode == "keyword_balanced_random"
    assert len(selected) == 12
    assert sum("有货+批批检" in rule["corpus"] for rule in selected) == 4
    assert sum("批批检+转奶" in rule["corpus"] for rule in selected) == 4
    assert sum("有货+转奶" in rule["corpus"] for rule in selected) == 4


def test_a2_profile_rule_selection_repeats_balanced_keywords_for_large_count():
    profile = resolve_quality_guard_profile("a2_sentiment_comment_202606")
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rules = [
        {"business_rule": "A2舆情改善评论", "corpus": f"有货+批批检规则-{index}\n有货补货物流码报告0.03"}
        for index in range(2)
    ] + [
        {"business_rule": "A2舆情改善评论", "corpus": f"批批检+转奶规则-{index}\n转奶肚肚物流码报告0.03"}
        for index in range(2)
    ] + [
        {"business_rule": "A2舆情改善评论", "corpus": f"有货+转奶规则-{index}\n有货转奶物流码报告0.03"}
        for index in range(2)
    ]
    asset = SimpleNamespace(metadata_json={}, content_json={})

    limit = service._generation_limit(asset, rules, requested_count=20)
    selected, selection_mode = service._select_rules_for_batch(
        rules,
        limit,
        focus_business_rule=None,
        profile=profile,
    )

    assert limit == 20
    assert selection_mode == "keyword_balanced_random_with_replacement"
    assert len(selected) == 20
    counts = [
        sum("有货+批批检" in rule["corpus"] for rule in selected),
        sum("批批检+转奶" in rule["corpus"] for rule in selected),
        sum("有货+转奶" in rule["corpus"] for rule in selected),
    ]
    assert min(counts) >= 6
    assert max(counts) <= 7


def test_comment_rule_selection_repeats_every_rule_evenly_when_requested_multiple():
    profile = resolve_quality_guard_profile("a2_sentiment_comment_202606")
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rules = [
        {
            "business_rule": "A2舆情改善评论",
            "corpus": f"业务规则-{index}\n关键词方向是有货+批批检。",
            "source_row_no": index,
        }
        for index in range(1, 24)
    ]
    asset = SimpleNamespace(metadata_json={}, content_json={})

    limit = service._generation_limit(asset, rules, requested_count=46)
    selected, selection_mode = service._select_rules_for_batch(
        rules,
        limit,
        focus_business_rule=None,
        profile=profile,
    )

    assert limit == 46
    assert selection_mode == "even_rule_repetition"
    assert len(selected) == 46
    counts = Counter(rule["source_row_no"] for rule in selected)
    assert set(counts) == set(range(1, 24))
    assert set(counts.values()) == {2}


def test_comment_focus_selection_repeats_rules_for_requested_count():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rules = [
        {
            "business_rule": "整体适应",
            "corpus": "像妈妈在评论区聊刚开始喝源悦的观察。",
            "examples": ["我家刚开始也在看源悦，想蹲蹲真实反馈"],
            "source_row_no": 1,
        }
    ]
    asset = SimpleNamespace(metadata_json={"default_generation_count": 1}, content_json={})

    limit = service._generation_limit(asset, rules, requested_count=5, allow_repeat=True)
    selected = service._select_rules_with_replacement(rules, limit)

    assert limit == 5
    assert len(selected) == 5
    assert all(rule["business_rule"] == "整体适应" for rule in selected)


def test_comment_multi_rule_probe_repeats_rules_evenly_with_remainder():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rules = [
        {
            "rule_id": f"a2_direct_{index}",
            "business_rule": f"会员权益-方向{index}",
            "corpus": f"方向{index}",
        }
        for index in range(28, 35)
    ]

    selected_rules = service._rules_for_multiple_items(
        rules,
        rule_ids=[rule["rule_id"] for rule in rules],
    )
    selected = service._select_rules_even_repetition_with_remainder(selected_rules, 30)

    counts = Counter(rule["rule_id"] for rule in selected)
    assert len(selected) == 30
    assert set(counts) == {rule["rule_id"] for rule in rules}
    assert max(counts.values()) - min(counts.values()) == 1
    assert sorted(counts.values()) == [4, 4, 4, 4, 4, 5, 5]


def test_comment_prompt_slots_override_does_not_mutate_source_rules():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rules = [{"rule_id": "a2_direct_28", "business_rule": "会员权益-集罐换礼", "corpus": "集罐"}]
    slots = {"开头方式": ["直接从自己的动作切入。"]}

    updated = service._rules_with_prompt_slots_override(rules, slots)

    assert updated[0]["prompt_slots"] == slots
    assert "prompt_slots" not in rules[0]


def test_comment_post_context_override_does_not_mutate_source_rules():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rules = [{"rule_id": "a2_direct_01", "business_rule": "有货-直给到货情绪"}]

    updated = service._rules_with_post_context_override(
        rules,
        "你正在回复一篇解读a2恢复供货消息的帖子。",
    )

    assert updated[0]["scenario_post_context"] == "你正在回复一篇解读a2恢复供货消息的帖子。"
    assert "scenario_post_context" not in rules[0]


def test_member_rule_prompt_examples_do_not_mix_other_member_benefits():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rule = {
        "business_rule": "会员权益-集罐换礼",
        "corpus": "集罐换奶粉",
        "examples": [
            "a2集罐能换奶粉，我先留空罐。",
            "a2积分换礼还挺实在。",
            "a2老客礼可以问问条件。",
            "a2会员有抽奖活动。",
            "a2集罐换正装奶粉。",
        ],
    }

    selected, meta = service._selected_prompt_examples(rule)

    assert selected == ["a2集罐能换奶粉，我先留空罐。"]
    assert meta["selected_example_source"] == "examples"


def test_a2_route_keeps_asset_format_selection_without_business_specific_override():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rule = {
        "business_rule": "A2舆情改善评论",
        "corpus": "写什么：妈妈说自己等到或买到 a2 了，所以先继续喝 a2，转奶先放一放。\n怎么说：像评论区接一句或顺手报个信，可以很短，不用把流程都讲全。",
        "examples": ["终于到了", "先不转了", "有底了"],
        "source_row_no": 16,
    }
    asset = SimpleNamespace(
        id=7,
        version_no=3,
        asset_key="a2_sentiment_comment_activity",
        content_json={
            "keyword_selection": {
                "comment_format_control": ["comment_short_clean", "comment_21_35"],
                "comment_speaking_style": ["pass_info"],
            }
        },
        metadata_json={},
    )

    plans = [service._plan_from_rule(rule, asset=asset, item_no=item_no) for item_no in range(1, 7)]

    assert all(
        plan["keyword_selection"]["comment_format_control"] == ["comment_short_clean", "comment_21_35"]
        for plan in plans
    )
    assert all(plan["keyword_selection"]["comment_speaking_style"] == ["pass_info"] for plan in plans)
    assert all("keyword_selection_override" not in plan for plan in plans)
    assert asset.content_json["keyword_selection"]["comment_format_control"] == ["comment_short_clean", "comment_21_35"]


def test_comment_single_rule_filter_supports_by_case_generation():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rules = [
        {"rule_id": "business_rule_001", "source_row_no": 1, "business_rule": "剧情讨论", "corpus": "剧情事实"},
        {"rule_id": "business_rule_002", "source_row_no": 2, "business_rule": "剧情讨论", "corpus": "对讲机"},
    ]
    asset = SimpleNamespace(metadata_json={}, content_json={})

    filtered = service._rules_for_single_item(rules, rule_id="business_rule_001", source_row_no=None)
    limit = service._generation_limit(asset, filtered, requested_count=10, allow_repeat=True)
    selected, selection_mode = service._select_rules_for_batch(
        filtered,
        limit,
        focus_business_rule="剧情讨论",
        profile=None,
    )

    assert limit == 10
    assert selection_mode == "random_with_replacement"
    assert {rule["rule_id"] for rule in selected} == {"business_rule_001"}


def test_comment_draft_rule_override_keeps_active_rule_unchanged():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rules = [
        {
            "rule_id": "business_rule_002",
            "source_row_no": 2,
            "business_rule": "剧情讨论",
            "corpus": "旧语料",
            "examples": ["旧示例"],
        }
    ]
    draft = {
        "rule_id": "business_rule_002",
        "source_row_no": None,
        "corpus": "新语料：\n\n示例：\n- 新示例1\n- 新示例2\n\n注意：示例只作为语义素材。",
    }

    updated = service._rules_with_draft_override(rules, draft)

    assert rules[0]["corpus"] == "旧语料"
    assert rules[0]["examples"] == ["旧示例"]
    assert updated[0]["corpus"].startswith("新语料")
    assert "示例：" not in updated[0]["corpus"]
    assert updated[0]["examples"] == ["新示例1", "新示例2"]
    assert updated[0]["draft_rule_override"]["enabled"] is True


def test_comment_prompt_bundle_draft_override_updates_rendered_bundle_only_for_test():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    original_bundle = {
        "generation_instruction": "生成一条真实用户评论。",
        "content_direction": "旧内容方向。",
        "activity_material": ["旧活动素材。"],
        "writing_requirements": ["旧写法。"],
        "notes": ["旧注意。"],
    }
    draft_bundle = {
        "generation_instruction": "生成一条小红书母婴社区真实用户评论。",
        "content_direction": "写看到有货后的即时反应。",
        "activity_material": ["a2已经到货。"],
        "writing_requirements": ["字数在20字以内"],
        "notes": ["不要说消极词。"],
    }
    rules = [
        {
            "rule_id": "a2_direct_01",
            "source_row_no": 1,
            "business_rule": "有货-直给简单报喜",
            "prompt_mode": "comment_prompt_bundle",
            "comment_prompt_bundle": original_bundle,
            "corpus": original_bundle["content_direction"],
            "examples": ["a2终于到货了"],
        }
    ]
    draft = {
        "rule_id": "a2_direct_01",
        "source_row_no": 1,
        "corpus": draft_bundle["content_direction"],
        "comment_prompt_bundle": draft_bundle,
    }

    updated = service._rules_with_draft_override(rules, draft)

    assert rules[0]["comment_prompt_bundle"] == original_bundle
    assert updated[0]["comment_prompt_bundle"] == draft_bundle
    assert updated[0]["corpus"] == draft_bundle["content_direction"]
    assert updated[0]["examples"] == ["a2终于到货了"]


def test_comment_plan_injects_three_reference_examples_from_pool():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rule = {
        "rule_id": "business_rule_001",
        "business_rule": "便便问题",
        "corpus": "像评论区宝妈接话，聊便便频率和软硬。",
        "examples": ["同款，崽崽都是香蕉软便便", "加一，现在固定一天一次", "我家最近也顺了", "蹲一个真实反馈"],
        "supplements": ["我家拉得挺轻松的，没有那么费劲"],
        "source_row_no": 1,
    }
    asset = SimpleNamespace(asset_key="yuanyue", id=7, version_no=3)

    plan = service._plan_from_rule(rule, asset=asset, item_no=1)

    assert plan["render_reference_examples"] is True
    assert len(plan["examples"]) == 3
    assert set(plan["examples"]).issubset(set(rule["examples"]))
    assert plan["supplements"] == []
    assert plan["example_pool_count"] == 4
    assert plan["supplement_pool_count"] == 1
    assert plan["example_sample_count"] == 3
    assert plan["selected_example_source"] == "examples"
    assert len(plan["selected_example_indices"]) == 3


def test_comment_plan_preserves_prompt_slots_from_rule():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rule = {
        "rule_id": "business_rule_001",
        "business_rule": "有货-直给到货情绪",
        "corpus": "像妈妈看到 a2 到货后顺手接一句。",
        "examples": ["a2终于到货了"],
        "prompt_slots": {
            "说话风格": [
                "适当加几个网络热词，不要过度。",
                "像评论区接楼，短一点，顺手补一句。",
            ]
        },
        "source_row_no": 1,
    }
    asset = SimpleNamespace(asset_key="a2_sentiment_comment_activity", id=7, version_no=3)

    plan = service._plan_from_rule(rule, asset=asset, item_no=1)

    assert plan["prompt_slots"] == rule["prompt_slots"]


def test_layered_comment_plan_keeps_activity_material_out_of_scenario_requirement():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rule = {
        "rule_id": "a2_direct_01",
        "business_rule": "有货-直给到货情绪",
        "corpus": "旧规则语料",
        "content_direction": "写看到供货恢复消息后的即时反应。",
        "activity_material": ["a2已经到货或来货", "可以写自己也买到了新货"],
        "scenario_generation_requirements": "直接说线上线下看到货或刚买到。",
        "examples": ["a2终于到货了"],
        "source_row_no": 1,
    }
    asset = SimpleNamespace(asset_key="a2_sentiment_comment_activity", id=7, version_no=65)

    plan = service._plan_from_rule(rule, asset=asset, item_no=1)

    assert plan["content_direction"] == rule["content_direction"]
    assert plan["activity_material"] == rule["activity_material"]
    assert plan["generation_requirements"] is None


def test_comment_prompt_bundle_plan_does_not_mix_examples_tones_or_slots():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    bundle = {
        "generation_instruction": "生成一条小红书母婴社区真实用户评论。",
        "content_direction": "写看到有货后的简单报喜。",
        "activity_material": ["a2已经到货。"],
        "writing_requirements": ["字数在20字以内"],
        "notes": ["不要说消极词。"],
    }
    rule = {
        "rule_id": "a2_direct_01",
        "business_rule": "有货-直给简单报喜",
        "prompt_mode": "comment_prompt_bundle",
        "comment_prompt_bundle": bundle,
        "corpus": bundle["content_direction"],
        "examples": ["不应该进入Prompt"],
        "variation_slots": [
            {"slot_code": "entry", "slot_name": "接法", "options": ["不应该进入Prompt"]}
        ],
        "source_row_no": 1,
    }
    asset = SimpleNamespace(
        asset_key="a2_sentiment_comment_activity",
        id=7,
        version_no=67,
        content_json={
            "comment_tone_options": {
                "stock": [
                    {"tone_code": "confirm", "tone_label": "确认", "prompt": "不应该进入Prompt"}
                ]
            }
        },
        metadata_json={},
    )

    plan = service._plan_from_rule(rule, asset=asset, item_no=1)

    assert plan["prompt_mode"] == "comment_prompt_bundle"
    assert plan["comment_prompt_bundle"] == bundle
    assert plan["render_reference_examples"] is False
    assert plan["examples"] == []
    assert plan["example_sample_count"] == 0
    assert "comment_tone_options" not in plan
    assert "variation_slots" not in plan


def test_comment_prompt_bundle_rotates_only_explicit_batch_prompt_slots():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    bundle = {
        "generation_instruction": "生成一条真实用户评论。",
        "content_direction": "写看到有货后的自然反应。",
        "activity_material": ["a2已经到货。"],
        "writing_requirements": ["字数在30字以内"],
        "notes": ["不要说消极词。"],
    }
    rules = service._rules_with_prompt_slots_override(
        [
            {
                "rule_id": "a2_direct_43",
                "business_rule": "有货-渠道-不提产品",
                "prompt_mode": "comment_prompt_bundle",
                "comment_prompt_bundle": bundle,
                "corpus": bundle["content_direction"],
                "examples": ["不应该进入Prompt"],
                "variation_slots": [
                    {"slot_code": "old", "slot_name": "旧槽", "options": ["不应该进入Prompt"]}
                ],
                "source_row_no": 2,
            }
        ],
        {
            "本条表达路径": [
                "不用时间词，从常买渠道切入，用询问句收尾。",
                "从生活动作切入，不以刚或终于开头。",
            ]
        },
    )
    asset = SimpleNamespace(
        asset_key="a2_sentiment_comment_activity",
        id=7,
        version_no=76,
        content_json={},
        metadata_json={},
    )

    first = service._plan_from_rule(rules[0], asset=asset, item_no=1, rule_occurrence_no=0)
    second = service._plan_from_rule(rules[0], asset=asset, item_no=2, rule_occurrence_no=1)

    assert first["prompt_slots"] == {
        "本条表达路径": ["不用时间词，从常买渠道切入，用询问句收尾。"]
    }
    assert second["prompt_slots"] == {
        "本条表达路径": ["从生活动作切入，不以刚或终于开头。"]
    }
    assert first["bundle_prompt_slots_source"] == "batch_override"
    assert first["examples"] == []
    assert "variation_slots" not in first


def test_comment_prompt_bundle_rotates_rule_asset_prompt_slots():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    bundle = {
        "generation_instruction": "生成一条真实用户评论。",
        "content_direction": "写看到有货后的自然反应。",
        "activity_material": ["a2已经到货。"],
        "writing_requirements": ["字数在30字以内"],
        "notes": ["不要说消极词。"],
    }
    rule = {
        "rule_id": "a2_direct_43",
        "business_rule": "有货-直给-不提产品",
        "prompt_mode": "comment_prompt_bundle",
        "comment_prompt_bundle": bundle,
        "corpus": bundle["content_direction"],
        "prompt_slots": {
            "本条表达路径": [
                "从常买渠道看到消息起句。",
                "像评论区顺手报信。",
            ]
        },
        "prompt_slot_selection_mode": "round_robin",
        "bundle_prompt_slots_source": "rule_asset",
        "variation_slots": [
            {"slot_code": "old", "slot_name": "旧槽", "options": ["不应该进入Prompt"]}
        ],
        "source_row_no": 2,
    }
    asset = SimpleNamespace(
        asset_key="a2_sentiment_comment_activity",
        id=7,
        version_no=77,
        content_json={},
        metadata_json={},
    )

    first = service._plan_from_rule(rule, asset=asset, item_no=1, rule_occurrence_no=0)
    second = service._plan_from_rule(rule, asset=asset, item_no=2, rule_occurrence_no=1)

    assert first["prompt_slots"] == {"本条表达路径": ["从常买渠道看到消息起句。"]}
    assert second["prompt_slots"] == {"本条表达路径": ["像评论区顺手报信。"]}
    assert first["bundle_prompt_slots_source"] == "rule_asset"
    assert "variation_slots" not in first


def test_comment_plan_preserves_variation_slots_from_rule():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rule = {
        "rule_id": "a2_direct_02",
        "business_rule": "有货-渠道线索",
        "corpus": "像看到 a2 到货渠道后顺手接一句。",
        "examples": ["我在山姆看到a2到货了"],
        "variation_slots": [
            {
                "slot_code": "comment_entry",
                "slot_name": "接法槽",
                "options": ["追问：先问一句", "报信：直接补一句信息"],
            }
        ],
        "source_row_no": 2,
    }
    asset = SimpleNamespace(asset_key="a2_sentiment_comment_activity", id=7, version_no=54)

    plan = service._plan_from_rule(rule, asset=asset, item_no=1)

    assert plan["variation_slots"] == [
        {
            "slot_code": "comment_entry",
            "slot_name": "接法槽",
            "options": ["追问：先问一句"],
        }
    ]
    assert plan["preselected_variation_slots"]["comment_entry"]["candidate_count"] == 2


def test_comment_plan_preserves_seed_expand_output_config_from_rule():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rule = {
        "rule_id": "business_rule_001",
        "business_rule": "有货-直给到货情绪",
        "corpus": "像妈妈看到 a2 到货后顺手接一句。",
        "examples": ["a2终于到货了"],
        "output_format_mode": "json_string_array",
        "expansion_count": 20,
        "source_row_no": 1,
    }
    asset = SimpleNamespace(
        asset_key="a2_sentiment_comment_activity",
        id=7,
        version_no=3,
        content_json={},
        metadata_json={},
    )

    plan = service._plan_from_rule(rule, asset=asset, item_no=1)

    assert plan["output_format"] == {"mode": "json_string_array", "count": 20}
    assert plan["output_format_mode"] == "json_string_array"
    assert plan["expansion_count"] == 20


def test_comment_plan_copies_generation_requirements_from_rule_asset():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    rule = {
        "rule_id": "business_rule_001",
        "business_rule": "剧情讨论",
        "corpus": "剧情规则",
        "examples": ["剧情示例"],
        "source_row_no": 1,
    }
    asset = SimpleNamespace(
        asset_key="a2_plot_discussion_comment",
        id=7,
        version_no=3,
        content_json={"generation_requirements": "只输出21到35字剧情讨论评论。"},
        metadata_json={},
    )

    plan = service._plan_from_rule(rule, asset=asset, item_no=1)

    assert plan["generation_requirements"] == "只输出21到35字剧情讨论评论。"


def test_comment_plan_copies_batch_variation_review_from_rule_asset():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    config = {
        "enabled": True,
        "expression_frequency": [
            {
                "group_key": "activity_status_question",
                "terms": ["活动还在不在", "活动还有没有"],
                "max_ratio": 0.2,
            }
        ],
    }
    asset = SimpleNamespace(
        asset_key="a2_plot_discussion_comment",
        id=7,
        version_no=3,
        content_json={"batch_variation_review": config},
        metadata_json={},
    )

    plan = service._plan_from_rule({"business_rule": "剧情讨论", "corpus": "剧情规则"}, asset=asset, item_no=1)

    assert plan["batch_variation_review"] == config


def test_comment_plan_prefers_explicit_batch_variation_review_override():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    asset_config = {"enabled": True, "expression_frequency": []}
    override = {
        "enabled": True,
        "expression_frequency": [
            {
                "group_key": "opener_just",
                "terms": ["刚"],
                "match_mode": "prefix",
                "max_ratio": 0.2,
            }
        ],
    }
    rule = service._rules_with_batch_variation_review_override(
        [{"business_rule": "有货-渠道", "corpus": "写自然反应。"}],
        override,
    )[0]
    asset = SimpleNamespace(
        asset_key="a2_sentiment_comment_activity",
        id=7,
        version_no=76,
        content_json={"batch_variation_review": asset_config},
        metadata_json={},
    )

    plan = service._plan_from_rule(rule, asset=asset, item_no=1)

    assert plan["batch_variation_review"] == override


def test_comment_plan_copies_delivery_selection_from_rule_asset():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    config = {
        "enabled": True,
        "target_count": 105,
        "max_similarity": 0.45,
    }
    asset = SimpleNamespace(
        asset_key="a2_sentiment_comment_activity",
        id=7,
        version_no=76,
        content_json={"delivery_selection": config},
        metadata_json={},
    )

    plan = service._plan_from_rule(
        {"business_rule": "有货-直给", "corpus": "写自然反应。"},
        asset=asset,
        item_no=1,
    )

    assert plan["delivery_selection"] == config


def test_comment_plan_prefers_explicit_delivery_selection_override():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    asset_config = {"enabled": True, "target_count": 100}
    override = {"enabled": True, "target_count": 105, "max_similarity": 0.45}
    rule = service._rules_with_delivery_selection_override(
        [{"business_rule": "有货-直给", "corpus": "写自然反应。"}],
        override,
    )[0]
    asset = SimpleNamespace(
        asset_key="a2_sentiment_comment_activity",
        id=7,
        version_no=76,
        content_json={"delivery_selection": asset_config},
        metadata_json={},
    )

    plan = service._plan_from_rule(rule, asset=asset, item_no=1)

    assert plan["delivery_selection"] == override


def test_comment_length_fallback_keeps_short_natural_clause():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)

    comment = service._fit_comment_length("从旧奶转源悦，我家娃皮肤敏感，先少量掺着喝。", max_chars=20)

    assert comment == "从旧奶转源悦，我家娃皮肤敏感"
    assert len(comment) <= 20


def test_comment_length_fallback_avoids_dangling_clause():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)

    comment = service._fit_comment_length("我家从转奶第三天开始，拉臭臭就没那么费劲了。", max_chars=20)

    assert comment != "我家从转奶第三天开始"
    assert not comment.endswith("开始")
    assert len(comment) <= 20


def test_comment_length_default_caps_at_thirty_five_chars():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)

    comment = service._fit_comment_length(
        "我们喝这款好几个月了，拉的时候没费过劲，前面也不见羊屎蛋，换纸尿裤时软软的挺舒坦"
    )

    assert len(comment) <= 35


def test_a2_comment_length_caps_at_eighty_chars():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    item = ContentBatchItem(plan_json={"quality_guard_profile_key": "a2_sentiment_comment_202606"})
    comment = "刚转奶先看物流码报告，爱他美样批也看，a2报告里蜡样那项再结合肚肚便便状态"

    fitted = service._fit_comment_length(comment, max_chars=service._comment_max_chars(item))

    assert fitted == comment
    assert len(fitted) > 35
    assert service._comment_max_chars(item) == 80

    overlong = comment * 3
    normalized = service._normalize_comment_length(item, overlong)
    assert len(normalized) <= 80

    low_information = service._low_information_rewrite_input(item)
    delivery_duplicate = service._delivery_duplicate_rewrite_input(item, {"body": "历史评论"})
    similarity = service._similarity_rewrite_input(
        item,
        {"item_no": 1, "body": "相似评论", "score": 0.8, "scope": "current_batch"},
    )
    assert "只输出一条80字以内的评论正文" in low_information["rewrite_instructions"]
    assert "只输出一条80字以内的评论正文" in delivery_duplicate["rewrite_instructions"]
    assert "只输出一条80字以内的评论正文" in similarity["rewrite_instructions"]


def test_comment_micro_reply_caps_generated_comment_at_ten_chars():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    item = ContentBatchItem(
        plan_json={
            "quality_guard_profile_key": "a2_negative_post_comment_202606",
            "keyword_selection": {"comment_format_control": ["comment_micro_reply"]},
        }
    )

    normalized = service._normalize_comment_length(item, "刚好能补上，我赶紧去问问门店。")

    assert normalized == "刚好能补上"
    assert len(normalized) <= 10
    assert service._comment_max_chars(item) == 10


def test_comment_micro_reply_caps_from_selected_keyword_snapshot():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    item = ContentBatchItem(
        plan_json={
            "unified_generation": {
                "selected_keywords": [
                    {
                        "category_code": "comment_format_control",
                        "keyword_code": "comment_micro_reply",
                    }
                ]
            }
        }
    )

    assert service._comment_max_chars(item) == 10


def test_comment_micro_reply_does_not_inject_business_fallback_for_short_output():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    item = ContentBatchItem(
        item_no=3,
        plan_json={
            "keyword_selection": {"comment_format_control": ["comment_micro_reply"]},
        },
    )

    assert service._normalize_comment_length(item, "嗯") == "嗯"


def test_comment_thread_short_reply_only_controls_length():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    item = ContentBatchItem(
        item_no=8,
        plan_json={
            "keyword_selection": {"comment_format_control": ["comment_thread_short_reply"]},
        },
    )

    normalized = service._normalize_comment_length(item, "问了几家店，先继续喝这个吧，转奶的事后面再说。")

    assert normalized == "问了几家店"
    assert len(normalized) <= 12


def test_a2_plot_discussion_length_is_not_trimmed_before_guard():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    item = ContentBatchItem(plan_json={"quality_guard_profile_key": "a2_plot_discussion_comment_202606"})
    comment = "娃刚看完山洞求援那段正演得起劲，我去店里续奶粉就打听有没有巴克队长款对讲机"

    normalized = service._normalize_comment_length(item, comment)

    assert normalized == comment
    assert len(normalized) > 35
    assert service._comment_max_chars(item) == 50


def test_a2_sentiment_post_length_is_not_trimmed_before_guard():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    item = ContentBatchItem(plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"})
    content = "今天路过常买母婴店，看到a2至初常喝段数已经上架，家里刚好快喝完，就先补了两罐。熟悉口粮能接上，心里踏实不少。"

    normalized = service._normalize_comment_length(item, content)

    assert normalized == content
    assert len(normalized) > 45
    assert service._comment_max_chars(item) == 100


def test_a2_sentiment_post_guard_rejects_incomplete_tail():
    item = ContentBatchItem(
        body="给娃囤奶粉的时候顺手翻了下报告，",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_incomplete_post" for issue in payload["issues"])


def test_a2_sentiment_post_guard_accepts_complete_shunshou_sentence():
    item = ContentBatchItem(
        body="逛超市看到a2至初又补货了，突然就觉得不想再看其他牌子。还是买回来一罐放着吧，还是熟悉的更顺手。",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_sentiment_post_guard_rejects_incomplete_data_sentence():
    item = ContentBatchItem(
        body="今天记录选奶粉的一点观察。a2至初有三方检测数据和多项检测信息列出来，虽然不会每条细",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_incomplete_post" for issue in payload["issues"])


def test_a2_sentiment_post_guard_rejects_incomplete_for_me_tail():
    item = ContentBatchItem(
        body="刚转a2至初，这次留意到能查到的数据挺全。除了三方检测数据，里面还有多项检测信息列出来。对我们这种爱较真的妈妈来说",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_incomplete_post" for issue in payload["issues"])


def test_a2_sentiment_post_guard_rejects_single_character_residual_tail():
    item = ContentBatchItem(
        body="刚看到线上店有a2了，等收到货我也扫一下罐底码，对上号就更安心了。其",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_incomplete_post" for issue in payload["issues"])


def test_a2_sentiment_post_guard_rejects_malformed_drinking_phrase():
    item = ContentBatchItem(
        body="熟悉的牌子愿意喝上，不用临时换",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" and "愿意喝上" in issue["evidence"] for issue in payload["issues"])


def test_a2_sentiment_post_guard_rejects_habitual_scan_wording():
    item = ContentBatchItem(
        body="刚补到a2至初，拿到手就习惯性扫了罐底码看报告",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" and "习惯性" in issue["evidence"] for issue in payload["issues"])


@pytest.mark.parametrize(
    "title",
    [
        "转奶期妈妈别焦虑，扫一下罐底就行",
        "别再纠结转奶了，扫罐底就安心",
        "报告看完就懂，转奶不用慌",
    ],
)
def test_a2_sentiment_post_guard_rejects_ai_advice_title(title):
    item = ContentBatchItem(
        title=title,
        body="今天才知道罐底码能扫出报告，我就顺手试了一下手上这罐。质检报告能看到，蜡样那项也有，心里踏实点。",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_title_ai_advice_tone" for issue in payload["issues"])


def test_a2_sentiment_post_guard_rejects_json_leak_output():
    item = ContentBatchItem(
        title="{",
        body='"title": "有姐妹一起转a2吗？刚补了一罐试试🧐",\n"body": "家里奶粉快见底了。"',
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_output_json_leak" for issue in payload["issues"])


def test_a2_sentiment_post_guard_rejects_transfer_decision_advice():
    item = ContentBatchItem(
        title="求姐妹帮看，我该开始转a2吗",
        body="家里奶粉快见底了，先补一罐a2至初试试，有姐妹帮我看看这样合适不？",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_transfer_decision_advice" for issue in payload["issues"])


def test_a2_sentiment_post_guard_rejects_report_safety_decision():
    item = ContentBatchItem(
        title="刚扫出来报告",
        body="罐底扫码看蜡样未检出，姐妹帮我看看是不是没问题？",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_report_safety_decision" for issue in payload["issues"])


def test_a2_sentiment_post_guard_accepts_report_regular_question():
    item = ContentBatchItem(
        title="扫出来报告了",
        body="罐底码扫出来报告了，但我看不太明白，有姐妹帮我看看这算正规报告吗？",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


@pytest.mark.parametrize(
    "body",
    [
        "质检报告跳出来一堆数字，有姐妹懂这个报告的吗？帮我看看是不是没问题呀？",
        "蜡样那项显示未检出，求帮看下是不是这样就行。",
    ],
)
def test_a2_sentiment_post_guard_rejects_loose_report_safety_decision(body):
    item = ContentBatchItem(
        title="刚扫出来报告",
        body=body,
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_report_safety_decision" for issue in payload["issues"])


def test_a2_sentiment_post_guard_rejects_feeding_decision_advice():
    item = ContentBatchItem(
        title="店员说能扫码看报告",
        body="报告我也看不太明白，宝妈们你们会给宝宝喝这款吗？我也在纠结中。",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_feeding_decision_advice" for issue in payload["issues"])


def test_a2_sentiment_post_guard_rejects_heavy_supply_negative_word():
    item = ContentBatchItem(
        title="a2终于补上了",
        body="之前断档了一阵，今天店员说a2至初到货了，先拿一罐回来。",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" and "断档" in issue["evidence"] for issue in payload["issues"])


def test_a2_sentiment_post_replaces_bad_zhichu_stock_typo():
    item = ContentBatchItem(
        title="补货扫到罐底报告了",
        body="今天去门店拿了两罐a2至补货，店员说现在罐底物流码能扫出批次报告。",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "a2至补货" not in item.body
    assert "a2至初" in item.body


def test_a2_sentiment_post_replacement_does_not_trim_article_body():
    body = "我们家宝贝最近准备换奶粉，做功课的时候留意到A2至初新批次，每罐都能查到检测报告和三方数据。感觉信息更透明了，让我这个谨慎型妈妈比较安心。没有复杂教程，就是觉得这点挺好，先补一罐回来试试看。"
    item = ContentBatchItem(
        title="看了一圈还是想换A2",
        body=body,
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert item.title == "看了一圈还是想换a2"
    assert item.body == body.replace("A2", "a2")
    assert item.body.endswith("先补一罐回来试试看。")


def test_a2_comment_replacement_still_trims_comment_body():
    body = "a2这次到货我会先看报告，蜡样检测和三方检测报告这些都能看到，评论区说法不用写得太像公告，简单说自己看完更有底就行。"
    item = ContentBatchItem(
        body=body,
        plan_json={"quality_guard_profile_key": "a2_sentiment_comment_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert item.body.startswith("a2这次到货")
    assert "三方质检报告" in item.body
    assert len(item.body) <= 60


def test_a2_sentiment_post_guard_allows_wax_toxin_wording():
    item = ContentBatchItem(
        body="等货到了我想扫罐底看看，质检和蜡毒那项都能查到就更放心",
        plan_json={"quality_guard_profile_key": "a2_sentiment_post_202606"},
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert item.body == "等货到了我想扫罐底看看，质检和蜡毒那项都能查到就更放心"
    assert not payload["issues"]


def test_comment_length_fallback_leaves_short_comment_unchanged():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)

    assert service._fit_comment_length("纸尿裤里不吓人") == "纸尿裤里不吓人"


def test_realness_rewrite_prefers_specific_can_continue_wording():
    text = "熟悉口粮能接上，心里踏实不少"
    hits = find_comment_realness_hits(text)

    rewritten = _remove_or_replace_realness_terms(text, hits, STATIC_COMMENT_REALNESS_REPLACEMENTS)

    assert "愿意喝上" not in rewritten
    assert "熟悉口粮接着喝" in rewritten


def test_low_information_comment_detection_only_catches_empty_openers():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)

    assert service._looks_low_information_comment("我们家")
    assert service._looks_low_information_comment("同款！")
    assert not service._looks_low_information_comment("我家不硬")
    assert not service._looks_low_information_comment("同款，便便软多了")


def test_yuanyue_brand_hallucination_guard_is_asset_scoped():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    item = ContentBatchItem(
        body="一直喝星飞帆，后来还看过a2、A2和爱他美。",
        plan_json={"asset_key": "yuanyue_comment_activity"},
        quality_json={},
    )

    service._sanitize_brand_hallucinations(item)

    assert "星飞帆" not in item.body
    assert "a2" not in item.body
    assert "A2" not in item.body
    assert "爱他美" not in item.body
    assert set(item.quality_json["brand_hallucination_guard"]["hits"]) == {"星飞帆", "爱他美", "a2", "A2"}

    other = ContentBatchItem(
        body="一直喝星飞帆，后来还看过a2、A2和爱他美。",
        plan_json={"asset_key": "other_activity"},
        quality_json={},
    )
    service._sanitize_brand_hallucinations(other)

    assert other.body == "一直喝星飞帆，后来还看过a2、A2和爱他美。"


def _a2_guard_plan(corpus: str) -> dict:
    return {
        "quality_guard_profile_key": "a2_sentiment_comment_202606",
        "business_rule": "A2舆情改善评论",
        "corpus": corpus,
        "unified_generation": {
            "selected_keywords": [
                {"category_code": "persona", "category_name": "人设", "keyword_name": "普通妈妈"},
                {"category_code": "comment_writing_instruction", "category_name": "生文指令", "keyword_name": "评论-短句"},
                {"category_code": "perturbation_rule", "category_name": "扰动规则", "keyword_name": "通用"},
                {"category_code": "comment_format_control", "category_name": "生文输出格式", "keyword_name": "生文输出格式-评论"},
            ]
        },
    }


def _a2_stock_direct_guard_plan(rule_id: str, business_rule: str) -> dict:
    plan = _a2_guard_plan("有货直给")
    plan["rule_id"] = rule_id
    plan["business_rule"] = business_rule
    return plan


@pytest.mark.parametrize(
    ("rule_id", "business_rule", "body"),
    [
        ("a2_direct_01", "有货-直给-提产品", "a2到货了，我先冲一波"),
        ("a2_direct_43", "有货-直给-不提产品", "终于补货了，我直接冲去下单"),
    ],
)
def test_a2_stock_direct_rules_allow_natural_purchase_wording(rule_id, business_rule, body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_stock_direct_guard_plan(rule_id, business_rule),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_brand_bad_stock_wording" for issue in payload["issues"])


@pytest.mark.parametrize(
    ("rule_id", "business_rule", "body"),
    [
        ("a2_direct_01", "有货-直给-提产品", "a2到货了，顺手拿了两袋"),
        ("a2_direct_43", "有货-直给-不提产品", "刚补货了，先冲一包"),
        ("a2_direct_01", "有货-直给-提产品", "a2来货了，我先拿三盒"),
        ("a2_direct_43", "有货-直给-不提产品", "门店到了，直接补一件"),
        ("a2_direct_01", "有货-直给-提产品", "a2能买到了，先搬一桶"),
        ("a2_direct_43", "有货-直给-不提产品", "刚发现能买了，先下单两瓶"),
    ],
)
def test_a2_stock_direct_rules_reject_invalid_formula_quantity_units(rule_id, business_rule, body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_stock_direct_guard_plan(rule_id, business_rule),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(
        issue["code"] == "activity_body_stock_direct_invalid_quantity_unit"
        for issue in payload["issues"]
    )


@pytest.mark.parametrize(
    ("rule_id", "business_rule", "body"),
    [
        ("a2_direct_01", "有货-直给-提产品", "a2到货了，我先拿两罐"),
        ("a2_direct_43", "有货-直给-不提产品", "刚看到有货，先带一箱"),
        ("a2_direct_01", "有货-直给-提产品", "a2来货了，家里放两箱子"),
    ],
)
def test_a2_stock_direct_rules_allow_valid_formula_quantity_units(rule_id, business_rule, body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_stock_direct_guard_plan(rule_id, business_rule),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert not any(
        issue["code"] == "activity_body_stock_direct_invalid_quantity_unit"
        for issue in payload["issues"]
    )


@pytest.mark.parametrize(
    ("rule_id", "business_rule", "body", "expected_code"),
    [
        (
            "a2_direct_01",
            "有货-直给-提产品",
            "终于补货了，先囤两罐",
            "activity_body_stock_direct_missing_product_name",
        ),
        (
            "a2_direct_43",
            "有货-直给-不提产品",
            "a2到货了，先囤两罐",
            "activity_body_stock_direct_unexpected_product_name",
        ),
        (
            "a2_direct_01",
            "有货-直给-提产品",
            "这边能调货，我先补两罐a2",
            "activity_body_brand_bad_stock_wording",
        ),
    ],
)
def test_a2_stock_direct_rules_keep_product_split_and_channel_risk_boundaries(
    rule_id,
    business_rule,
    body,
    expected_code,
):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_stock_direct_guard_plan(rule_id, business_rule),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == expected_code for issue in payload["issues"])


def _a2_plot_guard_plan(corpus: str = "剧情讨论业务规则") -> dict:
    return {
        "quality_guard_profile_key": "a2_plot_discussion_comment_202606",
        "business_rule": "业务规则-剧情讨论",
        "corpus": corpus,
        "unified_generation": {
            "selected_keywords": [
                {"category_code": "persona", "category_name": "人设", "keyword_name": "家庭妈妈"},
                {"category_code": "comment_writing_instruction", "category_name": "生文指令", "keyword_name": "评论-短句"},
                {"category_code": "perturbation_rule", "category_name": "扰动规则", "keyword_name": "通用"},
                {"category_code": "comment_format_control", "category_name": "生文输出格式", "keyword_name": "生文输出格式-评论"},
            ]
        },
    }


def test_a2_activity_guard_rejects_zhichu_a2_product_split_after_batch_generation():
    job = SimpleNamespace(strategy_json={"quality_guard_profile_key": "a2_sentiment_comment_202606"})
    item = ContentBatchItem(
        status="generated",
        body="至初这边有a2现货,我先给同款妈妈说一下。",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={"hard_pass": True},
    )

    ActivityQualityGuardService().review_batch(job, [item])

    guard = item.quality_json["activity_quality_guard"]
    assert guard["pass"] is False
    assert any(issue["code"] == "activity_body_bad_a2_zhichu_product_split" for issue in guard["issues"])
    assert item.quality_json["hard_pass"] is False
    assert item.quality_json["review_report"]["rewrite_required"] is True


def test_a2_activity_guard_accepts_zhichu_as_full_product_name():
    item = ContentBatchItem(
        body="a2至初这边有货，我先扫物流码看报告再补。",
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_bad_a2_zhichu_product_split" for issue in payload["issues"])


def test_a2_activity_guard_accepts_uppercase_a2_zhichu_product_name():
    item = ContentBatchItem(
        body="A2至初到货了，我先扫物流码看报告再补。",
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_bad_a2_zhichu_product_split" for issue in payload["issues"])


@pytest.mark.parametrize(
    ("body", "expected_code"),
    [
        ("至初这边有a2现货,我先给同款妈妈说一下。", "activity_body_bad_a2_zhichu_product_split"),
        ("官方通知a2全国补货了，大家放心买", "activity_body_bad_stock_claim"),
        ("转奶先半勺，三天后全换", "activity_body_transfer_tutorial"),
        ("喝a2后便便就好了", "activity_body_product_effect_claim"),
        ("检测报告证明这批绝对安全", "activity_body_report_overclaim"),
        ("本来被缺货逼到想转，a2能调货后又稳住", "activity_body_supply_pressure_transfer_memory"),
        ("缺货那阵子我真想换，现在a2有货又踏实了", "activity_body_supply_pressure_transfer_memory"),
        ("之前被a2货少逼得做功课，现在先不看了", "activity_body_supply_pressure_transfer_memory"),
        ("外面没货那阵子真想换奶，现在a2有货又踏实了", "activity_body_supply_pressure_transfer_memory"),
        ("家里快断顿了，a2能买到先接一罐", "activity_body_feeding_anxiety_stock"),
        ("奶粉接不上了，a2有货就先补一罐", "activity_body_feeding_anxiety_stock"),
        ("这波货源应该是在慢慢动起来", "activity_body_supply_trend_claim"),
        ("至初陆续有货的话，我真不想转来转去", "activity_body_supply_trend_claim"),
        ("店员让我别急，说a2会陆续到", "activity_body_supply_trend_claim"),
        ("至初有点回货，我决定再观望一下", "activity_body_supply_trend_claim"),
        ("我对放货还抱点希望，先不急着转", "activity_body_supply_trend_claim"),
        ("本来想换，看到有货立刻打消念头", "activity_body_brand_bad_transfer_framing"),
        ("转奶这几天已经够乱了，买到至初就回去", "activity_body_brand_bad_transfer_framing"),
        ("本来都说服自己换了，结果至初又有货", "activity_body_brand_bad_transfer_framing"),
        ("刚刚看到可下单，直接付款不犹豫", "activity_body_brand_bad_stock_wording"),
        ("看到a2有货就先拍下，后面再说后面", "activity_body_brand_bad_stock_wording"),
        ("我这边刚问到至初能拿现货了，赶紧来报个信", "activity_body_brand_bad_stock_wording"),
        ("看到a2有货可以冲了", "activity_body_brand_bad_stock_wording"),
        ("a2能下单就闭眼冲", "activity_body_brand_bad_stock_wording"),
        ("这边能调货，我先补货原来的", "activity_body_brand_bad_stock_wording"),
        ("问了下同城店可以调货，先补货拿回去接着喝", "activity_body_brand_bad_stock_wording"),
        ("母婴店说能帮我留这罐，就先不换啦", "activity_body_brand_bad_stock_wording"),
        ("能买到a2的话，是不是先继续原来的就行", "activity_body_ai_template_conditional_question"),
        ("如果a2能买到，是不是就先继续原来的", "activity_body_ai_template_conditional_question"),
        ("如果能补到a2，是不是就没必要急着换", "activity_body_ai_template_conditional_question"),
        ("能买到a2的话，是不是先拍一罐比较稳", "activity_body_ai_template_conditional_question"),
        ("囤这罐接着喝就安心了，先不急着换", "activity_body_brand_bad_stock_wording"),
        ("旧的那罐囤货喝完再说", "activity_body_brand_bad_stock_wording"),
        ("宝妈们可以多跑两家店看看，我这边也能订到了", "activity_body_brand_bad_stock_wording"),
        ("我这边a2还没消息，你们问到货了吗", "activity_body_brand_bad_stock_wording"),
        ("我这边还没看到有a2，大家有消息吗", "activity_body_brand_bad_stock_wording"),
        ("我家宝宝那罐还够喝两天，愿意喝着买a2就不换啦", "activity_forbidden_terms"),
        ("我们这儿a2能买到了吗，怕断档想先要这罐", "activity_forbidden_terms"),
    ],
)
def test_a2_activity_guard_rejects_high_confidence_logic_issues(body, expected_code):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == expected_code for issue in payload["issues"])
    assert item.quality_json["hard_pass"] is False
    assert item.quality_json["review_report"]["rewrite_required"] is True


@pytest.mark.parametrize(
    "body",
    [
        "a2至初这边有货，我先扫物流码看报告再补",
        "刚好补到a2，扫物流码看报告，慢慢转",
        "转奶前先看a2这罐报告，过渡慢慢来",
        "a2报告能看，我先观望下",
    ],
)
def test_a2_activity_guard_accepts_natural_logic_short_comments(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(
        issue["code"]
        in {
            "activity_body_bad_a2_zhichu_product_split",
            "activity_body_bad_stock_claim",
            "activity_body_transfer_tutorial",
            "activity_body_product_effect_claim",
            "activity_body_report_overclaim",
        }
        for issue in payload["issues"]
    )


@pytest.mark.parametrize(
    "body",
    [
        "刚好a2有货，先补一罐扫物流码看报告，慢慢转",
        "a2有货了，准备先补一罐，扫物流码看报告慢慢转",
        "家里没货了，刚好a2有货，先补一罐扫物流码看报告慢慢转",
        "手上那罐快空了，a2到货就先补一罐扫物流码看报告慢慢转",
        "家里那罐快见底了，刚好a2有货，先补一罐扫物流码看报告慢慢转",
        "大家可以问问附近母婴店，我刚买到a2了",
        "我这边刚问到a2能拿现货了",
        "我买到a2了，确实是有货了",
        "刚刚看到可下单，直接下单了",
        "看到a2能下单，我就先拍一罐",
        "能下单就先买原来的，换奶先放一边",
        "看到a2有货就先拍下，我最近线上线下都补到了一些",
        "能买到原来的a2，我就先继续喝",
        "不要随便转奶，对小朋友不好，货都有了继续喝吧",
        "刚联系上店里还有a2，我赶紧拿了这罐",
        "刚问到a2到货了，立马囤上两件接着喝",
        "a2有货了，我赶紧囤几罐接着喝",
        "今天问了几家店，说陆续有货了",
        "我这边导购也说快了，有底了也算是",
        "这两天看大家说的觉得还是能买到就先不换，我还是对放货抱点希望的",
        "那能买到a2就先不转了",
    ],
)
def test_a2_activity_guard_accepts_supply_positive_transfer_comments(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_supply_pressure_transfer_memory" for issue in payload["issues"])


def test_a2_activity_guard_keeps_natural_stockpile_quantity_wording():
    item = ContentBatchItem(
        body="a2有货了，我赶紧囤几罐接着喝",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert item.body == "a2有货了，我赶紧囤几罐接着喝"
    assert not any(repair.get("source") == "几罐" for repair in payload["repairs"])


def test_a2_activity_guard_keeps_take_one_can_wording_natural():
    item = ContentBatchItem(
        body="店里说有货，我先拿一罐a2接着喝",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert "补货拿" not in item.body
    assert not any(repair.get("replacement") == "补货拿" for repair in payload["repairs"])


@pytest.mark.asyncio
async def test_a2_comment_realness_allows_confirmed_sisters_rush_wording():
    item = ContentBatchItem(
        title="有货+转奶",
        body="a2终于到了！姐妹们冲",
        plan_json={
            "asset_key": "a2_sentiment_comment_activity",
            "quality_guard_profile_key": "a2_sentiment_comment_202606",
        },
        quality_json={"hard_pass": True},
    )

    payload = await CommentRealnessReviewService().review_and_rewrite_item(
        item=item,
        orchestrator=None,
        executor_code=None,
    )

    assert "姐妹们冲" not in payload["initial_hits"]
    assert item.body == "a2终于到了！姐妹们冲"


def _a2_negative_post_guard_plan(business_rule: str, corpus: str) -> dict:
    return {
        "quality_guard_profile_key": "a2_negative_post_comment_202606",
        "business_rule": business_rule,
        "corpus": corpus,
        "unified_generation": {
            "selected_keywords": [
                {"category_code": "persona", "category_name": "人设", "keyword_name": "老客妈妈"},
                {"category_code": "comment_writing_instruction", "category_name": "生文指令", "keyword_name": "评论-短句"},
                {"category_code": "perturbation_rule", "category_name": "扰动规则", "keyword_name": "通用"},
                {"category_code": "comment_format_control", "category_name": "生文输出格式", "keyword_name": "生文输出格式-评论"},
            ]
        },
    }


def test_a2_negative_post_guard_accepts_arrival_soft_comment_without_report_wording():
    item = ContentBatchItem(
        body="门店说a2新批次到了，家里快喝完的可以先问问",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，门店到货顺手提醒",
            "到货安抚，门店到货顺手提醒：\n像老客顺手提醒门店到货。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "到货安抚"
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_transfer_soft_comment_and_not_anxious_wording():
    item = ContentBatchItem(
        body="转奶不着急，我家一直喝a2，先试一罐慢慢来就行",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，老客说慢慢来",
            "转奶安抚，老客说慢慢来：\n像老客接一句自己的转奶节奏。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


def test_a2_negative_post_guard_rejects_direct_negative_wording():
    item = ContentBatchItem(
        body="最近断货焦虑，大家都很担心",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，门店到货顺手提醒",
            "到货安抚，门店到货顺手提醒：\n像老客顺手提醒门店到货。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])
    assert {"断货", "焦虑", "担心"}.issubset(set(payload["issues"][0]["evidence"]))


def test_a2_negative_post_guard_rejects_inappropriate_maternal_slang():
    item = ContentBatchItem(
        body="给娃囤屁粮的时候，看到a2的奶粉也补上架了",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，母婴日常口吻带到货消息",
            "到货安抚，母婴日常口吻带到货消息：\n像妈妈顺嘴带出a2到货消息。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])


def test_a2_negative_post_guard_rejects_unnatural_reminder_wording():
    item = ContentBatchItem(
        body="带娃路过母婴店，本来不买奶粉，看到a2上架还是记了一下",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，母婴日常口吻带到货消息",
            "到货安抚，母婴日常口吻带到货消息：\n像妈妈顺嘴带出a2到货消息。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])


def test_a2_negative_post_guard_rejects_stiff_cross_scene_wording():
    item = ContentBatchItem(
        body="现在选奶粉跟投资一样要看硬数据，a2有货又能看到检测报告，我准备先转过去试试",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，跨场景判断逻辑带a2",
            "转奶安抚，跨场景判断逻辑带a2：\n像妈妈把日常判断逻辑带到选奶粉里。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])


def test_a2_negative_post_guard_rejects_note_like_daily_purchase_wording():
    item = ContentBatchItem(
        body="今天买宝宝沐浴露发现a2也在货架上了，刚好快喝完我就顺手带了一罐。",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，母婴日常口吻带到货消息",
            "到货安抚，母婴日常口吻带到货消息：\n像妈妈顺嘴带出a2到货消息。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_note_like_daily_purchase" for issue in payload["issues"])


def test_a2_negative_post_guard_rejects_bad_batch_check_attribution():
    item = ContentBatchItem(
        body="我们这边刚补货，带批批检的姐妹可以冲了",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，姐妹互相传递有货信息",
            "到货安抚，姐妹互相传递有货信息：\n像评论区姐妹互相报个到货消息。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_bad_batch_check_attribution" for issue in payload["issues"])


def test_a2_negative_post_guard_replaces_uppercase_a2_in_body():
    item = ContentBatchItem(
        body="我这边店员说A2能订了，姐妹们可以问问自己常去的店",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，姐妹互相传递有货信息",
            "到货安抚，姐妹互相传递有货信息：\n像评论区姐妹互相报个到货消息。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "A2" not in item.body
    assert "a2" in item.body


def test_a2_negative_post_guard_rejects_detection_sheet_wording():
    item = ContentBatchItem(
        body="我们这儿店员刚说a2能拿货了，有检测单了",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，姐妹互相传递有货信息",
            "到货安抚，姐妹互相传递有货信息：\n像评论区姐妹互相报个到货消息。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" and "检测单" in issue["evidence"] for issue in payload["issues"])


def test_a2_negative_post_guard_replaces_third_party_report_wording():
    item = ContentBatchItem(
        body="a2有货了，三方检测报告能看到，我准备转奶先试试",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，看到有货和检测优势后想试a2",
            "转奶安抚，看到有货和检测优势后想试a2：\n像妈妈看到有货和检测优势后想试a2。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "三方检测报告" not in item.body
    assert "三方检测数据" in item.body


def test_a2_negative_post_guard_replaces_awkward_try_one_can_wording():
    item = ContentBatchItem(
        body="刚好a2能买到了，准备先转一罐试试，晚点再屯一罐，也想换一罐试一下",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，跨场景判断逻辑带a2",
            "有货后看检测信息再试a2：\na2有货后准备先试一罐。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "转一罐" not in item.body
    assert "屯一罐" not in item.body
    assert "换一罐" not in item.body
    assert "先试一罐" in item.body
    assert "囤一罐" in item.body


def test_a2_negative_post_guard_does_not_reject_context_only_cross_scene_terms():
    item = ContentBatchItem(
        body="宝宝入口的东西我习惯多看一眼，a2有检测报告就踏实些",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，跨场景判断逻辑带a2",
            "转奶安抚，跨场景判断逻辑带a2：\n不要直接写投资、职场、租房或上学。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert not any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])


def test_a2_negative_post_guard_allows_soft_do_not_hurry_transfer_wording():
    item = ContentBatchItem(
        body="转奶别太着急，按自己的节奏慢慢来就行",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，老客说慢慢来",
            "转奶安抚，老客说慢慢来：\n像老客接一句自己的转奶节奏。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_negative_post_guard_allows_cannot_hurry_transfer_wording():
    item = ContentBatchItem(
        body="转奶这事不能着急，慢慢来更稳一点",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，老客说慢慢来",
            "转奶安抚，老客说慢慢来：\n像老客接一句自己的转奶节奏。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_slow_switch_transfer_wording():
    item = ContentBatchItem(
        body="真的别急，慢慢换就好，我之前就是一点一点来的",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，老客说慢慢来",
            "转奶安抚，老客说慢慢来：\n像老客接一句自己的转奶节奏。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_transfer_back_to_a2_wording():
    item = ContentBatchItem(
        body="看到a2补货了，之前换过一阵的我又想转回来试试",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，看到到货想转回a2",
            "转奶安抚，看到到货想转回a2：\n像老客看到到货后想转回a2。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


def test_a2_negative_post_guard_rejects_hesitant_transfer_back_wording():
    item = ContentBatchItem(
        body="刚换了奶粉没几天就看到a2有货，犹豫要不要再转回来试试",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，看到到货想转回a2",
            "转奶安抚，看到到货想转回a2：\n像老客看到到货后想转回a2。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])


def test_a2_negative_post_guard_rejects_urgent_restock_wording():
    item = ContentBatchItem(
        body="a2补到货了，急着续的姐妹可以去问问",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，门店到货顺手提醒",
            "到货安抚，门店到货顺手提醒：\n像老客顺手提醒门店到货。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])


def test_a2_negative_post_guard_rejects_malformed_drinking_phrase():
    item = ContentBatchItem(
        body="看到a2愿意喝着囤就先不换牌子了",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，有货续上不转别的",
            "有货续上，不转别的：\na2有货后能续上，就不急着转别的。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" and "愿意喝着" in issue["evidence"] for issue in payload["issues"])


def test_a2_negative_post_guard_accepts_arrived_wording_as_arrival_marker():
    item = ContentBatchItem(
        body="我家楼下店也到了，可以先去补点",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，门店到货顺手提醒",
            "到货安抚，门店到货顺手提醒：\n像老客顺手提醒门店到货。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_can_restock_wording_as_arrival_marker():
    item = ContentBatchItem(
        body="老客了，a2能补到就接着喝，图个省心。",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，老客续喝稳住节奏",
            "到货安抚，老客续喝稳住节奏：\na2一直喝或已经喝习惯的场景，能补到就继续续上。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "到货安抚"
    assert not payload["issues"]


def test_a2_negative_post_guard_allows_duanliang_colloquial_wording():
    item = ContentBatchItem(
        body="我家这边门店刚来新批次，快去问问别断粮",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，门店到货顺手提醒",
            "到货安抚，门店到货顺手提醒：\n像老客顺手提醒门店到货。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_just_arrived_wording_as_arrival_marker():
    item = ContentBatchItem(
        body="尿不湿囤货时顺便问了下a2，说这批刚到",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，母婴日常口吻带到货消息",
            "到货安抚，母婴日常口吻带到货消息：\n像妈妈顺嘴带出a2到货消息。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_shelf_wording_as_arrival_marker():
    item = ContentBatchItem(
        body="给娃囤奶瓶清洁剂的时候看到a2新货已经上架了",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，母婴日常口吻带到货消息",
            "到货安抚，母婴日常口吻带到货消息：\n像妈妈顺嘴带出a2到货消息。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_display_shelf_wording_as_arrival_marker():
    item = ContentBatchItem(
        body="今天货架上a2摆出来了，家里快喝完的可以去看看",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，母婴日常口吻带到货消息",
            "到货安抚，母婴日常口吻带到货消息：\n像妈妈顺嘴带出a2到货消息。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_negative_post_guard_rejects_transfer_spoon_tutorial_wording():
    item = ContentBatchItem(
        body="转奶先加半勺，后面再慢慢来",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，换奶不写教程只写个人节奏",
            "转奶安抚，换奶不写教程只写个人节奏：\n像妈妈聊自己的节奏，不写教程。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])
    assert "半勺" in payload["issues"][0]["evidence"]


def test_a2_negative_post_guard_rejects_arrival_without_arrival_scene():
    item = ContentBatchItem(
        body="我家一直喝a2，心里稳一点",
        plan_json=_a2_negative_post_guard_plan(
            "到货安抚，老客续喝稳住节奏",
            "到货安抚，老客续喝稳住节奏：\n像老客在负面帖下稳一下节奏。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_missing_keyword_marker" for issue in payload["issues"])


def test_a2_negative_post_guard_rejects_transfer_without_transfer_scene():
    item = ContentBatchItem(
        body="门店说a2到货了，家里快喝完可以补上",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，老客说慢慢来",
            "转奶安抚，老客说慢慢来：\n像老客接一句自己的转奶节奏。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_missing_keyword_marker" for issue in payload["issues"])


def test_a2_negative_post_guard_accepts_report_then_continue_transfer_wording():
    item = ContentBatchItem(
        body="看到检测报告对上了，我还是打算续罐吧",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，老客说慢慢来",
            "破除转奶焦虑，稳老客拉新客：\n看报告后继续相信a2。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_batch_info_before_following_transfer_mood():
    item = ContentBatchItem(
        body="先别急着跟风转，看看自己家这罐的批次信息再说",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，老客说慢慢来",
            "破除转奶焦虑，稳老客拉新客：\n先看清原因和批次信息。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_report_search_transfer_wording():
    item = ContentBatchItem(
        body="我家也喝a2，看到报告能查到批次信息就挺安心",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，老客说慢慢来",
            "破除转奶焦虑，稳老客拉新客：\n看报告后继续相信a2。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_switching_other_brand_with_batch_report_wording():
    item = ContentBatchItem(
        body="要换别的牌子的话，建议先看自己手上这罐的批次报告",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，老客说慢慢来",
            "破除转奶焦虑，稳老客拉新客：\n先看清原因和批次信息。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_do_not_follow_switching_wording():
    item = ContentBatchItem(
        body="先别急着跟风换，看看自家娃适应情况再说",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，别被评论区带乱",
            "看清评论区信息再转奶：\n先看清原因再说，别跟着几条评论马上换。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


@pytest.mark.parametrize(
    "body",
    [
        "别跟着几条评论就别着急转奶，先稳住自己节奏",
        "宝宝喝得好好的就不用急着换，看别人说啥先别慌",
        "先看清楚宝宝的情况再决定吧，别被带偏了",
    ],
)
def test_a2_negative_post_guard_accepts_stabilize_before_switching_wording(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，别被评论区带乱",
            "看清评论区信息再转奶：\n别听他们带节奏，一直喝着没问题就不用转。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


@pytest.mark.parametrize(
    "body",
    [
        "我家是准备先搞一罐试试，急不来",
        "我家先拿一罐试试，看娃喝得怎么样",
    ],
)
def test_a2_negative_post_guard_accepts_one_can_trial_wording(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，换奶不写教程只写个人节奏",
            "转奶小步过渡：\n转奶或换奶时先小步试，不一下子全换。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_buy_one_can_for_trial_wording():
    item = ContentBatchItem(
        body="看到a2又能买了，每批检测结果也能查到，打算先买一罐让宝宝试试看",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，从观望到准备试",
            "看到有货和批批检后先试：\na2有货后准备先试一罐。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_long_term_stable_quality_wording():
    item = ContentBatchItem(
        body="喝久了a2就不太想换，质量一直挺稳的",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，老客说慢慢来",
            "破除转奶焦虑，稳老客拉新客：\n长期喝后不想换。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


@pytest.mark.parametrize(
    "body",
    [
        "我家一直喝这个，看到能补货就没急着换别的",
        "本来都准备试下别的牌子，一看能买到a2就又囤了几罐",
        "还好a2又能续上，接着喝放心",
    ],
)
def test_a2_negative_post_guard_accepts_restock_then_continue_wording(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，有货续上不转别的",
            "有货续上，不转别的：\na2有货后能续上，就不急着转别的。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


@pytest.mark.parametrize(
    "body",
    [
        "刚看到楼下店说到货了，那还是先喝至初吧",
        "看到店里到货了至初，先不换，省得宝宝又要适应",
    ],
)
def test_a2_negative_post_guard_accepts_stock_then_no_switch_short_wording(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，有货就先不转",
            "有货就先不转：\n看到附近有货后，先继续至初或先不换。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


def test_a2_negative_post_guard_accepts_stock_then_pulled_back_wording():
    item = ContentBatchItem(
        body="刚转了一周，看到有货又拉回来了",
        plan_json=_a2_negative_post_guard_plan(
            "转奶安抚，有货就先不转",
            "有货就先不转：\n看到有货后，从想换或已转的状态又被拉回来。",
        ),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "转奶安抚"
    assert not payload["issues"]


def test_a2_plot_discussion_guard_accepts_plot_with_store_activity():
    item = ContentBatchItem(
        body="第3集奶宝找到妈妈了，我下班顺路去门店把奶粉续上",
        plan_json=_a2_plot_guard_plan(),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "剧情讨论+门店引流"
    assert not payload["issues"]


@pytest.mark.parametrize(
    "body",
    [
        "找到了找到了，我家看到小奶宝扑进妈妈怀里才肯好好喝奶，奶粉罐子也快空了，正好明天路过母婴店去续上。",
        "后面就团聚了，我娃看得眼泪汪汪才肯睡，明天路过母婴店把奶粉续上",
    ],
)
def test_a2_plot_discussion_guard_accepts_maternity_store_and_xushang(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_plot_guard_plan(),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_plot_discussion_guard_flags_pure_plot_and_repairs_a2_cow_wording():
    item = ContentBatchItem(
        body="娃也认出奶宝是稀有A2奶牛了",
        plan_json=_a2_plot_guard_plan(),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert item.body == "娃也认出奶宝是稀有A2型奶牛了"
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_length_out_of_range" for issue in payload["issues"])
    assert any(issue["code"] == "activity_body_missing_store_activity" for issue in payload["issues"])


def test_a2_plot_discussion_guard_repairs_loose_a2_and_cave_wording():
    item = ContentBatchItem(
        body="娃说奶宝是稀有A2牛，山洞救援那段还想回看，它是稀有A2型的，我去门店续奶粉",
        plan_json=_a2_plot_guard_plan(),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert item.body == "娃说奶宝是稀有A2型奶牛，山洞求援那段还想回看，它是稀有A2型奶牛，我去门店续奶粉"
    assert payload["pass"] is True
    assert not payload["issues"]


@pytest.mark.parametrize(
    ("body", "expected_code"),
    [
        ("娃看完山洞求援还念叨奶宝，补货时问对讲机还有没有，愿意喝着演剧情", "activity_body_odd_drinking_play_phrase"),
        ("周末陪娃看这集热气球救小奶宝，奶粉快没了顺路去店里看看", "activity_body_unsupported_plot_detail"),
        ("顺路补奶粉带回对讲机，娃回来就开始演雪地救援那集", "activity_body_unsupported_plot_detail"),
        ("娃看完一直喊奶宝是A2牛牛，我准备去门店补奶粉顺便问活动", "activity_body_invalid_plot_wording"),
        ("娃说所有奶粉都叫A2型，周末去母婴店补货时我问问活动", "activity_body_invalid_plot_wording"),
        ("宝宝走散那段看得我家着急，明天去门店续奶粉顺便问活动", "activity_body_invalid_plot_wording"),
        ("宝宝找到妈妈啦，我明天去店里补奶粉顺便问活动", "activity_body_invalid_plot_wording"),
        ("我娃看完总念叨艾尔博士打A1怪兽，我续奶粉顺便问店员活动", "activity_body_invalid_plot_wording"),
        ("我娃说要看艾尔博士打A1大怪兽，我续奶粉顺便问店员活动", "activity_body_invalid_plot_wording"),
        ("娃最近刚看完第6集山洞求援，正好去店里补奶粉", "activity_body_unsupported_plot_detail"),
        ("娃最近刚看完第六集山洞求援，正好去店里补奶粉", "activity_body_unsupported_plot_detail"),
        ("最新一集看完娃还在念叨奶宝，我下班去门店补奶粉", "activity_body_unsupported_plot_detail"),
        ("奶宝救援队那段娃很爱看，我准备去母婴店续奶粉", "activity_body_unsupported_plot_detail"),
        ("奶宝团聚后娃喊洞穴救援开始，我到店把奶粉续上", "activity_body_unsupported_plot_detail"),
        ("山洞求援看完娃还在演，我去超市买奶粉时顺便问活动", "activity_body_activity_misstatement"),
        ("奶宝团聚那段娃放心了，我补两罐顺便领对讲机", "activity_body_activity_misstatement"),
        ("奶宝团聚后我去店里问有没有小礼物", "activity_body_activity_misstatement"),
        ("奶宝团聚后娃还在念叨，我补货前顺便问问盲盒活动", "activity_body_activity_misstatement"),
        ("奶宝团聚后娃拉着汪汪队喊支援，我到店把奶粉续上", "activity_body_invalid_plot_wording"),
        ("我娃看完山洞求援说想去店里找同款对讲机，我准备补奶粉", "activity_child_knows_store_gift"),
        ("我娃看完救援非要组队，奶粉快见底了正好群里说店里搞活动", "activity_body_missing_plot_anchor"),
    ],
)
def test_a2_plot_discussion_guard_rejects_odd_play_and_unsupported_plot(body, expected_code):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_plot_guard_plan(),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == expected_code for issue in payload["issues"])


def test_a2_plot_discussion_guard_allows_store_phrase_for_batch_frequency_review():
    item = ContentBatchItem(
        body="奶牛群转移那段娃看紧张了，我去门店看看活动还在不在",
        plan_json=_a2_plot_guard_plan(),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_bad_store_phrase" for issue in payload["issues"])


def test_a2_plot_discussion_guard_rejects_over_fifty_chars_without_trimming():
    body = "娃刚看完山洞求援那段正演得起劲，我去店里续奶粉就打听有没有巴克队长款对讲机，顺便看看活动和补货安排详情"
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_plot_guard_plan(),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert item.body == body
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_length_out_of_range" for issue in payload["issues"])


def test_comment_batch_variation_review_flags_only_overflow_items():
    config = {
        "enabled": True,
        "expression_frequency": [
            {
                "group_key": "activity_status_question",
                "label": "活动状态提问",
                "terms": ["活动还在不在"],
                "max_ratio": 0.4,
            }
        ],
    }
    items = [
        ContentBatchItem(item_no=1, status="generated", body="奶宝团聚了，我去问活动还在不在", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=2, status="generated", body="山洞求援看完，顺路问活动还在不在", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=3, status="generated", body="艾尔博士讲完，我再问活动还在不在", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=4, status="generated", body="巴克队长同款对讲机，我补货时看看", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=5, status="generated", body="第三集奶宝找妈妈，我到店续奶粉", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
    ]

    result = CommentBatchVariationReviewService().review_batch(items)

    assert result is not None
    assert result["pass"] is False
    assert items[0].quality_json["hard_pass"] is True
    assert items[1].quality_json["hard_pass"] is True
    assert items[2].quality_json["hard_pass"] is False
    assert items[2].quality_json["batch_variation_review"]["expression_frequency"]["issues"][0]["code"] == "batch_expression_frequency_cap_exceeded"
    assert items[4].quality_json["batch_variation_review"]["expression_frequency"]["metrics"][0]["max_allowed_count"] == 2


def test_comment_batch_variation_review_can_warn_without_changing_business_hard_pass():
    config = {
        "enabled": True,
        "affects_hard_pass": False,
        "expression_frequency": [
            {
                "group_key": "opener_just",
                "label": "刚字开头",
                "terms": ["刚"],
                "match_mode": "prefix",
                "max_ratio": 0.5,
            }
        ],
    }
    items = [
        ContentBatchItem(item_no=1, status="generated", body="刚看到补货了", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=2, status="generated", body="刚发现常买店有货", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=3, status="generated", body="刚刷到群里消息", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=4, status="generated", body="楼下店里来货了", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
    ]

    result = CommentBatchVariationReviewService().review_batch(items)

    assert result is not None
    assert result["pass"] is False
    assert result["affects_hard_pass"] is False
    assert items[2].quality_json["hard_pass"] is True
    assert items[2].quality_json["review_report"].get("rewrite_required") is not True
    advisory = items[2].quality_json["review_report"]["advisory_results"]
    assert advisory[0]["ae_code"] == "batch_variation.batch_expression_frequency_cap_exceeded"
    assert advisory[0]["affects_hard_pass"] is False


def test_comment_batch_delivery_selection_keeps_business_pass_separate_from_delivery_choice():
    config = {
        "enabled": True,
        "target_count": 3,
        "max_similarity": 0.7,
        "opening_first_char_frequency": {"max_count": 1},
    }
    items = [
        ContentBatchItem(item_no=1, status="generated", body="刚看到补货了", plan_json={"delivery_selection": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=2, status="generated", body="刚发现常买店有货", plan_json={"delivery_selection": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=3, status="generated", body="楼下店里来货了", plan_json={"delivery_selection": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=4, status="generated", body="群里说现在能买到了", plan_json={"delivery_selection": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=5, status="generated", body="页面显示已经到货", plan_json={"delivery_selection": config}, quality_json={"hard_pass": True}),
    ]

    result = CommentBatchDeliverySelectionService().select_batch(items)

    assert result is not None
    assert result["selected_count"] == 3
    assert result["shortfall_count"] == 0
    selected = [item for item in items if item.quality_json["delivery_selection"]["selected"]]
    assert len(selected) == 3
    assert sum(item.body.startswith("刚") for item in selected) == 1
    assert all(item.quality_json["hard_pass"] is True for item in items)


def test_comment_batch_delivery_selection_reports_bulk_refill_without_selecting_hard_failures_or_duplicates():
    config = {
        "enabled": True,
        "target_count": 4,
        "max_similarity": 0.7,
        "min_bulk_refill_count": 30,
        "bulk_refill_multiplier": 3,
    }
    items = [
        ContentBatchItem(item_no=1, status="generated", body="到货了，可以买了", plan_json={"delivery_selection": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=2, status="generated", body="到货了，可以买了", plan_json={"delivery_selection": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=3, status="generated", body="楼下店里来货了", plan_json={"delivery_selection": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=4, status="generated", body="正文业务错误", plan_json={"delivery_selection": config}, quality_json={"hard_pass": False}),
    ]

    result = CommentBatchDeliverySelectionService().select_batch(items)

    assert result is not None
    assert result["eligible_count"] == 2
    assert result["exact_duplicate_count"] == 1
    assert result["business_ineligible_count"] == 1
    assert result["selected_count"] == 2
    assert result["shortfall_count"] == 2
    assert result["suggested_bulk_refill_count"] == 30
    assert items[1].quality_json["delivery_selection"]["non_selection_reason"] == "exact_duplicate"
    assert items[3].quality_json["delivery_selection"]["non_selection_reason"] == "business_hard_pass_required"


def test_comment_batch_report_summary_exposes_delivery_selection_and_variation_counts():
    delivery_summary = {
        "eligible_count": 4,
        "selected_count": 3,
        "shortfall_count": 2,
        "suggested_bulk_refill_count": 30,
    }
    items = [
        ContentBatchReportItem(
            item_id=1,
            item_no=1,
            status="generated",
            body="到货了",
            body_chars=4,
            hard_pass=True,
            batch_variation_pass=False,
            delivery_selected=True,
            quality={"delivery_selection": delivery_summary},
        ),
        ContentBatchReportItem(
            item_id=2,
            item_no=2,
            status="generated",
            body="现在有货了",
            body_chars=6,
            hard_pass=True,
            batch_variation_pass=True,
            delivery_selected=True,
            quality={"delivery_selection": delivery_summary},
        ),
    ]

    summary = ContentBatchReportService.__new__(ContentBatchReportService)._summary(items)

    assert summary.batch_variation_warning_count == 1
    assert summary.delivery_candidate_count == 4
    assert summary.delivery_selected_count == 3
    assert summary.delivery_shortfall_count == 2
    assert summary.suggested_bulk_refill_count == 30


def test_comment_batch_variation_review_prefix_match_ignores_term_in_middle():
    config = {
        "enabled": True,
        "expression_frequency": [
            {
                "group_key": "opener_just",
                "label": "刚字开头",
                "terms": ["刚"],
                "match_mode": "prefix",
                "max_ratio": 0.4,
            }
        ],
    }
    items = [
        ContentBatchItem(item_no=1, status="generated", body="刚看到补货了", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=2, status="generated", body="今天刚看到补货了", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=3, status="generated", body="刚发现常买店有货", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=4, status="generated", body="刚刷到群里消息", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=5, status="generated", body="楼下店里来货了", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
    ]

    result = CommentBatchVariationReviewService().review_batch(items)

    assert result is not None
    metric = result["expression_frequency"]["metrics"][0]
    assert metric["match_mode"] == "prefix"
    assert metric["hit_item_nos"] == [1, 3, 4]
    assert metric["overflow_item_nos"] == [4]
    assert items[1].quality_json["hard_pass"] is True
    assert items[3].quality_json["hard_pass"] is False


def test_comment_batch_variation_review_flags_repeated_opening_prefix_and_clause():
    config = {
        "enabled": True,
        "opening_prefix_frequency": {"prefix_chars": 3, "max_count": 2},
        "opening_clause_frequency": {"max_count": 2},
    }
    items = [
        ContentBatchItem(item_no=1, status="generated", body="刚看到补货了，我去看看", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=2, status="generated", body="刚看到补货了，先问问", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=3, status="generated", body="刚看到补货了，顺路看看", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
        ContentBatchItem(item_no=4, status="generated", body="楼下店里来货了", plan_json={"batch_variation_review": config}, quality_json={"hard_pass": True}),
    ]

    result = CommentBatchVariationReviewService().review_batch(items)

    assert result is not None
    assert result["pass"] is False
    assert len(result["opening_frequency"]["metrics"]) == 2
    opening_issues = items[2].quality_json["batch_variation_review"]["opening_frequency"]["issues"]
    assert len(opening_issues) == 2
    assert all(issue["code"] == "batch_opening_frequency_cap_exceeded" for issue in opening_issues)
    assert items[2].quality_json["hard_pass"] is False


def test_a2_activity_guard_repairs_marker_and_entry_terms():
    item = ContentBatchItem(
        body="爱他美0.03有货，今天扫罐底批次物流码那个物流码的码查报告，截图保存了，新的一罐先看报告，纸尿裤和擦屁屁总有点红先不聊，我没慌",
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert payload["context_list"]["关键词"] == "有货+批批检"
    assert "爱他美" not in item.body
    assert "其他品牌" in item.body
    assert "物流码" not in item.body
    assert "罐底码" in item.body
    assert "物流码物流码" not in item.body
    assert "物流码那个物流码" not in item.body
    assert "记录" in item.body
    assert "一罐" not in item.body
    assert "纸尿裤" not in item.body
    assert "擦屁屁" not in item.body
    assert "慌" not in item.body
    assert "担心" not in item.body
    assert "补货" in item.body or "有货" in item.body or "到货" in item.body
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])
    assert payload["repairs"]


def test_a2_activity_guard_repairs_60_plus_and_detection_project_wording():
    item = ContentBatchItem(
        body="有货了先补a2，60+检测项目能看到，报告出来就放心了",
        plan_json=_a2_guard_plan("多项质检信息：\n关键词方向是有货+批批检，像妈妈看到报告信息后补货。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert "60+" in item.body
    assert "检测项目" not in item.body
    assert any(issue["code"] == "activity_forbidden_terms" and "60+" in issue["evidence"] for issue in payload["issues"])
    assert any(repair["source"] == "检测项目" for repair in payload["repairs"])


def test_a2_activity_guard_still_rejects_professional_indicator_wording():
    item = ContentBatchItem(
        body="有货了先补a2，专业指标能看到，报告出来就放心了",
        plan_json=_a2_guard_plan("多项质检信息：\n关键词方向是有货+批批检，像妈妈看到报告信息后补货。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])


def test_a2_activity_guard_rejects_batch_passed_inspection_overclaim():
    item = ContentBatchItem(
        body="刷到补货通知，看到a2每批都过检，这我就敢下手了",
        plan_json=_a2_guard_plan("不是偶尔抽查：\n关键词方向是有货+批批检，像妈妈看到报告后补货。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_report_overclaim" for issue in payload["issues"])


def test_a2_activity_guard_accepts_real_transfer_back_wording():
    item = ContentBatchItem(
        body="每批次都检测这波操作打动我了，刚好转奶不顺，我还是转回来吧！",
        plan_json=_a2_guard_plan("从皇美转回a2至初：\n关键词方向是批批检+转奶，像妈妈看报告后个人选择。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_repairs_bad_kan_shuo_wording():
    item = ContentBatchItem(
        body="同款刚补货，看说新批次有检测报告就安心，慢慢来",
        plan_json=_a2_guard_plan("补货后慢慢转：\n关键词方向是有货+转奶，像妈妈补货后慢慢转。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "看说" not in item.body
    assert "看到新批次有检测报告" in item.body
    assert not payload["issues"]
    assert any(repair["source"] == "看说" and repair["replacement"] == "看到" for repair in payload["repairs"])


def test_a2_activity_guard_repairs_bad_a2_code_wording():
    item = ContentBatchItem(
        body="这罐a2码一查报告信息都在，比导购说得具体",
        plan_json=_a2_guard_plan("批次质检信息看得见：\n关键词方向是有货+批批检，像妈妈看到报告信息后补货。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "a2码" not in item.body
    assert "这罐a2一查报告信息都在" in item.body
    assert not payload["issues"]
    assert any(repair["source"] == "a2码" and repair["replacement"] == "a2" for repair in payload["repairs"])


def test_a2_activity_guard_repairs_incomplete_scan_bottom_wording():
    item = ContentBatchItem(
        body="我家宝快喝完了，刚好看到a2新到货，扫完罐底才安心带回家",
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检，像妈妈补货前看这罐报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert item.body == "我家宝快喝完了，刚好看到a2新到货，扫完罐底才安心带回家"
    assert not payload["issues"]


def test_a2_activity_guard_repairs_bad_waxy_report_detail_but_blocks_003():
    item = ContentBatchItem(
        body="刚转a2，扫批次报告看到蜡样报告细节是小于0.03，心里有底。",
        plan_json=_a2_guard_plan("转奶前看蜡样检测：\n关键词方向是批批检+转奶，像妈妈转奶前看这罐蜡样检测报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert "蜡样报告细节" not in item.body
    assert "蜡样检测标准是小于0.03" in item.body
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])
    assert any(repair["source"] == "蜡样报告细节" for repair in payload["repairs"])


def test_a2_activity_guard_accepts_batch_quality_data_wording():
    item = ContentBatchItem(
        body="我们家转奶时也对比过雀巢，a2这罐能直接扫出这批的质检数据，感觉更透明些。",
        plan_json=_a2_guard_plan("对比后看新西兰三方检测信息：\n关键词方向是批批检+转奶。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])


def test_a2_activity_guard_accepts_contextual_this_batch_detection_short_comment():
    item = ContentBatchItem(
        body="有货了！这批也有检测，买得安心",
        plan_json=_a2_guard_plan("不是偶尔抽查看着更稳：\n关键词方向是有货+批批检，像妈妈补货时聊每批检测。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_accepts_contextual_this_can_tested_short_comment():
    item = ContentBatchItem(
        body="这次补a2看到这罐都测过，比抽查更放心些，毕竟宝宝天天喝呢。",
        plan_json=_a2_guard_plan("不是偶尔抽查看着更稳：\n关键词方向是有货+批批检，像妈妈补货时聊每批检测。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_accepts_stock_only_comment_without_batch_report():
    item = ContentBatchItem(
        body="刚看到a2能买了，我去瞅瞅。",
        plan_json=_a2_guard_plan("有货直给到货情绪：\n像评论区里刷到 a2 到货、能拍、能买之后的短句接话。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "有货"
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_stock_rule_is_not_misrouted_by_negative_member_benefit_boundary():
    item = ContentBatchItem(
        title="有货-直给到货情绪",
        body="a2到货了，我也终于买到了！",
        plan_json={
            "quality_guard_profile_key": "a2_sentiment_comment_202606",
            "business_rule": "有货-直给到货情绪",
            "corpus": "只写到货反应，不混入报告、转奶、会员权益等别的内容。",
        },
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "有货"
    assert not payload["issues"]


@pytest.mark.parametrize(
    "body",
    [
        "a2到货了，太好了！",
        "a2到货了！",
        "a2至初来货了！",
        "a2到了！终于等到了！",
        "终于等到了！",
    ],
)
def test_a2_stock_guard_accepts_complete_short_arrival_comments(body):
    item = ContentBatchItem(
        title="有货-直给到货情绪",
        body=body,
        plan_json={
            "quality_guard_profile_key": "a2_sentiment_comment_202606",
            "business_rule": "有货-直给到货情绪",
        },
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "有货"
    assert not any(issue["code"] == "activity_body_incomplete_comment" for issue in payload["issues"])


@pytest.mark.parametrize("body", ["到了！", "太好了！"])
def test_a2_stock_guard_rejects_short_comment_below_five_chars_or_without_supply_word(body):
    item = ContentBatchItem(
        title="有货-直给到货情绪",
        body=body,
        plan_json={
            "quality_guard_profile_key": "a2_sentiment_comment_202606",
            "business_rule": "有货-直给到货情绪",
        },
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_incomplete_comment" for issue in payload["issues"])


def test_a2_activity_guard_still_requires_batch_report_for_stock_batch_combo():
    item = ContentBatchItem(
        body="刚看到a2能买了，我去瞅瞅。",
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert payload["context_list"]["关键词"] == "有货+批批检"
    assert any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])


def test_a2_activity_guard_accepts_member_benefit_comment_without_batch_report():
    item = ContentBatchItem(
        body="长期喝a2的，空罐攒起来换奶粉挺实在。",
        plan_json=_a2_guard_plan("会员权益-集罐换礼：\n写妈妈看到 a2 会员活动里可以集罐换奶粉后的评论。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "会员权益"
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_requires_brand_anchor_for_member_benefit_comment():
    item = ContentBatchItem(
        body="长期喝的话，空罐攒起来换奶粉挺实在。",
        plan_json=_a2_guard_plan("会员权益-集罐换礼：\n写妈妈看到 a2 会员活动里可以集罐换奶粉后的评论。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_missing_a2_member_brand_anchor" for issue in payload["issues"])


def test_a2_activity_guard_normalizes_uppercase_brand_but_preserves_a2_protein():
    item = ContentBatchItem(
        body="长期喝A2的，会员活动里也提到A2蛋白，空罐攒起来换奶粉挺实在。",
        plan_json=_a2_guard_plan("会员权益-集罐换礼：\n写妈妈看到 a2 会员活动里可以集罐换奶粉后的评论。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert item.body == "长期喝a2的，会员活动里也提到A2蛋白，空罐攒起来换奶粉挺实在。"
    assert any(repair["code"] == "activity_body_a2_brand_case_normalized" for repair in payload["repairs"])


def test_a2_activity_guard_does_not_treat_a2_protein_as_member_brand_anchor():
    item = ContentBatchItem(
        body="看到A2蛋白这个说法了，空罐攒起来换奶粉挺实在。",
        plan_json=_a2_guard_plan("会员权益-集罐换礼：\n写妈妈看到 a2 会员活动里可以集罐换奶粉后的评论。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert item.body == "看到A2蛋白这个说法了，空罐攒起来换奶粉挺实在。"
    assert any(issue["code"] == "activity_body_missing_a2_member_brand_anchor" for issue in payload["issues"])


@pytest.mark.parametrize(
    "body",
    [
        "这次a2的检测报告列得挺清楚的。",
        "刚对比了下，a2这批的检测报告挺细的，看到这样心里踏实些。",
        "a2这次能直接扫到报告，省心点。",
        "这批检测报告挺详细的，蜡样检测也有，透明多了。",
    ],
)
def test_a2_activity_guard_accepts_report_detail_short_comments(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_guard_plan("检测报告更透明：\n关键词方向是有货+批批检，像妈妈短短聊这批检测报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_accepts_generic_batch_report_confidence_sentence():
    item = ContentBatchItem(
        body="这个批次有检测报告，心里有数了。",
        plan_json=_a2_guard_plan("检测报告更透明：\n关键词方向是有货+批批检，像妈妈短短聊这批检测报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_accepts_short_report_detail_without_subject():
    item = ContentBatchItem(
        body="这次检测报告挺详细的。",
        plan_json=_a2_guard_plan("检测报告更透明：\n关键词方向是有货+批批检，像妈妈短短聊这批检测报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_soft_repairs_that_batch_report_wording():
    item = ContentBatchItem(
        body="补货前扫一下罐底，那批报告就跳出来了",
        plan_json=_a2_guard_plan("检测报告更透明：\n关键词方向是有货+批批检，像妈妈短短聊这批检测报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "那批报告" not in item.body
    assert "这批报告" in item.body
    assert any(repair["source"] == "那批报告" and repair["replacement"] == "这批报告" for repair in payload["repairs"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])


def test_a2_activity_guard_accepts_every_batch_check_short_comment_without_emotion_word():
    item = ContentBatchItem(
        body="这次补a2是因为听店员说他们每批都查，不是抽测",
        plan_json=_a2_guard_plan("不是偶尔抽查看着更稳：\n关键词方向是有货+批批检，像妈妈补货时聊每批检测。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_accepts_batch_by_batch_check_short_comment():
    item = ContentBatchItem(
        body="今天又补了a2，主要是看中它批批都检，不是偶尔查",
        plan_json=_a2_guard_plan("不是偶尔抽查看着更稳：\n关键词方向是有货+批批检，像妈妈补货时聊每批检测。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_accepts_batch_by_batch_inspection_short_comment():
    item = ContentBatchItem(
        body="囤货时候看a2批批都验，就放心入啦",
        plan_json=_a2_guard_plan("不是偶尔抽查看着更稳：\n关键词方向是有货+批批检，像妈妈补货时聊每批检测。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_accepts_new_batch_detection_without_brand_when_context_is_clear():
    item = ContentBatchItem(
        body="看了下新批次也都有检测，这下补货不纠结了",
        plan_json=_a2_guard_plan("不是偶尔抽查看着更稳：\n关键词方向是有货+批批检，像妈妈补货时聊每批检测。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_accepts_new_zealand_third_party_data_without_report_wording():
    item = ContentBatchItem(
        body="功课做了几圈，a2的三方检测数据列得明明白白，选起来更踏实",
        plan_json=_a2_guard_plan("对比后看新西兰三方检测信息：\n关键词方向是批批检+转奶。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_normalizes_third_party_report_wording():
    item = ContentBatchItem(
        body="功课做了几圈，a2的三方检测报告列得明明白白，选起来更踏实",
        plan_json=_a2_guard_plan("对比后看新西兰三方检测信息：\n关键词方向是批批检+转奶。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "三方检测报告" not in item.body
    assert "三方质检报告" in item.body


def test_a2_activity_guard_blocks_003_even_with_waxy_detection_standard_wording():
    item = ContentBatchItem(
        body="刚转a2，扫批次报告看到蜡样检测标准是小于0.03，心里有底。",
        plan_json=_a2_guard_plan("转奶前看蜡样检测：\n关键词方向是批批检+转奶，像妈妈转奶前看这罐蜡样检测报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert "蜡样检测标准是小于0.03" in item.body
    assert "报告细节" not in item.body
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_accepts_short_complete_report_comment():
    item = ContentBatchItem(
        body="补上a2了，看到报告就安心些",
        plan_json=_a2_guard_plan("因为a2的优势转奶选择a2：\n关键词方向是有货+转奶，像妈妈补货后看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_incomplete_comment" for issue in payload["issues"])
    assert not payload["issues"]


@pytest.mark.parametrize(
    "body",
    [
        "店里还有a2我就先拍了",
        "求问姐妹是哪家店呀，我也想接着买a2",
    ],
)
def test_a2_activity_guard_accepts_short_thread_supply_comment(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_guard_plan("有货后继续原来的：\n关键词方向是有货+转奶，像妈妈在评论区问店铺、接着买原来的。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_incomplete_comment" for issue in payload["issues"])


def test_a2_activity_guard_rejects_vague_deictic_comment_without_product():
    item = ContentBatchItem(
        body="刚收到了，家里正好喝完这罐",
        plan_json=_a2_guard_plan("有货后继续原来的：\n关键词方向是有货+转奶，像妈妈在评论区说自己买到了。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_vague_deictic_without_product" for issue in payload["issues"])


def test_a2_activity_guard_accepts_contextual_can_reference_with_scan_report_detail():
    item = ContentBatchItem(
        body="刚在门店拿了这罐，扫物流码能看到检测报告，虽然那些专业词没记住",
        plan_json=_a2_guard_plan("报告查询互动：\n关键词方向是有货+批批检，像妈妈在a2评论区顺手接一句。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_vague_deictic_without_product" for issue in payload["issues"])


@pytest.mark.parametrize(
    "body",
    [
        "终于到了",
        "我的也快到了",
        "先不转了",
        "有底了",
        "我也等发货",
        "等发货中",
        "能不换就不换",
        "转奶先放放",
        "不折腾了",
        "a2终于到了",
        "刚看到a2到了",
        "a2到了，不过不着急转",
        "刚看到a2能拍了",
        "我的也到了，先喝着旧的",
        "到货了先不急着换",
        "a2刚到了，先不急着换",
        "我的到货了，先喝着a2",
        "刚看到a2到货了",
        "a2到货了，转奶先缓缓",
        "刚刷到a2有货了",
        "刚问了一下，a2到了",
        "刚问了客服说a2有货了",
        "刚问柜姐说有货了，立马去下单",
        "刚问了几家店都说有货",
        "刚问了导购，说a2到了",
        "昨天导购说到货了",
        "我订的那罐也到了",
    ],
)
def test_a2_activity_guard_accepts_contextual_supply_transfer_short_replies(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_guard_plan("有货后继续原来的：\n关键词方向是有货+转奶，像妈妈在评论区短接楼。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_incomplete_comment" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_vague_deictic_without_product" for issue in payload["issues"])


@pytest.mark.parametrize("body", ["导", "这个挺好", "先喝着", "再等等"])
def test_a2_activity_guard_rejects_empty_short_replies(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_guard_plan("有货后继续原来的：\n关键词方向是有货+转奶，像妈妈在评论区短接楼。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_incomplete_comment" for issue in payload["issues"])


def test_a2_activity_guard_rejects_scan_before_can_in_hand():
    item = ContentBatchItem(
        body="下单后扫一下物流码，这罐的报告就有了",
        plan_json=_a2_guard_plan("有货后看报告：\n关键词方向是有货+批批检，像妈妈到手后扫罐底码。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_scan_before_can_in_hand" for issue in payload["issues"])


@pytest.mark.parametrize("competitor", ["爱他美", "达能", "雀巢", "超启能恩", "美素", "皇家美素", "皇美"])
def test_a2_activity_guard_accepts_competitor_group_aliases(competitor):
    competitor_phrase = "爱他美样批也看" if competitor == "爱他美" else f"{competitor}也看"
    item = ContentBatchItem(
        body=f"有货了想转奶，先看物流码报告，{competitor_phrase}，a2能查到自己手里这批次更踏实",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "有货+转奶"


@pytest.mark.parametrize(
    "body",
    [
        "我们家刚转奶的时候也是扫罐底物流码看报告，看到蜡样检测那项才放心",
        "今天记录转奶第5天，会顺手扫罐底物流码看下这罐报告，蜡样检测那项0.03的数值心里比较有谱",
        "转奶前特意扫了罐底看报告，蜡样检测这关过了才放心",
        "之前看别的牌子没注意批次报告，转a2才留意到罐底能扫出来",
    ],
)
def test_a2_activity_guard_handles_natural_scan_report_chain_without_forced_brand(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_guard_plan("转奶前看蜡样检测：\n关键词方向是批批检+转奶，像妈妈转奶前看这罐蜡样检测报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["context_list"]["关键词"] == "批批检+转奶"
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    if "0.03" in body:
        assert payload["pass"] is False
        assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])
    else:
        assert payload["pass"] is True
        assert not payload["issues"]


@pytest.mark.parametrize(
    "body",
    [
        "刚到手先扫了下物流码，报告出来就感觉靠谱多了",
        "刚收到就扫了码，这份报告能对上就省事",
    ],
)
def test_a2_activity_guard_accepts_arrival_scan_report_short_comments(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_guard_plan("自己这批能扫出报告：\n关键词方向是有货+批批检，像妈妈拿到a2后扫物流码看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


@pytest.mark.parametrize(
    "body",
    [
        "最近这批检测报告都能查到了吗？要是能看见我就打算先补货试试",
        "同款，这次补货刚好看到检测报告，转奶选它更踏实些。",
        "有货先补一罐，看见报告就能踏实转奶了",
        "正准备转奶就赶上补货，看到有批次报告，放心先囤了。",
        "前两天还在看奶粉，正好看到a2新批次有报告，补货时就顺手带了罐先试试",
        "新一批能看报告确实安心，先补上慢慢转。",
        "刚好这次补货a2有批次报告，我打算补上了慢慢换",
        "a2这波补货看到报告有蜡样检测那项，就打算慢慢转了",
        "看到报告显示最新批次都有蜡样检测，就补上慢慢转了",
    ],
)
def test_a2_activity_guard_accepts_contextual_report_advantage_without_forced_brand(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert not payload["issues"]


def test_a2_activity_guard_repairs_duplicate_negative_prefix_wording():
    item = ContentBatchItem(
        body="囤货收到先扫罐底物流码看报告，批次和报告都能查到心里就踏实，慢慢转也不不悬",
        plan_json=_a2_guard_plan("补货后慢慢转：\n关键词方向是有货+转奶，像妈妈补货后慢慢转。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "不不悬" not in item.body
    assert "踏实些" in item.body
    assert not payload["issues"]


def test_a2_activity_guard_accepts_huangmei_transfer_back_to_a2():
    item = ContentBatchItem(
        body="我家之前转奶喝皇美，后来更在意a2至初每批蜡样检测能不能看到，还是准备慢慢换回去。",
        plan_json=_a2_guard_plan("从皇美转回a2至初：\n关键词方向是批批检+转奶，像妈妈转奶前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "批批检+转奶"
    assert not payload["issues"]


def test_a2_activity_guard_does_not_require_competitor_for_general_combo_keyword():
    general_combo = ContentBatchItem(
        body="有货了想转奶，先看物流码报告，a2能查到自己手里这批次更踏实",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(general_combo)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_allows_implicit_competitor_when_angle_names_competitor_comparison():
    item = ContentBatchItem(
        body="有货了想转奶，先看物流码报告，a2能查到自己手里这批次更踏实",
        plan_json=_a2_guard_plan("爱他美/达能转奶对比：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_body_missing_competitor_group" for issue in payload["issues"])


def test_a2_activity_guard_repairs_lab_notation():

    repaired = ContentBatchItem(
        body="有货了想转奶，先看物流码报告，a2蜡样检测标准是<0.03μg/kg我也会看",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    repaired_payload = ActivityQualityGuardService().review_item(repaired)

    assert repaired_payload is not None
    assert repaired_payload["pass"] is False
    assert "μg/kg" not in repaired.body
    assert "<0.03" in repaired.body
    assert "0.03" in repaired.body
    assert any(
        issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"]
        for issue in repaired_payload["issues"]
    )
    assert repaired_payload["repairs"]


def test_a2_activity_guard_accepts_natural_batch_report_comment_without_003():
    item = ContentBatchItem(
        body="补货时我也对比过爱他美样批和美素，最后还是更看重a2能不能查到手里这罐批次报告。",
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "0.03" not in item.body
    assert not payload["issues"]


def test_a2_activity_guard_accepts_split_batch_report_action_chain():
    item = ContentBatchItem(
        body="一直喝爱他美这次换a2了，店里有货我先扫物流码看这罐报告，确实能查到每批的信息放心点",
        plan_json=_a2_guard_plan("门店有货现场扫码看报告：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_blocks_003_in_check_logistics_code_wording():
    item = ContentBatchItem(
        body="昨天刚转a2，睡前喝奶剩半瓶看物流码查报告，蜡样检测那项0.03挺清楚。",
        plan_json=_a2_guard_plan("蜡样检测0.03轻提：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_accepts_can_bottom_scan_wording():
    item = ContentBatchItem(
        body="我们家宝宝刚转a2，罐底扫出来看到自己那罐的报告，比雀巢的更细一点。",
        plan_json=_a2_guard_plan("对比后看新西兰三方检测信息：\n关键词方向是批批检+转奶。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_accepts_aptamil_seen_when_a2_advantage_is_clear():
    item = ContentBatchItem(
        body="爱他美也看过，但a2罐底能扫自己这罐报告，这点更直观。",
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_generalizes_direct_competitor_brand_names():
    item = ContentBatchItem(
        body="之前喝爱他美，现在换a2后会先扫罐底物流码看报告，也对比过雀巢。",
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "爱他美" not in item.body
    assert "雀巢" not in item.body
    assert "之前的奶粉" in item.body
    assert "其他品牌" in item.body
    assert any(repair["code"] == "activity_body_direct_competitor_generalized" for repair in payload["repairs"])


def test_a2_activity_guard_accepts_switch_to_a2_scan_report_wording():
    item = ContentBatchItem(
        body="之前喝爱他美，现在换a2后会先扫罐底物流码看报告。",
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_accepts_aptamil_no_every_batch_wording():
    item = ContentBatchItem(
        body="之前喝爱他美没太看到每批检测，听说a2可以扫物流码看这罐报告，准备慢慢转过来。",
        plan_json=_a2_guard_plan("转奶前看肚肚和报告：\n关键词方向是批批检+转奶，像妈妈转奶前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_rejects_vague_batch_report_comparison_without_a2_advantage():
    item = ContentBatchItem(
        body="补到货了准备转奶，先看宝宝便便稳不稳，再把爱他美和a2的批次报告对一遍。",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])


def test_a2_activity_guard_accepts_aptamil_mentioned_when_a2_scan_advantage_is_clear():
    item = ContentBatchItem(
        body="店员说到货了，先问批次报告，爱他美也问过，a2能扫物流码直接看自己这罐更踏实。",
        plan_json=_a2_guard_plan("门店有货先问报告：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_accepts_aptamil_also_seen_when_a2_every_batch_is_clear():
    item = ContentBatchItem(
        body="刚转奶，爱他美也看过，a2每批能扫物流码查报告，那个蜡样检测数值挺清楚的。",
        plan_json=_a2_guard_plan("转奶前看肚肚和报告：\n关键词方向是批批检+转奶，像妈妈转奶前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_accepts_precise_aptamil_sample_batch_comparison():
    item = ContentBatchItem(
        body="导购通知到货我会先问批次报告，爱他美是样批，a2每批检能扫物流码更省事。",
        plan_json=_a2_guard_plan("门店有货先问报告：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_accepts_aptamil_platform_public_comparison():
    item = ContentBatchItem(
        body="补货前看了下，爱他美跨境是经销商平台公开，a2罐底扫码能看自己这罐报告更省心。",
        plan_json=_a2_guard_plan("竞品报告获取方式对比：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


@pytest.mark.parametrize("term", ["急", "担心", "不确定", "断粮", "直接着喝"])
def test_a2_activity_guard_rejects_soft_negative_words_without_repair(term):
    item = ContentBatchItem(
        body=f"有货了先看报告再转奶，{term}这个点我也会留意。",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert term in item.body
    assert any(issue["code"] == "activity_forbidden_terms" and term in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_allows_jiejie_decision_wording():
    item = ContentBatchItem(
        body="纠结转奶的话我会先看a2这罐批次报告，能扫罐底物流码查到就踏实些。",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])


def test_a2_activity_guard_allows_group_context_wording():
    item = ContentBatchItem(
        body="群里姐妹说这批报告能查，我刚好补a2就顺手看了下。",
        plan_json=_a2_guard_plan("门店有货现场扫码看报告：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])


def test_a2_activity_guard_allows_not_jiejie_positive_wording():
    item = ContentBatchItem(
        body="刚好门店到货，能看到检测报告这点挺加分，转奶就不纠结了",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])


def test_a2_activity_guard_allows_not_zhaji_positive_wording():
    item = ContentBatchItem(
        body="刚好家里奶快喝完了，先补上a2试试，报告能查到就不着急。",
        plan_json=_a2_guard_plan("因为a2的优势转奶选择a2：\n关键词方向是有货+转奶，像妈妈补货后看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not any(issue["code"] == "activity_forbidden_terms" for issue in payload["issues"])


def test_a2_activity_guard_rejects_zhaji_without_negative_prefix():
    item = ContentBatchItem(
        body="刚好家里奶快喝完了，先补上a2试试，着急这个点我会留意。",
        plan_json=_a2_guard_plan("因为a2的优势转奶选择a2：\n关键词方向是有货+转奶，像妈妈补货后看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" and "急" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_rejects_truncated_comment_tail():
    item = ContentBatchItem(
        body="到货先扫物流码，爱他美我也比过，a2报告里那项0.03标得细，心里",
        plan_json=_a2_guard_plan("门店有货先问报告：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_incomplete_comment" for issue in payload["issues"])


def test_a2_activity_guard_generalizes_out_of_scope_competitor_terms():
    item = ContentBatchItem(
        body="补货时我看美素和美赞臣，也看a2物流码报告。",
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "美素" not in item.body
    assert "美赞臣" not in item.body
    assert "其他品牌" in item.body
    assert any(repair["code"] == "activity_body_direct_competitor_generalized" for repair in payload["repairs"])


def test_a2_activity_guard_repairs_unconfirmed_nestle_value_attribution():
    item = ContentBatchItem(
        body="转奶功课里超启能恩和a2的0.03报告我都会翻，物流码和便便也留意。",
        plan_json=_a2_guard_plan("雀巢组转奶对照：\n关键词方向是批批检+转奶，像妈妈转奶前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert "超启能恩和a2的0.03" not in item.body
    assert "超启能恩" not in item.body
    assert "其他品牌" in item.body
    assert "蜡样检测报告" in item.body
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_repairs_ambiguous_competitor_003_comparison():
    item = ContentBatchItem(
        body="刚补货先扫物流码确认批次，转奶那几天我还会拿达能和0.03对比肚肚反应。",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert "达能和0.03" not in item.body
    assert "拿达能也看" not in item.body
    assert "达能" not in item.body
    assert "还会看其他品牌" in item.body
    assert "0.03对比肚肚" not in item.body
    assert "a2报告里蜡样那项" in item.body
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_repairs_competitor_comparison_and_blocks_003():
    item = ContentBatchItem(
        body="刚转a2，到货扫物流码看报告，蜡样检测蜡样检测蜡样检测蜡样检测0.03这条线比美素0.2细，心里有数。",
        plan_json=_a2_guard_plan("0.03轻对比：\n关键词方向是有货+批批检，像妈妈补货前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert "蜡样检测蜡样检测" not in item.body
    assert "美素" not in item.body
    assert "其他品牌" in item.body
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_blocks_plain_003_and_02_standard_comparison():
    item = ContentBatchItem(
        body="也看过其他品牌小于0.2的标准，a2蜡样检测标准是<0.03，这点在报告里能看到更有底。",
        plan_json=_a2_guard_plan("蜡样检测0.03轻提：\n关键词方向是有货+批批检，像妈妈补货前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert not any(issue["code"] == "activity_body_bad_003_competitor_comparison" for issue in payload["issues"])
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_blocks_003_in_wax_standard_angle_without_batch_report_wording():
    item = ContentBatchItem(
        body="我看奶粉会特意留意一下蜡毒那项，a2是<0.03，看到这个数字会放心些。",
        plan_json=_a2_guard_plan("蜡样检测0.03轻提：\n关键词方向是有货+批批检，重点讲蜡样检测标准。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert not any(issue["code"] == "activity_body_missing_combo_marker" for issue in payload["issues"])
    assert not any(issue["code"] == "activity_body_missing_a2_specific_advantage" for issue in payload["issues"])
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_does_not_break_less_than_003_wording():
    item = ContentBatchItem(
        body="补货前会看蜡样检测标准，a2是<0.03这点我记住了",
        plan_json=_a2_guard_plan("蜡样检测0.03轻提：\n关键词方向是有货+批批检，重点讲蜡样检测标准。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert "a2是<蜡样检测0.03" not in item.body
    assert "a2是<0.03这点" in item.body
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_repairs_bare_003_point_wording_but_keeps_initial_hard_hit():
    item = ContentBatchItem(
        body="补货前看报告，a2 0.03这点我会记一下",
        plan_json=_a2_guard_plan("蜡样检测0.03轻提：\n关键词方向是有货+批批检，重点讲蜡样检测标准。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert "蜡样检测这点" in item.body
    assert "0.03" not in item.body
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_accepts_combo_keyword_without_forcing_both_scene_words():
    item = ContentBatchItem(
        body="达能也在看，转奶前我会先扫a2物流码查报告，自己这批次能对上更踏实。",
        plan_json=_a2_guard_plan("有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert payload["context_list"]["关键词"] == "有货+转奶"
    assert "有货转奶" not in item.body


def test_a2_activity_guard_rejects_tail_cut_at_a2_brand():
    item = ContentBatchItem(
        body="最近准备转奶，门店到货先扫物流码，爱他美和a2",
        plan_json=_a2_guard_plan("门店到货转奶前问清：\n关键词方向是有货+转奶，像妈妈转奶前问报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_incomplete_comment" for issue in payload["issues"])


def test_a2_activity_guard_rejects_tail_cut_at_modal_verb():
    item = ContentBatchItem(
        body="到货先问物流码，爱他美样批也看过，a2每批检能",
        plan_json=_a2_guard_plan("门店有货先问报告：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_incomplete_comment" for issue in payload["issues"])


def test_a2_activity_guard_repairs_competitor_report_003_attribution():
    item = ContentBatchItem(
        body="到货先看物流码确认批批检，转奶这段跟爱他美报告里那项0.03对了下。",
        plan_json=_a2_guard_plan("门店到货转奶前问清：\n关键词方向是有货+转奶，像妈妈转奶前问报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert "爱他美报告里那项0.03" not in item.body
    assert "爱他美" not in item.body
    assert "其他品牌的样批也看，a2报告里蜡样那项" in item.body
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_generalizes_competitor_batch_report_wording():
    item = ContentBatchItem(
        body="刚看到有货，转奶前先扫a2物流码，顺手对比下爱他美每批报告。",
        plan_json=_a2_guard_plan("门店到货转奶前问清：\n关键词方向是有货+转奶，像妈妈转奶前问报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "爱他美" not in item.body
    assert "其他品牌每批报告" in item.body
    assert any(repair["code"] == "activity_body_direct_competitor_generalized" for repair in payload["repairs"])


def test_a2_activity_guard_accepts_nestle_every_batch_check():
    item = ContentBatchItem(
        body="转奶前也看过雀巢每批检，a2能扫罐底看自己这罐报告，便便也慢慢观察。",
        plan_json=_a2_guard_plan("雀巢组转奶对照：\n关键词方向是批批检+转奶，像妈妈转奶前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_accepts_royal_friso_sample_batch():
    item = ContentBatchItem(
        body="补货前问过皇家美素样批，a2罐底扫码能看自己这罐报告，这点我更习惯。",
        plan_json=_a2_guard_plan("竞品报告获取方式对比：\n关键词方向是有货+批批检，像妈妈补货前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_accepts_friso_short_name_sample_batch():
    item = ContentBatchItem(
        body="补货前问过美素样批，a2罐底扫码能看自己这罐报告，这点我更习惯。",
        plan_json=_a2_guard_plan("对美素打a2可查到自己这罐：\n关键词方向是有货+批批检，像妈妈补货前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert not payload["issues"]


def test_a2_activity_guard_generalizes_royal_friso_every_batch_check():
    item = ContentBatchItem(
        body="补货前问过皇家美素每批检测，a2罐底扫码能看自己这罐报告。",
        plan_json=_a2_guard_plan("竞品报告获取方式对比：\n关键词方向是有货+批批检，像妈妈补货前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "皇家美素" not in item.body
    assert "其他品牌每批检测" in item.body
    assert any(repair["code"] == "activity_body_direct_competitor_generalized" for repair in payload["repairs"])


def test_a2_activity_guard_generalizes_competitor_sample_batch_and_blocks_003():
    item = ContentBatchItem(
        body="刚收到通知有货了，雀巢那边看过样批，a2每批报告能查到蜡样检测0.03。",
        plan_json=_a2_guard_plan("补货前先扫物流码：\n关键词方向是有货+批批检，像妈妈补货前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert "雀巢" not in item.body
    assert "其他品牌那边看过样批" in item.body
    assert any(repair["code"] == "activity_body_direct_competitor_generalized" for repair in payload["repairs"])
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_does_not_require_combo_scene_marker():
    item = ContentBatchItem(
        body="爱他美样批也看过，a2罐底物流码能看自己这罐报告，便便状态我会继续观察。",
        plan_json=_a2_guard_plan("便便状态和蜡样那项：\n关键词方向是批批检+转奶，像妈妈转奶前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "物流码" not in item.body
    assert "罐底码" in item.body
    assert not any(issue["code"] == "activity_body_missing_combo_scene" for issue in payload["issues"])


def test_a2_activity_guard_repairs_duplicate_report_reference_and_batch_wording():
    item = ContentBatchItem(
        body="爱他美样批也看，转奶前我会扫物流码，a2报告里报告里那项0.03那项和a2每罐报告都要核。",
        plan_json=_a2_guard_plan("便便状态和蜡样那项：\n关键词方向是批批检+转奶，像妈妈转奶前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert "报告里报告里那项" not in item.body
    assert "0.03那项" not in item.body
    assert "a2每罐" not in item.body
    assert "a2报告里蜡样那项" in item.body
    assert "a2每批报告" in item.body
    assert any(issue["code"] == "activity_forbidden_terms" and "0.03" in issue["evidence"] for issue in payload["issues"])


def test_a2_activity_guard_repairs_duplicate_logistics_code_suffix():
    item = ContentBatchItem(
        body="a2有货了，我家快喝完，导购说可以先扫物流码码看报告再买",
        plan_json=_a2_guard_plan("快喝完时先看报告：\n关键词方向是有货+转奶，像妈妈补货前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "物流码码" not in item.body
    assert "扫罐底码看报告" in item.body
    assert any(repair["code"] == "activity_body_can_bottom_code_normalized" for repair in payload["repairs"])


def test_a2_activity_guard_repairs_redundant_logistics_code_reference():
    item = ContentBatchItem(
        body="想先转这罐试试，扫罐底物流码那个码就能看到这罐的检测报告对吧？",
        plan_json=_a2_guard_plan("问货时看扫码入口：\n关键词方向是有货+转奶，像妈妈问到货后看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert "物流码那个码" not in item.body
    assert "扫罐底码就能看到" in item.body
    assert any(repair["code"] == "activity_body_can_bottom_code_normalized" for repair in payload["repairs"])


def test_a2_activity_guard_repairs_logistics_code_repeated_after_can_bottom_replacement():
    item = ContentBatchItem(
        body="原来扫罐底的物流码就能看到检测报告啊",
        plan_json=_a2_guard_plan("报告查询互动：\n关键词方向是有货+批批检，像妈妈问报告怎么查。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert item.body == "原来扫罐底码就能看到检测报告啊"
    assert "物流码" not in item.body
    assert any(repair["code"] == "activity_body_can_bottom_code_normalized" for repair in payload["repairs"])


def test_a2_activity_guard_normalizes_qr_code_and_third_party_report_name():
    item = ContentBatchItem(
        body="刚开新罐扫了二维码，第三方检测报告能查到。",
        plan_json=_a2_guard_plan("开新罐查看报告：\n关键词方向是有货+批批检，像妈妈开罐时看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is True
    assert item.body == "刚开新罐扫了罐底码，三方质检报告能查到。"
    assert any(repair["code"] == "activity_body_can_bottom_code_normalized" for repair in payload["repairs"])


@pytest.mark.parametrize("body", ["看到熟悉的蓝罐子又有货了", "门店摆着蓝色的罐子，a2终于到了"])
def test_a2_activity_guard_rejects_can_color_description(body):
    item = ContentBatchItem(
        body=body,
        plan_json=_a2_guard_plan("有货直给：\n关键词方向是有货，像妈妈顺手报喜。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_can_color_description" for issue in payload["issues"])


def test_a2_activity_guard_limits_wax_term_to_once_per_comment():
    item = ContentBatchItem(
        body="转奶前看物流码报告，达能也看，a2报告里那项0.03我会比，蜡毒那项和蜡毒报告别重复说",
        plan_json=_a2_guard_plan("便便状态和蜡样那项：\n关键词方向是批批检+转奶，像妈妈转奶前看报告。"),
        quality_json={},
    )

    payload = ActivityQualityGuardService().review_item(item)

    assert payload is not None
    assert payload["pass"] is False
    assert any(issue["code"] == "activity_body_wax_term_overexposed" for issue in payload["issues"])


def test_article_pool_export_keeps_only_usable_generated_items():
    items = [
        ContentBatchReportItem(item_id=1, item_no=1, status="generated", body="可用", hard_pass=True, rewrite_required=False),
        ContentBatchReportItem(item_id=2, item_no=2, status="generated", body="硬校验失败", hard_pass=False, rewrite_required=True),
        ContentBatchReportItem(item_id=3, item_no=3, status="failed", body="", hard_pass=None, rewrite_required=None),
    ]

    exported = _article_pool_export_items(items)

    assert [item.item_id for item in exported] == [1]


def test_article_pool_context_omits_empty_and_comment_only_fields_for_articles():
    item = ContentBatchItem(
        title="能坐住一点了",
        body="孩子最近拼图能多坐一会儿，家里喝的是旺玥。",
        plan_json={
            "output_fields": ["title", "body"],
            "business_rule": "V3M-18｜眼脑营养｜使用反馈｜注意力不集中",
            "unified_generation": {
                "selected_keywords": [
                    {
                        "category_code": "perturbation_rule",
                        "keyword_name": "v2发散提醒",
                    }
                ]
            },
        },
        quality_json={},
    )

    assert build_article_pool_context_list(item) == {
        "扰动规则": "v2发散提醒",
        "业务规则": "V3M-18｜眼脑营养｜使用反馈｜注意力不集中",
    }


def test_article_pool_csv_writes_title_when_item_has_title():
    report = ContentBatchReportResponse(
        batch_id=1,
        batch_code="batch_title",
        asset_key="a2_sentiment_post_activity",
        product_topic="A2舆情相关帖子",
        status="generated",
        count=1,
        summary=ContentBatchReportSummary(),
        items=[
            ContentBatchReportItem(
                item_id=1,
                item_no=1,
                status="generated",
                title="补货顺便扫了下罐底码",
                body="去门店拿了一罐a2至初，店员说这批报告都能查到。",
                hard_pass=True,
                rewrite_required=False,
            )
        ],
    )
    rows = list(csv.DictReader(StringIO(_build_article_pool_csv(report).decode("utf-8-sig"))))

    assert rows[0]["标题"] == "补货顺便扫了下罐底码"
    assert rows[0]["正文"] == "去门店拿了一罐a2至初，店员说这批报告都能查到。"


def test_article_pool_filename_uses_generated_topic_and_time():
    report = ContentBatchReportResponse(
        batch_id=1,
        batch_code="comment_abcd",
        asset_key="a2_sentiment_comment_activity",
        product_topic="A2舆情改善评论",
        status="generated",
        count=0,
        summary=ContentBatchReportSummary(),
        items=[],
    )

    filename = _article_pool_csv_filename(report)

    assert re.fullmatch(r"生成A2舆情改善评论-\d{8}-\d{4}\.csv", filename)


def test_comment_realness_hits_catch_smooth_variants():
    hits = find_comment_realness_hits(
        "喝完奶拉臭臭不费劲，比之前喝的顺多了，软硬也稳，换季没中招，继续观察中，老母亲谁懂啊🍼，同款稳，金黄软。"
    )

    assert "顺多了" in hits
    assert "喝的顺" in hits
    assert "软硬也稳" in hits
    assert "换季没中招" in hits
    assert "继续观察中" in hits
    assert "老母亲" in hits
    assert "谁懂啊" in hits
    assert "🍼" in hits
    assert "同款稳" in hits
    assert "金黄软" in hits


def test_comment_realness_sanitize_keeps_poop_wording_natural():
    text = _remove_or_replace_realness_terms(
        "粑粑黄软的，家里一直金黄软，软一点软的",
        ["金黄软", "黄软"],
        {"金黄软": "金黄色，软软的", "黄软": "黄黄软软"},
    )

    assert "黄黄黄" not in text
    assert "软一点软的" not in text
    assert "黄黄软软" in text
    assert "金黄色，软软的" in text


def test_member_benefit_realness_rewrite_preserves_activity_fact_boundary():
    item = ContentBatchItem(
        body="a2抽奖有宝宝夏凉被，这个挺顺手",
        plan_json={
            "asset_key": "a2_sentiment_comment_activity",
            "business_rule": "会员权益-抽奖活动",
            "scenario_guard_keyword": "会员权益",
        },
    )

    payload = _rewrite_input_payload(
        item,
        hits=["挺顺"],
        replacements={"挺顺": "没那么费劲"},
        rewrite_round=1,
    )

    instructions = "\n".join(payload["rewrite_instructions"])
    assert "保留原评论里的会员活动、集罐、抽奖和本条礼品" in instructions
    assert "不要补转奶、其他品牌、宝宝状态或喝奶体验" in instructions
    assert "拉的时候没那么费劲" not in instructions


@pytest.mark.parametrize(
    ("business_rule", "guard_keyword", "preserve_text", "blocked_text"),
    [
        ("有货-渠道线索", "有货", "到货、能买、门店、下单或发货事实", "不要补检测报告、转奶过程、宝宝状态或喝奶效果"),
        ("批批检-自己这批报告可查", "有货+批批检", "扫码、批次、报告、检测、工艺或市场信息", "不要补到货抢购、转奶过程、宝宝状态或喝奶效果"),
        ("转奶-按自家节奏慢慢试", "有货+转奶", "转奶决定、个体差异、生活变量、报告依据或自家观察", "不要新增排便、胃口、睡眠、宝宝状态或喝奶效果"),
    ],
)
def test_a2_realness_rewrite_preserves_route_facts(
    business_rule,
    guard_keyword,
    preserve_text,
    blocked_text,
):
    item = ContentBatchItem(
        body="原评论",
        plan_json={
            "asset_key": "a2_sentiment_comment_activity",
            "business_rule": business_rule,
            "scenario_guard_keyword": guard_keyword,
        },
    )

    payload = _rewrite_input_payload(
        item,
        hits=["挺顺"],
        replacements={"挺顺": "没那么费劲"},
        rewrite_round=1,
    )

    instructions = "\n".join(payload["rewrite_instructions"])
    assert preserve_text in instructions
    assert blocked_text in instructions
    assert "拉的时候没那么费劲" not in instructions


@pytest.mark.asyncio
async def test_a2_comment_realness_review_catches_low_quality_thread_wording():
    item = ContentBatchItem(
        title="有货+转奶",
        body="我这罐快喝完了，姐妹有推荐的吗",
        plan_json={
            "asset_key": "a2_sentiment_comment_activity",
            "quality_guard_profile_key": "a2_sentiment_comment_202606",
        },
        quality_json={"hard_pass": True},
    )

    payload = await CommentRealnessReviewService().review_and_rewrite_item(
        item=item,
        orchestrator=None,
        executor_code=None,
    )

    assert payload["initial_hits"]
    assert "姐妹有推荐的吗" in payload["initial_hits"]
    assert payload["final_hits"] == []
    assert "有推荐的吗" not in item.body
    assert "哪买的啊" in item.body
    assert item.quality_json["hard_pass"] is True


@pytest.mark.asyncio
async def test_a2_comment_realness_terms_do_not_affect_other_assets():
    item = ContentBatchItem(
        title="普通评论",
        body="我这罐快喝完了，姐妹有推荐的吗",
        plan_json={"asset_key": "yuanyue_comment_activity"},
        quality_json={"hard_pass": True},
    )

    payload = await CommentRealnessReviewService().review_and_rewrite_item(
        item=item,
        orchestrator=None,
        executor_code=None,
    )

    assert payload["initial_hits"] == []
    assert item.body == "我这罐快喝完了，姐妹有推荐的吗"


@pytest.mark.asyncio
async def test_a2_comment_realness_review_catches_marketing_and_rush_wording():
    item = ContentBatchItem(
        title="有货+转奶",
        body="同款妈妈冲，刚抢到这罐a2，赶紧先下一单",
        plan_json={
            "asset_key": "a2_sentiment_comment_activity",
            "quality_guard_profile_key": "a2_sentiment_comment_202606",
        },
        quality_json={"hard_pass": True},
    )

    payload = await CommentRealnessReviewService().review_and_rewrite_item(
        item=item,
        orchestrator=None,
        executor_code=None,
    )

    assert {"同款妈妈冲", "抢到", "赶紧先下一单"}.issubset(set(payload["initial_hits"]))
    assert payload["final_hits"] == []
    assert "冲" not in item.body
    assert "抢到" not in item.body
    assert "赶紧" not in item.body


@pytest.mark.asyncio
async def test_a2_comment_realness_review_catches_stiff_report_and_baby_status_wording():
    item = ContentBatchItem(
        title="有货+转奶",
        body="刚收到a2，宝宝喝得挺痛快，安心值+1，就是自己这这罐报告要看看",
        plan_json={
            "asset_key": "a2_sentiment_comment_activity",
            "quality_guard_profile_key": "a2_sentiment_comment_202606",
        },
        quality_json={"hard_pass": True},
    )

    payload = await CommentRealnessReviewService().review_and_rewrite_item(
        item=item,
        orchestrator=None,
        executor_code=None,
    )

    assert {"宝宝喝得挺痛快", "安心值+1", "这这罐"}.issubset(set(payload["initial_hits"]))
    assert payload["final_hits"] == []
    assert "宝宝喝得挺痛快" not in item.body
    assert "安心值+1" not in item.body
    assert "这这罐" not in item.body


@pytest.mark.asyncio
async def test_a2_comment_realness_review_allows_real_report_sentiment_wording():
    item = ContentBatchItem(
        title="有货+批批检",
        body="Not Detected就是未检出，刚查完安心点了，有报告能自己查到这点确实加分",
        plan_json={
            "asset_key": "a2_sentiment_comment_activity",
            "quality_guard_profile_key": "a2_sentiment_comment_202606",
        },
        quality_json={"hard_pass": True},
    )

    payload = await CommentRealnessReviewService().review_and_rewrite_item(
        item=item,
        orchestrator=None,
        executor_code=None,
    )

    assert payload["initial_hits"] == []
    assert payload["final_hits"] == []
    assert "安心点" in item.body
    assert "加分" in item.body


@pytest.mark.asyncio
async def test_a2_comment_realness_rewrites_stiff_003_value_wording():
    item = ContentBatchItem(
        title="批批检+转奶",
        body="a2物流码扫出来能看蜡样那项，数值0.03，算是个参考依据",
        plan_json={
            "asset_key": "a2_sentiment_comment_activity",
            "quality_guard_profile_key": "a2_sentiment_comment_202606",
        },
        quality_json={"hard_pass": True},
    )

    payload = await CommentRealnessReviewService().review_and_rewrite_item(
        item=item,
        orchestrator=None,
        executor_code=None,
    )

    assert "数值0.03" in payload["initial_hits"]
    assert "数值0.03" not in item.body
    assert "蜡样那项" in item.body
    assert "0.03" not in item.body


@pytest.mark.asyncio
async def test_a2_comment_realness_review_catches_stiff_report_summary_wording():
    item = ContentBatchItem(
        title="有货+批批检",
        body="新入先确认报告，a2这罐能顺手核批次，就冲这点，报告都齐，也比问店员方便",
        plan_json={
            "asset_key": "a2_sentiment_comment_activity",
            "quality_guard_profile_key": "a2_sentiment_comment_202606",
        },
        quality_json={"hard_pass": True},
    )

    payload = await CommentRealnessReviewService().review_and_rewrite_item(
        item=item,
        orchestrator=None,
        executor_code=None,
    )

    assert {"新入", "确认报告", "顺手核批次", "就冲这点", "报告都齐", "比问店员方便"}.issubset(
        set(payload["initial_hits"])
    )
    assert payload["final_hits"] == []
    assert "新入" not in item.body
    assert "核批次" not in item.body
    assert "就冲这点" not in item.body
    assert "报告都齐" not in item.body
    assert "比问店员方便" not in item.body


@pytest.mark.asyncio
async def test_a2_comment_realness_review_catches_odd_thread_action_wording():
    item = ContentBatchItem(
        title="有货+批批检",
        body="美素也做了功课，a2到货后先扫物流码子看报告，我这瓶见底了，主要图它挺省事的",
        plan_json={
            "asset_key": "a2_sentiment_comment_activity",
            "quality_guard_profile_key": "a2_sentiment_comment_202606",
        },
        quality_json={"hard_pass": True},
    )

    payload = await CommentRealnessReviewService().review_and_rewrite_item(
        item=item,
        orchestrator=None,
        executor_code=None,
    )

    assert {"物流码子", "这瓶见底", "主要图它", "挺省事的"}.issubset(set(payload["initial_hits"]))
    assert payload["final_hits"] == []
    assert "物流码子" not in item.body
    assert "这瓶见底" not in item.body
    assert "主要图它" not in item.body


@pytest.mark.asyncio
async def test_comment_similarity_rewrite_updates_quality_metadata():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    service.executor_code = "maga_direct_llm_executor"
    item = ContentBatchItem(
        batch_id=1,
        item_no=2,
        status="generated",
        run_id=11,
        body="你们都在哪买的，多少钱一罐",
        quality_json={"hard_pass": True, "stage_call_count": 1, "run_status": "succeeded"},
        plan_json={"unified_generation": {"selected_keywords": []}},
    )
    similar_item = {
        "batch_id": 1,
        "item_no": 1,
        "body": "姐妹们都在哪买的，多少钱一罐呀",
        "score": 0.72,
        "scope": "current_batch",
    }
    orchestrator = _CommentRewriteOrchestrator("先问正品渠道，别急着囤。")

    await service._rewrite_item_for_similarity(item=item, similar_item=similar_item, orchestrator=orchestrator)

    assert item.body == "先问正品渠道，别急着囤。"
    rewrite = item.quality_json["similarity_rewrites"][0]
    assert rewrite["similar_item_no"] == 1
    assert rewrite["pre_rewrite_similarity_score"] == 0.72
    assert rewrite["similarity_rewrite_passed"] is True
    assert item.quality_json["review_report"]["rewrite_required"] is False
    assert orchestrator.input_payload["content_type"] == "comment"
    assert "避开相似评论" in " ".join(orchestrator.input_payload["rewrite_instructions"])


@pytest.mark.asyncio
async def test_comment_similarity_rewrite_rechecks_candidates_after_rewrite():
    service = ContentCommentBatchService.__new__(ContentCommentBatchService)
    service.executor_code = "maga_direct_llm_executor"
    item = ContentBatchItem(
        batch_id=1,
        item_no=3,
        status="generated",
        run_id=11,
        body="一直喝这款，家里省心少折腾。",
        quality_json={"hard_pass": True, "stage_call_count": 1, "run_status": "succeeded"},
        plan_json={"unified_generation": {"selected_keywords": []}},
    )
    previous_items = [
        ContentBatchItem(batch_id=1, item_no=1, status="generated", body="一直喝这款，没折腾换奶。"),
        ContentBatchItem(batch_id=1, item_no=2, status="generated", body="半夜冲奶就拿这罐，不用想"),
    ]

    async def fake_previous_items(db, current_item):  # noqa: ANN001
        return previous_items

    async def fake_history_items(db, current_item):  # noqa: ANN001
        return []

    service._previous_generated_items = fake_previous_items
    service._history_items_for_similarity = fake_history_items
    orchestrator = _CommentRewriteSequenceOrchestrator(
        [
            "半夜冲奶还拿这罐",
            "临睡那顿肯喝，我就放心。",
        ]
    )

    await service._review_and_rewrite_similarity(db=None, item=item, orchestrator=orchestrator)

    assert item.body == "临睡那顿肯喝，我就放心。"
    assert len(item.quality_json["similarity_rewrites"]) == 2
    assert item.quality_json["review_report"]["rewrite_required"] is False
    assert item.quality_json["hard_pass"] is True


class _CommentRewriteOrchestrator:
    def __init__(self, comment: str):
        self.comment = comment
        self.input_payload = None

    async def run_content_rewrite_stage(self, *, run_id, executor_code, input_payload):  # noqa: ANN001
        self.input_payload = input_payload
        return SimpleNamespace(
            output={"comment": self.comment},
            stage_calls=[object()],
            run=SimpleNamespace(status="succeeded"),
        )


class _CommentRewriteSequenceOrchestrator:
    def __init__(self, comments: list[str]):
        self.comments = list(comments)
        self.input_payloads = []

    async def run_content_rewrite_stage(self, *, run_id, executor_code, input_payload):  # noqa: ANN001
        self.input_payloads.append(input_payload)
        return SimpleNamespace(
            output={"comment": self.comments.pop(0)},
            stage_calls=[object()],
            run=SimpleNamespace(status="succeeded"),
        )


@pytest.mark.asyncio
async def test_ppl_run_can_start_royal_article_profile(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        session.add_all(
            [
                AssetRegistry(
                    asset_type="article_business_rule_set",
                    asset_key="royal_friso_ugc_post_rules_v1",
                    display_name="皇家美素佳儿UGC规则",
                    version_no=1,
                    status="active",
                    asset_stage="production",
                    content_json={
                        "rule_type": "business_rule",
                        "activity_name": "皇家美素佳儿UGC活动",
                        "default_generation_count": 10,
                        "items": [
                            {
                                "rule_id": "business_rule_001",
                                "business_rule": "日常口粮记录",
                                "topic": "日常口粮记录",
                                "corpus": "写皇家美素佳儿作为家里口粮自然出现。",
                                "examples": ["家里这罐皇家美素佳儿先继续喝着。"],
                                "source_row_no": 1,
                            }
                        ],
                    },
                    metadata_json={
                        "rule_type": "business_rule",
                        "keyword_asset_key": "royal_friso_ugc_post_keywords_v1",
                        "default_generation_count": 10,
                    },
                ),
                AssetRegistry(
                    asset_type="content_generation_keywords",
                    asset_key="royal_friso_ugc_post_keywords_v1",
                    display_name="皇家美素佳儿UGC表达扩散语料",
                    version_no=1,
                    status="active",
                    asset_stage="production",
                    content_json={
                        "asset_type": "content_generation_keywords",
                        "categories": [
                            {
                                "category_code": "article_generation_requirement",
                                "category_name": "生成要求",
                                "enabled": True,
                                "required": True,
                                "selection_mode": "one",
                                "applicable_content_types": ["article"],
                                "sub_keywords": [
                                    {
                                        "keyword_code": "royal_requirement",
                                        "keyword_name": "皇家要求",
                                        "corpus": ["只写一篇皇家UGC帖子。"],
                                        "enabled": True,
                                    }
                                ],
                            }
                        ],
                    },
                ),
            ]
        )
        await session.commit()

    response = await client.post(
        "/api/v1/content-agent/ppl-runs/start",
        json={"profile_code": "royal", "count": 1, "created_by": "ops"},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["profile"]["profile_code"] == "royal_friso_ugc_article"
    assert data["profile"]["content_type"] == "article"
    assert data["report"]["asset_key"] == "royal_friso_ugc_post_rules_v1"
    assert data["execution"]["requested_limit"] == 1

    async with session_factory() as session:
        item = (
            await session.execute(
                select(ContentBatchItem)
                .where(ContentBatchItem.batch_id == data["batch_id"])
                .order_by(ContentBatchItem.item_no)
            )
        ).scalars().first()

    assert item.plan_json["asset_key"] == "royal_friso_ugc_post_rules_v1"
    assert item.plan_json["keyword_asset_key"] == "royal_friso_ugc_post_keywords_v1"
    assert item.plan_json["unified_generation"]["keyword_asset"]["asset_key"] == "royal_friso_ugc_post_keywords_v1"


@pytest.mark.asyncio
async def test_ppl_run_can_start_a2_comment_profile(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        session.add(
            AssetRegistry(
                asset_type="comment_business_rule_set",
                asset_key="a2_sentiment_comment_activity",
                display_name="A2舆情改善评论规则",
                version_no=1,
                status="active",
                asset_stage="production",
                content_json={
                    "rule_type": "business_rule",
                    "activity_name": "A2舆情改善评论",
                    "default_generation_count": 1,
                    "quality_guard_profile_key": "a2_sentiment_comment_202606",
                    "items": [
                        {
                            "rule_id": "a2_direct_001",
                            "business_rule": "有货-直给到货情绪",
                            "corpus": "像妈妈看到 a2 到货后顺手接一句，不回讲以前买不到。",
                            "examples": ["a2终于到货了，我去看看", "我也买到a2新货了"],
                            "source_row_no": 1,
                        }
                    ],
                },
                metadata_json={
                    "rule_type": "business_rule",
                    "default_generation_count": 1,
                    "quality_guard_profile_key": "a2_sentiment_comment_202606",
                },
            )
        )
        await session.commit()

    response = await client.post(
        "/api/v1/content-agent/ppl-runs/start",
        json={"profile_code": "a2_comment", "count": 1, "created_by": "ops"},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["profile"]["profile_code"] == "a2_sentiment_comment"
    assert data["profile"]["content_type"] == "comment"
    assert data["report"]["asset_key"] == "a2_sentiment_comment_activity"
    assert data["execution"]["requested_limit"] == 1

    async with session_factory() as session:
        item = (
            await session.execute(
                select(ContentBatchItem)
                .where(ContentBatchItem.batch_id == data["batch_id"])
                .order_by(ContentBatchItem.item_no)
            )
        ).scalars().first()

    assert item.plan_json["asset_key"] == "a2_sentiment_comment_activity"
    assert item.plan_json["quality_guard_profile_key"] == "a2_sentiment_comment_202606"
    assert item.plan_json["unified_generation"]["capability"] == "content.generate"
    rendered_prompt = item.plan_json["unified_generation"]["rendered_prompt"]
    assert "a2终于到货了，我去看看" in rendered_prompt
    assert "有货-直给到货情绪" not in rendered_prompt
    assert "【业务规则】" not in rendered_prompt
    assert "业务规则" not in rendered_prompt
    assert "- 业务规则：" not in rendered_prompt
    assert "- 业务语料：" not in rendered_prompt


@pytest.mark.asyncio
async def test_ppl_run_a2_batch10_profile_persists_prompt_events_and_strategy_items(
    content_agent_workbench_client,
):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        session.add(
            AssetRegistry(
                asset_type="comment_business_rule_set",
                asset_key="a2_sentiment_comment_activity",
                display_name="A2舆情改善评论规则",
                version_no=1,
                status="active",
                asset_stage="production",
                content_json={
                    "rule_type": "business_rule",
                    "activity_name": "A2舆情改善评论",
                    "default_generation_count": 1,
                    "quality_guard_profile_key": "a2_sentiment_comment_202606",
                    "items": [
                        {
                            "rule_id": "a2_direct_001",
                            "business_rule": "有货-直给到货情绪",
                            "corpus": "像妈妈看到 a2 到货后顺手接一句。",
                            "examples": ["a2终于到货了，我去看看"],
                            "source_row_no": 1,
                        }
                    ],
                },
                metadata_json={
                    "rule_type": "business_rule",
                    "default_generation_count": 1,
                    "quality_guard_profile_key": "a2_sentiment_comment_202606",
                },
            )
        )
        await session.commit()

    response = await client.post(
        "/api/v1/content-agent/ppl-runs/start",
        json={"profile_code": "a2_comment_batch10", "created_by": "ops"},
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["profile"]["profile_code"] == "a2_stock_comment_batch10"
    assert data["profile"]["default_comments_per_prompt"] == 10
    assert data["execution"]["requested_limit"] == 10
    assert data["execution"]["generated_count"] == 10

    async with session_factory() as session:
        items = list(
            (
                await session.execute(
                    select(ContentBatchItem)
                    .where(ContentBatchItem.batch_id == data["batch_id"])
                    .order_by(ContentBatchItem.item_no)
                )
            ).scalars().all()
        )
        run_ids = {item.run_id for item in items}
        events = list(
            (
                await session.execute(
                    select(ContentAgentEvent)
                    .where(ContentAgentEvent.run_id.in_(run_ids))
                    .order_by(ContentAgentEvent.id)
                )
            ).scalars().all()
        )

    assert len(items) == 10
    assert len(run_ids) == 1
    assert [item.plan_json["generated_strategy"]["strategy_id"] for item in items] == [
        f"S{index:02d}" for index in range(1, 11)
    ]
    assert {event.event_type for event in events} == {
        "experiment/profile_resolved",
        "prompt/rendered",
        "output/parsed",
    }
    prompt_event = next(event for event in events if event.event_type == "prompt/rendered")
    assert "最重要的是整组差异性" in prompt_event.input_snapshot["rendered_prompt"]
    output_event = next(event for event in events if event.event_type == "output/parsed")
    assert len(output_event.output_snapshot["items"]) == 10


@pytest.mark.asyncio
async def test_batch_workbench_can_start_batch_then_list_and_open_report(content_agent_workbench_client):
    client, _session_factory = content_agent_workbench_client
    response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 2,
            "created_by": "ops",
        },
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["batch_id"]
    assert data["execution"] == {
        "requested_limit": 2,
        "generated_count": 2,
        "failed_count": 0,
        "item_ids": data["execution"]["item_ids"],
    }
    assert len(data["execution"]["item_ids"]) == 2
    assert data["report"]["summary"]["generated_count"] == 2
    assert [item["status"] for item in data["report"]["items"]] == ["generated", "generated"]
    assert data["report"]["items"][0]["title"]
    assert data["report"]["items"][0]["body"]

    list_response = await client.get("/api/v1/content-agent/batches", params={"limit": 10})
    assert list_response.status_code == 200
    list_data = list_response.json()["data"]
    assert list_data["total"] == 1
    assert list_data["items"][0]["batch_id"] == data["batch_id"]
    assert list_data["items"][0]["summary"]["generated_count"] == 2

    report_response = await client.get(
        f"/api/v1/content-agent/batches/{data['batch_id']}/report"
    )
    assert report_response.status_code == 200
    report = report_response.json()["data"]
    assert report["batch_code"] == data["batch_code"]
    assert report["summary"]["hard_pass_count"] == 2


@pytest.mark.asyncio
async def test_batch_workbench_exposes_business_usability_review_endpoint(
    content_agent_workbench_client,
    monkeypatch,
):
    client, session_factory = content_agent_workbench_client

    async with session_factory() as session:
        job = ContentBatchJob(
            batch_code="batch_business_review_api",
            asset_key="wangyue_v3_core_storyline_article_rules",
            product_topic="0705旺玥活动",
            count=1,
            status="generated",
        )
        session.add(job)
        await session.flush()
        session.add(
            ContentBatchItem(
                batch_id=job.id,
                item_no=1,
                status="generated",
                title="饭后那杯奶",
                body="孩子喝旺玥快一个月了，晚饭后那杯还是照常冲好。",
                plan_json={"asset_key": "wangyue_v3_core_storyline_article_rules"},
                quality_json={
                    "hard_pass": True,
                    "review_report": {"rewrite_required": False},
                    "product_experience_llm_quality_review": {
                        "business_usability_tier": "direct_pool",
                        "business_usability_reason": "业务判断稳定入池",
                    },
                },
            )
        )
        batch_id = job.id
        await session.commit()

    class FakeExecutionService:
        def __init__(self, *args, **kwargs):
            pass

        async def review_business_usability_items(self, batch_id, *, force=False, limit=None, concurrency=10):
            assert force is True
            assert limit == 1
            assert concurrency == 3
            return SimpleNamespace(
                batch_id=batch_id,
                reviewed_count=1,
                skipped_count=0,
                failed_count=0,
                reviewed_item_nos=[1],
                skipped_item_nos=[],
                failed_items=[],
                tier_counts={"direct_pool": 1},
            )

    monkeypatch.setattr(
        "app.api.v1.endpoints.content_agent.ContentBatchExecutionService",
        FakeExecutionService,
    )

    response = await client.post(
        f"/api/v1/content-agent/batches/{batch_id}/business-usability-review",
        json={"force": True, "limit": 1, "concurrency": 3},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["reviewed_count"] == 1
    assert data["tier_counts"] == {"direct_pool": 1}
    assert data["report"]["batch_id"] == batch_id
    assert data["report"]["items"][0]["business_usability_tier"] == "direct_pool"


@pytest.mark.asyncio
async def test_batch_workbench_exposes_wangyue_deferred_repair_endpoint(
    content_agent_workbench_client,
    monkeypatch,
):
    client, session_factory = content_agent_workbench_client

    async with session_factory() as session:
        job = ContentBatchJob(
            batch_code="batch_wangyue_deferred_repair_api",
            asset_key="wangyue_v3_core_storyline_article_rules",
            product_topic="0705旺玥活动",
            count=1,
            status="generated",
        )
        session.add(job)
        await session.flush()
        session.add(
            ContentBatchItem(
                batch_id=job.id,
                item_no=1,
                status="generated",
                title="新罐开封",
                body="新罐开封后，顺手复盘了一下为什么继续喝旺玥。",
                plan_json={"asset_key": "wangyue_v3_core_storyline_article_rules"},
                quality_json={
                    "hard_pass": True,
                    "review_report": {"rewrite_required": False},
                    "wangyue_focused_pipeline_review": {
                        "decision": "block",
                        "status": "manual_review",
                        "can_auto_pool": False,
                    },
                },
            )
        )
        batch_id = job.id
        await session.commit()

    class FakeExecutionService:
        def __init__(self, *args, **kwargs):
            pass

        async def repair_wangyue_holdout_items(self, batch_id, *, limit=None, concurrency=10):
            assert limit == 1
            assert concurrency == 2
            return SimpleNamespace(
                batch_id=batch_id,
                selected_count=1,
                repaired_count=1,
                released_count=1,
                held_count=0,
                skipped_count=0,
                failed_count=0,
                selected_item_nos=[1],
                repaired_item_nos=[1],
                released_item_nos=[1],
                held_item_nos=[],
                skipped_item_nos=[],
                failed_items=[],
            )

    monkeypatch.setattr(
        "app.api.v1.endpoints.content_agent.ContentBatchExecutionService",
        FakeExecutionService,
    )

    response = await client.post(
        f"/api/v1/content-agent/batches/{batch_id}/wangyue-deferred-repair",
        json={"limit": 1, "concurrency": 2},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["selected_count"] == 1
    assert data["repaired_count"] == 1
    assert data["released_count"] == 1
    assert data["released_item_nos"] == [1]
    assert data["report"]["batch_id"] == batch_id


@pytest.mark.asyncio
async def test_batch_workbench_exposes_temporal_logic_shadow_review_endpoint(
    content_agent_workbench_client,
    monkeypatch,
):
    client, session_factory = content_agent_workbench_client

    async with session_factory() as session:
        job = ContentBatchJob(
            batch_code="batch_temporal_shadow_api",
            asset_key="wangyue_v3_core_storyline_article_rules",
            product_topic="0705旺玥活动",
            count=1,
            status="generated",
        )
        session.add(job)
        await session.flush()
        session.add(
            ContentBatchItem(
                batch_id=job.id,
                item_no=1,
                status="generated",
                title="出门前找袜子想起这段",
                body="这段时间他三天两头不舒服，回看这段时间状态倒是挺稳。",
                plan_json={"asset_key": "wangyue_v3_core_storyline_article_rules"},
                quality_json={
                    "hard_pass": True,
                    "review_report": {"rewrite_required": False},
                },
            )
        )
        batch_id = job.id
        await session.commit()

    class FakeExecutionService:
        def __init__(self, *args, **kwargs):
            pass

        async def review_temporal_logic_shadow_items(self, batch_id, *, force=False, limit=None, concurrency=10):
            assert force is True
            assert limit == 1
            assert concurrency == 4
            return SimpleNamespace(
                batch_id=batch_id,
                reviewed_count=1,
                skipped_count=0,
                failed_count=0,
                reviewed_item_nos=[1],
                skipped_item_nos=[],
                failed_items=[],
                label_counts={"block": 1},
                usage_totals={"input_tokens": 20, "output_tokens": 10, "total_tokens": 30},
                latency_totals={
                    "total_latency_ms": 120,
                    "average_latency_ms": 120,
                    "max_latency_ms": 120,
                },
            )

    monkeypatch.setattr(
        "app.api.v1.endpoints.content_agent.ContentBatchExecutionService",
        FakeExecutionService,
    )

    response = await client.post(
        f"/api/v1/content-agent/batches/{batch_id}/temporal-logic-shadow-review",
        json={"force": True, "limit": 1, "concurrency": 4},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["reviewed_count"] == 1
    assert data["label_counts"] == {"block": 1}
    assert data["usage_totals"]["total_tokens"] == 30
    assert data["latency_totals"]["total_latency_ms"] == 120
    assert data["report"]["items"][0]["hard_pass"] is True


@pytest.mark.asyncio
async def test_batch_workbench_exposes_claim_public_disease_shadow_review_endpoint(
    content_agent_workbench_client,
    monkeypatch,
):
    client, session_factory = content_agent_workbench_client

    async with session_factory() as session:
        job = ContentBatchJob(
            batch_code="batch_claim_public_disease_shadow_api",
            asset_key="wangyue_v3_core_storyline_article_rules",
            product_topic="0705旺玥活动",
            count=1,
            status="generated",
        )
        session.add(job)
        await session.flush()
        session.add(
            ContentBatchItem(
                batch_id=job.id,
                item_no=1,
                status="generated",
                title="喝了就不会生病",
                body="孩子喝了旺玥就不会生病。",
                plan_json={"asset_key": "wangyue_v3_core_storyline_article_rules"},
                quality_json={"hard_pass": True, "review_report": {"rewrite_required": False}},
            )
        )
        batch_id = job.id
        await session.commit()

    class FakeExecutionService:
        def __init__(self, *args, **kwargs):
            pass

        async def review_claim_public_disease_shadow_items(
            self,
            batch_id,
            *,
            force=False,
            limit=None,
            concurrency=10,
        ):
            assert force is True
            assert limit == 1
            assert concurrency == 5
            return SimpleNamespace(
                batch_id=batch_id,
                reviewed_count=1,
                skipped_count=0,
                failed_count=0,
                reviewed_item_nos=[1],
                skipped_item_nos=[],
                failed_items=[],
                label_counts={"block": 1},
                usage_totals={"input_tokens": 18, "output_tokens": 9, "total_tokens": 27},
                latency_totals={
                    "total_latency_ms": 110,
                    "average_latency_ms": 110,
                    "max_latency_ms": 110,
                },
            )

    monkeypatch.setattr(
        "app.api.v1.endpoints.content_agent.ContentBatchExecutionService",
        FakeExecutionService,
    )

    response = await client.post(
        f"/api/v1/content-agent/batches/{batch_id}/claim-public-disease-shadow-review",
        json={"force": True, "limit": 1, "concurrency": 5},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["reviewed_count"] == 1
    assert data["label_counts"] == {"block": 1}
    assert data["usage_totals"]["total_tokens"] == 27
    assert data["report"]["items"][0]["hard_pass"] is True


@pytest.mark.asyncio
async def test_batch_workbench_exposes_content_fit_shadow_review_endpoint(
    content_agent_workbench_client,
    monkeypatch,
):
    client, session_factory = content_agent_workbench_client

    async with session_factory() as session:
        job = ContentBatchJob(
            batch_code="batch_content_fit_shadow_api",
            asset_key="wangyue_v3_core_storyline_article_rules",
            product_topic="0705旺玥活动",
            count=1,
            status="generated",
        )
        session.add(job)
        await session.flush()
        session.add(
            ContentBatchItem(
                batch_id=job.id,
                item_no=1,
                status="generated",
                title="安排进日常奶粉里",
                body="后来把旺玥安排进日常奶粉里。",
                plan_json={
                    "asset_key": "wangyue_v3_core_storyline_article_rules",
                    "post_type": "使用反馈",
                },
                quality_json={"hard_pass": True, "review_report": {"rewrite_required": False}},
            )
        )
        batch_id = job.id
        await session.commit()

    class FakeExecutionService:
        def __init__(self, *args, **kwargs):
            pass

        async def review_content_fit_shadow_items(
            self,
            batch_id,
            *,
            force=False,
            limit=None,
            concurrency=10,
        ):
            assert force is True
            assert limit == 1
            assert concurrency == 6
            return SimpleNamespace(
                batch_id=batch_id,
                reviewed_count=1,
                skipped_count=0,
                failed_count=0,
                reviewed_item_nos=[1],
                skipped_item_nos=[],
                failed_items=[],
                label_counts={"block": 1},
                usage_totals={"input_tokens": 17, "output_tokens": 8, "total_tokens": 25},
                latency_totals={
                    "total_latency_ms": 105,
                    "average_latency_ms": 105,
                    "max_latency_ms": 105,
                },
            )

    monkeypatch.setattr(
        "app.api.v1.endpoints.content_agent.ContentBatchExecutionService",
        FakeExecutionService,
    )

    response = await client.post(
        f"/api/v1/content-agent/batches/{batch_id}/content-fit-shadow-review",
        json={"force": True, "limit": 1, "concurrency": 6},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["reviewed_count"] == 1
    assert data["label_counts"] == {"block": 1}
    assert data["usage_totals"]["total_tokens"] == 25
    assert data["report"]["items"][0]["hard_pass"] is True


@pytest.mark.asyncio
async def test_batch_workbench_exposes_fluency_shadow_review_endpoint(
    content_agent_workbench_client,
    monkeypatch,
):
    client, session_factory = content_agent_workbench_client

    async with session_factory() as session:
        job = ContentBatchJob(
            batch_code="batch_fluency_shadow_api",
            asset_key="wangyue_v3_core_storyline_article_rules",
            product_topic="0705旺玥活动",
            count=1,
            status="generated",
        )
        session.add(job)
        await session.flush()
        session.add(
            ContentBatchItem(
                batch_id=job.id,
                item_no=1,
                status="generated",
                title="饭菜经常不稳定",
                body="孩子最近饭菜经常不稳定。",
                plan_json={"asset_key": "wangyue_v3_core_storyline_article_rules"},
                quality_json={"hard_pass": True, "review_report": {"rewrite_required": False}},
            )
        )
        batch_id = job.id
        await session.commit()

    class FakeExecutionService:
        def __init__(self, *args, **kwargs):
            pass

        async def review_fluency_shadow_items(
            self,
            batch_id,
            *,
            force=False,
            limit=None,
            concurrency=10,
        ):
            assert force is True
            assert limit == 1
            assert concurrency == 7
            return SimpleNamespace(
                batch_id=batch_id,
                reviewed_count=1,
                skipped_count=0,
                failed_count=0,
                reviewed_item_nos=[1],
                skipped_item_nos=[],
                failed_items=[],
                label_counts={"block": 1},
                usage_totals={"input_tokens": 16, "output_tokens": 7, "total_tokens": 23},
                latency_totals={
                    "total_latency_ms": 101,
                    "average_latency_ms": 101,
                    "max_latency_ms": 101,
                },
            )

    monkeypatch.setattr(
        "app.api.v1.endpoints.content_agent.ContentBatchExecutionService",
        FakeExecutionService,
    )

    response = await client.post(
        f"/api/v1/content-agent/batches/{batch_id}/fluency-shadow-review",
        json={"force": True, "limit": 1, "concurrency": 7},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["reviewed_count"] == 1
    assert data["label_counts"] == {"block": 1}
    assert data["usage_totals"]["total_tokens"] == 23
    assert data["report"]["items"][0]["hard_pass"] is True


@pytest.mark.asyncio
async def test_batch_workbench_exposes_focused_pipeline_shadow_validation_endpoint(
    content_agent_workbench_client,
    monkeypatch,
):
    client, session_factory = content_agent_workbench_client

    async with session_factory() as session:
        job = ContentBatchJob(
            batch_code="batch_focused_pipeline_shadow_api",
            asset_key="wangyue_v3_core_storyline_article_rules",
            product_topic="0705旺玥活动",
            count=1,
            status="generated",
        )
        session.add(job)
        await session.flush()
        session.add(
            ContentBatchItem(
                batch_id=job.id,
                item_no=1,
                status="generated",
                title="饭菜经常不稳定",
                body="孩子最近饭菜经常不稳定。",
                plan_json={"asset_key": "wangyue_v3_core_storyline_article_rules"},
                quality_json={"hard_pass": True, "review_report": {"rewrite_required": False}},
            )
        )
        batch_id = job.id
        await session.commit()

    class FakeExecutionService:
        def __init__(self, *args, **kwargs):
            pass

        async def review_focused_pipeline_shadow_items(
            self,
            batch_id,
            *,
            force=False,
            limit=None,
            concurrency=10,
            rehearse_rewrites=False,
        ):
            assert force is True
            assert limit == 1
            assert concurrency == 5
            assert rehearse_rewrites is True
            return SimpleNamespace(
                batch_id=batch_id,
                reviewed_count=1,
                skipped_count=0,
                failed_count=0,
                reviewed_item_nos=[1],
                skipped_item_nos=[],
                failed_items=[],
                decision_counts={"block": 1},
                rewrite_mode_counts={"fluency_humanize": 1},
                comparison_counts={"mismatch": 1},
                mismatch_item_nos=[1],
                action_comparison_counts={"mismatch": 1},
                action_mismatch_item_nos=[1],
                rewrite_rehearsal_counts={"accepted": 1},
                accepted_rewrite_item_nos=[1],
                manual_review_item_nos=[],
                usage_totals={"input_tokens": 60, "output_tokens": 20, "total_tokens": 80},
                latency_totals={
                    "total_latency_ms": 400,
                    "average_latency_ms": 100,
                    "max_latency_ms": 130,
                },
            )

    monkeypatch.setattr(
        "app.api.v1.endpoints.content_agent.ContentBatchExecutionService",
        FakeExecutionService,
    )

    response = await client.post(
        f"/api/v1/content-agent/batches/{batch_id}/focused-pipeline-shadow-review",
        json={"force": True, "limit": 1, "concurrency": 5, "rehearse_rewrites": True},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["decision_counts"] == {"block": 1}
    assert data["rewrite_mode_counts"] == {"fluency_humanize": 1}
    assert data["comparison_counts"] == {"mismatch": 1}
    assert data["mismatch_item_nos"] == [1]
    assert data["action_comparison_counts"] == {"mismatch": 1}
    assert data["action_mismatch_item_nos"] == [1]
    assert data["rewrite_rehearsal_counts"] == {"accepted": 1}
    assert data["accepted_rewrite_item_nos"] == [1]
    assert data["report"]["items"][0]["hard_pass"] is True


@pytest.mark.asyncio
async def test_batch_workbench_uses_default_executor_when_form_sends_blank_code(content_agent_workbench_client):
    client, _session_factory = content_agent_workbench_client
    response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 1,
            "executor_code": "   ",
            "created_by": "ops",
        },
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["execution"]["generated_count"] == 1
    assert data["report"]["items"][0]["status"] == "generated"


@pytest.mark.asyncio
async def test_a2_batch_generation_queues_independent_audit_without_running_sync_postprocess(
    content_agent_workbench_client,
    monkeypatch,
):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        session.add(
            AssetRegistry(
                asset_type="article_business_rule_set",
                asset_key="a2_reiyu_ugc_post_rules_v1",
                display_name="a2礼遇生文规则",
                version_no=1,
                status="active",
                asset_stage="production",
                content_json={
                    "rule_type": "business_rule",
                    "activity_name": "a2礼遇",
                    "default_generation_count": 1,
                    "items": [
                        {
                            "rule_id": "a2_reiyu_001",
                            "business_rule": "礼遇分享",
                            "topic": "a2礼遇",
                            "corpus": "先讲活动，再自然带出a2至初现在每批都有检测。",
                            "source_row_no": 1,
                        }
                    ],
                },
            )
        )
        await session.commit()

    dispatched: list[int] = []

    class FakeDispatcher:
        @classmethod
        def dispatch(cls, batch_id: int) -> bool:
            dispatched.append(batch_id)
            return True

    monkeypatch.setattr(
        "app.api.v1.endpoints.content_agent.A2ReiyuBatchAuditDispatcher",
        FakeDispatcher,
    )

    response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "a2_reiyu_ugc_post_rules_v1",
            "product_topic": "a2礼遇",
            "count": 1,
            "created_by": "ops",
        },
    )

    assert response.status_code == 200
    batch_id = response.json()["data"]["batch_id"]
    assert dispatched == [batch_id]
    async with session_factory() as session:
        job = await session.get(ContentBatchJob, batch_id)
        item = (
            await session.execute(
                select(ContentBatchItem).where(ContentBatchItem.batch_id == batch_id)
            )
        ).scalar_one()
    assert job.strategy_json["postprocess_mode"] == "generate_only"
    assert job.strategy_json["a2_reiyu_audit"]["status"] == "queued"
    assert item.quality_json["audit_skipped"] is True
    assert "product_experience_llm_quality_review" not in item.quality_json

    explicit_generate_only = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "a2_reiyu_ugc_post_rules_v1",
            "product_topic": "a2礼遇",
            "count": 1,
            "postprocess_mode": "generate_only",
            "created_by": "ops",
        },
    )
    assert explicit_generate_only.status_code == 200
    assert dispatched == [batch_id]
    explicit_batch_id = explicit_generate_only.json()["data"]["batch_id"]
    async with session_factory() as session:
        explicit_job = await session.get(ContentBatchJob, explicit_batch_id)
    assert "a2_reiyu_audit" not in explicit_job.strategy_json


@pytest.mark.asyncio
async def test_comment_batch_can_start_from_rule_asset_key_only(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    response = await client.post(
        "/api/v1/content-agent/comment-batches/start",
        json={"asset_key": "yuanyue_comment_activity", "concurrency": 10, "created_by": "ops"},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["execution"]["requested_limit"] == 3
    assert data["execution"]["generated_count"] == 3
    assert data["execution"]["failed_count"] == 0
    report = data["report"]
    assert report["asset_key"] == "yuanyue_comment_activity"
    assert report["product_topic"] == "美素佳儿源悦活动评论"
    assert report["items"][0]["title"] == "整体适应"
    assert report["items"][0]["body"].startswith("我家刚开始也在看源悦")
    assert report["items"][0]["generation_snapshot"] is None

    async with session_factory() as session:
        job = await session.get(ContentBatchJob, data["batch_id"])
        assert job.strategy_json["execution_concurrency"] == 10

    full_response = await client.get(f"/api/v1/content-agent/batches/{data['batch_id']}/report?full=true")
    full_first = full_response.json()["data"]["items"][0]
    assert full_first["quality"]["rule_type"] == "business_rule"
    assert full_first["generation_snapshot"]["rule_type"] == "business_rule"
    assert full_first["generation_snapshot"]["business_rule"]["business_rule"] == "整体适应"
    assert full_first["generation_snapshot"]["expert"]["expert_config_code"] == "comment_generator_v1"
    assert "整体适应" in full_first["generation_snapshot"]["rendered_prompt"]

    async with session_factory() as session:
        item = (
            await session.execute(
                select(ContentBatchItem)
                .where(ContentBatchItem.batch_id == data["batch_id"])
                .order_by(ContentBatchItem.item_no)
            )
        ).scalars().first()

    assert item.plan_json["rule_type"] == "business_rule"
    assert item.plan_json["render_reference_examples"] is True
    assert item.plan_json["business_rule"] == "整体适应"
    assert "像妈妈在评论区聊刚开始喝源悦" in item.plan_json["corpus"]
    assert item.plan_json["examples"] == ["我家刚开始也在看源悦，想蹲蹲真实反馈"]
    assert "以下参考示例仅供参考" in item.plan_json["unified_generation"]["rendered_prompt"]
    assert "我家刚开始也在看源悦，想蹲蹲真实反馈" in item.plan_json["unified_generation"]["rendered_prompt"]
    assert item.plan_json["unified_generation"]["capability"] == "content.generate"
    assert [kw["category_code"] for kw in item.plan_json["unified_generation"]["selected_keywords"]] == [
        "persona",
        "comment_writing_instruction",
        "perturbation_rule",
        "comment_speaking_style",
        "writing_method",
        "comment_format_control",
    ]


@pytest.mark.asyncio
async def test_comment_batch_can_oversample_then_select_delivery_without_changing_business_pass(
    content_agent_workbench_client,
):
    client, session_factory = content_agent_workbench_client
    response = await client.post(
        "/api/v1/content-agent/comment-batches/start",
        json={
            "asset_key": "yuanyue_comment_activity",
            "count": 3,
            "comment_batch_variation_review": {
                "enabled": True,
                "affects_hard_pass": False,
                "expression_frequency": [
                    {
                        "group_key": "opener_me",
                        "label": "我字开头",
                        "terms": ["我"],
                        "match_mode": "prefix",
                        "max_ratio": 0.2,
                    }
                ],
            },
            "comment_delivery_selection": {
                "enabled": True,
                "target_count": 2,
                "max_similarity": 1.0,
            },
            "created_by": "ops",
        },
    )

    assert response.status_code == 200
    report = response.json()["data"]["report"]
    assert report["summary"]["delivery_candidate_count"] == 3
    assert report["summary"]["delivery_selected_count"] == 2
    assert report["summary"]["delivery_shortfall_count"] == 0
    assert sum(item["delivery_selected"] is True for item in report["items"]) == 2
    assert all(item["hard_pass"] is True for item in report["items"])

    async with session_factory() as session:
        job = await session.get(ContentBatchJob, response.json()["data"]["batch_id"])
        assert job.strategy_json["comment_delivery_selection_override"]["target_count"] == 2


@pytest.mark.asyncio
async def test_comment_batch_can_focus_on_one_business_rule_for_testing(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    response = await client.post(
        "/api/v1/content-agent/comment-batches/start",
        json={
            "asset_key": "yuanyue_comment_activity",
            "business_rule": "整体适应",
            "count": 5,
            "created_by": "ops",
        },
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["execution"]["requested_limit"] == 5
    assert data["execution"]["generated_count"] == 5
    report = data["report"]
    assert [item["title"] for item in report["items"]] == ["整体适应"] * 5
    assert report["items"][0]["generation_snapshot"] is None
    full_response = await client.get(f"/api/v1/content-agent/batches/{data['batch_id']}/report?full=true")
    full_first = full_response.json()["data"]["items"][0]
    assert full_first["generation_snapshot"]["business_rule"]["business_rule"] == "整体适应"

    async with session_factory() as session:
        items = (
            await session.execute(
                select(ContentBatchItem)
                .where(ContentBatchItem.batch_id == data["batch_id"])
                .order_by(ContentBatchItem.item_no)
            )
        ).scalars().all()

    assert len(items) == 5
    assert all(item.plan_json["business_rule"] == "整体适应" for item in items)
    assert all(item.plan_json["examples"] == ["我家刚开始也在看源悦，想蹲蹲真实反馈"] for item in items)


@pytest.mark.asyncio
async def test_comment_batch_can_use_dedicated_keyword_package(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        session.add(
            AssetRegistry(
                asset_type="content_generation_keywords",
                asset_key="a2_plot_discussion_comment_keywords",
                display_name="A2剧情讨论评论语料包",
                version_no=1,
                status="active",
                asset_stage="production",
                content_json={
                    "categories": [
                        {
                            "category_code": "persona",
                            "category_name": "人设",
                            "sub_keywords": [
                                {
                                    "keyword_code": "plot_mom",
                                    "keyword_name": "剧情接话妈妈",
                                    "corpus": ["像妈妈在评论区接剧情，不从全局活动池里乱抽格式。"],
                                }
                            ],
                        }
                    ]
                },
            )
        )
        await session.commit()

    response = await client.post(
        "/api/v1/content-agent/comment-batches/start",
        json={
            "asset_key": "yuanyue_comment_activity",
            "keyword_asset_key": "a2_plot_discussion_comment_keywords",
            "created_by": "ops",
        },
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["report"]["items"][0]["generation_snapshot"] is None
    full_response = await client.get(f"/api/v1/content-agent/batches/{data['batch_id']}/report?full=true")
    first = full_response.json()["data"]["items"][0]
    assert first["generation_snapshot"]["keyword_asset"]["asset_key"] == "a2_plot_discussion_comment_keywords"
    assert first["generation_snapshot"]["selected_keywords"][0]["keyword_name"] == "剧情接话妈妈"

    async with session_factory() as session:
        item = (
            await session.execute(select(ContentBatchItem).order_by(ContentBatchItem.item_no))
        ).scalars().first()

    assert item.plan_json["keyword_asset_key"] == "a2_plot_discussion_comment_keywords"
    assert item.plan_json["unified_generation"]["keyword_asset"]["asset_key"] == "a2_plot_discussion_comment_keywords"


@pytest.mark.asyncio
async def test_batch_report_can_export_generated_results_excel(content_agent_workbench_client):
    client, _session_factory = content_agent_workbench_client
    start_response = await client.post(
        "/api/v1/content-agent/comment-batches/start",
        json={"asset_key": "yuanyue_comment_activity", "created_by": "ops"},
    )
    assert start_response.status_code == 200
    batch_id = start_response.json()["data"]["batch_id"]

    response = await client.get(f"/api/v1/content-agent/batches/{batch_id}/export.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*=UTF-8''" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["生文结果"]
    result = workbook["生文结果"]
    assert result["A1"].value == "标题"
    assert result["B1"].value == "正文"
    assert result["C1"].value == "业务规则"
    assert result["B2"].value.startswith("我家刚开始也在看源悦")
    headers = [cell.value for cell in result[1]]
    assert "Run ID" in headers
    assert "Task ID" in headers
    assert "系统语料包" in headers


@pytest.mark.asyncio
async def test_a2_comment_batch_applies_activity_quality_guard_and_article_pool_export(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        session.add(
            AssetRegistry(
                asset_type="comment_business_rule_set",
                asset_key="a2_sentiment_comment_activity",
                display_name="A2舆情改善业务规则规则",
                version_no=1,
                status="active",
                asset_stage="production",
                content_json={
                    "rule_type": "business_rule",
                    "activity_name": "A2舆情改善评论",
                    "default_generation_count": 3,
                    "quality_guard_profile_key": "a2_sentiment_comment_202606",
                    "keyword_selection": {
                        "comment_generation_requirement": ["xhs_maternal_comment_requirement"],
                        "persona": ["family_mom", "experienced_mom", "rational_comparer", "chatty_mom", "new_mom", "working_mom"],
                        "comment_writing_instruction": ["natural_comment"],
                        "perturbation_rule": ["random_thinking_shift"],
                        "writing_method": ["question_hook", "plain_explain"],
                        "comment_format_control": [
                            "comment_short_clean",
                            "comment_light_emoji",
                            "comment_two_sentence",
                            "comment_21_35",
                            "comment_21_50",
                        ],
                    },
                    "items": [
                        {
                            "rule_id": "a2_001",
                            "business_rule": "A2舆情改善评论",
                            "corpus": "补货前先扫物流码：\n关键词方向是有货+批批检，像妈妈分享到货后先看报告。",
                            "examples": ["有货我先扫罐底码，爱他美样批也看，a2这批报告能点开"],
                            "source_row_no": 1,
                        },
                        {
                            "rule_id": "a2_002",
                            "business_rule": "A2舆情改善评论",
                            "corpus": "转奶前看肚肚和报告：\n关键词方向是批批检+转奶，像妈妈转奶前看报告。",
                            "examples": ["转奶前我先看这罐报告，娃肚肚适应再慢慢来"],
                            "source_row_no": 2,
                        },
                        {
                            "rule_id": "a2_003",
                            "business_rule": "A2舆情改善评论",
                            "corpus": "有货后准备转奶：\n关键词方向是有货+转奶，像妈妈看到有货后先做转奶功课。",
                            "examples": ["到货先拿一罐，转奶这事按家里节奏来"],
                            "source_row_no": 3,
                        },
                    ],
                },
                metadata_json={
                    "rule_type": "business_rule",
                    "default_generation_count": 3,
                    "quality_guard_profile_key": "a2_sentiment_comment_202606",
                    "keyword_selection": {
                        "comment_generation_requirement": ["xhs_maternal_comment_requirement"],
                        "persona": ["family_mom", "experienced_mom", "rational_comparer", "chatty_mom", "new_mom", "working_mom"],
                        "comment_writing_instruction": ["natural_comment"],
                        "perturbation_rule": ["random_thinking_shift"],
                        "writing_method": ["question_hook", "plain_explain"],
                        "comment_format_control": [
                            "comment_short_clean",
                            "comment_light_emoji",
                            "comment_two_sentence",
                            "comment_21_35",
                            "comment_21_50",
                        ],
                    },
                },
            )
        )
        await session.commit()

    response = await client.post(
        "/api/v1/content-agent/comment-batches/start",
        json={"asset_key": "a2_sentiment_comment_activity", "created_by": "ops"},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    report = data["report"]
    assert report["product_topic"] == "A2舆情改善评论"
    first, second, third = report["items"]
    assert first["hard_pass"] is True
    assert second["hard_pass"] is True
    assert third["hard_pass"] is True

    full_report_response = await client.get(
        f"/api/v1/content-agent/batches/{data['batch_id']}/report",
        params={"full": True},
    )

    assert full_report_response.status_code == 200
    full_report = full_report_response.json()["data"]
    full_first, full_second, full_third = full_report["items"]
    first_guard = full_first["quality"]["activity_quality_guard"]
    second_guard = full_second["quality"]["activity_quality_guard"]
    third_guard = full_third["quality"]["activity_quality_guard"]
    assert first_guard["profile_key"] == "a2_sentiment_comment_202606"
    assert first_guard["context_list"]["关键词"] == "有货+批批检"
    assert "蜡毒" not in json.dumps(first_guard["context_list"], ensure_ascii=False)
    assert first_guard["pass"] is True
    assert not first_guard["issues"]
    assert second_guard["context_list"]["关键词"] == "批批检+转奶"
    assert second_guard["pass"] is True
    assert not second_guard["issues"]
    assert third_guard["context_list"]["关键词"] == "有货+转奶"
    assert third_guard["pass"] is True

    async with session_factory() as session:
        persisted_items = (
            await session.execute(
                select(ContentBatchItem)
                .where(ContentBatchItem.batch_id == data["batch_id"])
                .order_by(ContentBatchItem.item_no)
            )
        ).scalars().all()

    assert [item.quality_json["activity_quality_guard"] for item in persisted_items] == [
        first_guard,
        second_guard,
        third_guard,
    ]

    first_plan = persisted_items[0].plan_json
    assert first_plan["keyword_selection"]["persona"] == [
        "family_mom",
        "experienced_mom",
        "rational_comparer",
        "chatty_mom",
        "new_mom",
        "working_mom",
    ]
    selected_codes = {
        keyword["keyword_code"]
        for keyword in first_plan["unified_generation"]["selected_keywords"]
    }
    assert "careful_observer" not in selected_codes
    assert "specific_comment_question" not in selected_codes
    assert "scene_detail" not in selected_codes

    export_response = await client.get(f"/api/v1/content-agent/batches/{data['batch_id']}/export-article-pool.csv")

    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith("text/csv")
    rows = list(csv.DictReader(StringIO(export_response.content.decode("utf-8-sig"))))
    assert list(rows[0]) == ["标题", "正文", "上下文变量(context_list)"]
    assert rows[0]["标题"] == first["title"]
    assert rows[0]["正文"] == first["body"]
    exported_context = json.loads(rows[0]["上下文变量(context_list)"])
    assert exported_context["关键词"] == "有货+批批检"
    assert exported_context["业务规则"].startswith("A2舆情改善评论，有货+批批检-")
    assert "蜡毒" not in rows[0]["上下文变量(context_list)"]


@pytest.mark.asyncio
async def test_article_batch_can_start_from_article_business_rule_asset_key_only(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={"asset_key": "yuanyue_product_experience", "created_by": "ops"},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["execution"]["requested_limit"] == 2
    assert data["execution"]["generated_count"] == 2
    report = data["report"]
    assert report["asset_key"] == "yuanyue_product_experience"
    assert report["product_topic"] == "美素佳儿源悦活动生文"
    assert report["items"][0]["title"]
    assert report["items"][0]["body"]

    async with session_factory() as session:
        item = (
            await session.execute(
                select(ContentBatchItem)
                .where(ContentBatchItem.batch_id == data["batch_id"])
                .order_by(ContentBatchItem.item_no)
            )
        ).scalars().first()

    assert item.plan_json["rule_type"] == "business_rule"
    assert item.plan_json["business_rule"] == "奶量补充"
    assert "product_experience" not in item.plan_json
    assert item.plan_json["unified_generation"]["capability"] == "content.generate"
    assert [kw["category_code"] for kw in item.plan_json["unified_generation"]["selected_keywords"]] == [
        "persona",
        "writing_instruction",
        "article_speaking_style",
        "perturbation_rule",
        "writing_method",
        "article_format_control",
    ]


@pytest.mark.asyncio
async def test_article_batch_start_supports_two_articles_per_prompt(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue_product_experience",
            "articles_per_prompt": 2,
            "count": 2,
            "created_by": "ops",
        },
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["execution"]["requested_limit"] == 2
    assert data["execution"]["generated_count"] == 2
    assert data["execution"]["failed_count"] == 0
    assert data["report"]["status"] == "generated"
    assert data["report"]["summary"]["generated_count"] == 2

    async with session_factory() as session:
        job = (
            await session.execute(
                select(ContentBatchJob).where(ContentBatchJob.id == data["batch_id"])
            )
        ).scalar_one()
        items = (
            await session.execute(
                select(ContentBatchItem)
                .where(ContentBatchItem.batch_id == data["batch_id"])
                .order_by(ContentBatchItem.item_no)
            )
        ).scalars().all()

    assert job.status == "generated"
    assert len({item.run_id for item in items}) == 1
    assert [item.quality_json["multi_output"]["selected_index"] for item in items] == [0, 1]
    assert all(item.quality_json["multi_output"]["materialized_to_batch_items"] for item in items)
    assert "一次生成 2 篇" not in items[0].plan_json["unified_generation"]["rendered_prompt"]
    assert "items 必须正好 2 个" in items[0].plan_json["unified_generation"]["rendered_prompt"]
    assert items[0].plan_json["multi_output_group"]["group_id"] == items[1].plan_json["multi_output_group"]["group_id"]


@pytest.mark.asyncio
async def test_article_batch_start_applies_draft_rule_corpus(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue_product_experience",
            "rule_id": "business_rule_001",
            "source_row_no": 1,
            "draft_corpus": "草稿里的奶量补充语料，只用于本次测试。",
            "draft_selling_painpoint_group": "营养丰富+营养不足-ugc",
            "draft_rule_id": "business_rule_001",
            "draft_source_row_no": 1,
            "count": 1,
            "created_by": "ops",
        },
    )

    assert response.status_code == 200
    data = response.json()["data"]
    async with session_factory() as session:
        item = (
            await session.execute(
                select(ContentBatchItem).where(ContentBatchItem.batch_id == data["batch_id"])
            )
        ).scalar_one()

    assert item.plan_json["corpus"] == "草稿里的奶量补充语料，只用于本次测试。"
    assert item.plan_json["selling_painpoint_group"] == "营养丰富+营养不足-ugc"
    assert item.plan_json["draft_rule_override"] == {
        "enabled": True,
        "rule_id": "business_rule_001",
        "source_row_no": 1,
        "selling_painpoint_group": "营养丰富+营养不足-ugc",
    }


@pytest.mark.asyncio
async def test_article_batch_start_keeps_generated_status_when_postprocess_fails(
    content_agent_workbench_client,
    monkeypatch,
):
    client, session_factory = content_agent_workbench_client

    async def broken_title_repair(self, batch_id, job):  # noqa: ANN001
        raise RuntimeError("synthetic postprocess failure")

    monkeypatch.setattr(
        "app.services.content_batch_execution_service.ContentBatchExecutionService._repair_generated_titles",
        broken_title_repair,
    )

    response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue_product_experience",
            "articles_per_prompt": 2,
            "count": 2,
            "created_by": "ops",
        },
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["execution"]["generated_count"] == 2
    assert data["report"]["status"] == "generated"

    async with session_factory() as session:
        job = (
            await session.execute(
                select(ContentBatchJob).where(ContentBatchJob.id == data["batch_id"])
            )
        ).scalar_one()
        items = (
            await session.execute(
                select(ContentBatchItem).where(ContentBatchItem.batch_id == data["batch_id"])
            )
        ).scalars().all()

    assert job.status == "generated"
    assert job.strategy_json["postprocess_errors"] == [
        {
            "step": "title_repair",
            "error": "synthetic postprocess failure",
            "error_type": "RuntimeError",
        }
    ]
    assert [item.status for item in items] == ["generated", "generated"]


@pytest.mark.asyncio
async def test_comment_batch_runs_forbidden_term_review_and_rewrite(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        session.add(
            AssetRegistry(
                asset_type="business_forbidden_terms",
                asset_key="yuanyue_comment_activity",
                display_name="源悦评论业务违禁词",
                version_no=1,
                status="active",
                asset_stage="production",
                content_json={
                    "schema_version": "1",
                    "terms": [{"term": "源悦", "enabled": True}],
                },
            )
        )
        await session.commit()

    response = await client.post(
        "/api/v1/content-agent/comment-batches/start",
        json={"asset_key": "yuanyue_comment_activity", "created_by": "ops"},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    report = data["report"]
    first = report["items"][0]
    assert "源悦" not in first["body"]
    assert first["forbidden_hits"] == []
    assert first["generation_snapshot"] is None

    full_response = await client.get(f"/api/v1/content-agent/batches/{data['batch_id']}/report?full=true")
    report = full_response.json()["data"]
    first = report["items"][0]
    assert first["quality"]["forbidden_terms_review"]["initial_hits"] == ["源悦"]
    assert first["quality"]["forbidden_terms_review"]["final_hits"] == []
    assert first["quality"]["review_report"]["hard_results"][-1]["ae_code"] == "forbidden_terms_guard"
    assert first["generation_snapshot"]["forbidden_terms_review"]["initial_hits"] == ["源悦"]
    assert first["generation_snapshot"]["rewrite_records"][0]["capability"] == "content.rewrite"
    assert "源悦" in first["generation_snapshot"]["rewrite_records"][0]["before"]["comment"]
    assert "源悦" not in first["generation_snapshot"]["rewrite_records"][0]["after"]["comment"]
    assert report["summary"]["rewrite_item_count"] >= 1

    async with session_factory() as session:
        stage_calls = (await session.execute(select(ContentAgentStageCall))).scalars().all()

    assert any(stage.capability == "content.rewrite" for stage in stage_calls)


@pytest.mark.asyncio
async def test_comment_batch_runs_realness_review_and_rewrite(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        session.add(
            AssetRegistry(
                asset_type="comment_business_rule_set",
                asset_key="yuanyue_comment_activity",
                display_name="源悦活动业务规则规则",
                version_no=2,
                status="active",
                asset_stage="production",
                    content_json={
                        "rule_type": "business_rule",
                        "activity_name": "美素佳儿源悦活动评论",
                        "default_generation_count": 1,
                        "keyword_selection": {"comment_format_control": ["comment_short_clean"]},
                        "items": [
                        {
                            "rule_id": "business_rule_realness_001",
                            "business_rule": "便便问题",
                            "corpus": "便便问题：\n像评论区妈妈随手反馈便便软硬和拉的时候费不费劲。",
                            "examples": ["刚换源悦，拉得挺顺畅"],
                            "supplements": [],
                            "source_row_no": 1,
                        }
                    ],
                },
                metadata_json={
                    "rule_type": "business_rule",
                    "default_generation_count": 1,
                    "rule_count": 1,
                    "example_count": 1,
                },
            )
        )
        await session.commit()

    response = await client.post(
        "/api/v1/content-agent/comment-batches/start",
        json={"asset_key": "yuanyue_comment_activity", "created_by": "ops"},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    report = data["report"]
    first = report["items"][0]
    assert "挺顺畅" not in first["body"]
    assert "顺畅" not in first["body"]
    assert "拉的时候没那么费劲" in first["body"]
    assert first["generation_snapshot"] is None

    full_response = await client.get(f"/api/v1/content-agent/batches/{data['batch_id']}/report?full=true")
    report = full_response.json()["data"]
    first = report["items"][0]
    realness_review = first["quality"]["comment_realness_review"]
    assert realness_review["initial_hits"][0] == "拉得挺顺畅"
    assert realness_review["final_hits"] == []
    assert first["quality"]["review_report"]["hard_results"][-1]["ae_code"] == "comment_realness_guard"
    assert first["generation_snapshot"]["comment_realness_review"]["initial_hits"][0] == "拉得挺顺畅"
    rewrite_record = first["generation_snapshot"]["rewrite_records"][0]
    assert rewrite_record["rewrite_source"] == "comment_realness_review"
    assert "拉得挺顺畅" in rewrite_record["style_hits"]
    assert "拉得挺顺畅" in " ".join(rewrite_record["rewrite_instructions"])

    async with session_factory() as session:
        stage_calls = (await session.execute(select(ContentAgentStageCall))).scalars().all()

    assert any(stage.capability == "content.rewrite" for stage in stage_calls)


@pytest.mark.asyncio
async def test_batch_workbench_uses_maga_default_provider_model(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        session.add(
            LLMProviderConfig(
                id=1,
                provider_code="aihubmix",
                provider_name="AIHubMix",
                provider_type="openai_compatible",
                base_url="https://api.example.test/v1",
                api_key="test-key",
                default_model="deepseek-v4-flash",
                priority=100,
                enabled=1,
            )
        )
        await session.commit()

    response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 1,
            "created_by": "ops",
        },
    )

    assert response.status_code == 200
    async with session_factory() as session:
        item = (await session.execute(select(ContentBatchItem))).scalars().first()

    assert item.plan_json["model_config"] == {
        "provider_code": "aihubmix",
        "model_code": "deepseek-v4-flash",
        "ge_model": "deepseek-v4-flash",
        "ae_model": "deepseek-v4-flash",
    }


@pytest.mark.asyncio
async def test_batch_workbench_rotates_models_under_same_provider(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        session.add(
            LLMProviderConfig(
                id=1,
                provider_code="aihubmix",
                provider_name="AIHubMix",
                provider_type="openai_compatible",
                base_url="https://api.example.test/v1",
                api_key="test-key",
                default_model="deepseek-v4-flash",
                priority=100,
                enabled=1,
            )
        )
        await session.commit()

    response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 4,
            "model_config": {"provider_code": "aihubmix", "temperature": 0.9},
            "model_config_rotation": [
                {"model_code": "deepseek-v4-flash"},
                {"model_code": "glm-5.2"},
                {"model_code": "qwen3.7-plus"},
                {"model_code": "doubao-seed-2-0-pro"},
            ],
            "created_by": "ops",
        },
    )

    assert response.status_code == 200
    async with session_factory() as session:
        items = (
            await session.execute(select(ContentBatchItem).order_by(ContentBatchItem.item_no))
        ).scalars().all()

    assert [item.plan_json["model_config"]["provider_code"] for item in items] == ["aihubmix"] * 4
    assert [item.plan_json["model_config"]["model_code"] for item in items] == [
        "deepseek-v4-flash",
        "glm-5.2",
        "qwen3.7-plus",
        "doubao-seed-2-0-pro",
    ]
    assert [item.plan_json["model_config"]["ge_model"] for item in items] == [
        "deepseek-v4-flash",
        "glm-5.2",
        "qwen3.7-plus",
        "doubao-seed-2-0-pro",
    ]
    assert all(item.plan_json["model_config"]["temperature"] == 0.9 for item in items)


@pytest.mark.asyncio
async def test_batch_workbench_can_record_operator_feedback_and_manual_edit(content_agent_workbench_client):
    client, _session_factory = content_agent_workbench_client
    start_response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 1,
            "created_by": "ops",
        },
    )
    assert start_response.status_code == 200
    item = start_response.json()["data"]["report"]["items"][0]

    feedback_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={
            "action": "request_revision",
            "feedback_text": "开头再具体一点，少一点总结腔。",
            "created_by": "reviewer-a",
        },
    )
    assert feedback_response.status_code == 200
    feedback_data = feedback_response.json()["data"]
    assert feedback_data["review_status"] == "needs_revision"
    assert feedback_data["version_no"] == 1
    assert feedback_data["item"]["human_feedback_text"] == "开头再具体一点，少一点总结腔。"

    manual_edit_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={
            "action": "manual_edit",
            "title": "我家便便不规律那阵子，转源悦后的真实记录",
            "body": "这是运营人工改后的正文，保留真实经历，也避免医疗化表达。",
            "feedback_text": "运营直接人工改稿。",
            "created_by": "reviewer-a",
        },
    )
    assert manual_edit_response.status_code == 200
    manual_data = manual_edit_response.json()["data"]
    assert manual_data["review_status"] == "manual_edited"
    assert manual_data["version_no"] == 2
    assert manual_data["item"]["title"] == "我家便便不规律那阵子，转源悦后的真实记录"
    assert manual_data["item"]["body"] == "这是运营人工改后的正文，保留真实经历，也避免医疗化表达。"
    assert manual_data["item"]["version_compare"]["compare_type"] == "manual_edit"
    assert manual_data["item"]["version_compare"]["before"]["body"] == item["body"]
    assert manual_data["item"]["version_compare"]["after"]["body"] == "这是运营人工改后的正文，保留真实经历，也避免医疗化表达。"
    assert manual_data["item"]["version_compare"]["body_changed"] is True

    approve_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={"action": "approve", "feedback_text": "可发布", "created_by": "reviewer-a"},
    )
    assert approve_response.status_code == 200
    approve_data = approve_response.json()["data"]
    assert approve_data["review_status"] == "approved"
    assert approve_data["version_no"] == 3
    assert approve_data["item"]["feedback_count"] == 3

    report_response = await client.get(
        f"/api/v1/content-agent/batches/{start_response.json()['data']['batch_id']}/report"
    )
    report_item = report_response.json()["data"]["items"][0]
    assert report_response.json()["data"]["summary"]["feedback_count"] == 3
    assert report_item["status"] == "approved"
    assert report_item["review_status"] == "approved"
    assert report_item["latest_version_no"] == 3
    assert report_item["human_feedback_text"] == "可发布"
    assert report_item["feedback_count"] == 3
    assert report_item["version_compare"]["compare_type"] == "manual_edit"
    assert report_item["version_compare"]["after"]["body"] == "这是运营人工改后的正文，保留真实经历，也避免医疗化表达。"


@pytest.mark.asyncio
async def test_batch_feedback_can_auto_rewrite_from_operator_revision(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    start_response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 1,
            "created_by": "ops",
        },
    )
    item = start_response.json()["data"]["report"]["items"][0]
    original_body = item["body"]

    feedback_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={
            "action": "request_revision",
            "feedback_text": "开头再具体一点，少一点总结腔。",
            "quoted_text": "原正文比较总结。",
            "feedback_categories": ["unnatural", "too_ad_like", "unknown"],
            "issue_codes": ["corporate_summary_tone", "none", "corporate_summary_tone"],
            "responsibility_layer": "source_corpus",
            "auto_rewrite": True,
            "model_config": {
                "provider_code": "aliyun",
                "model_code": "qwen-plus",
                "ge_model": "qwen-plus",
                "ae_model": "qwen-plus",
            },
            "created_by": "reviewer-a",
        },
    )

    assert feedback_response.status_code == 200
    feedback_data = feedback_response.json()["data"]
    assert feedback_data["review_status"] == "needs_revision"
    assert feedback_data["version_no"] == 2
    rewritten = feedback_data["item"]
    assert rewritten["status"] == "needs_revision"
    assert rewritten["body"] != original_body
    assert "按运营反馈调整：开头再具体一点" in rewritten["body"]
    assert rewritten["quality"]["human_review"]["auto_rewrite"]["source"] == "operator_feedback"
    assert rewritten["quality"]["human_review"]["issue_codes"] == ["corporate_summary_tone"]
    assert rewritten["quality"]["human_review"]["responsibility_layer"] == "source_corpus"
    assert rewritten["quality"]["review_report"]["rewrite_reason"] == "operator_feedback"
    assert rewritten["generation_snapshot"]["rewrite_records"][-1]["capability"] == "content.rewrite"
    assert rewritten["version_compare"]["compare_type"] == "auto_rewrite"
    assert rewritten["version_compare"]["before"]["body"] == original_body
    assert rewritten["version_compare"]["after"]["body"] == rewritten["body"]
    assert rewritten["version_compare"]["body_changed"] is True

    async with session_factory() as session:
        versions = (
            await session.execute(
                select(ContentBatchItemVersion)
                .where(ContentBatchItemVersion.item_id == item["item_id"])
                .order_by(ContentBatchItemVersion.version_no)
            )
        ).scalars().all()
        feedback = (await session.execute(select(ContentFeedback))).scalar_one()
        stage_calls = (await session.execute(select(ContentAgentStageCall))).scalars().all()

    assert [version.source_action for version in versions] == ["request_revision", "auto_rewrite"]
    assert feedback.metadata_json["auto_rewrite"] is True
    assert feedback.metadata_json["auto_rewrite_version_id"] == versions[-1].id
    assert any(stage.capability == "content.rewrite" for stage in stage_calls)
    rewrite_stage = next(
        stage
        for stage in stage_calls
        if stage.capability == "content.rewrite"
        and (stage.input_snapshot or {}).get("rewrite_source") == "operator_feedback"
    )
    rewrite_input = rewrite_stage.input_snapshot or {}
    rewrite_instructions = "\n".join(rewrite_input.get("rewrite_instructions") or [])
    assert rewrite_input["rewrite_source"] == "operator_feedback"
    assert rewrite_input["content_type"] == "article"
    assert rewrite_input["output_fields"] == ["title", "body"]
    assert "只输出 JSON" in rewrite_input["rendered_prompt"]
    assert "只输出改写后的评论正文" not in rewrite_input["rendered_prompt"]
    assert rewrite_input["quoted_text"] == "原正文比较总结。"
    assert rewrite_input["feedback_categories"] == ["unnatural", "too_ad_like"]
    assert rewrite_input["model_config"]["provider_code"] == "aliyun"
    assert rewrite_input["model_config"]["model_code"] == "qwen-plus"
    assert rewrite_input["model_config"]["ge_model"] == "qwen-plus"
    assert rewrite_input["model_config"]["ae_model"] == "qwen-plus"
    assert rewrite_input["model_config"]["temperature"] >= 0.55
    assert "不是违禁词替换" in rewrite_instructions
    assert "不要只做同义替换" in rewrite_instructions
    assert "运营圈选的原文片段：原正文比较总结。" in rewrite_instructions
    assert "不自然/生硬" in rewrite_instructions
    assert feedback.quoted_text == "原正文比较总结。"
    assert feedback.metadata_json["feedback_categories"] == ["unnatural", "too_ad_like"]
    assert versions[0].metadata_json["issue_codes"] == ["corporate_summary_tone"]
    assert versions[0].metadata_json["responsibility_layer"] == "source_corpus"
    assert feedback.metadata_json["issue_codes"] == ["corporate_summary_tone"]
    assert feedback.metadata_json["responsibility_layer"] == "source_corpus"


@pytest.mark.asyncio
async def test_batch_feedback_can_auto_rewrite_failed_item_with_generated_content(
    content_agent_workbench_client,
):
    client, session_factory = content_agent_workbench_client
    start_response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 1,
            "created_by": "ops",
        },
    )
    item = start_response.json()["data"]["report"]["items"][0]

    async with session_factory() as session:
        stored_item = await session.get(ContentBatchItem, item["item_id"])
        stored_item.status = "failed"
        stored_item.error_message = "自动审核失败，但已保留生成正文"
        await session.commit()

    feedback_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={
            "action": "request_revision",
            "feedback_text": "去掉页面承接，保留已有事实。",
            "auto_rewrite": True,
            "created_by": "reviewer-a",
        },
    )

    assert feedback_response.status_code == 200
    rewritten = feedback_response.json()["data"]["item"]
    assert rewritten["status"] == "needs_revision"
    assert rewritten["body"] != item["body"]
    assert rewritten["quality"]["review_report"]["rewrite_reason"] == "operator_feedback"


@pytest.mark.asyncio
async def test_batch_feedback_insights_summarize_operator_feedback(content_agent_workbench_client):
    client, _session_factory = content_agent_workbench_client
    start_response = await client.post(
        "/api/v1/content-agent/comment-batches/start",
        json={"asset_key": "yuanyue_comment_activity", "created_by": "ops"},
    )
    item = start_response.json()["data"]["report"]["items"][0]

    feedback_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={
            "action": "request_revision",
            "feedback_text": "这句有点像广告口吻，不够像真实评论。",
            "quoted_text": "想蹲蹲真实反馈",
            "feedback_categories": ["too_ad_like", "unnatural", "unknown"],
            "created_by": "reviewer-a",
        },
    )
    assert feedback_response.status_code == 200

    insight_response = await client.get(
        f"/api/v1/content-agent/batches/{start_response.json()['data']['batch_id']}/feedback-insights"
    )

    assert insight_response.status_code == 200
    insights = insight_response.json()["data"]
    assert insights["total_feedback_count"] == 1
    assert insights["category_stats"] == [
        {"code": "unnatural", "label": "不自然/生硬", "count": 1},
        {"code": "too_ad_like", "label": "广告感太强", "count": 1},
    ]
    assert insights["action_stats"] == [{"code": "request_revision", "label": "要求修改", "count": 1}]
    assert insights["samples"][0]["quoted_text"] == "想蹲蹲真实反馈"
    assert insights["samples"][0]["feedback_categories"] == ["too_ad_like", "unnatural"]
    assert insights["suggestions"][0]["suggestion_type"] == "system_keyword"
    assert insights["suggestions"][0]["target"] == "表达扩散语料 / 生文指令"
    assert "广告口吻" in insights["suggestions"][0]["evidence"][0]


@pytest.mark.asyncio
async def test_batch_feedback_can_accept_auto_rewrite(content_agent_workbench_client):
    client, _session_factory = content_agent_workbench_client
    start_response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 1,
            "created_by": "ops",
        },
    )
    item = start_response.json()["data"]["report"]["items"][0]

    rewrite_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={
            "action": "request_revision",
            "feedback_text": "开头再具体一点，少一点总结腔。",
            "auto_rewrite": True,
            "created_by": "reviewer-a",
        },
    )
    rewritten = rewrite_response.json()["data"]["item"]

    accept_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={"action": "accept_rewrite", "created_by": "reviewer-a"},
    )

    assert accept_response.status_code == 200
    accepted = accept_response.json()["data"]["item"]
    assert accepted["status"] == "approved"
    assert accepted["review_status"] == "approved"
    assert accepted["body"] == rewritten["body"]
    assert accepted["version_compare"]["compare_type"] == "accept_rewrite"
    assert accepted["version_compare"]["after"]["body"] == rewritten["body"]


@pytest.mark.asyncio
async def test_batch_feedback_can_reject_auto_rewrite_and_restore_source(content_agent_workbench_client):
    client, _session_factory = content_agent_workbench_client
    start_response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 1,
            "created_by": "ops",
        },
    )
    item = start_response.json()["data"]["report"]["items"][0]
    original_body = item["body"]

    rewrite_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={
            "action": "request_revision",
            "feedback_text": "开头再具体一点，少一点总结腔。",
            "auto_rewrite": True,
            "created_by": "reviewer-a",
        },
    )
    rewritten_body = rewrite_response.json()["data"]["item"]["body"]

    reject_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={"action": "reject_rewrite", "created_by": "reviewer-a"},
    )

    assert reject_response.status_code == 200
    rejected = reject_response.json()["data"]["item"]
    assert rejected["status"] == "needs_revision"
    assert rejected["review_status"] == "needs_revision"
    assert rejected["body"] == original_body
    assert rejected["version_compare"]["compare_type"] == "reject_rewrite"
    assert rejected["version_compare"]["before"]["body"] == rewritten_body
    assert rejected["version_compare"]["after"]["body"] == original_body


@pytest.mark.asyncio
async def test_batch_feedback_is_persisted_for_training(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    start_response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 1,
            "created_by": "ops",
        },
    )
    item = start_response.json()["data"]["report"]["items"][0]

    feedback_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={
            "action": "request_revision",
            "feedback_text": "开头像真实妈妈一点，少一点口号。",
            "created_by": "reviewer-a",
        },
    )

    assert feedback_response.status_code == 200
    feedback_data = feedback_response.json()["data"]
    assert feedback_data["item"]["feedback_count"] == 1

    async with session_factory() as session:
        feedback = (await session.execute(select(ContentFeedback))).scalar_one()
    assert feedback.item_id == item["item_id"]
    assert feedback.version_id == feedback_data["version_id"]
    assert feedback.action == "request_revision"
    assert feedback.review_status == "needs_revision"
    assert feedback.comment == "开头像真实妈妈一点，少一点口号。"
    assert feedback.submitter == "reviewer-a"
    assert feedback.metadata_json["source"] == "content_batch_workbench"
    assert feedback.metadata_json.get("asset_change_request_id") is None


@pytest.mark.asyncio
async def test_operator_feedback_can_add_business_forbidden_term(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    start_response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 1,
            "created_by": "ops",
        },
    )
    item = start_response.json()["data"]["report"]["items"][0]
    term = item["body"][:4]

    feedback_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={
            "action": "request_revision",
            "feedback_text": f"加入业务违禁词：{term}",
            "business_forbidden_terms": [term],
            "created_by": "ops",
        },
    )

    assert feedback_response.status_code == 200
    feedback_data = feedback_response.json()["data"]
    assert feedback_data["item"]["review_status"] == "needs_revision"
    assert term not in feedback_data["item"]["body"]
    assert term not in feedback_data["item"]["forbidden_hits"]
    assert feedback_data["item"]["quality"]["forbidden_terms_review"]["initial_hits"] == [term]
    assert feedback_data["item"]["quality"]["forbidden_terms_review"]["final_hits"] == []

    report_response = await client.get(
        f"/api/v1/content-agent/batches/{start_response.json()['data']['batch_id']}/report"
    )
    report = report_response.json()["data"]
    assert term not in report["items"][0]["body"]
    assert term not in report["items"][0]["forbidden_hits"]
    assert report["summary"]["forbidden_hit_count"] == 0

    async with session_factory() as session:
        asset = (
            await session.execute(
                select(AssetRegistry).where(
                    AssetRegistry.asset_type == "business_forbidden_terms",
                    AssetRegistry.asset_key == "yuanyue",
                    AssetRegistry.status == "active",
                )
            )
        ).scalar_one()
        feedback = (await session.execute(select(ContentFeedback))).scalar_one()
        change_request = (await session.execute(select(AssetChangeRequest))).scalar_one_or_none()

    assert asset.content_json["terms"][-1]["term"] == term
    assert feedback.metadata_json["business_forbidden_terms"] == [term]
    assert feedback.metadata_json["business_forbidden_terms_added"] == [term]
    assert feedback.metadata_json["forbidden_terms_review"]["initial_hits"] == [term]
    assert feedback.metadata_json["forbidden_terms_review"]["final_hits"] == []
    assert change_request is None


@pytest.mark.asyncio
async def test_fact_rule_feedback_creates_asset_change_request(content_agent_workbench_client):
    client, session_factory = content_agent_workbench_client
    start_response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验复盘",
            "count": 1,
            "created_by": "ops",
        },
    )
    item = start_response.json()["data"]["report"]["items"][0]

    feedback_response = await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={
            "action": "request_revision",
            "feedback_text": "源悦和 a2 蛋白、a2 公司没有关系，是完全独立的两款奶粉。禁止提及a2 蛋白。",
            "created_by": "ops",
        },
    )

    assert feedback_response.status_code == 200
    feedback_data = feedback_response.json()["data"]
    assert feedback_data["item"]["review_status"] == "needs_revision"

    async with session_factory() as session:
        change_request = (await session.execute(select(AssetChangeRequest))).scalar_one()
        feedback = (await session.execute(select(ContentFeedback))).scalar_one()

    assert "禁止提及a2 蛋白" in change_request.source_text
    assert change_request.status == "pending"
    assert change_request.requester == "ops"
    assert change_request.context_json["asset_key"] == "yuanyue"
    assert change_request.context_json["intent"] == "fact_or_compliance_rule"
    assert "compliance_rules" in change_request.context_json["affected_asset_types"]
    assert feedback.metadata_json["asset_change_request_id"] == change_request.id
    assert feedback.metadata_json["asset_change_intent"] == "fact_or_compliance_rule"


@pytest.mark.asyncio
async def test_training_feedback_samples_list_returns_cross_batch_feedback(content_agent_workbench_client):
    client, _session_factory = content_agent_workbench_client
    start_response = await client.post(
        "/api/v1/content-agent/batches/start",
        json={
            "asset_key": "yuanyue",
            "product_topic": "宝宝便便不规律",
            "target_audience": "新手妈妈",
            "style": "经验老道型",
            "count": 1,
            "created_by": "ops",
        },
    )
    item = start_response.json()["data"]["report"]["items"][0]

    await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={
            "action": "request_revision",
            "feedback_text": "开头像真实妈妈一点，少一点口号。",
            "created_by": "reviewer-a",
        },
    )
    await client.post(
        f"/api/v1/content-agent/batch-items/{item['item_id']}/feedback",
        json={"action": "approve", "feedback_text": "修改后可发布", "created_by": "reviewer-b"},
    )

    response = await client.get("/api/v1/content-agent/feedback-samples")
    assert response.status_code == 200
    data = response.json()["data"]
    assert data["total"] == 2
    assert data["items"][0]["review_status"] == "approved"
    assert data["items"][0]["comment"] == "修改后可发布"
    assert data["items"][0]["product_topic"] == "宝宝便便不规律"
    assert data["items"][0]["body_preview"]

    filtered = await client.get(
        "/api/v1/content-agent/feedback-samples",
        params={"review_status": "needs_revision"},
    )
    assert filtered.status_code == 200
    filtered_data = filtered.json()["data"]
    assert filtered_data["total"] == 1
    assert filtered_data["items"][0]["review_status"] == "needs_revision"
    assert filtered_data["items"][0]["comment"] == "开头像真实妈妈一点，少一点口号。"


@pytest.mark.asyncio
async def test_comment_review_replay_reuses_generated_body_and_updates_guard_result(
    content_agent_workbench_client,
):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        job = ContentBatchJob(
            batch_code="comment_review_replay_test",
            asset_key="yuanyue_comment_activity",
            product_topic="A2舆情改善评论",
            count=1,
            status="generated",
            strategy_json={"quality_guard_profile_key": "a2_sentiment_comment_202606"},
        )
        session.add(job)
        await session.flush()
        body = "刚在门店拿了这罐，扫物流码能看到检测报告，虽然那些专业词没记住"
        plan = _a2_guard_plan(
            "报告查询互动：\n关键词方向是有货+批批检，像妈妈在a2评论区顺手接一句。"
        )
        plan["output_fields"] = ["comment"]
        item = ContentBatchItem(
            batch_id=job.id,
            item_no=1,
            status="generated",
            plan_json=plan,
            title="批批检-报告查询互动",
            body=body,
            quality_json={
                "hard_pass": False,
                "review_report": {
                    "hard_results": [
                        {
                            "ae_code": "activity_quality_guard.activity_body_vague_deictic_without_product",
                            "pass": False,
                            "feedback": "旧规则误判",
                        }
                    ],
                    "rewrite_required": True,
                    "rewrite_reason": "活动专项质量守卫未通过",
                },
                "activity_quality_guard": {
                    "pass": False,
                    "issues": [{"code": "activity_body_vague_deictic_without_product"}],
                },
            },
        )
        session.add(item)
        await session.commit()
        batch_id = job.id
        item_id = item.id

    response = await client.post(
        f"/api/v1/content-agent/comment-batches/{batch_id}/review-replay",
        json={"item_nos": [1], "created_by": "reviewer-a"},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["reviewed_item_nos"] == [1]
    assert data["changed_pass_item_nos"] == [1]
    assert data["body_changed_item_nos"] == []
    report_item = data["report"]["items"][0]
    assert report_item["body"] == body
    assert report_item["hard_pass"] is True

    async with session_factory() as session:
        persisted = await session.get(ContentBatchItem, item_id)
        versions = list(
            (
                await session.execute(
                    select(ContentBatchItemVersion).where(ContentBatchItemVersion.item_id == item_id)
                )
            ).scalars().all()
        )
    assert persisted is not None
    assert persisted.body == body
    assert persisted.quality_json["review_replay_history"][-1]["after_hard_pass"] is True
    assert len(versions) == 1
    assert versions[0].source_action == "comment_review_replay"
    assert versions[0].body == body


@pytest.mark.asyncio
async def test_comment_review_replay_persists_a2_brand_case_repair(
    content_agent_workbench_client,
):
    client, session_factory = content_agent_workbench_client
    async with session_factory() as session:
        job = ContentBatchJob(
            batch_code="comment_review_replay_a2_case_test",
            asset_key="a2_sentiment_comment_activity",
            product_topic="A2舆情改善评论",
            count=1,
            status="generated",
            strategy_json={"quality_guard_profile_key": "a2_sentiment_comment_202606"},
        )
        session.add(job)
        await session.flush()
        plan = _a2_guard_plan(
            "会员权益-集罐换礼：\n写妈妈看到 a2 会员活动里可以集罐换奶粉后的评论。"
        )
        plan["output_fields"] = ["comment"]
        item = ContentBatchItem(
            batch_id=job.id,
            item_no=1,
            status="generated",
            plan_json=plan,
            title="会员权益-集罐换礼",
            body="长期喝A2的，空罐攒起来换奶粉挺实在。",
            quality_json={"hard_pass": True, "review_report": {"rewrite_required": False}},
        )
        session.add(item)
        await session.commit()
        batch_id = job.id
        item_id = item.id

    response = await client.post(
        f"/api/v1/content-agent/comment-batches/{batch_id}/review-replay",
        json={"item_nos": [1], "created_by": "reviewer-a"},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["body_changed_item_nos"] == [1]
    assert data["report"]["items"][0]["body"] == "长期喝a2的，空罐攒起来换奶粉挺实在。"

    async with session_factory() as session:
        persisted = await session.get(ContentBatchItem, item_id)
        versions = list(
            (
                await session.execute(
                    select(ContentBatchItemVersion).where(ContentBatchItemVersion.item_id == item_id)
                )
            ).scalars().all()
        )
    assert persisted is not None
    assert persisted.body == "长期喝a2的，空罐攒起来换奶粉挺实在。"
    assert len(versions) == 1
    assert versions[0].body == persisted.body


def _yuanyue_assets() -> list[AssetRegistry]:
    return [
        AssetRegistry(
            asset_type="painpoint_model",
            asset_key="yuanyue",
            display_name="源悦痛点模型",
            version_no=1,
            status="active",
            content_json={
                "items": [
                    {"painpoint": "便便不规律", "symptom": "拉臭费劲", "description": "便便状态不稳定", "selling_point": "好消化易吸收"},
                    {"painpoint": "肚肚不舒服", "symptom": "胀气", "description": "喝奶后肚肚闹腾", "selling_point": "温和"},
                ]
            },
        ),
        AssetRegistry(
            asset_type="product_selling_points",
            asset_key="yuanyue",
            display_name="源悦产品卖点",
            version_no=1,
            status="active",
            content_json={
                "items": [
                    {"selling_point": "好消化易吸收", "advantage": "软凝乳"},
                    {"selling_point": "温和", "advantage": "亲和宝宝肚肚"},
                ]
            },
        ),
        AssetRegistry(
            asset_type="reference_examples",
            asset_key="yuanyue",
            display_name="源悦参考例文",
            version_no=1,
            status="active",
            content_json={
                "items": [
                    {"example_id": "ex1", "title": "过来人经验", "body": "先观察宝宝便便状态", "style_tags": ["经验笔记"]},
                    {"example_id": "ex2", "title": "场景共鸣", "body": "新手妈妈别焦虑", "style_tags": ["场景共鸣"]},
                ]
            },
        ),
        AssetRegistry(
            asset_type="compliance_rules",
            asset_key="yuanyue",
            display_name="源悦审核规则",
            version_no=1,
            status="active",
            content_json={"items": [{"dimension": "禁止治疗便秘", "risk_level": "high"}]},
        ),
        AssetRegistry(
            asset_type="comment_business_rule_set",
            asset_key="yuanyue_comment_activity",
            display_name="源悦活动业务规则规则",
            version_no=1,
            status="active",
            content_json={
                "rule_type": "business_rule",
                "activity_name": "美素佳儿源悦活动评论",
                "default_generation_count": 10,
                "items": [
                    {
                        "rule_id": "business_rule_001",
                        "business_rule": "整体适应",
                        "corpus": "整体适应：\n像妈妈在评论区聊刚开始喝源悦的观察，语气自然一点。",
                        "examples": ["我家刚开始也在看源悦，想蹲蹲真实反馈"],
                        "supplements": [],
                        "source_row_no": 1,
                    },
                    {
                        "rule_id": "business_rule_002",
                        "business_rule": "成分讨论",
                        "corpus": "成分讨论：\n像在确认信息，别写成科普长文。",
                        "examples": ["软分子蛋白这个点我也想了解下"],
                        "supplements": [],
                        "source_row_no": 2,
                    },
                    {
                        "rule_id": "business_rule_003",
                        "business_rule": "同款求反馈",
                        "corpus": "同款求反馈：\n像同阶段妈妈顺手问一句。",
                        "examples": ["有同月龄宝宝喝过吗，想看看大家怎么说"],
                        "supplements": [],
                        "source_row_no": 3,
                    },
                ],
            },
            metadata_json={
                "rule_type": "business_rule",
                "default_generation_count": 10,
                "rule_count": 3,
                "example_count": 3,
            },
        ),
        AssetRegistry(
            asset_type="article_business_rule_set",
            asset_key="yuanyue_product_experience",
            display_name="源悦生文业务规则",
            version_no=1,
            status="active",
            asset_stage="production",
            content_json={
                "rule_type": "business_rule",
                "activity_name": "美素佳儿源悦活动生文",
                "default_generation_count": 10,
                "items": [
                    {
                        "rule_id": "business_rule_001",
                        "business_rule": "奶量补充",
                        "topic": "奶量补充",
                        "corpus": "围绕宝宝的奶量补充体验自然展开。",
                        "examples": ["刚换源悦那阵子，喂奶没之前那么拉扯。"],
                        "source_row_no": 1,
                    },
                    {
                        "rule_id": "business_rule_002",
                        "business_rule": "消化吸收",
                        "topic": "消化吸收",
                        "corpus": "围绕喝完后的肚肚状态和便便节奏自然展开。",
                        "examples": ["主要看喝完后的肚肚状态和便便节奏。"],
                        "source_row_no": 2,
                    },
                ],
            },
            metadata_json={
                "rule_type": "business_rule",
                "default_generation_count": 10,
                "rule_count": 2,
                "example_count": 2,
            },
        ),
    ]
