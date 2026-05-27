"""API tests for MAGA Asset Steward surfaces."""

import pytest
import pytest_asyncio
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.api.v1.endpoints.assets import router
from app.core.database import get_db
from app.models.base import Base
from app.models.content_agent import ExecutorRegistry
from app.models.maga_assets import AssetChangeProposal, AssetChangeRequest, AssetImportRun, AssetRegistry


@pytest_asyncio.fixture
async def asset_client():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(
            Base.metadata.create_all,
            tables=[
                AssetRegistry.__table__,
                AssetImportRun.__table__,
                AssetChangeRequest.__table__,
                AssetChangeProposal.__table__,
                ExecutorRegistry.__table__,
            ],
        )
    session_factory = async_sessionmaker(engine, expire_on_commit=False)

    async with session_factory() as session:
        session.add(
            AssetRegistry(
                asset_type="compliance_rules",
                asset_key="yuanyue",
                display_name="源悦审核规则",
                version_no=1,
                status="active",
                content_json={"items": [{"dimension": "不得宣称治疗便秘"}]},
                created_by="test",
            )
        )
        session.add(
            AssetImportRun(
                source_name="源悦种草活动-ai训练规则.xlsx",
                source_hash="abc",
                status="succeeded",
                imported_assets=3,
                summary_json={"asset_key": "yuanyue"},
                created_by="test",
            )
        )
        session.add_all(
            [
                ExecutorRegistry(
                    executor_code="hermes_maga_worker",
                    executor_type="hermes_profile",
                    profile_name="maga-worker",
                    invoke_url="mock://maga-worker/invoke",
                    supported_capabilities_json=[{"capability": "asset.import", "schema_version": "1"}],
                    config_json={"executor_token": "test-token"},
                    enabled=1,
                ),
                AssetRegistry(
                    asset_type="painpoint_model",
                    asset_key="yuanyue",
                    display_name="源悦旧痛点模型",
                    version_no=1,
                    status="active",
                    content_json={"items": [{"painpoint": "历史错误痛点"}]},
                    created_by="test",
                ),
                AssetRegistry(
                    asset_type="painpoint_model",
                    asset_key="yuanyue",
                    display_name="源悦痛点模型",
                    version_no=2,
                    status="active",
                    content_json={
                        "topics": [
                            {
                                "baby_stage": "转奶期宝宝",
                                "topic": "肠胃弱/奶量上不去",
                                "descriptions": ["便便不规律"],
                            }
                        ]
                    },
                    created_by="test",
                ),
                AssetRegistry(
                    asset_type="brand_profile",
                    asset_key="yuanyue",
                    display_name="源悦品牌资料",
                    version_no=1,
                    status="active",
                    content_json={"content_style": "高质量真实用户ugc"},
                    created_by="test",
                ),
                AssetRegistry(
                    asset_type="reference_examples",
                    asset_key="yuanyue",
                    display_name="源悦参考例文",
                    version_no=1,
                    status="active",
                    content_json={
                        "items": [
                            {
                                "direction": "吃",
                                "painpoint": "转奶/消化吸收",
                                "post_format": "用后分享",
                                "style_tags": ["用后分享", "吃"],
                                "body": "新手妈妈别急着焦虑，先看宝宝喝奶和便便状态。",
                            }
                        ]
                    },
                    created_by="test",
                ),
                AssetRegistry(
                    asset_type="brand_profile",
                    asset_key="other-brand",
                    display_name="其他品牌资料",
                    version_no=1,
                    status="active",
                    content_json={"content_style": "其他品牌风格"},
                    created_by="test",
                ),
                AssetRegistry(
                    asset_type="painpoint_model",
                    asset_key="topic-brand",
                    display_name="主题结构资产",
                    version_no=1,
                    status="active",
                    content_json={"topics": [{"topic": "睡眠倒退", "descriptions": ["夜醒频繁"]}]},
                    created_by="test",
                ),
                AssetRegistry(
                    asset_type="painpoint_model",
                    asset_key="yuanyue",
                    display_name="源悦候选扩展主题",
                    version_no=3,
                    status="active",
                    asset_stage="candidate",
                    content_json={"topics": [{"topic": "AI扩展候选痛点", "descriptions": ["候选表达"]}]},
                    created_by="test",
                ),
                AssetRegistry(
                    asset_type="painpoint_expression_candidates",
                    asset_key="yuanyue",
                    display_name="源悦候选痛点描述",
                    version_no=1,
                    status="active",
                    asset_stage="candidate",
                    content_json={
                        "items": [
                            {
                                "topic": "便便不规律",
                                "symptom": "羊屎蛋/干硬",
                                "expression": "便便又干又硬，一粒粒的像羊屎",
                                "expression_source": "xlsx_seed",
                            },
                            {
                                "topic": "便便不规律",
                                "symptom": None,
                                "expression": "每天换尿不湿都会忍不住看便便状态，怕又是一粒粒、硬硬的那种。",
                                "expression_source": "ai_expanded",
                            },
                        ]
                    },
                    created_by="test",
                ),
            ]
        )
        await session.commit()

    app = FastAPI()
    app.include_router(router, prefix="/api/v1/assets")

    async def override_get_db():
        async with session_factory() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as client:
        yield client

    await engine.dispose()


@pytest.mark.asyncio
async def test_extract_reference_elements_preview(asset_client):
    response = await asset_client.post(
        "/api/v1/assets/reference-elements/extract",
        json={"asset_key": "yuanyue", "limit": 1, "persist": False},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["source_item_count"] == 1
    assert data["extracted_count"] == 1
    assert data["persisted_asset_id"] is None
    item = data["items"][0]
    assert item["source_example_id"] == "reference_example_1"
    assert item["title_hook"]["hook_type"] in {"生活经验", "情绪困扰", "场景冲突"}
    assert "奶量" in item["content_atoms"]["painpoint_signals"] or "便便" in item["content_atoms"]["painpoint_signals"]
    assert item["writing_strategy"]["proof_style"]
    assert item["safety"]["avoid_copy_phrases"]


@pytest.mark.asyncio
async def test_extract_reference_elements_persist_candidate_asset(asset_client):
    response = await asset_client.post(
        "/api/v1/assets/reference-elements/extract",
        json={"asset_key": "yuanyue", "limit": 1, "persist": True, "created_by": "tester"},
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["persisted_asset_id"] is not None
    assert data["persisted_asset_version"] == 1

    detail_response = await asset_client.get(
        "/api/v1/assets/reference_content_elements/yuanyue",
        params={"asset_stage": "candidate"},
    )
    assert detail_response.status_code == 200
    asset = detail_response.json()["data"]
    assert asset["asset_stage"] == "candidate"
    assert asset["asset_type"] == "reference_content_elements"
    assert asset["content_json"]["items"][0]["element_source"] == "rules_v1"


@pytest.mark.asyncio
async def test_list_assets_and_get_latest_asset(asset_client):
    list_response = await asset_client.get("/api/v1/assets", params={"asset_key": "yuanyue"})
    assert list_response.status_code == 200
    list_payload = list_response.json()
    assert list_payload["code"] == 200
    listed = list_payload["data"]
    compliance_asset = next(item for item in listed if item["asset_type"] == "compliance_rules")
    assert compliance_asset["asset_key"] == "yuanyue"

    get_response = await asset_client.get("/api/v1/assets/compliance_rules/yuanyue")
    assert get_response.status_code == 200
    get_payload = get_response.json()
    assert get_payload["code"] == 200
    asset = get_payload["data"]
    assert asset["version_no"] == 1
    assert asset["content_json"]["items"][0]["dimension"] == "不得宣称治疗便秘"


@pytest.mark.asyncio
async def test_asset_summary_and_import_runs(asset_client):
    summary_response = await asset_client.get("/api/v1/assets/summary", params={"asset_key": "yuanyue"})
    assert summary_response.status_code == 200
    summary_payload = summary_response.json()
    assert summary_payload["code"] == 200
    data = summary_payload["data"]
    compliance = next(item for item in data if item["asset_type"] == "compliance_rules")
    assert compliance["item_count"] == 1
    assert compliance["version_no"] == 1

    runs_response = await asset_client.get("/api/v1/assets/import-runs")
    assert runs_response.status_code == 200
    runs_payload = runs_response.json()
    assert runs_payload["code"] == 200
    runs = runs_payload["data"]
    assert runs[0]["source_name"] == "源悦种草活动-ai训练规则.xlsx"
    assert runs[0]["imported_assets"] == 3


@pytest.mark.asyncio
async def test_generation_options_are_extracted_from_uploaded_assets(asset_client):
    response = await asset_client.get("/api/v1/assets/generation-options", params={"asset_key": "yuanyue"})
    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    data = payload["data"]
    assert data["asset_keys"] == ["other-brand", "topic-brand", "yuanyue"]
    assert "肠胃弱/奶量上不去" in data["product_topics"]
    assert "AI扩展候选痛点" not in data["product_topics"]
    assert "历史错误痛点" not in data["product_topics"]
    assert "奶量上不去" not in data["product_topics"]
    assert "便便不规律" not in data["product_topics"]
    assert "吃" not in data["product_topics"]
    assert "转奶期宝宝" in data["target_audiences"]
    assert "新手妈妈别急着焦虑" not in data["target_audiences"]
    assert "新手妈妈" not in data["target_audiences"]
    assert "高质量真实用户ugc" not in data["styles"]
    assert "其他品牌风格" not in data["styles"]
    assert "用后分享" not in data["styles"]
    assert "吃" not in data["styles"]
    assert "长" not in data["styles"]
    assert "经验复盘" in data["styles"]
    assert "情绪共情" in data["styles"]
    assert data["persona_profiles"] == []

    topic_response = await asset_client.get("/api/v1/assets/generation-options", params={"asset_key": "topic-brand"})
    assert topic_response.status_code == 200
    topic_payload = topic_response.json()
    assert topic_payload["code"] == 200
    assert topic_payload["data"]["product_topics"] == ["睡眠倒退"]
    assert topic_payload["data"]["target_audiences"] == ["新手妈妈"]


@pytest.mark.asyncio
async def test_asset_stage_can_list_candidate_assets_without_polluting_production(asset_client):
    production_response = await asset_client.get("/api/v1/assets/summary", params={"asset_key": "yuanyue"})
    production_payload = production_response.json()
    assert production_payload["code"] == 200
    assert all(item["asset_stage"] == "production" for item in production_payload["data"])

    all_stage_response = await asset_client.get(
        "/api/v1/assets/summary",
        params={"asset_key": "yuanyue", "asset_stage": ""},
    )
    all_stage_payload = all_stage_response.json()
    assert all_stage_payload["code"] == 200
    assert {"production", "candidate"}.issubset({item["asset_stage"] for item in all_stage_payload["data"]})

    candidate_response = await asset_client.get(
        "/api/v1/assets/summary",
        params={"asset_key": "yuanyue", "asset_stage": "candidate"},
    )
    candidate_payload = candidate_response.json()
    assert candidate_payload["code"] == 200
    candidates = candidate_payload["data"]
    assert all(item["asset_stage"] == "candidate" for item in candidates)
    assert any(item["asset_type"] == "painpoint_model" for item in candidates)

    candidate_detail_response = await asset_client.get(
        "/api/v1/assets/painpoint_model/yuanyue",
        params={"asset_stage": "candidate"},
    )
    candidate_detail = candidate_detail_response.json()["data"]
    assert candidate_detail["asset_stage"] == "candidate"
    assert candidate_detail["content_json"]["topics"][0]["topic"] == "AI扩展候选痛点"


@pytest.mark.asyncio
async def test_painpoint_expression_candidate_detail_auto_classifies_missing_symptom(asset_client):
    response = await asset_client.get(
        "/api/v1/assets/painpoint_expression_candidates/yuanyue",
        params={"asset_stage": "candidate"},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    items = payload["data"]["content_json"]["items"]
    ai_item = next(item for item in items if item["expression_source"] == "ai_expanded")
    assert ai_item["symptom"] == "羊屎蛋/干硬"
    assert ai_item["symptom_source"] == "auto_inferred"


@pytest.mark.asyncio
async def test_create_candidate_asset_api_keeps_generation_options_production_only(asset_client):
    response = await asset_client.post(
        "/api/v1/assets/candidates",
        json={
            "asset_type": "painpoint_model",
            "asset_key": "yuanyue",
            "display_name": "源悦候选痛点描述",
            "source_name": "seed-expansion",
            "source_uri": "file:///tmp/source.xlsx",
            "source_hash": "hash-1",
            "content_json": {
                "items": [
                    {
                        "candidate_id": "c1",
                        "topic": "AI候选新主题",
                        "expression": "宝宝状态忽高忽低，妈妈只能每天细看变化。",
                    }
                ]
            },
            "metadata_json": {"source_kind": "ai_seed_expansion"},
            "created_by": "codex",
        },
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    created = payload["data"]
    assert created["asset_stage"] == "candidate"
    assert created["version_no"] == 4

    candidate_response = await asset_client.get(
        "/api/v1/assets/summary",
        params={"asset_key": "yuanyue", "asset_stage": "candidate"},
    )
    candidates = candidate_response.json()["data"]
    painpoint_candidates = [item for item in candidates if item["asset_type"] == "painpoint_model"]
    assert len(painpoint_candidates) == 1
    assert painpoint_candidates[0]["version_no"] == 4

    options_response = await asset_client.get("/api/v1/assets/generation-options", params={"asset_key": "yuanyue"})
    options = options_response.json()["data"]
    assert "AI候选新主题" not in options["product_topics"]


@pytest.mark.asyncio
async def test_asset_steward_creates_request_proposal_and_applies_new_asset_version(asset_client):
    request_response = await asset_client.post(
        "/api/v1/assets/change-requests",
        json={
            "source_text": "新增宝宝便便不规律方向，避免治疗便秘",
            "requester": "运营A",
            "context_json": {"asset_key": "yuanyue"},
        },
    )
    assert request_response.status_code == 200
    request_payload = request_response.json()
    assert request_payload["code"] == 200
    request_data = request_payload["data"]
    assert request_data["status"] == "pending"

    proposal_response = await asset_client.post(
        "/api/v1/assets/change-proposals",
        json={
            "request_id": request_data["id"],
            "risk_level": "high",
            "summary": "补充便便不规律表达规则",
            "affected_assets_json": [{"asset_type": "compliance_rules", "asset_key": "yuanyue"}],
            "proposed_changes_json": {
                "assets": [
                    {
                        "asset_type": "compliance_rules",
                        "asset_key": "yuanyue",
                        "display_name": "源悦审核规则",
                        "content_json": {"items": [{"dimension": "禁止治疗便秘、改善便秘等医疗化表述"}]},
                    }
                ]
            },
            "risk_notes_json": ["母婴消化相关默认高风险，需要人工确认"],
            "smoke_test_json": {
                "product_topic": "宝宝便便不规律",
                "target_audience": "新手妈妈",
                "style": "经验老道型",
            },
        },
    )
    assert proposal_response.status_code == 200
    proposal_payload = proposal_response.json()
    assert proposal_payload["code"] == 200
    proposal_data = proposal_payload["data"]
    assert proposal_data["status"] == "proposed"

    apply_response = await asset_client.post(f"/api/v1/assets/change-proposals/{proposal_data['id']}/apply")
    assert apply_response.status_code == 200
    apply_payload = apply_response.json()
    assert apply_payload["code"] == 200
    applied = apply_payload["data"]
    assert applied["status"] == "applied"
    assert applied["created_asset_ids"]

    latest_response = await asset_client.get("/api/v1/assets/compliance_rules/yuanyue")
    latest_payload = latest_response.json()
    assert latest_payload["code"] == 200
    latest = latest_payload["data"]
    assert latest["version_no"] == 2
    assert latest["content_json"]["items"][0]["dimension"] == "禁止治疗便秘、改善便秘等医疗化表述"


@pytest.mark.asyncio
async def test_change_request_can_generate_and_apply_candidate_compliance_rule(asset_client):
    request_response = await asset_client.post(
        "/api/v1/assets/change-requests",
        json={
            "source_text": "运营反馈：源悦和 a2 蛋白、a2 公司没有关系。禁止提及 a2 蛋白。",
            "requester": "ops",
            "context_json": {
                "asset_key": "yuanyue",
                "detected_terms": ["a2", "蛋白", "公司没有关系"],
                "item_no": 1,
            },
        },
    )
    request_data = request_response.json()["data"]

    list_response = await asset_client.get("/api/v1/assets/change-requests", params={"limit": 5})
    list_data = list_response.json()["data"]
    assert list_data[0]["id"] == request_data["id"]
    assert list_data[0]["status"] == "pending"

    proposal_response = await asset_client.post(
        f"/api/v1/assets/change-requests/{request_data['id']}/propose-compliance-rule"
    )
    assert proposal_response.status_code == 200
    proposal = proposal_response.json()["data"]
    assert proposal["status"] == "proposed"
    assert proposal["risk_level"] == "high"
    proposed_asset = proposal["proposed_changes_json"]["assets"][0]
    assert proposed_asset["asset_type"] == "compliance_rules"
    assert proposed_asset["asset_stage"] == "candidate"
    assert proposed_asset["content_json"]["items"][-1]["forbidden_terms"] == ["a2", "a2蛋白", "a2公司"]

    proposals_response = await asset_client.get("/api/v1/assets/change-proposals", params={"limit": 5})
    proposals = proposals_response.json()["data"]
    assert proposals[0]["id"] == proposal["id"]

    updated_requests_response = await asset_client.get("/api/v1/assets/change-requests", params={"limit": 5})
    assert updated_requests_response.json()["data"][0]["status"] == "proposed"

    apply_response = await asset_client.post(f"/api/v1/assets/change-proposals/{proposal['id']}/apply")
    assert apply_response.status_code == 200
    applied = apply_response.json()["data"]
    assert applied["status"] == "applied"

    candidate_response = await asset_client.get(
        "/api/v1/assets/compliance_rules/yuanyue",
        params={"asset_stage": "candidate"},
    )
    candidate = candidate_response.json()["data"]
    assert candidate["asset_stage"] == "candidate"
    assert candidate["content_json"]["items"][-1]["rule_type"] == "forbidden_product_fact"
    assert "a2蛋白" in candidate["content_json"]["items"][-1]["forbidden_terms"]


@pytest.mark.asyncio
async def test_upload_yuanyue_training_rules_imports_assets(asset_client, tmp_path):
    from openpyxl import Workbook

    workbook_path = tmp_path / "源悦种草活动-ai训练规则.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "品牌资料整理"
    ws["C3"] = "好消化易吸收"
    ws["C4"] = "高质量真实用户ugc"
    ws["C9"] = "好消化易吸收"

    ws2 = wb.create_sheet("内容模型")
    ws2.append(["序号", "宝宝阶段", "核心痛点", "具体表现", "痛点描述", "对应卖点"])
    ws2.append([None, None, "便便不规律", "羊屎蛋/干硬", "便便又干又硬", "好消化易吸收"])

    ws3 = wb.create_sheet("ugc常规-卖点表述")
    ws3.append(["序号", "对应卖点", "卖点描述", "负责人"])
    ws3.append([None, "便便不规律", "常规表述", "东昕"])

    ws4 = wb.create_sheet("审核规则")
    ws4.append(["序号", "审核内容", "分类", "审核维度（问题分类）", "审核意见（返回给用户的）"])
    ws4.append([1, "文案审核", "草稿审核", "夸大产品效果或虚构使用经历", "文本不符合活动要求"])
    wb.save(workbook_path)

    with workbook_path.open("rb") as file_obj:
        response = await asset_client.post(
            "/api/v1/assets/imports/yuanyue-training-rules",
            data={"asset_key": "yuanyue", "created_by": "ops"},
            files={"file": ("源悦种草活动-ai训练规则.xlsx", file_obj, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    data = payload["data"]
    assert data["imported_assets"] >= 5
    assert ["ugc_expression_corpus", "yuanyue"] in data["asset_keys"]
