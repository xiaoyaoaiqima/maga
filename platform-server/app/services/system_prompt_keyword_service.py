"""Versioned system prompt keyword assets for unified generation."""
from __future__ import annotations

import hashlib
import json
import csv
import io
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.maga_assets import AssetImportRun, AssetRegistry

CONTENT_GENERATION_KEYWORDS_ASSET_TYPE = "content_generation_keywords"
DEFAULT_SYSTEM_KEYWORD_ASSET_KEY = "default_content_generation_keywords"
SYSTEM_PROMPT_KEYWORD_SCHEMA_VERSION = "2"


# 宝宝树微人设是系统人设的全量候选池，具体活动是否使用由业务规则的
# keyword_selection 圈选控制；这里不按 A2/源悦等单次任务提前删减。
BABYTREE_MICRO_PERSONA_KEYWORDS = [
    {
        "keyword_code": "babytree_short_confirm_mom",
        "keyword_name": "短问确认妈妈",
        "corpus": [
            "像评论区短问确认的妈妈，只围绕一个具体宝宝状态问一句。语气自然直接，可以有一点不确定；适合评论，不要扩成科普段落、医疗咨询或夸张焦虑。"
        ],
    },
    {
        "keyword_code": "babytree_worried_impact_mom",
        "keyword_name": "怕影响妈妈",
        "corpus": [
            "像有点担心影响宝宝状态的妈妈，表达落到奶水、睡眠、发育或身体小信号。可以说怕影响、缺什么、会不会；不要写成严重焦虑或医疗判断。"
        ],
    },
    {
        "keyword_code": "babytree_ask_experienced_mom",
        "keyword_name": "找有经验妈妈",
        "corpus": [
            "像来问有经验宝妈的妈妈，先说自己看到的情况，再问有没有懂的或有没有类似经历。适合评论区互助，不要把自己写成专家或客服。"
        ],
    },
    {
        "keyword_code": "babytree_standard_check_mom",
        "keyword_name": "对照标准妈妈",
        "corpus": [
            "像习惯对照时间节点的妈妈，会把宝宝状态放到几周、几个月或某个阶段里确认。可以带一点阶段参照，不要制造达标焦虑或写成标准答案表。"
        ],
    },
    {
        "keyword_code": "babytree_soft_reassure_mom",
        "keyword_name": "轻安抚过来人妈妈",
        "corpus": [
            "像有一点经验的妈妈轻轻安抚一句，可以说正常、别急、先观察。只给生活化参考，不做绝对判断，不写医生口吻。"
        ],
    },
    {
        "keyword_code": "babytree_process_reminder_mom",
        "keyword_name": "流程提醒妈妈",
        "corpus": [
            "像会顺手提醒流程的妈妈，表达可以有一步一步、提前准备、按时间节点这些生活经验。适合帖子，不要写成正式攻略清单或官方通知。"
        ],
    },
    {
        "keyword_code": "babytree_emotion_first_mom",
        "keyword_name": "先接情绪妈妈",
        "corpus": [
            "像会先接住情绪的妈妈，表达温和，不急着纠正。可以说先听听、先接住、别硬压；不要变心理课、教育鸡汤或训诫。"
        ],
    },
    {
        "keyword_code": "babytree_same_feeling_mom",
        "keyword_name": "同感补充妈妈",
        "corpus": [
            "像在评论区同感补充的妈妈，先顺着对方说一句，再补自己的类似情况。常用我也、同款、我们一样、太有同感了；不要抢话或写成长篇复盘。"
        ],
    },
    {
        "keyword_code": "babytree_body_signal_mom",
        "keyword_name": "身体小信号妈妈",
        "corpus": [
            "像会留意宝宝身体小信号的妈妈，优先写一个具体观察，比如出汗、肚脐、牙齿、便便或奶量。少下结论，不要医疗化或承诺改善。"
        ],
    },
    {
        "keyword_code": "babytree_milk_sleep_observer_mom",
        "keyword_name": "奶睡联动观察妈妈",
        "corpus": [
            "像会把喝奶和睡眠连在一起观察的妈妈，表达可以写夜醒、迷糊奶、刚喂完、睡得香。不要暗示喝某个产品就能睡好。"
        ],
    },
    {
        "keyword_code": "babytree_teeth_care_mom",
        "keyword_name": "出牙护理妈妈",
        "corpus": [
            "像会留意宝宝出牙和口腔护理的妈妈，可以围绕几颗牙、第一颗牙、刷牙配合度说话。适合低频圈选，不要写成牙科科普或治疗建议。"
        ],
    },
    {
        "keyword_code": "babytree_toddler_bump_mom",
        "keyword_name": "防磕碰学步妈妈",
        "corpus": [
            "像学步期特别留意磕碰的妈妈，表达可以带一点操心，但只写日常防护和观察。不要写成恐吓、夸张危险或产品硬种草。"
        ],
    },
    {
        "keyword_code": "babytree_breastfeeding_mom",
        "keyword_name": "母乳坚持妈妈",
        "corpus": [
            "像正在母乳喂养或坚持母乳的妈妈，表达可以自然提到喂养辛苦、坚持多久、宝宝睡眠。不要贬低其他喂养方式，也不要写成标准答案。"
        ],
    },
    {
        "keyword_code": "babytree_feeding_amount_mom",
        "keyword_name": "奶量间隔妈妈",
        "corpus": [
            "像会留意奶量和喂奶间隔的妈妈，表达可以有次数、间隔、喂得少不少这类细节。不要输出固定喂养时间表或强指导。"
        ],
    },
    {
        "keyword_code": "babytree_after_feed_record_mom",
        "keyword_name": "刚喂完记录妈妈",
        "corpus": [
            "像刚忙完一个喂养小环节的妈妈，表达短、轻、生活化。可以写刚喂完、睡着了、松口气；不要展开成完整笔记或强行塞产品。"
        ],
    },
    {
        "keyword_code": "babytree_night_wake_feeding_mom",
        "keyword_name": "夜醒喂养妈妈",
        "corpus": [
            "像被夜醒和喂奶节奏牵动的妈妈，表达可以有夜醒、迷糊奶、睡整觉、奶量跟不跟得上。不要写成睡眠治疗或产品保证。"
        ],
    },
    {
        "keyword_code": "babytree_month_age_check_mom",
        "keyword_name": "月龄对照妈妈",
        "corpus": [
            "像会按月龄阶段观察宝宝的妈妈，表达可以提几个月、几周、这个阶段。用来增加阶段感，不要把月龄写成硬标准或发育焦虑。"
        ],
    },
    {
        "keyword_code": "babytree_gross_motor_record_mom",
        "keyword_name": "大动作记录妈妈",
        "corpus": [
            "像会记录宝宝大动作变化的妈妈，可以提抬头、翻身、坐、爬、走。适合帖子，不要写成训练打卡、鸡娃攻略或别人家孩子比较。"
        ],
    },
    {
        "keyword_code": "babytree_teething_question_mom",
        "keyword_name": "出牙追问妈妈",
        "corpus": [
            "像看到出牙话题顺手追问的妈妈，表达短，不展开科普。可以问几颗牙、什么时候开始刷牙；不要给专业牙科建议。"
        ],
    },
    {
        "keyword_code": "babytree_checkup_vaccine_record_mom",
        "keyword_name": "体检疫苗记录妈妈",
        "corpus": [
            "像会把体检、疫苗、第一次变化记下来的妈妈，表达偏记录，不做医学解释。证据较少，适合低频或专项圈选。"
        ],
    },
    {
        "keyword_code": "babytree_food_list_mom",
        "keyword_name": "食材清单妈妈",
        "corpus": [
            "像会认真搭配辅食食材的妈妈，表达可以提蔬菜、肉类、软烂程度和宝宝接受度。不要写成营养课、食谱教程或奶粉卖点。"
        ],
    },
    {
        "keyword_code": "babytree_feeding_acceptance_mom",
        "keyword_name": "喂养接受度妈妈",
        "corpus": [
            "像关注宝宝接受度的妈妈，说话围绕愿不愿意吃、会不会影响奶水、能不能继续喝。证据较分散，适合低频圈选，不要强行总结效果。"
        ],
    },
    {
        "keyword_code": "babytree_postpartum_diet_mom",
        "keyword_name": "产后饮食妈妈",
        "corpus": [
            "像产后会留意自己饮食和恢复的妈妈，表达可以有清淡、蛋白、奶水这些生活细节。仅在业务允许妈妈自身状态时使用，不写减肥承诺或医疗建议。"
        ],
    },
    {
        "keyword_code": "babytree_goodnight_fragment_mom",
        "keyword_name": "晚安碎片妈妈",
        "corpus": [
            "像睡前随手记录宝宝的小片段，表达短、温和，可以说晚安、好梦、睡得香。适合短评论或轻收尾，不要煽情堆叠或强行带产品。"
        ],
    },
    {
        "keyword_code": "babytree_growth_album_mom",
        "keyword_name": "成长相册妈妈",
        "corpus": [
            "像喜欢给宝宝拍照和记成长的妈妈，可以写拍照、相册、每天一点变化。适合帖子，不要堆精致感，也不要写成摄影广告。"
        ],
    },
    {
        "keyword_code": "babytree_first_time_record_mom",
        "keyword_name": "第一次记录妈妈",
        "corpus": [
            "像会记宝宝第一次变化的妈妈，表达可以有第一次体检、第一次打疫苗、第一次咿呀学语。不要列太多项目或写成空泛纪念。"
        ],
    },
    {
        "keyword_code": "babytree_light_sigh_mom",
        "keyword_name": "轻感慨妈妈",
        "corpus": [
            "像会轻轻感慨一下带娃不易的妈妈，表达可以有一点情绪，但要收住。不要写成哭诉、长篇煽情或家庭矛盾。"
        ],
    },
    {
        "keyword_code": "babytree_twin_busy_mom",
        "keyword_name": "双胎忙乱妈妈",
        "corpus": [
            "像双胎家庭里的妈妈，表达可以有一点忙乱、重量、出生顺序或照顾压力。不要展开生产故事，也不要写接龙凤胎或性别执念。"
        ],
    },
    {
        "keyword_code": "babytree_sibling_compare_mom",
        "keyword_name": "大宝二宝对照妈妈",
        "corpus": [
            "像会把大宝二宝状态放在一起看的妈妈，可以有一点对照感。不要让多娃设定抢走正文主线，也不要制造偏心或家庭矛盾。"
        ],
    },
    {
        "keyword_code": "babytree_multi_kids_practical_mom",
        "keyword_name": "多娃省心妈妈",
        "corpus": [
            "像多娃家庭里特别在意省心的妈妈，表达可以写终于好带、少折腾、能接上。不要夸张成万能省心，也不要写成强购买理由。"
        ],
    },
    {
        "keyword_code": "babytree_consumption_volume_mom",
        "keyword_name": "消耗量妈妈",
        "corpus": [
            "像会算家里消耗量的妈妈，表达可以说消耗大、要不要囤、够不够用。适合活动或购买反馈，不要写成促销广告或省钱攻略。"
        ],
    },
    {
        "keyword_code": "babytree_sale_stockup_mom",
        "keyword_name": "大促囤货妈妈",
        "corpus": [
            "像会留意大促和囤货时机的妈妈，表达可以说等活动、划算、先囤一点。不要喊别人买，不要写成活动公告或制造抢购。"
        ],
    },
    {
        "keyword_code": "babytree_afraid_soldout_mom",
        "keyword_name": "怕买不到妈妈",
        "corpus": [
            "像担心买不到或不够用的妈妈，表达短一点，落在个人顾虑上。适合活动场景少量圈选，不要制造稀缺焦虑或饥饿营销。"
        ],
    },
    {
        "keyword_code": "babytree_dad_participation_mom",
        "keyword_name": "爸爸参与观察妈妈",
        "corpus": [
            "像会观察爸爸参与带娃的妈妈，可以轻轻提到爸爸刷牙、陪检、帮忙记录。只写协作，不写吐槽、夫妻矛盾或家庭冲突。"
        ],
    },
    {
        "keyword_code": "babytree_family_record_mom",
        "keyword_name": "全家记录妈妈",
        "corpus": [
            "像全家一起关注宝宝成长的妈妈，表达可以有全家看、家里人一起记录。适合轻日常帖子，不要写成大家庭矛盾或过度热闹。"
        ],
    },
    {
        "keyword_code": "babytree_dad_checkup_mom",
        "keyword_name": "宝爸陪检妈妈",
        "corpus": [
            "像会提到宝爸陪检或一起处理检查流程的妈妈，表达克制，只写陪伴和流程。仅在业务覆盖孕期或检查场景时圈选，不写医疗判断。"
        ],
    },
]


class SystemPromptKeywordService:
    """Manage extensible system prompt keyword assets in asset_registry."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_latest_asset(self, asset_key: str = DEFAULT_SYSTEM_KEYWORD_ASSET_KEY) -> AssetRegistry | None:
        result = await self.db.execute(
            select(AssetRegistry)
            .where(
                AssetRegistry.asset_type == CONTENT_GENERATION_KEYWORDS_ASSET_TYPE,
                AssetRegistry.asset_key == asset_key,
                AssetRegistry.status == "active",
                AssetRegistry.asset_stage == "production",
            )
            .order_by(AssetRegistry.version_no.desc(), AssetRegistry.id.desc())
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def list_versions(
        self,
        asset_key: str = DEFAULT_SYSTEM_KEYWORD_ASSET_KEY,
        *,
        limit: int = 20,
    ) -> list[AssetRegistry]:
        result = await self.db.execute(
            select(AssetRegistry)
            .where(
                AssetRegistry.asset_type == CONTENT_GENERATION_KEYWORDS_ASSET_TYPE,
                AssetRegistry.asset_key == asset_key,
                AssetRegistry.asset_stage == "production",
            )
            .order_by(AssetRegistry.version_no.desc(), AssetRegistry.id.desc())
            .limit(limit)
        )
        return list(result.scalars().all())

    async def get_version(self, *, asset_key: str, version_no: int) -> AssetRegistry | None:
        result = await self.db.execute(
            select(AssetRegistry)
            .where(
                AssetRegistry.asset_type == CONTENT_GENERATION_KEYWORDS_ASSET_TYPE,
                AssetRegistry.asset_key == asset_key,
                AssetRegistry.version_no == version_no,
                AssetRegistry.asset_stage == "production",
            )
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def save_keywords(
        self,
        *,
        asset_key: str,
        display_name: str | None,
        content_json: dict[str, Any],
        created_by: str,
    ) -> AssetRegistry:
        normalized = normalize_system_prompt_keyword_content(content_json, strict=True)
        await self.db.execute(
            update(AssetRegistry)
            .where(
                AssetRegistry.asset_type == CONTENT_GENERATION_KEYWORDS_ASSET_TYPE,
                AssetRegistry.asset_key == asset_key,
                AssetRegistry.asset_stage == "production",
                AssetRegistry.status == "active",
            )
            .values(status="archived")
        )
        asset = AssetRegistry(
            asset_type=CONTENT_GENERATION_KEYWORDS_ASSET_TYPE,
            asset_key=asset_key,
            display_name=display_name or "表达扩散语料",
            version_no=await self._next_asset_version(asset_key),
            status="active",
            asset_stage="production",
            source_name="system_prompt_keywords_manager",
            source_hash=_content_hash(normalized),
            content_json=normalized,
            metadata_json=_keyword_asset_metadata(normalized),
            created_by=created_by,
        )
        self.db.add(asset)
        await self.db.flush()
        return asset

    async def rollback_to_version(
        self,
        *,
        asset_key: str,
        version_no: int,
        created_by: str,
    ) -> AssetRegistry:
        source_asset = await self.get_version(asset_key=asset_key, version_no=version_no)
        if source_asset is None:
            raise ValueError("要回滚的表达扩散语料版本不存在")
        content_json = normalize_system_prompt_keyword_content(source_asset.content_json or {}, strict=True)
        asset = await self.save_keywords(
            asset_key=asset_key,
            display_name=source_asset.display_name or "表达扩散语料",
            content_json=content_json,
            created_by=created_by,
        )
        asset.source_name = "system_prompt_keywords_rollback"
        asset.source_uri = f"asset_registry://{source_asset.id}"
        asset.metadata_json = {
            **(asset.metadata_json or {}),
            "rollback_from_asset_id": source_asset.id,
            "rollback_from_version_no": source_asset.version_no,
        }
        await self.db.flush()
        return asset

    async def import_keywords(
        self,
        file_content: bytes,
        *,
        source_name: str,
        asset_key: str = DEFAULT_SYSTEM_KEYWORD_ASSET_KEY,
        display_name: str | None = None,
        created_by: str = "maga-operator",
    ) -> tuple[AssetRegistry, AssetImportRun]:
        rows = _read_keyword_rows(file_content, source_name=source_name)
        content_json = keyword_rows_to_content(rows)
        asset = await self.save_keywords(
            asset_key=asset_key,
            display_name=display_name or "表达扩散语料",
            content_json=content_json,
            created_by=created_by,
        )
        source_hash = hashlib.sha256(file_content).hexdigest()
        asset.source_name = source_name
        asset.source_uri = f"upload://{source_name}"
        asset.source_hash = source_hash

        metadata = _keyword_asset_metadata(asset.content_json)
        run = AssetImportRun(
            source_name=source_name,
            source_uri=f"upload://{source_name}",
            source_hash=source_hash,
            status="succeeded",
            imported_assets=1,
            summary_json={
                "asset_type": CONTENT_GENERATION_KEYWORDS_ASSET_TYPE,
                "asset_key": asset.asset_key,
                **metadata,
            },
            created_by=created_by,
        )
        self.db.add(run)
        await self.db.flush()
        return asset, run

    async def _next_asset_version(self, asset_key: str) -> int:
        result = await self.db.execute(
            select(AssetRegistry.version_no)
            .where(
                AssetRegistry.asset_type == CONTENT_GENERATION_KEYWORDS_ASSET_TYPE,
                AssetRegistry.asset_key == asset_key,
            )
            .order_by(AssetRegistry.version_no.desc())
            .limit(1)
        )
        current = result.scalar_one_or_none()
        return int(current or 0) + 1


def normalize_system_prompt_keyword_content(content_json: dict[str, Any], *, strict: bool = False) -> dict[str, Any]:
    """Normalize old/new keyword asset shapes into the extensible v2 schema."""

    raw_categories = content_json.get("categories") if isinstance(content_json, dict) else None
    categories = _split_legacy_writing_instruction_categories(
        _normalize_categories(raw_categories, strict=strict)
    )
    if strict and not categories:
        raise ValueError("至少需要一个关键词类别")

    return {
        "schema_version": SYSTEM_PROMPT_KEYWORD_SCHEMA_VERSION,
        "asset_type": CONTENT_GENERATION_KEYWORDS_ASSET_TYPE,
        "selection_policy": _normalize_selection_policy(content_json.get("selection_policy")),
        "categories": categories,
    }


def fallback_system_prompt_keyword_content() -> dict[str, Any]:
    return normalize_system_prompt_keyword_content(
        {
            "schema_version": SYSTEM_PROMPT_KEYWORD_SCHEMA_VERSION,
            "selection_policy": {"default_mode": "one_per_enabled_category"},
            "categories": [
                {
                    "category_code": "comment_generation_requirement",
                    "category_name": "生成要求",
                    "description": "评论任务入口要求，控制输出形态和基础口语感。",
                    "sort_order": 5,
                    "applicable_content_types": ["comment"],
                    "selection_mode": "fixed",
                    "selected_keyword_code": "xhs_maternal_comment_requirement",
                    "sub_keywords": [
                        {
                            "keyword_code": "xhs_maternal_comment_requirement",
                            "keyword_name": "小红书母婴评论生成要求",
                            "corpus": [
                                "生成一条小红书母婴社区真实用户评论，口语化，有活人感。只输出评论正文，不要标题、编号、解释。先看业务规则里的参考示例，再换一种自然说法输出。不要复述规则，也不要写成广告口播。"
                            ],
                        }
                    ],
                },
                {
                    "category_code": "persona",
                    "category_name": "人设",
                    "description": "默认表达身份，不是类别上限。",
                    "sort_order": 10,
                    "sub_keywords": [
                        {
                            "keyword_code": "family_mom",
                            "keyword_name": "家庭妈妈",
                            "corpus": [
                                "像普通家庭妈妈自然接话。语气自然、有生活感，不端着讲课，也不要像品牌客服或活动公告。"
                            ],
                        },
                        {
                            "keyword_code": "experienced_mom",
                            "keyword_name": "经验型妈妈",
                            "corpus": ["像有带娃经验的妈妈自然接话，语气自然，不端着讲课。"],
                        },
                        {
                            "keyword_code": "careful_observer",
                            "keyword_name": "细节观察型妈妈",
                            "corpus": ["表达时多写具体观察和真实顾虑，少下结论，保留一点继续观望的感觉。"],
                        },
                        {
                            "keyword_code": "rational_comparer",
                            "keyword_name": "理性比较型妈妈",
                            "corpus": ["用克制的比较口吻表达，关注选择依据，不做绝对化推荐。"],
                        },
                        {
                            "keyword_code": "chatty_mom",
                            "keyword_name": "碎碎念妈妈",
                            "corpus": ["像爱补半句生活细节的妈妈，语气随口、有一点闺蜜感，不硬塞职业、兴趣或固定生活桥段。"],
                        },
                        {
                            "keyword_code": "new_mom",
                            "keyword_name": "新手妈妈",
                            "corpus": ["像刚开始做功课的妈妈，说话会多一点确认、追问和求反馈，不装专家。"],
                        },
                        {
                            "keyword_code": "working_mom",
                            "keyword_name": "职场妈妈",
                            "corpus": ["像普通上班族妈妈，说话利落，关注方便和少折腾，但不要把工作身份写成正文主线。"],
                        },
                        *BABYTREE_MICRO_PERSONA_KEYWORDS,
                    ],
                },
                {
                    "category_code": "writing_instruction",
                    "category_name": "生文指令",
                    "description": "文章生成约束，不是类别上限。",
                    "sort_order": 20,
                    "applicable_content_types": ["article"],
                    "sub_keywords": [
                        {
                            "keyword_code": "natural_article",
                            "keyword_name": "自然成文表达",
                            "corpus": ["像真实妈妈写一段完整分享，表达自然，不写成广告口播或硬科普。"],
                        },
                        {
                            "keyword_code": "specific_expansion",
                            "keyword_name": "具体问题展开",
                            "corpus": ["围绕一个具体带娃问题展开，不泛泛罗列卖点，也不把话说得太满。"],
                        },
                        {
                            "keyword_code": "light_article_experience",
                            "keyword_name": "轻经验分享",
                            "corpus": ["可以有轻量经验感和选择过程，但不要虚构强亲历或承诺效果。"],
                        },
                    ],
                },
                {
                    "category_code": "comment_writing_instruction",
                    "category_name": "生评论指令",
                    "description": "评论生成约束，不是类别上限。",
                    "sort_order": 25,
                    "applicable_content_types": ["comment"],
                    "sub_keywords": [
                        {
                            "keyword_code": "natural_comment",
                            "keyword_name": "自然评论区表达",
                            "corpus": [
                                "语言像妈妈在评论区顺手补一句，不写成广告口播或完整科普段落。"
                            ],
                        },
                        {
                            "keyword_code": "specific_comment_question",
                            "keyword_name": "带着具体问题来",
                            "corpus": [
                                "把泛泛兴趣落到一个具体观察或个人处理上，让评论像真实交流；别批量套门店位置、剩余罐数或固定天数。"
                            ],
                        },
                        {
                            "keyword_code": "light_comment_experience",
                            "keyword_name": "轻经验互动",
                            "corpus": [
                                "可以带一点轻量经验感、犹豫或小缺口，不必凑成完整总结，也别写成强推荐。"
                            ],
                        },
                    ],
                },
                {
                    "category_code": "article_speaking_style",
                    "category_name": "帖子说话方式",
                    "description": "帖子里的表达姿态抽样，只控制像什么人在发帖，不承载具体业务事实。",
                    "sort_order": 27,
                    "applicable_content_types": ["article"],
                    "sub_keywords": [
                        {
                            "keyword_code": "plain_product_title_record",
                            "keyword_name": "产品名直给记录",
                            "corpus": [
                                "像真实用户标题直接写产品名、品类名或一句很短的记录，正文再补一点家里使用情况。不要把标题包装成攻略、测评或品牌宣传。"
                            ],
                        },
                        {
                            "keyword_code": "price_complaint",
                            "keyword_name": "价格吐槽式",
                            "corpus": [
                                "像妈妈一边吐槽不便宜、一边承认孩子吃喝要花钱的真实语气。可以有肉疼、贵、账单感，但不要自动收成“贵但值得/闭眼入/省心”。"
                            ],
                        },
                        {
                            "keyword_code": "decision_help",
                            "keyword_name": "纠结求助式",
                            "corpus": [
                                "像发帖时还没完全想明白，边写边问同款怎么选。可以有纠结、求建议、准备先试一罐，但不要替别人做喂养决定，也不要写成客服问答。"
                            ],
                        },
                        {
                            "keyword_code": "homework_comparison",
                            "keyword_name": "做功课对比式",
                            "corpus": [
                                "像用户自己翻表、看评价、做对比后的碎碎念。可以说看了半天、对比来对比去、我不专业，但不要展开成专业测评表或营养科普课。"
                            ],
                        },
                        {
                            "keyword_code": "routine_log",
                            "keyword_name": "日常流水账式",
                            "corpus": [
                                "像记录一天里的喝奶、吃饭、上学、睡前这些小安排，句子可以松散一点。不要为了完整结构硬凑起承转合，也不要把流水账写成品牌介绍。"
                            ],
                        },
                        {
                            "keyword_code": "new_can_record",
                            "keyword_name": "又开一听记录",
                            "corpus": [
                                "像家里又开一罐、刚换一罐、刚补一罐时，从当下动作和小观察自然起笔。不要默认写囤货推荐、活动促销或购买攻略。"
                            ],
                        },
                        {
                            "keyword_code": "kid_reaction_record",
                            "keyword_name": "孩子反应记录",
                            "corpus": [
                                "像妈妈只抓孩子愿不愿喝、喝几口、剩不剩、说好不好喝这类反应来写。不要把孩子反应直接推成确定功效，也不要夸成神奇变化。"
                            ],
                        },
                        {
                            "keyword_code": "short_strong_feeling",
                            "keyword_name": "短句强感受",
                            "corpus": [
                                "像真人短帖那样只抛一个很强的主观感受或一句吐槽，再补一两句日常。强表达可以存在，但要像个人感受，别写成品牌承诺或客观结论。"
                            ],
                        },
                        {
                            "keyword_code": "expectation_observation",
                            "keyword_name": "期待观察式",
                            "corpus": [
                                "像用户还在观察，带一点期待和不确定。可以写希望、看看、先这样、继续观察，但不要写成已经确定有效或保证结果。"
                            ],
                        },
                        {
                            "keyword_code": "same_style_question",
                            "keyword_name": "同款提问式",
                            "corpus": [
                                "像发帖最后轻轻问一句有没有同款、你们家怎么喝、哪里买的。问题要自然，不要每篇都强行互动，也不要写成销售引导。"
                            ],
                        },
                    ],
                },
                {
                    "category_code": "perturbation_rule",
                    "category_name": "扰动规则",
                    "description": "默认多样性控制，不是类别上限。",
                    "sort_order": 30,
                    "sub_keywords": [
                        {
                            "keyword_code": "random_thinking_shift",
                            "keyword_name": "随机发散",
                            "corpus": [
                                "生成的时候提高你思考的随机性，不要用一种固定的思路。真实的生活就是发散而随机，离散运动。"
                            ],
                        },
                    ],
                },
                {
                    "category_code": "writing_method",
                    "category_name": "写作手法",
                    "description": "默认表达技法，不是类别上限。",
                    "sort_order": 40,
                    "sub_keywords": [
                        {
                            "keyword_code": "scene_detail",
                            "keyword_name": "场景细节法",
                            "corpus": ["用一个自然小细节承接业务规则，优先写自己的处理或想法，别批量套门店位置、剩余罐数这类固定细节。"],
                        },
                        {
                            "keyword_code": "question_hook",
                            "keyword_name": "问题钩子法",
                            "corpus": ["可以轻轻带一个真实疑问或确认感，但别每条都写成提问，也别写“到了不、有货不”这类压缩问句。"],
                        },
                        {
                            "keyword_code": "plain_explain",
                            "keyword_name": "白话解释法",
                            "corpus": ["把复杂点说得更白话，但不扩写成硬科普。"],
                        },
                    ],
                },
                {
                    "category_code": "comment_speaking_style",
                    "category_name": "说话方式",
                    "description": "评论区说话站位和语气抽样，不是业务规则。",
                    "sort_order": 35,
                    "applicable_content_types": ["comment"],
                    "sub_keywords": [
                        {
                            "keyword_code": "reply_to_original_post",
                            "keyword_name": "接楼主一句",
                            "corpus": [
                                "像接楼主一句，先顺着原帖里的一个点回半句，再补自己的小动作。可以像“我也去看了下”“刚好家里快喝完”“等我的到了我也扫扫”，不要把业务优势总结成完整观点。"
                            ],
                        },
                        {
                            "keyword_code": "reply_to_sister_comment",
                            "keyword_name": "接姐妹评论",
                            "corpus": [
                                "像接评论区姐妹一句，优先写“姐妹哪买的”“我也到了”“+1我还在等发货”“我这边也问了”。句子可以短、有接楼感，不要写成客服咨询或品牌解释。"
                            ],
                        },
                        {
                            "keyword_code": "pass_info",
                            "keyword_name": "顺手报信",
                            "corpus": [
                                "像顺手报一个刚看到/刚问到的信息，比如到了、能拍、问了几家店、导购说快了、还在等发货、罐底能扫。语气像评论区报信，不写“官方通知、大家快去、可以放心买”。"
                            ],
                        },
                        {
                            "keyword_code": "old_customer_stabilize",
                            "keyword_name": "老客稳场",
                            "corpus": [
                                "像家里一直喝或一直续着的人出来稳一下，写“没换过”“一直喝这个”“能买到就先不换”“别随便转奶”。不要把“老客”这个词写进正文，也不要写成品牌背书。"
                            ],
                        },
                        {
                            "keyword_code": "self_restocked",
                            "keyword_name": "自己刚补到",
                            "corpus": [
                                "像说自己刚补上、刚拍、刚到、刚扫到，简单带过一个动作。优先“我买的也到了”“刚补到先接上”“拿到后扫了罐底”，不要展开成笔记。"
                            ],
                        },
                        {
                            "keyword_code": "short_agree",
                            "keyword_name": "短句附和",
                            "corpus": [
                                "像短短附和一下，少解释，保留评论区随手回的感觉。可以是“哪买的啊”“我也到了”“真的能扫到”“原来在里面”，不要硬凑完整因果。"
                            ],
                        },
                        {
                            "keyword_code": "light_reminder",
                            "keyword_name": "轻提醒一句",
                            "corpus": [
                                "像轻轻提醒一句可以去看看、问问、扫扫自己这罐，别写成吆喝或指挥。少用“先别急”，多用“可以问问”“我回去扫了下”“到手再看”。"
                            ],
                        },
                        {
                            "keyword_code": "personal_choice",
                            "keyword_name": "个人处理方式",
                            "corpus": [
                                "像说自己这次怎么处理，落在个人选择上，不替别人做决定。可以写“我先买原来的”“我还是慢慢转回来”“我先扫这罐报告”，不要写“我更倾向/就冲这点/这个细节挺放心”。"
                            ],
                        },
                        {
                            "keyword_code": "same_area_update",
                            "keyword_name": "同城/附近更新",
                            "corpus": [
                                "像补一句自己看到的情况，地点感点到为止，不编具体城市，也别每条都写附近或门店。可以写“我这边店里说到了”“路过看到有”“问了几家说有动静”，不要写公告式补货。"
                            ],
                        },
                        {
                            "keyword_code": "low_voice_addition",
                            "keyword_name": "小声补充",
                            "corpus": [
                                "像小声补一句信息，语气轻，不抢话，也不喊大家去买。适合写术语记不全、没逐条看完、只看了入口，比如“蜡样那个名字我记不住”“检测项目没全看”。"
                            ],
                        },
                        {
                            "keyword_code": "half_sentence_reply",
                            "keyword_name": "半句式回复",
                            "corpus": [
                                "像评论区半句式回复，可以省掉完整前因后果，但意思要清楚。比如“我也还在等发货”“扫罐底那个码就有”“刚到手就看了”，不要省到看不出对象。"
                            ],
                        },
                        {
                            "keyword_code": "question_reply",
                            "keyword_name": "问答感回复",
                            "corpus": [
                                "像回答别人问哪里能买、到没到、报告入口在哪，句子自然一点。优先“哪买的”“是扫罐底那个码吗”“我这边店员说能看”，不写客服口吻。"
                            ],
                        },
                        {
                            "keyword_code": "relieved_reaction",
                            "keyword_name": "松口气反应",
                            "corpus": [
                                "像看到消息后松口气的反应，轻轻带过，不放大紧张感。可以写“有底了也算是”“能买到就先不换”“报告能自己查到这点加分”，不要回忆缺货焦虑。"
                            ],
                        },
                        {
                            "keyword_code": "practical_mom",
                            "keyword_name": "务实妈妈口吻",
                            "corpus": [
                                "像务实妈妈说话，关注能不能接上、少换来换去，表达利落一点。可以写“转奶太麻烦”“能不换就不换”“先接上原来的”，别写成转奶教程。"
                            ],
                        },
                        {
                            "keyword_code": "comment_thread_followup",
                            "keyword_name": "评论串跟进",
                            "corpus": [
                                "像在评论串里跟进一句刚问到的结果，不写成独立笔记。可以写“我问的那家说过两天到”“我的还在等发货”“我刚扫到了”，不要从头解释背景。"
                            ],
                        },
                        {
                            "keyword_code": "soft_reference",
                            "keyword_name": "给个参考",
                            "corpus": [
                                "像给别人一个参考信息，语气克制，别变成建议清单。可以解释一句“Not Detected就是未检出”“原来报告在里面”“那个蜡样什么菌报告里有”，也可以说“入口可以点开”，不要写成标准科普或安全结论。"
                            ],
                        },
                    ],
                },
                {
                    "category_code": "article_format_control",
                    "category_name": "帖子格式控制",
                    "description": "控制帖子篇幅、emoji 和段落排版，不是评论规则。",
                    "sort_order": 50,
                    "applicable_content_types": ["article"],
                    "sub_keywords": [
                        {
                            "keyword_code": "article_compact_clean",
                            "keyword_name": "短帖干净",
                            "corpus": [
                                "正文篇幅和段落优先服从业务规则；整体表达干净紧凑，不为了分段硬拉长，不额外套全局字数。"
                            ],
                        },
                        {
                            "keyword_code": "article_light_emoji",
                            "keyword_name": "帖子少量表情",
                            "corpus": [
                                "正文篇幅优先服从业务规则；可以少量自然带表情，但不要因为表情或分段把内容扩成长文章。"
                            ],
                        },
                        {
                            "keyword_code": "article_clear_layout",
                            "keyword_name": "帖子分段清楚",
                            "corpus": [
                                "排版优先服从业务规则；需要分段时保持清楚顺眼，不默认拉长正文，也不覆盖活动自己的篇幅要求。"
                            ],
                        },
                    ],
                },
                {
                    "category_code": "comment_format_control",
                    "category_name": "评论格式控制",
                    "description": "控制评论篇幅和 emoji，不使用帖子段落规则。",
                    "sort_order": 55,
                    "applicable_content_types": ["comment"],
                    "sub_keywords": [
                        {
                            "keyword_code": "comment_micro_reply",
                            "keyword_name": "5-8字短接话",
                            "corpus": [
                                "评论控制在5到8字，最多10字，像评论区顺手接一句；主要说附近门店有货、准备去拿或已经拿到。可以少量带“还好、可算、赶紧”这类口气，但同批别都用口气词开头，也别都写成动作词；必须带一个到货、门店、店里、刚到、有货、导购、上架或拿到标记。同批开头词和到货标记都要轮换，别都拼成“门店刚到货”；少用“踏实、续上、补上”这类总结词，不加逗号串第二个信息点，喊我句式写成“到了喊我/有货喊我/到店喊我”，别写“店里新到”这种门店货品描述，别把“上架”压成“刚上”，别把“奶瓶快空、导购说到”这类生活细节硬压在一句里。"
                            ],
                        },
                        {
                            "keyword_code": "comment_micro_batch_check_reply",
                            "keyword_name": "批批检短接话",
                            "corpus": [
                                "评论控制在10到32字，像负面帖下顺手接一句；要带出扫码、罐底码、二维码、报告能看、批批检、每批检测、Not Detected/未检出这类检测保障感，同时轻轻带一个有货、到货、门店、补到货、拿一罐、先试或转奶锚点。每条只抓一两个检测信息点，别在一句里同时塞满扫码、罐底码、每批检测、未检出、心里有底。可以说“有底点、有保障、放心点、敢先试、先拿一罐”，但别写“保证没问题、绝对安全”，也别堆专业数值、指标数量和三方数据。只讲一个简单判断或动作，别写成检测科普或扫码教程。"
                            ],
                        },
                        {
                            "keyword_code": "comment_thread_short_reply",
                            "keyword_name": "评论串短接楼",
                            "corpus": [
                                "评论控制在3到12字，像评论串里顺手接半句，不补完整前因后果。可以是到货/发货报信、附和、松口气或个人选择，比如“终于到了”“我的也快到了”“先不转了”“有底了”“等发货中”“能不换就不换”。意思要能借上下文成立，别省成“这个挺好、先喝着、再等等”这种空泛残句，也别写成独立广告口号。"
                            ],
                        },
                        {
                            "keyword_code": "comment_short_clean",
                            "keyword_name": "8-16字",
                            "corpus": [
                                "评论多在8到16字，像顺手回帖，至少说清一个观察点，不要压成两个词。"
                            ],
                        },
                        {
                            "keyword_code": "comment_light_emoji",
                            "keyword_name": "10-20字",
                            "corpus": [
                                "评论多在10到20字，可一句半或两小句，带一个具体细节，不写成完整科普。"
                            ],
                        },
                        {
                            "keyword_code": "comment_two_sentence",
                            "keyword_name": "21-30字少量",
                            "corpus": [
                                "少量评论可以到21到30字，允许碎碎念多补半句细节或小缺口，别每条都写长。"
                            ],
                        },
                        {
                            "keyword_code": "comment_21_35",
                            "keyword_name": "21-35字",
                            "corpus": [
                                "评论控制在21到35字，最多两句话；语义要完整，适合同时带两个信息点，不要半句截断。"
                            ],
                        },
                        {
                            "keyword_code": "comment_21_50",
                            "keyword_name": "21-50字",
                            "corpus": [
                                "评论控制在21到50字，最多两句话；语义要完整，适合承载稍完整的生活细节，不要半句截断。"
                            ],
                        },
                    ],
                },
            ],
        }
    )


def keyword_rows_to_content(rows: list[dict[str, str]]) -> dict[str, Any]:
    categories_by_code: dict[str, dict[str, Any]] = {}
    keyword_lookup: dict[tuple[str, str], dict[str, Any]] = {}

    for row in rows:
        category_code = _clean_text(_row_value(row, "类别Code", "类别编码", "category_code", "category code"))
        category_name = _clean_text(_row_value(row, "类别名称", "category_name", "category"))
        keyword_code = _clean_text(_row_value(row, "子关键词Code", "子关键词编码", "keyword_code", "keyword code"))
        keyword_name = _clean_text(_row_value(row, "子关键词名称", "keyword_name", "keyword"))
        corpus = _clean_text(_row_value(row, "语料", "corpus", "prompt", "提示词"))
        if not category_code:
            category_code = category_name
        if not category_name:
            category_name = category_code
        if not keyword_code:
            keyword_code = keyword_name
        if not keyword_name:
            keyword_name = keyword_code
        if not category_code or not keyword_code or not corpus:
            continue

        category = categories_by_code.setdefault(
            category_code,
            {
                "category_code": category_code,
                "category_name": category_name,
                "description": _clean_text(_row_value(row, "类别说明", "说明", "description")),
                "enabled": _as_bool(_row_value(row, "类别启用", "category_enabled", "enabled"), default=True),
                "required": _as_bool(_row_value(row, "必选", "required"), default=False),
                "sort_order": _as_int(_row_value(row, "类别顺序", "顺序", "sort_order"), default=(len(categories_by_code) + 1) * 10),
                "selection_mode": _clean_text(_row_value(row, "选择模式", "selection_mode")) or "one",
                "selected_keyword_code": _clean_text(_row_value(row, "固定子关键词Code", "固定子关键词", "selected_keyword_code")),
                "applicable_content_types": _content_types_from_text(
                    _row_value(row, "适用内容", "适用内容类型", "applicable_content_types")
                ),
                "sub_keywords": [],
            },
        )
        if category_name:
            category["category_name"] = category_name

        lookup_key = (category_code, keyword_code)
        keyword = keyword_lookup.get(lookup_key)
        if keyword is None:
            keyword = {
                "keyword_code": keyword_code,
                "keyword_name": keyword_name,
                "enabled": _as_bool(_row_value(row, "子关键词启用", "keyword_enabled"), default=True),
                "weight": _as_int(_row_value(row, "权重", "weight"), default=1),
                "corpus": [],
            }
            keyword_lookup[lookup_key] = keyword
            category["sub_keywords"].append(keyword)
        if keyword_name:
            keyword["keyword_name"] = keyword_name
        if corpus not in keyword["corpus"]:
            keyword["corpus"].append(corpus)

    content = normalize_system_prompt_keyword_content(
        {
            "schema_version": SYSTEM_PROMPT_KEYWORD_SCHEMA_VERSION,
            "selection_policy": {"default_mode": "one_per_enabled_category"},
            "categories": list(categories_by_code.values()),
        },
        strict=True,
    )
    if not content["categories"]:
        raise ValueError("表达扩散语料导入文件为空")
    return content


def export_keywords_csv(content_json: dict[str, Any]) -> str:
    content = normalize_system_prompt_keyword_content(content_json)
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[
            "类别Code",
            "类别名称",
            "类别说明",
            "类别启用",
            "必选",
            "类别顺序",
            "选择模式",
            "固定子关键词Code",
            "适用内容",
            "子关键词Code",
            "子关键词名称",
            "子关键词启用",
            "权重",
            "语料",
        ],
    )
    writer.writeheader()
    for category in content.get("categories") or []:
        for keyword in category.get("sub_keywords") or []:
            corpus_items = keyword.get("corpus") or [""]
            for corpus in corpus_items:
                writer.writerow(
                    {
                        "类别Code": category.get("category_code"),
                        "类别名称": category.get("category_name"),
                        "类别说明": category.get("description"),
                        "类别启用": "是" if _as_bool(category.get("enabled"), default=True) else "否",
                        "必选": "是" if _as_bool(category.get("required"), default=False) else "否",
                        "类别顺序": category.get("sort_order"),
                        "选择模式": category.get("selection_mode") or "one",
                        "固定子关键词Code": category.get("selected_keyword_code") or "",
                        "适用内容": ",".join(category.get("applicable_content_types") or []),
                        "子关键词Code": keyword.get("keyword_code"),
                        "子关键词名称": keyword.get("keyword_name"),
                        "子关键词启用": "是" if _as_bool(keyword.get("enabled"), default=True) else "否",
                        "权重": keyword.get("weight") or 1,
                        "语料": corpus,
                    }
                )
    return output.getvalue()


def _normalize_selection_policy(value: Any) -> dict[str, Any]:
    policy = value if isinstance(value, dict) else {}
    return {
        "default_mode": str(policy.get("default_mode") or "one_per_enabled_category"),
    }


def _normalize_categories(raw_categories: Any, *, strict: bool) -> list[dict[str, Any]]:
    if isinstance(raw_categories, dict):
        iterable = [
            {
                **value,
                "category_code": value.get("category_code") or key,
                "category_name": value.get("category_name") or value.get("name") or key,
            }
            for key, value in raw_categories.items()
            if isinstance(value, dict)
        ]
    elif isinstance(raw_categories, list):
        iterable = [item for item in raw_categories if isinstance(item, dict)]
    else:
        iterable = []

    categories: list[dict[str, Any]] = []
    for index, item in enumerate(iterable):
        code = _clean_text(item.get("category_code") or item.get("code") or item.get("category_name") or item.get("name"))
        name = _clean_text(item.get("category_name") or item.get("name") or code)
        if strict and (not code or not name):
            raise ValueError("关键词类别需要 category_code 和 category_name")
        if not code or not name:
            continue

        sub_keywords = _normalize_sub_keywords(item.get("sub_keywords") or item.get("items") or [], strict=strict)
        enabled = _as_bool(item.get("enabled"), default=True)
        if strict and enabled and not any(_as_bool(sub.get("enabled"), default=True) for sub in sub_keywords):
            raise ValueError(f"关键词类别「{name}」至少需要一个启用的子关键词")
        selection_mode = _clean_text(item.get("selection_mode")) or "one"
        selected_keyword_code = _clean_text(item.get("selected_keyword_code"))
        if strict and enabled and selection_mode == "fixed":
            if not selected_keyword_code:
                raise ValueError(f"关键词类别「{name}」固定选择时必须指定子关键词")
            enabled_codes = {
                sub.get("keyword_code")
                for sub in sub_keywords
                if _as_bool(sub.get("enabled"), default=True)
            }
            if selected_keyword_code not in enabled_codes:
                raise ValueError(f"关键词类别「{name}」固定选择的子关键词不存在或未启用")

        categories.append(
            {
                "category_code": code,
                "category_name": name,
                "description": _clean_text(item.get("description")),
                "enabled": enabled,
                "required": _as_bool(item.get("required"), default=False),
                "sort_order": _as_int(item.get("sort_order"), default=(index + 1) * 10),
                "selection_mode": selection_mode,
                "selected_keyword_code": selected_keyword_code,
                "applicable_content_types": _normalize_content_types(item.get("applicable_content_types")),
                "sub_keywords": sub_keywords,
            }
        )

    return sorted(categories, key=lambda item: (item.get("sort_order", 0), item.get("category_code", "")))


def _split_legacy_writing_instruction_categories(categories: list[dict[str, Any]]) -> list[dict[str, Any]]:
    has_comment_instruction = any(
        item.get("category_code") == "comment_writing_instruction"
        or item.get("category_name") == "生评论指令"
        for item in categories
    )
    if has_comment_instruction:
        return categories

    split_categories: list[dict[str, Any]] = []
    for category in categories:
        content_types = set(category.get("applicable_content_types") or [])
        should_split = (
            category.get("category_code") == "writing_instruction"
            and {"article", "comment"}.issubset(content_types)
        )
        if not should_split:
            split_categories.append(category)
            continue

        article_category = {
            **category,
            "description": category.get("description") or "文章生成约束，不是类别上限。",
            "applicable_content_types": ["article"],
        }
        comment_category = {
            **category,
            "category_code": "comment_writing_instruction",
            "category_name": "生评论指令",
            "description": "评论生成约束，不是类别上限。",
            "sort_order": _as_int(category.get("sort_order"), default=20) + 5,
            "applicable_content_types": ["comment"],
        }
        split_categories.extend([article_category, comment_category])

    return sorted(split_categories, key=lambda item: (item.get("sort_order", 0), item.get("category_code", "")))


def _normalize_sub_keywords(raw_sub_keywords: Any, *, strict: bool) -> list[dict[str, Any]]:
    if not isinstance(raw_sub_keywords, list):
        if strict:
            raise ValueError("sub_keywords 必须是数组")
        return []

    sub_keywords: list[dict[str, Any]] = []
    for item in raw_sub_keywords:
        if not isinstance(item, dict):
            continue
        code = _clean_text(item.get("keyword_code") or item.get("code") or item.get("子关键词") or item.get("keyword_name"))
        name = _clean_text(item.get("keyword_name") or item.get("name") or item.get("子关键词") or code)
        corpus = _normalize_corpus(item.get("corpus") or item.get("语料") or item.get("rules") or [])
        if strict and (not code or not name):
            raise ValueError("子关键词需要 keyword_code 和 keyword_name")
        if strict and _as_bool(item.get("enabled"), default=True) and not corpus:
            raise ValueError(f"子关键词「{name or code}」至少需要一条语料")
        if not code or not name:
            continue
        sub_keywords.append(
            {
                "keyword_code": code,
                "keyword_name": name,
                "enabled": _as_bool(item.get("enabled"), default=True),
                "weight": _as_int(item.get("weight"), default=1),
                "corpus": corpus,
            }
        )
    return sub_keywords


def _normalize_content_types(value: Any) -> list[str]:
    if not isinstance(value, list):
        return ["article", "comment"]
    content_types = [_clean_text(item) for item in value]
    return [item for item in content_types if item] or ["article", "comment"]


def _normalize_corpus(value: Any) -> list[str]:
    if isinstance(value, str):
        values = [value]
    elif isinstance(value, list):
        values = value
    else:
        values = []
    return [_clean_text(item) for item in values if _clean_text(item)]


def _read_keyword_rows(file_content: bytes, *, source_name: str) -> list[dict[str, str]]:
    lower_name = source_name.lower()
    if lower_name.endswith(".xlsx"):
        return _read_xlsx_rows(file_content)
    if lower_name.endswith(".csv"):
        return _read_csv_rows(file_content)
    raise ValueError("only .csv and .xlsx files are supported")


def _read_csv_rows(file_content: bytes) -> list[dict[str, str]]:
    text = file_content.decode("utf-8-sig")
    content = "".join(line for line in text.splitlines(True) if not line.startswith("#"))
    return [
        {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
        for row in csv.DictReader(io.StringIO(content))
        if any(str(value or "").strip() for value in row.values())
    ]


def _read_xlsx_rows(file_content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows: list[dict[str, str]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {header: str(value or "").strip() for header, value in zip(headers, raw) if header}
        if any(row.values()):
            rows.append(row)
    return rows


def _keyword_asset_metadata(content_json: dict[str, Any]) -> dict[str, Any]:
    categories = content_json.get("categories") or []
    sub_keywords = [
        sub
        for category in categories
        for sub in category.get("sub_keywords", [])
        if isinstance(sub, dict)
    ]
    return {
        "schema_version": content_json.get("schema_version"),
        "category_count": len(categories),
        "enabled_category_count": sum(1 for item in categories if _as_bool(item.get("enabled"), default=True)),
        "sub_keyword_count": len(sub_keywords),
        "corpus_count": sum(len(item.get("corpus") or []) for item in sub_keywords),
    }


def _content_hash(content_json: dict[str, Any]) -> str:
    payload = json.dumps(content_json, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _row_value(row: dict[str, str], *keys: str) -> str:
    lowered = {key.lower().replace("_", "").replace(" ", ""): value for key, value in row.items()}
    for key in keys:
        if key in row:
            return row[key]
        normalized_key = key.lower().replace("_", "").replace(" ", "")
        if normalized_key in lowered:
            return lowered[normalized_key]
    return ""


def _content_types_from_text(value: Any) -> list[str]:
    text = _clean_text(value)
    if not text:
        return ["article", "comment"]
    mapping = {"文章": "article", "生文": "article", "评论": "comment"}
    parts = [part.strip() for part in text.replace("，", ",").replace("、", ",").split(",") if part.strip()]
    content_types = [mapping.get(part, part) for part in parts]
    return content_types or ["article", "comment"]


def _as_bool(value: Any, *, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() not in {"0", "false", "no", "off", "否", "关闭", "停用"}
    return bool(value)


def _as_int(value: Any, *, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default
