"""Build operator-facing reports for MAGA content batch jobs."""
from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime
from io import BytesIO, StringIO
from types import SimpleNamespace
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.content_agent import (
    ContentAgentRun,
    ContentAgentStageCall,
    ContentBatchItem,
    ContentBatchItemVersion,
    ContentBatchJob,
    ContentFeedback,
)
from app.schemas.content_batch_report import (
    ContentBatchListItem,
    ContentBatchFeedbackInsightResponse,
    ContentBatchFeedbackOptimizationSuggestion,
    ContentBatchFeedbackSample,
    ContentBatchFeedbackStat,
    ContentBatchRejectReason,
    ContentBatchListResponse,
    ContentBatchReportItem,
    ContentBatchReportResponse,
    ContentBatchSimilarityWarning,
    ContentBatchReportSummary,
    ContentBatchStageTrace,
    ContentBatchVersionCompare,
    ContentBatchVersionSnapshot,
    ContentFeedbackSample,
    ContentFeedbackSampleListResponse,
)
from app.services.forbidden_term_review_service import ForbiddenTermReviewService, find_forbidden_hits
from app.services.activity_quality_guard_service import build_article_pool_context_list
from app.services.comment_delivery_ledger_service import CommentDeliveryLedgerService

SIMILARITY_WARNING_THRESHOLD = 0.42
CLOSURE_CLUSTER_WINDOW_CHARS = 30
CLOSURE_CLUSTER_WARNING_RATIO = 0.2
CLOSURE_CLUSTER_WARNING_MIN_COUNT = 5
CLOSURE_CLUSTER_DEFINITIONS = [
    {
        "code": "peace_of_mind",
        "name": "安心簇",
        "phrases": ["省心", "踏实", "放心", "安心", "心里有底", "心里有数"],
    },
    {
        "code": "worth_it",
        "name": "值了簇",
        "phrases": ["这钱花得值", "花得值", "贵也认了", "肉疼但值", "没白花", "钱包肉疼但值"],
    },
    {
        "code": "right_choice",
        "name": "选对簇",
        "phrases": ["没选错", "选对了", "还好选了它", "感觉选对了", "判断没错", "选得还行", "没白选", "没白喝", "没白挑", "最好的证明"],
    },
    {
        "code": "keep_drinking",
        "name": "继续喝簇",
        "phrases": ["继续喝", "先喝着", "还会回购", "准备续上", "续上"],
    },
    {
        "code": "mom_satisfied",
        "name": "妈妈满足簇",
        "phrases": ["当妈的就图这个", "我也就满足了", "我也认了", "也就认了"],
    },
]
CONTENT_PATH_SKELETON_WARNING_RATIO = 0.25
CONTENT_PATH_SKELETON_WARNING_MIN_COUNT = 5
CONTENT_PATH_SKELETON_PARTS: dict[str, tuple[str, ...]] = {
    "selection": (
        "选奶",
        "挑奶",
        "换奶",
        "对比",
        "成分",
        "配方",
        "做功课",
        "攻略",
        "看了好几款",
        "看中",
        "最后选",
        "定了",
        "入手",
    ),
    "drinking_acceptance": (
        "喝得",
        "喝完",
        "爱喝",
        "愿意喝",
        "主动",
        "抱着杯子",
        "咕咚",
        "顺口",
        "不挑",
        "不抗拒",
        "不排斥",
        "喝光",
    ),
    "state_observation": (
        "状态",
        "精神",
        "身形",
        "结实",
        "小脸",
        "圆润",
        "长肉",
        "背上有肉",
        "小腿",
        "有劲",
        "个子",
        "身高",
        "裤子",
        "抱起来",
    ),
    "mom_closure": (
        "省心",
        "踏实",
        "放心",
        "安心",
        "心里有底",
        "心里有数",
        "值",
        "没选错",
        "选对",
        "继续喝",
        "回购",
        "续",
        "囤",
    ),
}

_FEEDBACK_CATEGORY_LABELS = {
    "unnatural": "不自然/生硬",
    "too_long": "太长/啰嗦",
    "too_ad_like": "广告感太强",
    "fact_issue": "信息不准确",
    "tone_mismatch": "语气不对",
    "forbidden_term": "有违禁词",
    "rule_mismatch": "不符合业务规则",
}
_FEEDBACK_ACTION_LABELS = {
    "approve": "通过",
    "request_revision": "要求修改",
    "manual_edit": "人工编辑",
    "accept_rewrite": "采纳改写",
    "reject_rewrite": "不采纳改写",
}
_FEEDBACK_REVIEW_STATUS_LABELS = {
    "approved": "已通过",
    "needs_revision": "待修改",
    "manual_edited": "人工编辑",
}
_REWRITE_DECISION_LABELS = {
    "auto_rewrite_requested": "触发系统改写",
    "accept_rewrite": "采纳改写",
    "reject_rewrite": "不采纳改写",
}


class ContentBatchReportService:
    """Return a compact batch result view suitable for operator review screens."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def list_batch_reports(
        self,
        *,
        limit: int = 20,
        offset: int = 0,
        asset_key: str | None = None,
        rule_id: str | None = None,
        source_row_no: int | None = None,
        keyword: str | None = None,
        product_topic: str | None = None,
    ) -> ContentBatchListResponse:
        conditions = []
        normalized_asset_key = str(asset_key or "").strip()
        if normalized_asset_key:
            conditions.append(ContentBatchJob.asset_key == normalized_asset_key)

        normalized_product_topic = str(product_topic or "").strip()
        if normalized_product_topic:
            conditions.append(ContentBatchJob.product_topic == normalized_product_topic)

        normalized_keyword = str(keyword or "").strip()
        if normalized_keyword:
            content_match = (
                select(ContentBatchItem.id)
                .where(
                    ContentBatchItem.batch_id == ContentBatchJob.id,
                    or_(
                        ContentBatchItem.title.contains(normalized_keyword, autoescape=True),
                        ContentBatchItem.body.contains(normalized_keyword, autoescape=True),
                    ),
                )
                .exists()
            )
            conditions.append(content_match)

        topic_result = await self.db.execute(
            select(ContentBatchJob.product_topic)
            .distinct()
            .order_by(ContentBatchJob.product_topic)
        )
        product_topics = [topic for topic in topic_result.scalars().all() if topic]

        rule_filter_enabled = (
            bool(str(rule_id or "").strip()) or source_row_no is not None
        )
        base_query = select(ContentBatchJob).order_by(
            ContentBatchJob.create_time.desc(),
            ContentBatchJob.id.desc(),
        )
        if conditions:
            base_query = base_query.where(*conditions)

        item_cache: dict[int, list[ContentBatchItem]] = {}
        if rule_filter_enabled:
            # 重要逻辑：批次与单条规则的关系目前只存在 item.plan_json，
            # 首版不新增表和迁移，所以列表按 job 拉取后用计划快照过滤。
            result = await self.db.execute(base_query)
            filtered_jobs: list[ContentBatchJob] = []
            for job in result.scalars().all():
                items = await self._batch_items(job.id)
                if self._items_match_rule_filter(items, rule_id=rule_id, source_row_no=source_row_no):
                    item_cache[job.id] = items
                    filtered_jobs.append(job)
            total = len(filtered_jobs)
            jobs = filtered_jobs[offset : offset + limit]
        else:
            total_query = select(func.count()).select_from(ContentBatchJob)
            if conditions:
                total_query = total_query.where(*conditions)
            total_result = await self.db.execute(total_query)
            total = int(total_result.scalar_one() or 0)
            result = await self.db.execute(base_query.offset(offset).limit(limit))
            jobs = list(result.scalars().all())

        list_items: list[ContentBatchListItem] = []
        for job in jobs:
            items = item_cache.get(job.id) or await self._batch_items(job.id)
            versions_by_item = await self._versions_for_items(items)
            feedback_counts = await self._feedback_counts_for_items(items)
            forbidden_terms = await self._forbidden_terms_for_job(job)
            report_items = [
                self._report_item(
                    item,
                    [],
                    versions=versions_by_item.get(item.id, []),
                    run=None,
                    feedback_count=feedback_counts.get(item.id, 0),
                    forbidden_terms=forbidden_terms,
                )
                for item in items
            ]
            self._attach_similarity_warnings(report_items)
            list_items.append(
                ContentBatchListItem(
                    batch_id=job.id,
                    batch_code=job.batch_code,
                    asset_key=job.asset_key,
                    product_topic=job.product_topic,
                    target_audience=job.target_audience,
                    persona_target=self._job_persona_target(job),
                    style=job.style,
                    status=job.status,
                    count=job.count,
                    summary=self._summary(report_items),
                    create_time=job.create_time,
                    update_time=job.update_time,
                )
            )
        return ContentBatchListResponse(
            total=total,
            items=list_items,
            product_topics=product_topics,
        )

    def _items_match_rule_filter(
        self,
        items: list[ContentBatchItem],
        *,
        rule_id: str | None,
        source_row_no: int | None,
    ) -> bool:
        normalized_rule_id = str(rule_id or "").strip()
        for item in items:
            plan = item.plan_json or {}
            plan_rule_id = str(plan.get("rule_id") or "").strip()
            plan_source_row_no = _int_or_none(plan.get("source_row_no"))
            if normalized_rule_id and plan_rule_id != normalized_rule_id:
                continue
            if source_row_no is not None and plan_source_row_no != source_row_no:
                continue
            return True
        return False

    async def get_batch_report(self, batch_id: int, *, include_details: bool = False) -> ContentBatchReportResponse:
        job = await self._require_job(batch_id)
        items = await self._batch_items(batch_id)
        stage_calls = await self._stage_calls_for_items(items)
        stages_by_run = self._group_stages_by_run(stage_calls)
        runs_by_id = await self._runs_for_items(items)
        versions_by_item = await self._versions_for_items(items)
        feedback_counts = await self._feedback_counts_for_items(items)
        forbidden_terms = await self._forbidden_terms_for_job(job)
        report_items = [
            self._report_item(
                item,
                stages_by_run.get(item.run_id or -1, []),
                versions=versions_by_item.get(item.id, []),
                run=runs_by_id.get(item.run_id or -1),
                feedback_count=feedback_counts.get(item.id, 0),
                forbidden_terms=forbidden_terms,
                include_details=include_details,
            )
            for item in items
        ]
        self._attach_similarity_warnings(report_items)
        return ContentBatchReportResponse(
            batch_id=job.id,
            batch_code=job.batch_code,
            asset_key=job.asset_key,
            product_topic=job.product_topic,
            target_audience=job.target_audience,
            persona_target=self._job_persona_target(job),
            style=job.style,
            status=job.status,
            count=job.count,
            summary=self._summary(report_items),
            items=report_items,
        )

    async def export_batch_report_excel(self, batch_id: int) -> tuple[str, bytes]:
        report = await self.get_batch_report(batch_id, include_details=True)
        workbook = Workbook()
        result_sheet = workbook.active
        result_sheet.title = "生文结果"

        _write_result_sheet(result_sheet, report)

        output = BytesIO()
        workbook.save(output)
        filename = _excel_filename(report)
        return filename, output.getvalue()

    async def export_article_pool_csv(self, batch_id: int) -> tuple[str, bytes]:
        report = await self.get_batch_report(batch_id, include_details=True)
        content = _build_article_pool_csv(report)
        filename = _article_pool_csv_filename(report)
        await self._record_article_pool_delivery(report, filename)
        return filename, content

    async def _record_article_pool_delivery(self, report: ContentBatchReportResponse, filename: str) -> None:
        items = _article_pool_export_items(report.items)
        if not items:
            return
        entries = [
            {
                "category": item.content_angle or item.asset_combo_key or "",
                "comment_text": item.body or "",
                "batch_id": report.batch_id,
                "item_id": item.item_id,
                "metadata_json": {
                    "batch_code": report.batch_code,
                    "product_topic": report.product_topic,
                    "item_no": item.item_no,
                },
            }
            for item in items
            if str(item.body or "").strip()
        ]
        await CommentDeliveryLedgerService(self.db).upsert_many(
            asset_key=report.asset_key,
            entries=entries,
            source_type="maga_batch",
            source_uri=f"batch:{report.batch_id}/{filename}",
            delivered_by="content_batch_export",
        )

    async def list_feedback_samples(
        self,
        *,
        limit: int = 50,
        offset: int = 0,
        review_status: str | None = None,
    ) -> ContentFeedbackSampleListResponse:
        conditions = []
        if review_status:
            conditions.append(ContentFeedback.review_status == review_status)
        total_query = (
            select(func.count())
            .select_from(ContentFeedback)
            .join(ContentBatchItem, ContentBatchItem.id == ContentFeedback.item_id)
            .join(ContentBatchJob, ContentBatchJob.id == ContentBatchItem.batch_id, isouter=True)
        )
        query = (
            select(ContentFeedback, ContentBatchItem, ContentBatchJob)
            .join(ContentBatchItem, ContentBatchItem.id == ContentFeedback.item_id)
            .join(ContentBatchJob, ContentBatchJob.id == ContentBatchItem.batch_id, isouter=True)
            .order_by(ContentFeedback.create_time.desc(), ContentFeedback.id.desc())
            .offset(offset)
            .limit(limit)
        )
        if conditions:
            total_query = total_query.where(*conditions)
            query = query.where(*conditions)
        total_result = await self.db.execute(total_query)
        result = await self.db.execute(query)
        return ContentFeedbackSampleListResponse(
            total=int(total_result.scalar_one() or 0),
            items=[
                self._feedback_sample(feedback, item, job)
                for feedback, item, job in result.all()
            ],
        )

    async def build_feedback_insights(self, batch_id: int) -> ContentBatchFeedbackInsightResponse:
        job = await self._require_job(batch_id)
        items = await self._batch_items(batch_id)
        item_by_id = {item.id: item for item in items}
        feedbacks = await self._feedbacks_for_items(items)
        category_counter: Counter[str] = Counter()
        action_counter: Counter[str] = Counter()
        review_status_counter: Counter[str] = Counter()
        rewrite_decision_counter: Counter[str] = Counter()
        evidence_by_category: dict[str, list[str]] = {}
        samples: list[ContentBatchFeedbackSample] = []

        for feedback in feedbacks:
            metadata = feedback.metadata_json if isinstance(feedback.metadata_json, dict) else {}
            categories = _feedback_categories(metadata)
            category_counter.update(categories)
            action_counter.update([feedback.action])
            review_status_counter.update([feedback.review_status])
            if metadata.get("auto_rewrite") is True:
                rewrite_decision_counter.update(["auto_rewrite_requested"])
            if feedback.action in {"accept_rewrite", "reject_rewrite"}:
                rewrite_decision_counter.update([feedback.action])

            item = item_by_id.get(feedback.item_id)
            evidence = _feedback_evidence(feedback, item)
            if evidence:
                for category in categories:
                    evidence_by_category.setdefault(category, []).append(evidence)
            if len(samples) < 8 and (feedback.comment or feedback.quoted_text or categories):
                samples.append(
                    ContentBatchFeedbackSample(
                        feedback_id=feedback.id,
                        item_id=feedback.item_id,
                        item_no=item.item_no if item else 0,
                        action=feedback.action,
                        review_status=feedback.review_status,
                        comment=feedback.comment,
                        quoted_text=feedback.quoted_text,
                        feedback_categories=categories,
                        create_time=self._format_time(feedback.create_time),
                    )
                )

        # 这里仅生成只读建议单，不写回规则资产，也不触发自动学习。
        return ContentBatchFeedbackInsightResponse(
            batch_id=job.id,
            batch_code=job.batch_code,
            asset_key=job.asset_key,
            product_topic=job.product_topic,
            total_feedback_count=len(feedbacks),
            category_stats=_counter_stats(category_counter, _FEEDBACK_CATEGORY_LABELS),
            action_stats=_counter_stats(action_counter, _FEEDBACK_ACTION_LABELS),
            review_status_stats=_counter_stats(review_status_counter, _FEEDBACK_REVIEW_STATUS_LABELS),
            rewrite_decision_stats=_counter_stats(rewrite_decision_counter, _REWRITE_DECISION_LABELS),
            samples=samples,
            suggestions=self._feedback_optimization_suggestions(
                category_counter=category_counter,
                rewrite_decision_counter=rewrite_decision_counter,
                evidence_by_category=evidence_by_category,
                content_type=_batch_content_type(items),
                total_feedback_count=len(feedbacks),
            ),
        )

    async def build_item_report(self, item: ContentBatchItem) -> ContentBatchReportItem:
        stage_calls = await self._stage_calls_for_items([item])
        versions = (await self._versions_for_items([item])).get(item.id, [])
        run = (await self._runs_for_items([item])).get(item.run_id or -1)
        feedback_count = (await self._feedback_counts_for_items([item])).get(item.id, 0)
        job = await self._job_for_item(item)
        forbidden_terms = await self._forbidden_terms_for_job(job)
        return self._report_item(
            item,
            stage_calls,
            versions=versions,
            run=run,
            feedback_count=feedback_count,
            forbidden_terms=forbidden_terms,
            include_details=True,
        )

    def _feedback_sample(
        self,
        feedback: ContentFeedback,
        item: ContentBatchItem,
        job: ContentBatchJob | None,
    ) -> ContentFeedbackSample:
        return ContentFeedbackSample(
            feedback_id=feedback.id,
            batch_id=feedback.batch_id,
            batch_code=job.batch_code if job else None,
            item_id=feedback.item_id,
            item_no=item.item_no,
            version_id=feedback.version_id,
            action=feedback.action,
            review_status=feedback.review_status,
            comment=feedback.comment,
            submitter=feedback.submitter,
            title=item.title,
            body_preview=(item.body or "")[:180] if item.body else None,
            product_topic=job.product_topic if job else None,
            target_audience=job.target_audience if job else None,
            persona_target=self._job_persona_target(job) if job else None,
            style=job.style if job else None,
            asset_key=job.asset_key if job else None,
            metadata=feedback.metadata_json,
            create_time=feedback.create_time.strftime("%Y-%m-%d %H:%M:%S") if feedback.create_time else None,
        )

    async def _require_job(self, batch_id: int) -> ContentBatchJob:
        result = await self.db.execute(select(ContentBatchJob).where(ContentBatchJob.id == batch_id))
        job = result.scalar_one_or_none()
        if job is None:
            raise ValueError("batch job not found")
        return job

    async def _job_for_item(self, item: ContentBatchItem) -> ContentBatchJob | None:
        if not item.batch_id:
            return None
        result = await self.db.execute(select(ContentBatchJob).where(ContentBatchJob.id == item.batch_id))
        return result.scalar_one_or_none()

    async def _forbidden_terms_for_job(self, job: ContentBatchJob | None) -> list[str]:
        return await ForbiddenTermReviewService(self.db).list_terms(asset_key=job.asset_key if job else None)

    @staticmethod
    def _job_persona_target(job: ContentBatchJob | None) -> str | None:
        if job is None:
            return None
        strategy = job.strategy_json or {}
        value = strategy.get("persona_target") if isinstance(strategy, dict) else None
        return value if isinstance(value, str) and value.strip() else None

    async def _batch_items(self, batch_id: int) -> list[ContentBatchItem]:
        result = await self.db.execute(
            select(ContentBatchItem).where(ContentBatchItem.batch_id == batch_id).order_by(ContentBatchItem.item_no)
        )
        return list(result.scalars().all())

    async def _stage_calls_for_items(self, items: list[ContentBatchItem]) -> list[ContentAgentStageCall]:
        run_ids = [item.run_id for item in items if item.run_id]
        if not run_ids:
            return []
        result = await self.db.execute(
            select(ContentAgentStageCall)
            .where(ContentAgentStageCall.run_id.in_(run_ids))
            .order_by(ContentAgentStageCall.run_id, ContentAgentStageCall.sequence_no)
        )
        return list(result.scalars().all())

    async def _runs_for_items(self, items: list[ContentBatchItem]) -> dict[int, ContentAgentRun]:
        run_ids = [item.run_id for item in items if item.run_id]
        if not run_ids:
            return {}
        result = await self.db.execute(select(ContentAgentRun).where(ContentAgentRun.id.in_(run_ids)))
        return {run.id: run for run in result.scalars().all()}

    async def _latest_versions_for_items(self, items: list[ContentBatchItem]) -> dict[int, ContentBatchItemVersion]:
        item_ids = [item.id for item in items if item.id]
        if not item_ids:
            return {}
        result = await self.db.execute(
            select(ContentBatchItemVersion)
            .where(ContentBatchItemVersion.item_id.in_(item_ids))
            .order_by(ContentBatchItemVersion.item_id, ContentBatchItemVersion.version_no.desc())
        )
        latest: dict[int, ContentBatchItemVersion] = {}
        for version in result.scalars().all():
            latest.setdefault(version.item_id, version)
        return latest

    async def _versions_for_items(self, items: list[ContentBatchItem]) -> dict[int, list[ContentBatchItemVersion]]:
        item_ids = [item.id for item in items if item.id]
        if not item_ids:
            return {}
        result = await self.db.execute(
            select(ContentBatchItemVersion)
            .where(ContentBatchItemVersion.item_id.in_(item_ids))
            .order_by(ContentBatchItemVersion.item_id, ContentBatchItemVersion.version_no)
        )
        versions: dict[int, list[ContentBatchItemVersion]] = {}
        for version in result.scalars().all():
            versions.setdefault(version.item_id, []).append(version)
        return versions

    async def _feedback_counts_for_items(self, items: list[ContentBatchItem]) -> dict[int, int]:
        item_ids = [item.id for item in items if item.id]
        if not item_ids:
            return {}
        result = await self.db.execute(
            select(ContentFeedback.item_id, func.count(ContentFeedback.id))
            .where(ContentFeedback.item_id.in_(item_ids))
            .group_by(ContentFeedback.item_id)
        )
        return {int(item_id): int(count or 0) for item_id, count in result.all()}

    async def _feedbacks_for_items(self, items: list[ContentBatchItem]) -> list[ContentFeedback]:
        item_ids = [item.id for item in items if item.id]
        if not item_ids:
            return []
        result = await self.db.execute(
            select(ContentFeedback)
            .where(ContentFeedback.item_id.in_(item_ids))
            .order_by(ContentFeedback.create_time.desc(), ContentFeedback.id.desc())
        )
        return list(result.scalars().all())

    def _feedback_optimization_suggestions(
        self,
        *,
        category_counter: Counter[str],
        rewrite_decision_counter: Counter[str],
        evidence_by_category: dict[str, list[str]],
        content_type: str,
        total_feedback_count: int,
    ) -> list[ContentBatchFeedbackOptimizationSuggestion]:
        suggestions: list[ContentBatchFeedbackOptimizationSuggestion] = []
        if total_feedback_count <= 0:
            return suggestions

        if category_counter.get("forbidden_term", 0):
            suggestions.append(
                ContentBatchFeedbackOptimizationSuggestion(
                    suggestion_type="business_forbidden_term",
                    target="业务违禁词",
                    title="把运营明确不希望出现的表达收敛到业务违禁词",
                    reason="本批次存在违禁词类反馈，适合走确定性审核闸口，不建议塞回生文提示词。",
                    evidence=_evidence_for_categories(evidence_by_category, ["forbidden_term"]),
                    priority=_suggestion_priority(category_counter["forbidden_term"], total_feedback_count),
                )
            )

        fact_rule_count = category_counter.get("fact_issue", 0) + category_counter.get("rule_mismatch", 0)
        if fact_rule_count:
            suggestions.append(
                ContentBatchFeedbackOptimizationSuggestion(
                    suggestion_type="business_rule",
                    target="业务规则包",
                    title="补充事实边界或业务规则约束",
                    reason="反馈指向信息准确性或业务规则不匹配，应由业务规则包承载，保持生成链路可追溯。",
                    evidence=_evidence_for_categories(evidence_by_category, ["fact_issue", "rule_mismatch"]),
                    priority=_suggestion_priority(fact_rule_count, total_feedback_count),
                )
            )

        expression_categories = ["unnatural", "too_ad_like", "too_long", "tone_mismatch"]
        expression_count = sum(category_counter.get(category, 0) for category in expression_categories)
        if expression_count:
            suggestions.append(
                ContentBatchFeedbackOptimizationSuggestion(
                    suggestion_type="system_keyword",
                    target=_system_keyword_target(category_counter, content_type),
                    title="补强表达扩散语料",
                    reason="反馈集中在自然度、广告感、篇幅或语气，适合补表达扩散语料示例，不应增加很硬的业务规则。",
                    evidence=_evidence_for_categories(evidence_by_category, expression_categories),
                    priority=_suggestion_priority(expression_count, total_feedback_count),
                )
            )

        if rewrite_decision_counter.get("reject_rewrite", 0):
            suggestions.append(
                ContentBatchFeedbackOptimizationSuggestion(
                    suggestion_type="expert_prompt",
                    target="审核改写 Expert",
                    title="复盘被拒绝的系统改写",
                    reason="存在不采纳系统改写，说明改写 Expert 可能仍在做机械替换，需要用被拒样例调整改写指令。",
                    evidence=[],
                    priority=_suggestion_priority(rewrite_decision_counter["reject_rewrite"], total_feedback_count),
                )
            )

        return suggestions

    def _group_stages_by_run(self, stage_calls: list[ContentAgentStageCall]) -> dict[int, list[ContentAgentStageCall]]:
        grouped: dict[int, list[ContentAgentStageCall]] = {}
        for stage in stage_calls:
            grouped.setdefault(stage.run_id, []).append(stage)
        return grouped

    def _report_item(
        self,
        item: ContentBatchItem,
        stage_calls: list[ContentAgentStageCall],
        latest_version: ContentBatchItemVersion | None = None,
        *,
        versions: list[ContentBatchItemVersion] | None = None,
        run: ContentAgentRun | None = None,
        feedback_count: int = 0,
        forbidden_terms: list[str] | None = None,
        include_details: bool = False,
    ) -> ContentBatchReportItem:
        ordered_versions = versions or []
        latest_version = latest_version or (ordered_versions[-1] if ordered_versions else None)
        quality = item.quality_json or {}
        review = quality.get("review_report") or {}
        diversity = item.diversity_json or {}
        runtime_result = self._runtime_result(stage_calls)
        generation_stage = self._generation_stage(stage_calls)
        text = f"{item.title or ''}\n{item.body or ''}"
        forbidden_hits = self._forbidden_hits_for_report(
            text,
            forbidden_terms,
            quality=quality,
        )
        final_state = _final_postprocess_state(quality)
        rewrite_required = final_state["rewrite_required"]
        hard_pass = final_state["hard_pass"]
        rewrite_reason = review.get("rewrite_reason") or final_state["rewrite_reason"]
        batch_variation = (
            quality.get("batch_variation_review")
            if isinstance(quality.get("batch_variation_review"), dict)
            else {}
        )
        delivery_selection = (
            quality.get("delivery_selection")
            if isinstance(quality.get("delivery_selection"), dict)
            else {}
        )
        business_usability = _business_usability_from_quality(quality)
        detail_value = include_details
        return ContentBatchReportItem(
            item_id=item.id,
            item_no=item.item_no,
            status=item.status,
            task_id=item.task_id,
            run_id=item.run_id,
            title=item.title,
            body=item.body,
            body_preview=(item.body or "")[:160] if detail_value and item.body else None,
            body_chars=len(item.body or "") if detail_value else 0,
            hard_pass=hard_pass,
            batch_variation_pass=batch_variation.get("pass"),
            delivery_selected=delivery_selection.get("selected"),
            delivery_rank=delivery_selection.get("delivery_rank"),
            delivery_non_selection_reason=delivery_selection.get("non_selection_reason"),
            audit_skipped=bool(quality.get("audit_skipped")),
            rewrite_required=rewrite_required,
            rewrite_reason=rewrite_reason,
            business_usability_tier=business_usability.get("tier"),
            business_usability_reason=business_usability.get("reason"),
            rewrite_rounds=review.get("rewrite_rounds") if detail_value else None,
            suggestion_count=len(review.get("suggestions") or []) if detail_value else 0,
            replacement_count=len(review.get("replacement_needed") or []) if detail_value else 0,
            forbidden_hits=forbidden_hits,
            final_path=runtime_result.get("final_path") if detail_value else None,
            debug_dir=runtime_result.get("debug_dir") if detail_value else None,
            review_status=(
                latest_version.review_status if latest_version else (quality.get("human_review") or {}).get("review_status")
            ),
            latest_version_no=latest_version.version_no if latest_version else None,
            human_feedback_text=(
                latest_version.feedback_text if latest_version else (quality.get("human_review") or {}).get("feedback_text")
            ),
            feedback_count=feedback_count,
            reject_reasons=self._reject_reasons(item, review, forbidden_hits),
            similarity_warnings=[],
            version_compare=self._version_compare(latest_version, ordered_versions),
            runtime_mode=runtime_result.get("mode") or quality.get("executor") if detail_value else None,
            generation_duration_ms=(
                self._stage_duration_ms(generation_stage) if detail_value and generation_stage else None
            ),
            total_duration_ms=self._total_duration_ms(run, stage_calls) if detail_value else None,
            trace_run_id=item.run_id if detail_value else None,
            trace_stage_calls=[self._stage_trace(stage) for stage in stage_calls],
            opening_type=diversity.get("opening_type") if detail_value else None,
            structure_type=diversity.get("structure_type") if detail_value else None,
            content_angle=diversity.get("content_angle") if detail_value else None,
            persona_lens=diversity.get("persona_lens") if detail_value else None,
            scene_type=diversity.get("scene_type") if detail_value else None,
            evidence_type=diversity.get("evidence_type") if detail_value else None,
            asset_combo_key=(item.plan_json or {}).get("asset_combo_key") if detail_value else None,
            asset_reuse_reason=(item.plan_json or {}).get("asset_reuse_reason") if detail_value else None,
            generation_snapshot=self._generation_snapshot(item, stage_calls, run, quality) if include_details else None,
            diversity=(diversity or None) if include_details else None,
            quality=(quality or None) if include_details else self._quality_summary(quality),
            error_message=item.error_message if item.status == "failed" else None,
        )

    def _quality_summary(self, quality: dict[str, Any]) -> dict[str, Any] | None:
        if not quality:
            return None
        review = quality.get("review_report") if isinstance(quality.get("review_report"), dict) else {}
        summary: dict[str, Any] = {
            "hard_pass": quality.get("hard_pass"),
            "final_postprocess_state": _final_postprocess_state(quality),
            "stage_call_count": quality.get("stage_call_count"),
            "review_report": {
                "rewrite_required": review.get("rewrite_required"),
                "rewrite_reason": review.get("rewrite_reason"),
            },
        }
        for key in (
            "forbidden_terms_review",
            "mouth_phrase_budget_guard",
            "product_experience_phrase_review",
            "product_experience_llm_review",
            "ai_flavor_review",
        ):
            value = review.get(key)
            if isinstance(value, dict):
                summary["review_report"][key] = self._review_summary(value)
        for key in (
            "product_experience_phrase_guard",
            "product_experience_llm_quality_review",
            "wangyue_focused_pipeline_review",
            "ai_flavor_humanizer",
            "mouth_phrase_budget_guard",
        ):
            value = quality.get(key)
            if isinstance(value, dict):
                summary[key] = self._review_summary(value)
        for key in ("batch_variation_review", "delivery_selection"):
            value = quality.get(key)
            if isinstance(value, dict):
                summary[key] = dict(value)
        for key in (
            "product_experience_phrase_rewrites",
            "product_experience_llm_quality_rewrites",
            "product_experience_llm_quality_failures",
            "ai_flavor_humanizer_rewrites",
            "mouth_phrase_budget_rewrites",
            "article_length_guard",
        ):
            value = quality.get(key)
            if isinstance(value, list):
                payload = {"count": len(value)}
                if key.endswith("_failures"):
                    payload["last_error"] = _last_failure_error(value)
                summary[key] = payload
            elif isinstance(value, dict):
                summary[key] = self._review_summary(value)
        return summary

    @staticmethod
    def _review_summary(value: dict[str, Any]) -> dict[str, Any]:
        keys = (
            "pass",
            "pass_",
            "rewrite_required",
            "rewrite_reason",
            "severity",
            "reasons",
            "issues",
            "final_hits",
            "initial_hits",
            "body_chars",
            "length_target",
            "mark_rewrite_required",
            "repair_count",
            "business_usability_tier",
            "business_usability_reason",
            "status",
            "decision",
            "unavailable_dimensions",
            "rewrite_modes",
            "requires_rewrite",
            "can_auto_pool",
            "blocked_by_code_hard",
            "affects_pool",
        )
        return {key: value.get(key) for key in keys if key in value}

    def _version_compare(
        self,
        latest_version: ContentBatchItemVersion | None,
        versions: list[ContentBatchItemVersion],
    ) -> ContentBatchVersionCompare | None:
        # “通过”等反馈也会生成版本；报告里应展示最近一次真正改变文本的版本差异。
        candidate_versions = list(reversed(versions))
        if latest_version is not None and latest_version not in versions:
            candidate_versions.insert(0, latest_version)
        if not candidate_versions:
            return None
        for version in candidate_versions:
            if version.source_action not in {"accept_rewrite", "auto_rewrite", "manual_edit", "reject_rewrite"}:
                continue
            before = self._compare_before_snapshot(version, versions)
            if before is None:
                continue
            after = self._version_snapshot(version)
            title_changed = (before.title or "") != (after.title or "")
            body_changed = (before.body or "") != (after.body or "")
            if not title_changed and not body_changed and version.source_action != "auto_rewrite":
                continue
            return ContentBatchVersionCompare(
                compare_type=version.source_action,
                before=before,
                after=after,
                title_changed=title_changed,
                body_changed=body_changed,
                body_before_chars=len(before.body or ""),
                body_after_chars=len(after.body or ""),
            )
        return None

    def _compare_before_snapshot(
        self,
        latest_version: ContentBatchItemVersion,
        versions: list[ContentBatchItemVersion],
    ) -> ContentBatchVersionSnapshot | None:
        metadata = latest_version.metadata_json if isinstance(latest_version.metadata_json, dict) else {}
        decision_version_id = metadata.get("decision_for_version_id")
        if decision_version_id is not None:
            for version in versions:
                if version.id == decision_version_id:
                    if latest_version.source_action == "accept_rewrite":
                        return self._compare_before_snapshot(version, versions)
                    return self._version_snapshot(version)

        source_version_id = metadata.get("source_version_id")
        if source_version_id is not None:
            for version in versions:
                if version.id == source_version_id:
                    return self._version_snapshot(version)

        if latest_version.source_action == "manual_edit":
            previous_content = metadata.get("previous_content") if isinstance(metadata.get("previous_content"), dict) else {}
            if previous_content:
                return ContentBatchVersionSnapshot(
                    version_id=None,
                    version_no=max((latest_version.version_no or 1) - 1, 0),
                    source_action="before_manual_edit",
                    review_status=None,
                    title=previous_content.get("title"),
                    body=previous_content.get("body"),
                    feedback_text=latest_version.feedback_text,
                    created_by=latest_version.created_by,
                    create_time=self._format_time(latest_version.create_time),
                )

        if latest_version.source_action not in {"accept_rewrite", "auto_rewrite", "manual_edit", "reject_rewrite"}:
            return None
        previous_versions = [version for version in versions if version.version_no < latest_version.version_no]
        if not previous_versions:
            return None
        return self._version_snapshot(previous_versions[-1])

    @staticmethod
    def _version_snapshot(version: ContentBatchItemVersion) -> ContentBatchVersionSnapshot:
        return ContentBatchVersionSnapshot(
            version_id=version.id,
            version_no=version.version_no,
            source_action=version.source_action,
            review_status=version.review_status,
            title=version.title,
            body=version.body,
            feedback_text=version.feedback_text,
            created_by=version.created_by,
            create_time=ContentBatchReportService._format_time(version.create_time),
        )

    @staticmethod
    def _format_time(value: Any) -> str | None:
        return value.strftime("%Y-%m-%d %H:%M:%S") if value else None

    def _runtime_result(self, stage_calls: list[ContentAgentStageCall]) -> dict[str, Any]:
        for stage in stage_calls:
            if stage.capability != "content.generate":
                continue
            output = stage.output_snapshot or {}
            runtime_result = output.get("runtime_result")
            if isinstance(runtime_result, dict):
                return runtime_result
        return {}

    def _generation_stage(self, stage_calls: list[ContentAgentStageCall]) -> ContentAgentStageCall | None:
        return next((stage for stage in stage_calls if stage.capability == "content.generate"), None)

    def _generation_snapshot(
        self,
        item: ContentBatchItem,
        stage_calls: list[ContentAgentStageCall],
        run: ContentAgentRun | None,
        quality: dict[str, Any],
    ) -> dict[str, Any] | None:
        plan = item.plan_json or {}
        if not isinstance(plan, dict):
            plan = {}
        unified = plan.get("unified_generation") if isinstance(plan.get("unified_generation"), dict) else {}
        generation_stage = self._generation_stage(stage_calls)
        generation_input = self._stage_input(generation_stage)
        source = self._generation_source(generation_input, unified)
        business_rule = self._business_rule_snapshot(source, plan)
        selected_keywords = self._selected_keywords_snapshot(source, unified, quality)
        expert = self._dict_value(source.get("expert")) or self._dict_value(unified.get("expert")) or {}
        model_config = (
            self._dict_value(source.get("model_config"))
            or self._dict_value(expert.get("model_config"))
            or self._dict_value(plan.get("model_config"))
            or {}
        )
        rendered_prompt = self._string_or_none(source.get("rendered_prompt") or unified.get("rendered_prompt"))
        selected_prompt_slots = source.get("selected_prompt_slots") or unified.get("selected_prompt_slots") or []
        raw_comment_tone = (
            self._dict_value(source.get("comment_tone"))
            or self._dict_value(unified.get("comment_tone"))
            or self._dict_value(source.get("comment_persona"))
            or self._dict_value(unified.get("comment_persona"))
            or {}
        )
        comment_tone = (
            {
                "tone_code": raw_comment_tone.get("tone_code") or raw_comment_tone.get("persona_code"),
                "tone_label": raw_comment_tone.get("tone_label") or raw_comment_tone.get("persona_label"),
                "prompt": raw_comment_tone.get("prompt"),
            }
            if raw_comment_tone
            else {}
        )
        forbidden_review = self._forbidden_review(quality)
        realness_review = self._comment_realness_review(quality)
        activity_quality_guard = self._activity_quality_guard(quality)
        rewrite_records = self._rewrite_records(stage_calls)
        if not any([business_rule, selected_keywords, expert, rendered_prompt, forbidden_review, realness_review, activity_quality_guard, rewrite_records, generation_stage]):
            return None
        capability = (
            self._string_or_none(source.get("capability"))
            or self._string_or_none(unified.get("capability"))
            or (generation_stage.capability if generation_stage else None)
        )
        output_fields = source.get("output_fields") or business_rule.get("output_fields") or plan.get("output_fields") or []
        return {
            "schema_version": "1",
            "rule_type": business_rule.get("rule_type"),
            "content_type": source.get("content_type") or ("comment" if output_fields == ["comment"] else "article"),
            "capability": capability,
            "output_fields": output_fields,
            "business_rule": business_rule,
            "selected_keywords": selected_keywords,
            "selected_prompt_slots": selected_prompt_slots,
            "comment_tone": comment_tone,
            "keyword_asset": self._dict_value(source.get("keyword_asset")) or self._dict_value(unified.get("keyword_asset")) or {},
            "expert": expert,
            "model_config": model_config,
            "model_route": self._model_route(
                run=run,
                capability=capability,
                runtime_mode=self._runtime_result(stage_calls).get("mode") or quality.get("executor"),
                model_config=model_config,
            ),
            "rendered_prompt": rendered_prompt,
            "forbidden_terms_review": forbidden_review,
            "comment_realness_review": realness_review,
            "activity_quality_guard": activity_quality_guard,
            "rewrite_records": rewrite_records,
            "execution_stages": [self._stage_trace(stage).model_dump(mode="json") for stage in stage_calls],
        }

    def _generation_source(self, generation_input: dict[str, Any], unified: dict[str, Any]) -> dict[str, Any]:
        nested_snapshot = generation_input.get("generation_snapshot")
        if isinstance(nested_snapshot, dict):
            return nested_snapshot
        if any(key in generation_input for key in ("business_rule", "selected_keywords", "expert", "rendered_prompt")):
            return generation_input
        return unified

    def _business_rule_snapshot(self, source: dict[str, Any], plan: dict[str, Any]) -> dict[str, Any]:
        business_rule = self._dict_value(source.get("business_rule")) or plan
        return {
            key: value
            for key, value in dict(business_rule).items()
            if key not in {"unified_generation", "batch_context", "model_config"}
        }

    def _selected_keywords_snapshot(
        self,
        source: dict[str, Any],
        unified: dict[str, Any],
        quality: dict[str, Any],
    ) -> list[dict[str, Any]]:
        candidates = source.get("selected_keywords") or unified.get("selected_keywords") or quality.get("selected_keywords")
        return [item for item in candidates or [] if isinstance(item, dict)]

    def _forbidden_review(self, quality: dict[str, Any]) -> dict[str, Any] | None:
        review_report = quality.get("review_report") if isinstance(quality.get("review_report"), dict) else {}
        review = quality.get("forbidden_terms_review") or review_report.get("forbidden_terms_review")
        return review if isinstance(review, dict) else None

    def _comment_realness_review(self, quality: dict[str, Any]) -> dict[str, Any] | None:
        review_report = quality.get("review_report") if isinstance(quality.get("review_report"), dict) else {}
        review = quality.get("comment_realness_review") or review_report.get("comment_realness_review")
        return review if isinstance(review, dict) else None

    def _activity_quality_guard(self, quality: dict[str, Any]) -> dict[str, Any] | None:
        review = quality.get("activity_quality_guard")
        return review if isinstance(review, dict) else None

    def _rewrite_records(self, stage_calls: list[ContentAgentStageCall]) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        for stage in stage_calls:
            if stage.capability != "content.rewrite":
                continue
            input_payload = stage.input_snapshot or {}
            output_payload = stage.output_snapshot or {}
            records.append(
                {
                    "stage_call_id": stage.stage_call_id,
                    "sequence_no": stage.sequence_no,
                    "capability": stage.capability,
                    "status": stage.status,
                    "duration_ms": self._stage_duration_ms(stage),
                    "before": self._rewrite_before(input_payload),
                    "after": self._rewrite_after(output_payload),
                    "forbidden_hits": self._list_of_strings(input_payload.get("forbidden_hits")),
                    "style_hits": self._list_of_strings(input_payload.get("style_hits")),
                    "rewrite_source": self._string_or_none(input_payload.get("rewrite_source")),
                    "rewrite_instructions": self._list_of_strings(input_payload.get("rewrite_instructions")),
                    "expert": self._dict_value(input_payload.get("expert")) or {},
                    "model_config": self._dict_value(input_payload.get("model_config")) or {},
                    "rendered_prompt": self._string_or_none(input_payload.get("rendered_prompt")),
                    "error_message": stage.error_message,
                }
            )
        return records

    def _rewrite_before(self, input_payload: dict[str, Any]) -> dict[str, str]:
        previous = input_payload.get("previous_content") or input_payload.get("previous_draft") or {}
        return self._content_summary(previous if isinstance(previous, dict) else {})

    def _rewrite_after(self, output_payload: dict[str, Any]) -> dict[str, str]:
        return self._content_summary(output_payload)

    def _content_summary(self, payload: dict[str, Any]) -> dict[str, str]:
        final = payload.get("final") if isinstance(payload.get("final"), dict) else {}
        comment = payload.get("comment") or final.get("comment")
        if comment:
            return {"comment": str(comment)}
        title = payload.get("title") or final.get("title")
        body = payload.get("body") or final.get("body")
        result: dict[str, str] = {}
        if title:
            result["title"] = str(title)
        if body:
            result["body"] = str(body)
        return result

    def _model_route(
        self,
        *,
        run: ContentAgentRun | None,
        capability: str | None,
        runtime_mode: str | None,
        model_config: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "executor_code": run.executor_code if run else None,
            "executor_type": run.executor_type if run else None,
            "capability": capability,
            "runtime_mode": runtime_mode,
            "provider_code": model_config.get("provider_code") or model_config.get("provider"),
            "model_code": model_config.get("model_code") or model_config.get("ge_model"),
            "temperature": model_config.get("temperature"),
            "max_tokens": model_config.get("max_tokens"),
        }

    def _stage_input(self, stage: ContentAgentStageCall | None) -> dict[str, Any]:
        if stage is None or not isinstance(stage.input_snapshot, dict):
            return {}
        return stage.input_snapshot

    def _stage_trace(self, stage: ContentAgentStageCall) -> ContentBatchStageTrace:
        return ContentBatchStageTrace(
            stage_call_id=stage.stage_call_id,
            sequence_no=stage.sequence_no,
            capability=stage.capability,
            status=stage.status,
            duration_ms=self._stage_duration_ms(stage),
            error_message=stage.error_message,
            stats=stage.stats_json,
        )

    def _reject_reasons(
        self,
        item: ContentBatchItem,
        review: dict[str, Any],
        forbidden_hits: list[str],
    ) -> list[ContentBatchRejectReason]:
        reasons: list[ContentBatchRejectReason] = []
        # 把结构化审核结果转成运营能直接看到的驳回原因。
        for hard_result in review.get("hard_results") or []:
            if not isinstance(hard_result, dict) or hard_result.get("pass") is not False:
                continue
            evidence = hard_result.get("evidence") or []
            reasons.append(
                ContentBatchRejectReason(
                    source="hard_review",
                    code=self._string_or_none(hard_result.get("ae_code")),
                    message=self._reject_message(
                        hard_result.get("feedback") or hard_result.get("reason"),
                        evidence=evidence,
                        fallback="硬性审核未通过",
                    ),
                    risk_level=self._string_or_none(hard_result.get("risk_level")),
                    evidence=[str(item) for item in evidence if item is not None],
                )
            )
        for failed in review.get("failed_aes") or []:
            if isinstance(failed, dict):
                reasons.append(
                    ContentBatchRejectReason(
                        source="failed_ae",
                        code=self._string_or_none(failed.get("ae_code") or failed.get("code")),
                        message=str(failed.get("feedback") or failed.get("reason") or failed.get("message") or "AE 审核失败"),
                        risk_level=self._string_or_none(failed.get("risk_level")),
                    )
                )
            elif failed:
                reasons.append(ContentBatchRejectReason(source="failed_ae", code=str(failed), message=str(failed)))
        for term in forbidden_hits:
            reasons.append(ContentBatchRejectReason(source="forbidden_term", code=term, message=f"命中禁用词：{term}"))
        if item.status == "failed" and item.error_message:
            reasons.append(ContentBatchRejectReason(source="executor_error", message=item.error_message))
        return reasons

    def _total_duration_ms(self, run: ContentAgentRun | None, stage_calls: list[ContentAgentStageCall]) -> int | None:
        # Run 时间可能被数据库取整，阶段 stats 通常更精确；总耗时不能小于阶段耗时。
        durations = [value for value in (self._run_duration_ms(run), self._stage_calls_duration_ms(stage_calls)) if value is not None]
        return max(durations) if durations else None

    def _reject_message(self, value: Any, *, evidence: list[Any], fallback: str) -> str:
        message = self._string_or_none(value)
        if message and message.lower() not in {"fail", "failed", "false"}:
            return message
        evidence_text = "、".join(str(item) for item in evidence if item is not None)
        return f"命中硬性审核红线：{evidence_text}" if evidence_text else fallback

    def _stage_duration_ms(self, stage: ContentAgentStageCall | None) -> int | None:
        if not stage:
            return None
        stats_duration = self._duration_from_stats(stage.stats_json or {})
        if stats_duration is not None:
            return stats_duration
        if stage.started_at and stage.finished_at:
            return max(0, int((stage.finished_at - stage.started_at).total_seconds() * 1000))
        return None

    def _run_duration_ms(self, run: ContentAgentRun | None) -> int | None:
        if not run or not run.started_at or not run.finished_at:
            return None
        return max(0, int((run.finished_at - run.started_at).total_seconds() * 1000))

    def _stage_calls_duration_ms(self, stage_calls: list[ContentAgentStageCall]) -> int | None:
        durations = [duration for stage in stage_calls if (duration := self._stage_duration_ms(stage)) is not None]
        return sum(durations) if durations else None

    def _duration_from_stats(self, stats: dict[str, Any]) -> int | None:
        for key in ("total_latency_ms", "latency_ms", "duration_ms", "elapsed_ms"):
            value = stats.get(key)
            if isinstance(value, (int, float)):
                return max(0, int(value))
        return None

    def _string_or_none(self, value: Any) -> str | None:
        if value is None:
            return None
        return str(value)

    def _dict_value(self, value: Any) -> dict[str, Any]:
        return value if isinstance(value, dict) else {}

    def _list_of_strings(self, value: Any) -> list[str]:
        if not isinstance(value, list):
            return []
        return [str(item).strip() for item in value if str(item).strip()]

    def _summary(self, items: list[ContentBatchReportItem]) -> ContentBatchReportSummary:
        generated_statuses = {"generated", "approved", "manual_edited", "needs_revision"}
        generated = [item for item in items if item.status in generated_statuses]
        body_lengths = [item.body_chars for item in generated if item.body_chars]
        forbidden_hit_count = sum(len(item.forbidden_hits) for item in items)
        delivery_summary = next(
            (
                item.quality.get("delivery_selection")
                for item in generated
                if isinstance(item.quality, dict)
                and isinstance(item.quality.get("delivery_selection"), dict)
            ),
            {},
        )
        return ContentBatchReportSummary(
            total_count=len(items),
            generated_count=len(generated),
            failed_count=sum(1 for item in items if item.status == "failed"),
            hard_pass_count=sum(1 for item in generated if item.hard_pass is True),
            batch_variation_warning_count=sum(
                1 for item in generated if item.batch_variation_pass is False
            ),
            delivery_candidate_count=int(delivery_summary.get("eligible_count") or 0),
            delivery_selected_count=int(delivery_summary.get("selected_count") or 0),
            delivery_shortfall_count=int(delivery_summary.get("shortfall_count") or 0),
            suggested_bulk_refill_count=int(
                delivery_summary.get("suggested_bulk_refill_count") or 0
            ),
            audit_skipped_count=sum(1 for item in generated if item.audit_skipped),
            rewrite_item_count=sum(1 for item in items if item.rewrite_reason or item.rewrite_rounds),
            remaining_rewrite_required_count=sum(1 for item in items if item.rewrite_required is True),
            forbidden_hit_count=forbidden_hit_count,
            feedback_count=sum(item.feedback_count for item in items),
            avg_body_chars=round(sum(body_lengths) / len(body_lengths), 2) if body_lengths else None,
            max_pairwise_jaccard_2gram=self._max_pairwise_jaccard([item.body or "" for item in generated]),
            similarity_warning_count=sum(1 for item in items if item.similarity_warnings),
            closure_cluster_stats=self._closure_cluster_stats(generated),
            content_path_skeleton_stats=self._content_path_skeleton_stats(generated),
            real_user_pool_stats=self._real_user_pool_stats(items),
            mouth_phrase_budget_stats=self._mouth_phrase_budget_stats(generated),
            business_usability_stats=self._business_usability_stats(generated),
        )

    def _forbidden_hits(self, text: str, business_terms: list[str] | None = None) -> list[str]:
        return find_forbidden_hits(text, business_terms)

    def _forbidden_hits_for_report(
        self,
        text: str,
        business_terms: list[str] | None,
        *,
        quality: dict[str, Any],
    ) -> list[str]:
        review = quality.get("forbidden_terms_review")
        if isinstance(review, dict) and isinstance(review.get("final_hits"), list):
            return self._list_of_strings(review["final_hits"])
        return self._forbidden_hits(text, business_terms)

    def _closure_cluster_stats(self, items: list[ContentBatchReportItem]) -> dict[str, Any]:
        checked_items = [item for item in items if str(item.body or "").strip()]
        clusters: list[dict[str, Any]] = []
        item_nos_with_hits: set[int] = set()
        for definition in CLOSURE_CLUSTER_DEFINITIONS:
            hits: list[dict[str, Any]] = []
            for item in checked_items:
                closing = self._closing_window(item.body or "")
                matched_phrases = self._matched_closure_phrases(closing, definition["phrases"])
                if not matched_phrases:
                    continue
                item_nos_with_hits.add(item.item_no)
                hits.append(
                    {
                        "item_no": item.item_no,
                        "phrases": matched_phrases,
                        "closing_text": closing,
                    }
                )
            count = len(hits)
            ratio = round(count / len(checked_items), 4) if checked_items else 0.0
            watch = count >= CLOSURE_CLUSTER_WARNING_MIN_COUNT or (
                len(checked_items) >= 5 and count >= 2 and ratio > CLOSURE_CLUSTER_WARNING_RATIO
            )
            clusters.append(
                {
                    "cluster_code": definition["code"],
                    "cluster_name": definition["name"],
                    "count": count,
                    "ratio": ratio,
                    "watch": bool(count and watch),
                    "warning": False,
                    "phrases": definition["phrases"],
                    "hits": hits,
                }
            )
        clusters.sort(key=lambda cluster: (-cluster["count"], cluster["cluster_code"]))
        return {
            "window_chars": CLOSURE_CLUSTER_WINDOW_CHARS,
            "total_checked": len(checked_items),
            "closing_hit_count": len(item_nos_with_hits),
            "warning_threshold": {
                "min_count": CLOSURE_CLUSTER_WARNING_MIN_COUNT,
                "ratio": CLOSURE_CLUSTER_WARNING_RATIO,
            },
            "clusters": clusters,
        }

    def _closing_window(self, text: str) -> str:
        clean = re.sub(r"\s+", "", text or "")
        return clean[-CLOSURE_CLUSTER_WINDOW_CHARS:]

    def _matched_closure_phrases(self, closing: str, phrases: list[str]) -> list[str]:
        normalized_closing = re.sub(r"\s+", "", closing or "")
        matched: list[str] = []
        for phrase in phrases:
            normalized_phrase = re.sub(r"\s+", "", phrase)
            if normalized_phrase and normalized_phrase in normalized_closing:
                matched.append(phrase)
        return matched

    def _content_path_skeleton_stats(self, items: list[ContentBatchReportItem]) -> dict[str, Any]:
        checked_items = [item for item in items if str(item.body or "").strip()]
        hits: list[dict[str, Any]] = []
        part_counts: Counter[str] = Counter()
        for item in checked_items:
            part_hits = self._content_path_part_hits(item.body or "")
            for part in part_hits:
                part_counts[part] += 1
            if set(CONTENT_PATH_SKELETON_PARTS).issubset(part_hits):
                hits.append(
                    {
                        "item_no": item.item_no,
                        "part_hits": part_hits,
                        "body_preview": (item.body or "")[:120],
                    }
                )
        count = len(hits)
        ratio = round(count / len(checked_items), 4) if checked_items else 0.0
        warning = count >= CONTENT_PATH_SKELETON_WARNING_MIN_COUNT or (
            len(checked_items) >= 5 and count >= 2 and ratio > CONTENT_PATH_SKELETON_WARNING_RATIO
        )
        return {
            "skeleton_name": "选奶/喝奶接受/状态观察/妈妈收口",
            "total_checked": len(checked_items),
            "complete_skeleton_count": count,
            "complete_skeleton_ratio": ratio,
            "warning": bool(count and warning),
            "warning_threshold": {
                "min_count": CONTENT_PATH_SKELETON_WARNING_MIN_COUNT,
                "ratio": CONTENT_PATH_SKELETON_WARNING_RATIO,
            },
            "part_counts": dict(part_counts),
            "part_phrases": {part: list(phrases) for part, phrases in CONTENT_PATH_SKELETON_PARTS.items()},
            "hits": hits,
        }

    def _content_path_part_hits(self, body: str) -> dict[str, list[str]]:
        text = re.sub(r"\s+", "", body or "")
        return {
            part: self._matched_closure_phrases(text, list(phrases))
            for part, phrases in CONTENT_PATH_SKELETON_PARTS.items()
            if self._matched_closure_phrases(text, list(phrases))
        }

    def _real_user_pool_stats(self, items: list[ContentBatchReportItem]) -> dict[str, Any]:
        used_items = [item for item in items if self._real_user_pool_snapshot(item)]
        source_type_counts: Counter[str] = Counter()
        tag_counts: Counter[str] = Counter()
        risk_tag_counts: Counter[str] = Counter()
        layer_counts: Counter[str] = Counter()
        route_family_counts: Counter[str] = Counter()
        detail_family_counts: Counter[str] = Counter()
        hash_counts: Counter[str] = Counter()
        pool_assets: Counter[str] = Counter()
        title_reference_counts: Counter[str] = Counter()
        route_text_counts: Counter[str] = Counter()
        texture_text_counts: Counter[str] = Counter()
        opening_text_counts: Counter[str] = Counter()
        for item in used_items:
            pool = self._real_user_pool_snapshot(item)
            if pool.get("asset_key"):
                pool_assets[str(pool.get("asset_key"))] += 1
            for source_type, count in (pool.get("source_type_counts") or {}).items():
                source_type_counts[str(source_type)] += int(count or 0)
            for tag, count in (pool.get("tag_counts") or {}).items():
                tag_counts[str(tag)] += int(count or 0)
            for tag, count in (pool.get("risk_tag_counts") or {}).items():
                risk_tag_counts[str(tag)] += int(count or 0)
            for layer, count in (pool.get("layer_counts") or {}).items():
                layer_counts[str(layer)] += int(count or 0)
            for family, count in (pool.get("route_family_counts") or {}).items():
                route_family_counts[str(family)] += int(count or 0)
            for family, count in (pool.get("detail_family_counts") or {}).items():
                detail_family_counts[str(family)] += int(count or 0)
            for dedupe_hash in pool.get("dedupe_hashes") or []:
                if dedupe_hash:
                    hash_counts[str(dedupe_hash)] += 1
            title_reference = pool.get("title_reference") if isinstance(pool.get("title_reference"), dict) else {}
            for title in title_reference.get("selected_titles") or []:
                if title:
                    title_reference_counts[str(title)] += 1
            prompt_text_by_layer = pool.get("prompt_text_by_layer") if isinstance(pool.get("prompt_text_by_layer"), dict) else {}
            for text in prompt_text_by_layer.get("title_shape") or []:
                if text:
                    title_reference_counts[str(text)] += 1
            for text in prompt_text_by_layer.get("route") or []:
                if text:
                    route_text_counts[str(text)] += 1
            for text in prompt_text_by_layer.get("texture") or []:
                if text:
                    texture_text_counts[str(text)] += 1
            for layer in ("opening_texture", "ending"):
                for text in prompt_text_by_layer.get(layer) or []:
                    if text:
                        opening_text_counts[str(text)] += 1
        repeated = [
            {"dedupe_hash": dedupe_hash, "count": count}
            for dedupe_hash, count in hash_counts.most_common()
            if count > 1
        ]
        return {
            "enabled_item_count": len(used_items),
            "pool_assets": dict(pool_assets),
            "source_type_counts": dict(source_type_counts),
            "layer_counts": dict(layer_counts),
            "route_family_counts": dict(route_family_counts),
            "detail_family_counts": dict(detail_family_counts),
            "tag_counts": dict(tag_counts),
            "risk_tag_counts": dict(risk_tag_counts),
            "repeated_dedupe_hashes": repeated,
            "title_reference_repeat_top": _repeat_top(title_reference_counts),
            "route_repeat_top": _repeat_top(route_text_counts),
            "texture_repeat_top": _repeat_top(texture_text_counts),
            "opening_phrase_repeat_top": _repeat_top(opening_text_counts),
        }

    def _mouth_phrase_budget_stats(self, items: list[ContentBatchReportItem]) -> dict[str, Any]:
        budget_items = [item for item in items if self._mouth_phrase_budget_snapshot(item)]
        if not budget_items:
            return {}
        term_to_groups: dict[str, set[str]] = {}
        group_limits: dict[str, dict[str, Any]] = {}
        allowed_terms_by_item: dict[int, list[str]] = {}
        for item in budget_items:
            budget = self._mouth_phrase_budget_snapshot(item)
            allowed_terms_by_item[item.item_no] = self._list_of_strings(budget.get("allowed_terms"))
            for group in budget.get("groups") or []:
                if not isinstance(group, dict):
                    continue
                group_code = str(group.get("code") or "").strip()
                if not group_code:
                    continue
                group_limits.setdefault(
                    group_code,
                    {
                        "group_code": group_code,
                        "group_name": str(group.get("name") or "").strip(),
                        "terms": self._list_of_strings(group.get("terms")),
                        "max_count": group.get("max_count"),
                        "term_limits": group.get("term_limits") if isinstance(group.get("term_limits"), dict) else {},
                    },
                )
                for term in self._list_of_strings(group.get("terms")):
                    term_to_groups.setdefault(term, set()).add(group_code)
        term_hits: dict[str, list[dict[str, Any]]] = {term: [] for term in term_to_groups}
        for item in budget_items:
            title = item.title or ""
            body = item.body or ""
            for term in term_to_groups:
                title_count = title.count(term)
                body_count = body.count(term)
                if title_count or body_count:
                    term_hits[term].append(
                        {
                            "item_no": item.item_no,
                            "title_count": title_count,
                            "body_count": body_count,
                            "total_count": title_count + body_count,
                            "allowed": term in set(allowed_terms_by_item.get(item.item_no) or []),
                            "body_preview": body[:120],
                        }
                    )
        term_stats = []
        for term, hits in term_hits.items():
            count = sum(hit["total_count"] for hit in hits)
            item_nos = [hit["item_no"] for hit in hits]
            term_stats.append(
                {
                    "term": term,
                    "count": count,
                    "item_count": len(set(item_nos)),
                    "item_nos": item_nos,
                    "hits": hits,
                }
            )
        term_stats.sort(key=lambda item: (-item["count"], item["term"]))

        group_stats = []
        for group_code, group in group_limits.items():
            terms = group.get("terms") or []
            group_count = sum(sum(hit["total_count"] for hit in term_hits.get(term, [])) for term in terms)
            term_limits = group.get("term_limits") if isinstance(group.get("term_limits"), dict) else {}
            over_terms = [
                {
                    "term": term,
                    "count": sum(hit["total_count"] for hit in term_hits.get(term, [])),
                    "max_count": int(limit or 0),
                }
                for term, limit in term_limits.items()
                if sum(hit["total_count"] for hit in term_hits.get(term, [])) > int(limit or 0)
            ]
            max_count = group.get("max_count")
            over_group = max_count is not None and group_count > int(max_count or 0)
            group_stats.append(
                {
                    **group,
                    "count": group_count,
                    "over_budget": bool(over_group or over_terms),
                    "over_terms": over_terms,
                }
            )
        group_stats.sort(key=lambda item: (-item["count"], item["group_code"]))
        return {
            "enabled_item_count": len(budget_items),
            "term_stats": term_stats,
            "group_stats": group_stats,
            "over_budget_groups": [item for item in group_stats if item["over_budget"]],
        }

    def _business_usability_stats(self, items: list[ContentBatchReportItem]) -> dict[str, Any]:
        eligible_items: list[ContentBatchReportItem] = []
        excluded: list[dict[str, Any]] = []
        for item in items:
            tier = str(item.business_usability_tier or "").strip()
            if not tier:
                continue
            final_state = _final_postprocess_state(item.quality) if item.quality else {}
            final_pass = bool(final_state.get("hard_pass")) and not bool(final_state.get("rewrite_required"))
            if final_pass:
                eligible_items.append(item)
                continue
            excluded.append(
                {
                    "item_no": item.item_no,
                    "business_usability_tier": tier,
                    "rewrite_reason": final_state.get("rewrite_reason") or item.rewrite_reason,
                    "reasons": final_state.get("reasons") or [],
                }
            )

        counts = Counter(str(item.business_usability_tier or "").strip() for item in eligible_items)
        if not counts:
            return {"excluded_by_final_postprocess": excluded} if excluded else {}
        tiers = ("direct_pool", "light_fix_usable", "hold_out")
        stats = {
            "counts": {tier: counts.get(tier, 0) for tier in tiers},
            "item_nos_by_tier": {
                tier: [
                    item.item_no
                    for item in eligible_items
                    if str(item.business_usability_tier or "").strip() == tier
                ]
                for tier in tiers
                if counts.get(tier, 0)
            },
        }
        if excluded:
            stats["excluded_by_final_postprocess"] = excluded
        return stats

    def _real_user_pool_snapshot(self, item: ContentBatchReportItem) -> dict[str, Any]:
        snapshot = item.generation_snapshot or {}
        business_rule = snapshot.get("business_rule") if isinstance(snapshot.get("business_rule"), dict) else {}
        pool = business_rule.get("real_user_pool") if isinstance(business_rule.get("real_user_pool"), dict) else {}
        return pool

    def _mouth_phrase_budget_snapshot(self, item: ContentBatchReportItem) -> dict[str, Any]:
        snapshot = item.generation_snapshot or {}
        business_rule = snapshot.get("business_rule") if isinstance(snapshot.get("business_rule"), dict) else {}
        budget = business_rule.get("mouth_phrase_budget") if isinstance(business_rule.get("mouth_phrase_budget"), dict) else {}
        return budget if budget.get("enabled") is True else {}

    def _max_pairwise_jaccard(self, bodies: list[str]) -> float:
        max_score = 0.0
        for i, left in enumerate(bodies):
            for right in bodies[i + 1 :]:
                max_score = max(max_score, self._jaccard_2gram(left, right))
        return round(max_score, 4)

    def _jaccard_2gram(self, left: str, right: str) -> float:
        left_tokens = self._text_2grams(left)
        right_tokens = self._text_2grams(right)
        if not left_tokens or not right_tokens:
            return 0.0
        return len(left_tokens & right_tokens) / len(left_tokens | right_tokens)

    def _text_2grams(self, text: str) -> set[str]:
        clean = re.sub(r"\s+", "", text or "")
        return {clean[i : i + 2] for i in range(max(len(clean) - 1, 0)) if clean[i : i + 2].strip()}

    def _attach_similarity_warnings(self, items: list[ContentBatchReportItem]) -> None:
        comparable = [item for item in items if item.body]
        for index, left in enumerate(comparable):
            for right in comparable[index + 1 :]:
                score = round(self._jaccard_2gram(left.body or "", right.body or ""), 4)
                if score < SIMILARITY_WARNING_THRESHOLD:
                    continue
                warning_for_left = ContentBatchSimilarityWarning(
                    item_no=right.item_no,
                    score=score,
                    reason="正文 2-gram 相似度偏高",
                    scope="current_batch",
                )
                warning_for_right = ContentBatchSimilarityWarning(
                    item_no=left.item_no,
                    score=score,
                    reason="正文 2-gram 相似度偏高",
                    scope="current_batch",
                )
                left.similarity_warnings.append(warning_for_left)
                right.similarity_warnings.append(warning_for_right)
        self._attach_rewrite_similarity_warnings(items)
        self._attach_quality_similarity_watch_warnings(items)

    def _attach_rewrite_similarity_warnings(self, items: list[ContentBatchReportItem]) -> None:
        for item in items:
            quality = item.quality or {}
            rewrites = quality.get("similarity_rewrites") or []
            if not isinstance(rewrites, list):
                continue
            for rewrite in rewrites:
                if not isinstance(rewrite, dict):
                    continue
                scope = rewrite.get("scope") or "current_batch"
                if scope != "history":
                    continue
                similar_item_no = rewrite.get("similar_item_no")
                score = rewrite.get("post_rewrite_similarity_score") or rewrite.get("similarity_score")
                if not isinstance(similar_item_no, int) or not isinstance(score, (int, float)):
                    continue
                if rewrite.get("similarity_rewrite_passed") is True and float(score) < SIMILARITY_WARNING_THRESHOLD:
                    continue
                item.similarity_warnings.append(
                    ContentBatchSimilarityWarning(
                        item_no=similar_item_no,
                        score=round(float(score), 4),
                        reason="与历史批次正文 2-gram 相似度偏高",
                        batch_id=rewrite.get("similar_batch_id"),
                        batch_code=rewrite.get("similar_batch_code"),
                        scope="history",
                    )
                )

    def _attach_quality_similarity_watch_warnings(self, items: list[ContentBatchReportItem]) -> None:
        for item in items:
            quality = item.quality or {}
            watches = quality.get("similarity_watch") or []
            if not isinstance(watches, list):
                continue
            for watch in watches:
                if not isinstance(watch, dict):
                    continue
                scope = str(watch.get("scope") or "current_batch")
                if scope != "history":
                    continue
                similar_item_no = watch.get("similar_item_no")
                score = watch.get("similarity_score")
                if not isinstance(similar_item_no, int) or not isinstance(score, (int, float)):
                    continue
                item.similarity_warnings.append(
                    ContentBatchSimilarityWarning(
                        item_no=similar_item_no,
                        score=round(float(score), 4),
                        reason="与历史批次正文 2-gram 相似度偏高",
                        batch_id=watch.get("similar_batch_id"),
                        batch_code=watch.get("similar_batch_code"),
                        scope="history",
                    )
                )


def _feedback_categories(metadata: dict[str, Any]) -> list[str]:
    values = metadata.get("feedback_categories")
    if not isinstance(values, list):
        return []
    categories: list[str] = []
    for value in values:
        code = str(value or "").strip()
        if code in _FEEDBACK_CATEGORY_LABELS and code not in categories:
            categories.append(code)
    return categories


def _feedback_evidence(feedback: ContentFeedback, item: ContentBatchItem | None) -> str | None:
    parts: list[str] = []
    if item is not None:
        parts.append(f"第 {item.item_no} 条")
    if feedback.quoted_text:
        parts.append(f"片段：{_truncate(feedback.quoted_text, 80)}")
    if feedback.comment:
        parts.append(f"反馈：{_truncate(feedback.comment, 100)}")
    return "；".join(parts) if parts else None


def _counter_stats(counter: Counter[str], labels: dict[str, str]) -> list[ContentBatchFeedbackStat]:
    label_order = {code: index for index, code in enumerate(labels)}
    return [
        ContentBatchFeedbackStat(code=code, label=labels.get(code, code), count=count)
        for code, count in sorted(
            counter.items(),
            key=lambda item: (-item[1], label_order.get(item[0], len(label_order)), labels.get(item[0], item[0])),
        )
        if count > 0
    ]


def _evidence_for_categories(evidence_by_category: dict[str, list[str]], categories: list[str]) -> list[str]:
    evidence: list[str] = []
    for category in categories:
        for item in evidence_by_category.get(category, []):
            if item not in evidence:
                evidence.append(item)
            if len(evidence) >= 4:
                return evidence
    return evidence


def _batch_content_type(items: list[ContentBatchItem]) -> str:
    for item in items:
        plan = item.plan_json or {}
        output_fields = plan.get("output_fields") or []
        if output_fields == ["comment"]:
            return "comment"
        if "body" in output_fields:
            return "article"
        if plan.get("rule_type") == "business_rule":
            return "comment"
    return "article"


def _system_keyword_target(category_counter: Counter[str], content_type: str) -> str:
    if category_counter.get("too_long", 0):
        return "表达扩散语料 / 评论格式控制" if content_type == "comment" else "表达扩散语料 / 帖子格式控制"
    if content_type == "comment":
        return "表达扩散语料 / 生文指令"
    return "表达扩散语料 / 写作手法"


def _suggestion_priority(count: int, total: int) -> str:
    ratio = count / max(total, 1)
    if count >= 3 or ratio >= 0.5:
        return "high"
    if count >= 2 or ratio >= 0.25:
        return "medium"
    return "low"


def _truncate(value: str, limit: int) -> str:
    text = str(value or "").strip()
    return text if len(text) <= limit else f"{text[:limit]}..."


def _write_overview_sheet(sheet: Any, report: ContentBatchReportResponse) -> None:
    summary = report.summary
    business_counts = summary.business_usability_stats.get("counts") or {}
    rows = [
        ("批次ID", report.batch_id),
        ("批次Code", report.batch_code or ""),
        ("资产Key", report.asset_key),
        ("主题", report.product_topic),
        ("人群", report.target_audience or "-"),
        ("人设/对象", report.persona_target or "-"),
        ("风格", report.style or "-"),
        ("状态", report.status),
        ("总数", summary.total_count),
        ("已生成", summary.generated_count),
        ("失败", summary.failed_count),
        ("最终机器通过", summary.hard_pass_count),
        ("批次同质化提醒", summary.batch_variation_warning_count),
        ("交付候选", summary.delivery_candidate_count),
        ("交付入选", summary.delivery_selected_count),
        ("交付缺口", summary.delivery_shortfall_count),
        ("建议整批补量", summary.suggested_bulk_refill_count),
        ("自动改写", summary.rewrite_item_count),
        ("仍需处理", summary.remaining_rewrite_required_count),
        (
            "业务可用性",
            (
                f"直接入池 {business_counts.get('direct_pool', 0)} / "
                f"轻修可用 {business_counts.get('light_fix_usable', 0)} / "
                f"暂不入池 {business_counts.get('hold_out', 0)}"
            )
            if business_counts
            else "-",
        ),
        ("禁用词命中", summary.forbidden_hit_count),
        ("相似提醒", summary.similarity_warning_count),
        ("平均字数", summary.avg_body_chars if summary.avg_body_chars is not None else "-"),
    ]
    sheet.append(["字段", "值"])
    for row in rows:
        sheet.append(list(row))
    _style_table_header(sheet, header_row=1, column_count=2)
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 48
    sheet.freeze_panes = "A2"


def _write_result_sheet(sheet: Any, report: ContentBatchReportResponse) -> None:
    headers = [
        "标题",
        "正文",
        "业务规则",
        "序号",
        "状态",
        "字数",
        "最终机器通过",
        "业务可用性",
        "业务可用性原因",
        "审核状态",
        "改写轮次",
        "改写原因",
        "禁用词",
        "相似提醒",
        "运行模式",
        "生文耗时ms",
        "总耗时ms",
        "Run ID",
        "Task ID",
        "系统语料包",
        "Expert",
        "模型",
        "错误信息",
        "批次多样性通过",
        "交付入选",
        "交付序号",
        "未入选原因",
    ]
    sheet.append(headers)
    for item in report.items:
        snapshot = item.generation_snapshot or {}
        sheet.append(
            [
                item.title or "",
                item.body or "",
                _business_rule_label(snapshot.get("business_rule")),
                item.item_no,
                item.status,
                item.body_chars,
                _bool_label(item.hard_pass),
                _business_usability_label(item.business_usability_tier),
                item.business_usability_reason or "",
                item.review_status or "",
                item.rewrite_rounds or 0,
                item.rewrite_reason or "",
                "、".join(item.forbidden_hits or []),
                _similarity_text(item.similarity_warnings),
                item.runtime_mode or "",
                item.generation_duration_ms if item.generation_duration_ms is not None else "",
                item.total_duration_ms if item.total_duration_ms is not None else "",
                item.trace_run_id or item.run_id or "",
                item.task_id or "",
                _keyword_asset_label(snapshot.get("keyword_asset")),
                _expert_label(snapshot.get("expert")),
                _model_label(snapshot),
                item.error_message or "",
                _bool_label(item.batch_variation_pass),
                _bool_label(item.delivery_selected),
                item.delivery_rank or "",
                item.delivery_non_selection_reason or "",
            ]
        )
    _style_table_header(sheet, header_row=1, column_count=len(headers))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 24,
        "B": 58,
        "C": 32,
        "D": 8,
        "E": 12,
        "F": 8,
        "G": 12,
        "H": 12,
        "I": 10,
        "J": 36,
        "K": 18,
        "L": 28,
        "M": 12,
        "N": 12,
        "O": 12,
        "P": 12,
        "Q": 12,
        "R": 28,
        "S": 26,
        "T": 24,
        "U": 36,
        "V": 26,
        "W": 24,
        "X": 16,
        "Y": 12,
        "Z": 12,
        "AA": 28,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 42


def _build_article_pool_csv(report: ContentBatchReportResponse) -> bytes:
    headers = ["标题", "正文", "上下文变量(context_list)"]
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(headers)
    for item in _article_pool_export_items(report.items):
        context_list = _article_pool_context_list(item)
        writer.writerow(
            [
                item.title or "",
                item.body or "",
                json.dumps(context_list, ensure_ascii=False),
            ]
        )
    return output.getvalue().encode("utf-8-sig")


def _article_pool_export_items(items: list[ContentBatchReportItem]) -> list[ContentBatchReportItem]:
    # 文章池导出只保留可直接入库的正文，审核失败或仍需重写的行留在批次报告里看。
    return [item for item in items if _article_pool_item_exportable(item)]


def _article_pool_item_exportable(item: ContentBatchReportItem) -> bool:
    if str(item.status or "") != "generated":
        return False
    if not str(item.body or "").strip():
        return False
    if item.hard_pass is not True:
        return False
    if item.rewrite_required is True:
        return False
    if item.quality and not _final_postprocess_state(item.quality)["hard_pass"]:
        return False
    return True


def _article_pool_context_list(item: ContentBatchReportItem) -> dict[str, str]:
    quality = item.quality or {}
    guard = quality.get("activity_quality_guard") if isinstance(quality, dict) else {}
    if isinstance(guard, dict) and isinstance(guard.get("context_list"), dict):
        return {
            str(key): str(value or "")
            for key, value in guard["context_list"].items()
        }

    snapshot = item.generation_snapshot or {}
    business_rule = snapshot.get("business_rule") if isinstance(snapshot.get("business_rule"), dict) else {}
    plan = dict(business_rule or {})
    plan["unified_generation"] = {"selected_keywords": snapshot.get("selected_keywords") or []}
    pseudo_item = SimpleNamespace(
        plan_json=plan,
        quality_json=quality,
        title=item.title,
        body=item.body,
    )
    return build_article_pool_context_list(pseudo_item)


def _final_postprocess_state(quality: dict[str, Any] | None) -> dict[str, Any]:
    quality = quality if isinstance(quality, dict) else {}
    if not quality:
        return {
            "hard_pass": False,
            "rewrite_required": False,
            "rewrite_reason": "quality_missing",
            "reasons": ["quality_missing"],
        }
    review_report = quality.get("review_report") if isinstance(quality.get("review_report"), dict) else {}
    reasons: list[str] = []
    rewrite_reasons: list[str] = []

    if quality.get("hard_pass") is False:
        reasons.append("hard_pass_false")
    if review_report.get("rewrite_required") is True:
        reason = str(review_report.get("rewrite_reason") or "review_report_rewrite_required")
        reasons.append(reason)
        rewrite_reasons.append(reason)
    if quality.get("postprocess_blocked"):
        reasons.append("postprocess_blocked")
        rewrite_reasons.append("postprocess_blocked")

    _append_blocking_review_reason(
        reasons,
        rewrite_reasons,
        "forbidden_terms_review",
        quality.get("forbidden_terms_review") or review_report.get("forbidden_terms_review"),
    )
    _append_blocking_review_reason(
        reasons,
        rewrite_reasons,
        "mouth_phrase_budget_guard",
        quality.get("mouth_phrase_budget_guard") or review_report.get("mouth_phrase_budget_guard"),
    )
    _append_blocking_review_reason(
        reasons,
        rewrite_reasons,
        "ai_flavor_humanizer",
        quality.get("ai_flavor_humanizer") or review_report.get("ai_flavor_review"),
    )
    _append_blocking_review_reason(
        reasons,
        rewrite_reasons,
        "product_experience_phrase_guard",
        quality.get("product_experience_phrase_guard") or review_report.get("product_experience_phrase_review"),
    )
    llm_review = quality.get("product_experience_llm_quality_review") or review_report.get(
        "product_experience_llm_review"
    )
    _append_blocking_review_reason(reasons, rewrite_reasons, "product_experience_llm_quality_review", llm_review)

    focused_review = quality.get("wangyue_focused_pipeline_review") or review_report.get(
        "wangyue_focused_pipeline_review"
    )
    if isinstance(focused_review, dict):
        unavailable_dimensions = focused_review.get("unavailable_dimensions") or []
        if unavailable_dimensions:
            reasons.append("wangyue_focused_pipeline_unavailable")
        elif focused_review.get("decision") == "block":
            reasons.append("wangyue_focused_pipeline_block")
            if focused_review.get("requires_rewrite") or focused_review.get("status") == "manual_review":
                rewrite_reasons.append("wangyue_focused_pipeline_block")
        elif focused_review.get("can_auto_pool") is False:
            reasons.append("wangyue_focused_pipeline_hold")

    length_guard = quality.get("article_length_guard")
    if isinstance(length_guard, dict) and length_guard.get("pass") is False:
        reasons.append("article_length_guard")
        rewrite_reasons.append("article_length_guard")

    # A failed LLM quality review is not a text issue by itself, but it means the
    # full audit chain did not complete. If no later successful review exists,
    # keep the item out of machine-pass/export pools.
    llm_failures = quality.get("product_experience_llm_quality_failures")
    llm_failure_mark_only = quality.get("product_experience_llm_quality_review_unavailable_mark_only") is True
    if (
        isinstance(llm_failures, list)
        and llm_failures
        and not _review_passed(llm_review)
        and not llm_failure_mark_only
    ):
        reasons.append("product_experience_llm_quality_review_failed")
        rewrite_reasons.append("product_experience_llm_quality_review_failed")

    return {
        "hard_pass": not reasons,
        "rewrite_required": bool(rewrite_reasons),
        "rewrite_reason": "；".join(_dedupe_reason_list(rewrite_reasons)) or None,
        "reasons": _dedupe_reason_list(reasons),
    }


def _append_blocking_review_reason(
    reasons: list[str],
    rewrite_reasons: list[str],
    source: str,
    review: Any,
) -> None:
    if not isinstance(review, dict):
        return
    if review.get("mark_rewrite_required") is False:
        return
    if review.get("final_hits"):
        reasons.append(source)
        rewrite_reasons.append(source)
        return
    if review.get("rewrite_required") is True or review.get("mark_rewrite_required") is True:
        reasons.append(source)
        rewrite_reasons.append(source)
        return
    if review.get("pass") is False or review.get("pass_") is False:
        reasons.append(source)


def _review_passed(review: Any) -> bool:
    if not isinstance(review, dict):
        return False
    if review.get("mark_rewrite_required") is False:
        return True
    if review.get("pass") is True or review.get("pass_") is True:
        return not review.get("rewrite_required")
    return False


def _last_failure_error(failures: list[Any]) -> str | None:
    for item in reversed(failures):
        if not isinstance(item, dict):
            if item:
                return str(item)
            continue
        message = item.get("error_message") or item.get("message") or item.get("reason")
        if message:
            return str(message)
    return None


def _dedupe_reason_list(reasons: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for reason in reasons:
        normalized = str(reason or "").strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        result.append(normalized)
    return result


def _style_table_header(sheet: Any, *, header_row: int, column_count: int) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    side = Side(style="thin", color="D9D9D9")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in sheet.iter_rows(min_row=header_row, max_row=sheet.max_row, max_col=column_count):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _bool_label(value: bool | None) -> str:
    if value is True:
        return "通过"
    if value is False:
        return "未通过"
    return "未知"


def _business_usability_from_quality(quality: dict[str, Any]) -> dict[str, str | None]:
    review = quality.get("product_experience_llm_quality_review")
    if not isinstance(review, dict):
        return {"tier": None, "reason": None}
    tier = str(review.get("business_usability_tier") or "").strip() or None
    reason = str(review.get("business_usability_reason") or "").strip() or None
    return {"tier": tier, "reason": reason}


def _business_usability_label(value: str | None) -> str:
    labels = {
        "direct_pool": "直接入池",
        "light_fix_usable": "轻修可用",
        "hold_out": "暂不入池",
    }
    return labels.get(str(value or "").strip(), str(value or "").strip())


def _similarity_text(warnings: list[ContentBatchSimilarityWarning]) -> str:
    if not warnings:
        return ""
    return "；".join(
        f"与第{warning.item_no}条 {round(warning.score * 100)}%"
        + (f"（{warning.batch_code}）" if warning.batch_code else "")
        for warning in warnings
    )


def _business_rule_label(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    for key in ("business_rule", "topic", "rule_id"):
        text = str(value.get(key) or "").strip()
        if text:
            return text
    return str(value.get("rule_type") or "")


def _int_or_none(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _repeat_top(counter: Counter[str], *, limit: int = 10) -> list[dict[str, Any]]:
    return [
        {"text": text, "count": count}
        for text, count in counter.most_common(limit)
        if text and count > 1
    ]


def _keyword_asset_label(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    asset_key = str(value.get("asset_key") or "").strip()
    source = str(value.get("source") or "").strip()
    return f"{asset_key}（{source}）" if source else asset_key


def _expert_label(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    code = str(value.get("expert_config_code") or "").strip()
    source = str(value.get("source") or "").strip()
    return f"{code}（{source}）" if source else code


def _model_label(snapshot: dict[str, Any]) -> str:
    route = snapshot.get("model_route") if isinstance(snapshot.get("model_route"), dict) else {}
    expert = snapshot.get("expert") if isinstance(snapshot.get("expert"), dict) else {}
    model_config = snapshot.get("model_config") if isinstance(snapshot.get("model_config"), dict) else {}
    for source in (route, model_config, expert.get("model_config") if isinstance(expert.get("model_config"), dict) else {}):
        for key in ("model_code", "ge_model", "model"):
            text = str(source.get(key) or "").strip()
            if text:
                return text
    return ""


def _excel_filename(report: ContentBatchReportResponse) -> str:
    batch_code = re.sub(r"[^0-9A-Za-z_-]+", "_", report.batch_code or f"batch_{report.batch_id}").strip("_")
    return f"生文结果_{batch_code or report.batch_id}.xlsx"


def _article_pool_csv_filename(report: ContentBatchReportResponse) -> str:
    topic = re.sub(r"[^0-9A-Za-z\u4e00-\u9fa5_-]+", "_", report.product_topic or "评论").strip("_")
    timestamp = datetime.now().strftime("%Y%m%d-%H%M")
    return f"生成{topic or '评论'}-{timestamp}.csv"
