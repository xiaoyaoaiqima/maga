"""Import operator-authored comment-angle rule sets into MAGA assets."""
from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.maga_assets import AssetImportRun, AssetRegistry

COMMENT_ANGLE_RULE_ASSET_TYPE = "comment_angle_rule_set"
DEFAULT_COMMENT_ANGLE_ASSET_KEY = "yuanyue_comment_activity"
DEFAULT_COMMENT_BATCH_TOPIC = "美素佳儿源悦活动评论"
DEFAULT_COMMENT_BATCH_LIMIT = 10


@dataclass(slots=True)
class CommentAngleRuleSetImportResult:
    import_run_id: int
    asset_id: int
    asset_key: str
    keyword_asset_key: str | None
    quality_guard_profile_key: str | None
    keyword_selection: dict[str, Any] | None
    source_hash: str
    rule_count: int
    example_count: int
    warnings: list[str]


async def import_comment_angle_rule_set(
    db: AsyncSession,
    file_content: bytes,
    *,
    source_name: str,
    asset_key: str = DEFAULT_COMMENT_ANGLE_ASSET_KEY,
    display_name: str | None = None,
    keyword_asset_key: str | None = None,
    quality_guard_profile_key: str | None = None,
    keyword_selection: dict[str, Any] | str | None = None,
    created_by: str = "maga-operator",
) -> CommentAngleRuleSetImportResult:
    """Persist a source-yue comment-angle CSV/XLSX as one versioned rule asset."""
    normalized_asset_key = (asset_key or DEFAULT_COMMENT_ANGLE_ASSET_KEY).strip() or DEFAULT_COMMENT_ANGLE_ASSET_KEY
    source_hash = hashlib.sha256(file_content).hexdigest()
    rows = _read_rule_rows(file_content, source_name=source_name)
    items = [_row_to_rule_item(row, index) for index, row in enumerate(rows, start=1)]
    items = [item for item in items if item is not None]
    if not items:
        raise ValueError("comment angle rule set is empty")

    example_count = sum(len(item.get("examples") or []) + len(item.get("supplements") or []) for item in items)
    warnings = _warnings_for_items(items)
    activity_name = _activity_name_for_import(normalized_asset_key, display_name)
    content_json = {
        "rule_type": "comment_angle",
        "rule_package_type": "comment_angle",
        "business_rule_package": True,
        "activity_name": activity_name,
        "default_generation_count": DEFAULT_COMMENT_BATCH_LIMIT,
        "items": items,
    }
    normalized_keyword_asset_key = _normalize_keyword_asset_key(keyword_asset_key)
    if normalized_keyword_asset_key:
        content_json["keyword_asset_key"] = normalized_keyword_asset_key
    normalized_quality_guard_profile_key = _normalize_keyword_asset_key(quality_guard_profile_key)
    if normalized_quality_guard_profile_key:
        content_json["quality_guard_profile_key"] = normalized_quality_guard_profile_key
    normalized_keyword_selection = _normalize_keyword_selection(keyword_selection)
    if normalized_keyword_selection:
        content_json["keyword_selection"] = normalized_keyword_selection

    await db.execute(
        update(AssetRegistry)
        .where(
            AssetRegistry.asset_type == COMMENT_ANGLE_RULE_ASSET_TYPE,
            AssetRegistry.asset_key == normalized_asset_key,
            AssetRegistry.status == "active",
        )
        .values(status="archived")
    )
    asset = AssetRegistry(
        asset_type=COMMENT_ANGLE_RULE_ASSET_TYPE,
        asset_key=normalized_asset_key,
        display_name=display_name or "源悦-业务规则（评论切角）",
        version_no=await _next_asset_version(db, COMMENT_ANGLE_RULE_ASSET_TYPE, normalized_asset_key),
        status="active",
        asset_stage="production",
        source_name=source_name,
        source_uri=f"upload://{source_name}",
        source_hash=source_hash,
        content_json=content_json,
        metadata_json={
            "rule_type": "comment_angle",
            "rule_package_type": "comment_angle",
            "business_rule_package": True,
            "activity_name": activity_name,
            "default_generation_count": DEFAULT_COMMENT_BATCH_LIMIT,
            "rule_count": len(items),
            "example_count": example_count,
            "keyword_asset_key": normalized_keyword_asset_key,
            "quality_guard_profile_key": normalized_quality_guard_profile_key,
            "keyword_selection": normalized_keyword_selection,
            "warnings": warnings,
        },
        created_by=created_by,
    )
    db.add(asset)
    await db.flush()

    run = AssetImportRun(
        source_name=source_name,
        source_uri=f"upload://{source_name}",
        source_hash=source_hash,
        status="succeeded",
        imported_assets=1,
        summary_json={
            "asset_key": normalized_asset_key,
            "asset_type": COMMENT_ANGLE_RULE_ASSET_TYPE,
            "rule_count": len(items),
            "example_count": example_count,
            "keyword_asset_key": normalized_keyword_asset_key,
            "quality_guard_profile_key": normalized_quality_guard_profile_key,
            "keyword_selection": normalized_keyword_selection,
            "warnings": warnings,
        },
        created_by=created_by,
    )
    db.add(run)
    await db.flush()
    return CommentAngleRuleSetImportResult(
        import_run_id=run.id,
        asset_id=asset.id,
        asset_key=normalized_asset_key,
        keyword_asset_key=normalized_keyword_asset_key,
        quality_guard_profile_key=normalized_quality_guard_profile_key,
        keyword_selection=normalized_keyword_selection,
        source_hash=source_hash,
        rule_count=len(items),
        example_count=example_count,
        warnings=warnings,
    )


def comment_angle_import_summary(result: CommentAngleRuleSetImportResult) -> dict[str, Any]:
    return {
        "asset_type": COMMENT_ANGLE_RULE_ASSET_TYPE,
        "asset_key": result.asset_key,
        "keyword_asset_key": result.keyword_asset_key,
        "quality_guard_profile_key": result.quality_guard_profile_key,
        "keyword_selection": result.keyword_selection,
        "rule_count": result.rule_count,
        "example_count": result.example_count,
        "warnings": result.warnings,
        "default_generation_count": DEFAULT_COMMENT_BATCH_LIMIT,
    }


def _normalize_keyword_asset_key(value: str | None) -> str | None:
    normalized = str(value or "").strip()
    return normalized or None


def _normalize_keyword_selection(value: dict[str, Any] | str | None) -> dict[str, Any] | None:
    if value is None:
        return None
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        try:
            value = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise ValueError("keyword_selection must be valid JSON") from exc
    if not isinstance(value, dict):
        raise ValueError("keyword_selection must be a JSON object")
    normalized: dict[str, list[str]] = {}
    for category_code, codes in value.items():
        key = str(category_code or "").strip()
        if not key:
            continue
        if isinstance(codes, dict):
            codes = codes.get("include") or codes.get("codes") or codes.get("keyword_codes")
        if isinstance(codes, str):
            codes = re.split(r"[,，\s]+", codes)
        if not isinstance(codes, list):
            raise ValueError("keyword_selection values must be arrays or comma-separated strings")
        normalized_codes = [str(code).strip() for code in codes if str(code).strip()]
        if normalized_codes:
            normalized[key] = normalized_codes
    return normalized or None


def _activity_name_for_import(asset_key: str, display_name: str | None) -> str:
    if asset_key == DEFAULT_COMMENT_ANGLE_ASSET_KEY:
        return DEFAULT_COMMENT_BATCH_TOPIC
    normalized = str(display_name or "").strip()
    if normalized:
        normalized = re.sub(r"(?:评论)?切角(?:规则)?$", "评论", normalized).strip()
        return normalized or str(display_name or "").strip()
    return asset_key


def _read_rule_rows(file_content: bytes, *, source_name: str) -> list[dict[str, str]]:
    lower_name = source_name.lower()
    if lower_name.endswith(".xlsx"):
        return _read_xlsx_rows(file_content)
    if lower_name.endswith(".csv"):
        return _read_csv_rows(file_content)
    raise ValueError("only .csv and .xlsx files are supported")


def _read_csv_rows(file_content: bytes) -> list[dict[str, str]]:
    text = file_content.decode("utf-8-sig")
    content = "".join(line for line in text.splitlines(True) if not line.startswith("#"))
    return [
        {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
        for row in csv.DictReader(io.StringIO(content))
        if any(str(value or "").strip() for value in row.values())
    ]


def _read_xlsx_rows(file_content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows: list[dict[str, str]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {
            header: str(value or "").strip()
            for header, value in zip(headers, raw)
            if header
        }
        if any(row.values()):
            rows.append(row)
    return rows


def _row_to_rule_item(row: dict[str, str], index: int) -> dict[str, Any] | None:
    comment_angle, corpus = _angle_and_corpus(row)
    if not comment_angle or not corpus:
        return None
    examples = _line_examples(row.get("评论示例")) or _examples_from_corpus(corpus)
    supplements = _line_examples(row.get("评论补充"))
    return {
        "rule_id": f"comment_angle_{index:03d}",
        "comment_angle": comment_angle,
        "corpus": corpus,
        "examples": examples,
        "supplements": supplements,
        "source_row_no": index,
    }


def _angle_and_corpus(row: dict[str, str]) -> tuple[str, str]:
    if row.get("语料"):
        return row.get("评论切角", "").strip(), row.get("语料", "").strip()
    if row.get("评论类型") and row.get("评论切角"):
        return row.get("评论类型", "").strip(), row.get("评论切角", "").strip()
    return row.get("评论切角", "").strip(), row.get("语料", "").strip()


def _line_examples(value: str | None) -> list[str]:
    examples: list[str] = []
    for raw in (value or "").splitlines():
        line = _clean_example_line(raw)
        if line:
            examples.append(line)
    return examples


def _examples_from_corpus(corpus: str) -> list[str]:
    if "示例" not in corpus:
        return []
    after = re.split(r"示例[:：]", corpus, maxsplit=1)
    if len(after) < 2:
        return []
    body = re.split(r"\n\s*注意[:：]", after[1], maxsplit=1)[0]
    return [_clean_example_line(line) for line in body.splitlines() if _clean_example_line(line)]


def _clean_example_line(value: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^[\-\*•]\s*", "", text)
    text = re.sub(r"^\d+[、.．]\s*", "", text)
    return text.strip()


def _warnings_for_items(items: list[dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    missing_examples = [item["comment_angle"] for item in items if not item.get("examples") and not item.get("supplements")]
    if missing_examples:
        warnings.append(f"{len(missing_examples)} 条规则缺少参考示例")
    hard_rule_rows = [
        item["comment_angle"]
        for item in items
        if re.search(r"必须|只写|固定数字|生成重点必须|可写方向", item.get("corpus") or "")
    ]
    if hard_rule_rows:
        warnings.append(f"{len(hard_rule_rows)} 条规则含较硬约束，后续可用示例扩展替代加规则")
    return warnings


async def _next_asset_version(db: AsyncSession, asset_type: str, asset_key: str) -> int:
    result = await db.execute(
        select(AssetRegistry.version_no)
        .where(AssetRegistry.asset_type == asset_type, AssetRegistry.asset_key == asset_key)
        .order_by(AssetRegistry.version_no.desc())
        .limit(1)
    )
    current = result.scalar_one_or_none()
    return int(current or 0) + 1
