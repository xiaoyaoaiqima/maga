"""Import marketing asset corpora into MAGA-owned asset tables."""
from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.maga_assets import AssetImportRun, AssetRegistry


@dataclass(slots=True)
class AssetImportResult:
    import_run_id: int | None
    imported_assets: int
    asset_keys: list[tuple[str, str]]
    source_hash: str


async def import_yuanyue_training_rules(
    db: AsyncSession,
    workbook_path: str | Path,
    *,
    source_name: str = "源悦种草活动-ai训练规则.xlsx",
    asset_key: str = "yuanyue",
    created_by: str = "maga-asset-steward",
) -> AssetImportResult:
    """Import the 源悦 training-rule workbook as versioned MAGA assets."""
    path = Path(workbook_path)
    source_hash = _sha256_file(path)
    wb = load_workbook(path, read_only=False, data_only=True)

    assets: list[AssetRegistry] = []
    assets.append(
        await _build_asset(
            db,
            asset_type="brand_profile",
            asset_key=asset_key,
            display_name="源悦品牌资料",
            content=_parse_brand_profile(wb),
            source_name=source_name,
            source_uri=str(path),
            source_hash=source_hash,
            created_by=created_by,
        )
    )
    assets.append(
        await _build_asset(
            db,
            asset_type="product_selling_points",
            asset_key=asset_key,
            display_name="源悦产品卖点",
            content=_parse_product_selling_points(wb),
            source_name=source_name,
            source_uri=str(path),
            source_hash=source_hash,
            created_by=created_by,
        )
    )
    assets.append(
        await _build_asset(
            db,
            asset_type="painpoint_model",
            asset_key=asset_key,
            display_name="源悦痛点模型",
            content={"items": _parse_content_model(wb)},
            source_name=source_name,
            source_uri=str(path),
            source_hash=source_hash,
            created_by=created_by,
        )
    )
    assets.append(
        await _build_asset(
            db,
            asset_type="ugc_expression_corpus",
            asset_key=asset_key,
            display_name="源悦 UGC 卖点表述语料",
            content={"items": _parse_expression_sheet(wb, "ugc卖点表述", label_field="painpoint_or_selling_point")},
            source_name=source_name,
            source_uri=str(path),
            source_hash=source_hash,
            created_by=created_by,
        )
    )
    premium = _parse_expression_sheet(wb, "精品-卖点表述", label_field="painpoint")
    if premium:
        assets.append(
            await _build_asset(
                db,
                asset_type="premium_expression_corpus",
                asset_key=asset_key,
                display_name="源悦精品卖点表述语料",
                content={"items": premium},
                source_name=source_name,
                source_uri=str(path),
                source_hash=source_hash,
                created_by=created_by,
            )
        )
    reference_examples = _parse_reference_examples(wb)
    if reference_examples:
        assets.append(
            await _build_asset(
                db,
                asset_type="reference_examples",
                asset_key=asset_key,
                display_name="源悦参考例文",
                content={"items": reference_examples},
                source_name=source_name,
                source_uri=str(path),
                source_hash=source_hash,
                created_by=created_by,
            )
        )
    assets.append(
        await _build_asset(
            db,
            asset_type="compliance_rules",
            asset_key=asset_key,
            display_name="源悦审核规则",
            content={"items": _parse_compliance_rules(wb)},
            source_name=source_name,
            source_uri=str(path),
            source_hash=source_hash,
            created_by=created_by,
        )
    )

    for asset in assets:
        db.add(asset)
    await db.flush()

    run = AssetImportRun(
        source_name=source_name,
        source_uri=str(path),
        source_hash=source_hash,
        status="succeeded",
        imported_assets=len(assets),
        summary_json={
            "asset_key": asset_key,
            "asset_types": [asset.asset_type for asset in assets],
            "sheet_names": wb.sheetnames,
        },
        created_by=created_by,
    )
    db.add(run)
    await db.flush()
    return AssetImportResult(
        import_run_id=run.id,
        imported_assets=len(assets),
        asset_keys=[(asset.asset_type, asset.asset_key) for asset in assets],
        source_hash=source_hash,
    )


async def _build_asset(
    db: AsyncSession,
    *,
    asset_type: str,
    asset_key: str,
    display_name: str,
    content: dict[str, Any],
    source_name: str,
    source_uri: str,
    source_hash: str,
    created_by: str,
) -> AssetRegistry:
    version = await _next_version(db, asset_type, asset_key)
    return AssetRegistry(
        asset_type=asset_type,
        asset_key=asset_key,
        display_name=display_name,
        version_no=version,
        status="active",
        source_name=source_name,
        source_uri=source_uri,
        source_hash=source_hash,
        content_json=content,
        metadata_json={"importer": "yuanyue_training_rules_v1"},
        created_by=created_by,
    )


async def _next_version(db: AsyncSession, asset_type: str, asset_key: str) -> int:
    result = await db.execute(
        select(AssetRegistry.version_no)
        .where(AssetRegistry.asset_type == asset_type, AssetRegistry.asset_key == asset_key)
        .order_by(AssetRegistry.version_no.desc())
        .limit(1)
    )
    current = result.scalar_one_or_none()
    return int(current or 0) + 1


def _parse_brand_profile(wb) -> dict[str, Any]:
    ws = wb["品牌资料整理"]
    return {
        "brand_key": "yuanyue",
        "brand_name": "源悦",
        "content_focus": _clean(ws["C3"].value),
        "content_style": _clean(ws["C4"].value),
    }


def _parse_product_selling_points(wb) -> dict[str, Any]:
    ws = wb["品牌资料整理"]
    items: list[dict[str, Any]] = []
    current_level = None
    current_selling_point = None
    for row in range(9, ws.max_row + 1):
        level = _clean(ws.cell(row, 2).value) or current_level
        selling_point = _clean(ws.cell(row, 3).value) or current_selling_point
        ingredient = _clean(ws.cell(row, 4).value)
        advantage = _clean(ws.cell(row, 5).value)
        vivid_explanation = _clean(ws.cell(row, 6).value)
        if _emptyish(ingredient) and _emptyish(advantage) and _emptyish(vivid_explanation) and not selling_point:
            continue
        current_level = level
        current_selling_point = selling_point
        items.append(
            {
                "level": level,
                "selling_point": selling_point,
                "ingredient": None if _emptyish(ingredient) else ingredient,
                "advantage": None if _emptyish(advantage) else advantage,
                "vivid_explanation": None if _emptyish(vivid_explanation) else vivid_explanation,
            }
        )
    return {"items": items}


def _parse_content_model(wb) -> list[dict[str, Any]]:
    ws = wb["内容模型"]
    items: list[dict[str, Any]] = []
    current_stage = None
    current_painpoint = None
    for row in range(2, ws.max_row + 1):
        baby_stage = _clean(ws.cell(row, 2).value) or current_stage
        painpoint = _clean(ws.cell(row, 3).value) or current_painpoint
        symptom = _clean(ws.cell(row, 4).value)
        description = _clean(ws.cell(row, 5).value)
        selling_point = _clean(ws.cell(row, 6).value)
        extras = [_clean(ws.cell(row, col).value) for col in range(7, 11)]
        extras = [item for item in extras if not _emptyish(item)]
        if _emptyish(painpoint) and _emptyish(symptom) and _emptyish(description) and _emptyish(selling_point):
            continue
        current_stage = baby_stage
        current_painpoint = painpoint
        items.append(
            {
                "baby_stage": None if _emptyish(baby_stage) else baby_stage,
                "painpoint": painpoint,
                "symptom": None if _emptyish(symptom) else symptom,
                "description": None if _emptyish(description) else description,
                "selling_point": None if _emptyish(selling_point) else selling_point,
                "extra_descriptions": extras,
            }
        )
    return items


def _parse_expression_sheet(wb, sheet_name: str, *, label_field: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    items: list[dict[str, Any]] = []
    current_label = None
    for row in range(2, ws.max_row + 1):
        label = _clean(ws.cell(row, 2).value) or current_label
        expression = _clean(ws.cell(row, 3).value)
        owner = _clean(ws.cell(row, 4).value)
        if _emptyish(label) and _emptyish(expression):
            continue
        current_label = label
        if _emptyish(expression):
            continue
        items.append(
            {
                label_field: label,
                "expression": expression,
                "owner": None if _emptyish(owner) else owner,
            }
        )
    return items


def _parse_reference_examples(wb) -> list[dict[str, Any]]:
    if "例文收集" not in wb.sheetnames:
        return []
    ws = wb["例文收集"]
    header = [_clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    header_map = {name: idx + 1 for idx, name in enumerate(header) if name}
    real_workbook_shape = {"来源", "方向", "ID昵称", "发布链接", "发布形式", "笔记标题", "笔记正文"}.issubset(
        set(header_map)
    )

    items: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        if real_workbook_shape:
            title = _clean(ws.cell(row, header_map["笔记标题"]).value)
            body = _clean(ws.cell(row, header_map["笔记正文"]).value)
            if _emptyish(title) and _emptyish(body):
                continue
            source = _clean(ws.cell(row, header_map["来源"]).value)
            direction = _clean(ws.cell(row, header_map["方向"]).value)
            author = _clean(ws.cell(row, header_map["ID昵称"]).value)
            post_format = _clean(ws.cell(row, header_map["发布形式"]).value)
            item = {
                "example_id": f"yuanyue_ref_{len(items) + 1:03d}",
                "source": None if _emptyish(source) else source,
                "direction": None if _emptyish(direction) else direction,
                "author_name": None if _emptyish(author) else author,
                "follower_count_w": None if "粉丝量" not in header_map else _clean(ws.cell(row, header_map["粉丝量"]).value),
                "post_url": None if "发布链接" not in header_map else _clean(ws.cell(row, header_map["发布链接"]).value),
                "post_format": None if _emptyish(post_format) else post_format,
                "title": None if _emptyish(title) else title,
                "body": None if _emptyish(body) else body,
                "painpoint": None if "痛点" not in header_map else _clean(ws.cell(row, header_map["痛点"]).value),
                "reference_type": None if _emptyish(post_format) else post_format,
                "owner": None if _emptyish(author) else author,
                "style_tags": [tag for tag in [post_format, direction] if not _emptyish(tag)],
                "structure_tags": [],
            }
            items.append(item)
            continue

        title = _clean(ws.cell(row, 2).value)
        body = _clean(ws.cell(row, 3).value)
        reference_type = _clean(ws.cell(row, 4).value)
        owner = _clean(ws.cell(row, 5).value)
        if _emptyish(title) and _emptyish(body):
            continue
        index = len(items) + 1
        items.append(
            {
                "example_id": f"yuanyue_ref_{index:03d}",
                "title": None if _emptyish(title) else title,
                "body": None if _emptyish(body) else body,
                "reference_type": None if _emptyish(reference_type) else reference_type,
                "owner": None if _emptyish(owner) else owner,
                "style_tags": [] if _emptyish(reference_type) else [reference_type],
                "structure_tags": [],
            }
        )
    return items


def _parse_compliance_rules(wb) -> list[dict[str, Any]]:
    ws = wb["审核规则"]
    items: list[dict[str, Any]] = []
    current_content = None
    current_category = None
    for row in range(2, ws.max_row + 1):
        sequence = _clean(ws.cell(row, 1).value)
        audit_content = _clean(ws.cell(row, 2).value) or current_content
        category = _clean(ws.cell(row, 3).value) or current_category
        dimension = _clean(ws.cell(row, 4).value)
        feedback = _clean(ws.cell(row, 5).value)
        if _emptyish(dimension) and _emptyish(feedback):
            continue
        current_content = audit_content
        current_category = category
        items.append(
            {
                "sequence": sequence,
                "audit_content": audit_content,
                "category": category,
                "dimension": dimension,
                "feedback": feedback,
            }
        )
    return items


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _emptyish(value: Any) -> bool:
    if value is None:
        return True
    return str(value).strip() in {"", "/", "\\"}
