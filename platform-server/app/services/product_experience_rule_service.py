"""Import article business-rule packages into MAGA assets."""
from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.maga_assets import AssetImportRun, AssetRegistry
from app.services.business_rule_asset_types import (
    ARTICLE_BUSINESS_RULE_ASSET_TYPE,
    ARTICLE_BUSINESS_RULE_ASSET_TYPES,
)

PRODUCT_EXPERIENCE_RULE_ASSET_TYPE = ARTICLE_BUSINESS_RULE_ASSET_TYPE
DEFAULT_PRODUCT_EXPERIENCE_ASSET_KEY = "yuanyue_product_experience"
DEFAULT_PRODUCT_EXPERIENCE_ACTIVITY_NAME = "文章业务规则生文"
DEFAULT_PRODUCT_EXPERIENCE_BATCH_LIMIT = 10


@dataclass(slots=True)
class ProductExperienceRuleSetImportResult:
    import_run_id: int
    asset_id: int
    asset_key: str
    keyword_asset_key: str | None
    keyword_selection: dict[str, Any] | None
    source_hash: str
    rule_count: int
    example_count: int
    warnings: list[str]


async def import_product_experience_rule_set(
    db: AsyncSession,
    file_content: bytes,
    *,
    source_name: str,
    asset_key: str = DEFAULT_PRODUCT_EXPERIENCE_ASSET_KEY,
    display_name: str | None = None,
    keyword_asset_key: str | None = None,
    keyword_selection: dict[str, Any] | str | None = None,
    created_by: str = "maga-operator",
) -> ProductExperienceRuleSetImportResult:
    """Persist article business-rule CSV/XLSX as a versioned asset."""
    normalized_asset_key = (
        asset_key or DEFAULT_PRODUCT_EXPERIENCE_ASSET_KEY
    ).strip() or DEFAULT_PRODUCT_EXPERIENCE_ASSET_KEY
    source_hash = hashlib.sha256(file_content).hexdigest()
    rows = _read_rule_rows(file_content, source_name=source_name)
    items = [_row_to_rule_item(row, index) for index, row in enumerate(rows, start=1)]
    items = [item for item in items if item is not None]
    if not items:
        raise ValueError("article business rule set is empty")

    example_count = sum(len(item.get("examples") or []) + len(item.get("supplements") or []) for item in items)
    warnings = _warnings_for_items(items)
    activity_name = (
        _infer_activity_name(items)
        or _infer_activity_name_from_display_name(display_name)
        or _infer_activity_name_from_asset_key(normalized_asset_key)
        or DEFAULT_PRODUCT_EXPERIENCE_ACTIVITY_NAME
    )
    word_count = _infer_word_count(items)
    content_json = {
        "rule_type": "business_rule",
        "rule_package_type": "business_rule",
        "business_rule_package": True,
        "activity_name": activity_name,
        "default_generation_count": DEFAULT_PRODUCT_EXPERIENCE_BATCH_LIMIT,
        "allow_repeat_generation": True,
        "items": items,
    }
    if word_count:
        content_json["word_count"] = word_count
    normalized_keyword_asset_key = _normalize_keyword_asset_key(keyword_asset_key)
    if normalized_keyword_asset_key:
        content_json["keyword_asset_key"] = normalized_keyword_asset_key
    normalized_keyword_selection = _normalize_keyword_selection(keyword_selection)
    if normalized_keyword_selection:
        content_json["keyword_selection"] = normalized_keyword_selection

    await db.execute(
        update(AssetRegistry)
        .where(
            AssetRegistry.asset_type.in_(ARTICLE_BUSINESS_RULE_ASSET_TYPES),
            AssetRegistry.asset_key == normalized_asset_key,
            AssetRegistry.status == "active",
        )
        .values(status="archived")
    )
    asset = AssetRegistry(
        asset_type=PRODUCT_EXPERIENCE_RULE_ASSET_TYPE,
        asset_key=normalized_asset_key,
        display_name=display_name or "源悦生文业务规则",
        version_no=await _next_asset_version(db, ARTICLE_BUSINESS_RULE_ASSET_TYPES, normalized_asset_key),
        status="active",
        asset_stage="production",
        source_name=source_name,
        source_uri=f"upload://{source_name}",
        source_hash=source_hash,
        content_json=content_json,
        metadata_json={
            "rule_type": "business_rule",
            "rule_package_type": "business_rule",
            "business_rule_package": True,
            "rule_count": len(items),
            "example_count": example_count,
            "keyword_asset_key": normalized_keyword_asset_key,
            "keyword_selection": normalized_keyword_selection,
            "activity_name": activity_name,
            "word_count": word_count,
            "default_generation_count": DEFAULT_PRODUCT_EXPERIENCE_BATCH_LIMIT,
            "allow_repeat_generation": True,
            "warnings": warnings,
        },
        created_by=created_by,
    )
    db.add(asset)
    await db.flush()

    run = AssetImportRun(
        source_name=source_name,
        source_uri=f"upload://{source_name}",
        source_hash=source_hash,
        status="succeeded",
        imported_assets=1,
        summary_json={
            "asset_key": normalized_asset_key,
            "asset_type": PRODUCT_EXPERIENCE_RULE_ASSET_TYPE,
            "rule_count": len(items),
            "example_count": example_count,
            "keyword_asset_key": normalized_keyword_asset_key,
            "keyword_selection": normalized_keyword_selection,
            "activity_name": activity_name,
            "word_count": word_count,
            "default_generation_count": DEFAULT_PRODUCT_EXPERIENCE_BATCH_LIMIT,
            "warnings": warnings,
        },
        created_by=created_by,
    )
    db.add(run)
    await db.flush()
    return ProductExperienceRuleSetImportResult(
        import_run_id=run.id,
        asset_id=asset.id,
        asset_key=normalized_asset_key,
        keyword_asset_key=normalized_keyword_asset_key,
        keyword_selection=normalized_keyword_selection,
        source_hash=source_hash,
        rule_count=len(items),
        example_count=example_count,
        warnings=warnings,
    )


def product_experience_import_summary(result: ProductExperienceRuleSetImportResult) -> dict[str, Any]:
    return {
        "asset_type": PRODUCT_EXPERIENCE_RULE_ASSET_TYPE,
        "asset_key": result.asset_key,
        "keyword_asset_key": result.keyword_asset_key,
        "keyword_selection": result.keyword_selection,
        "rule_count": result.rule_count,
        "example_count": result.example_count,
        "default_generation_count": DEFAULT_PRODUCT_EXPERIENCE_BATCH_LIMIT,
        "warnings": result.warnings,
    }


def _normalize_keyword_asset_key(value: str | None) -> str | None:
    normalized = str(value or "").strip()
    return normalized or None


def _normalize_keyword_selection(value: dict[str, Any] | str | None) -> dict[str, Any] | None:
    if value is None:
        return None
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        try:
            value = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise ValueError("keyword_selection must be valid JSON") from exc
    if not isinstance(value, dict):
        raise ValueError("keyword_selection must be a JSON object")
    normalized: dict[str, list[str]] = {}
    for category_code, codes in value.items():
        key = str(category_code or "").strip()
        if not key:
            continue
        if isinstance(codes, dict):
            codes = codes.get("include") or codes.get("codes") or codes.get("keyword_codes")
        if isinstance(codes, str):
            codes = re.split(r"[,，\s]+", codes)
        if not isinstance(codes, list):
            raise ValueError("keyword_selection values must be arrays or comma-separated strings")
        normalized_codes = [str(code).strip() for code in codes if str(code).strip()]
        if normalized_codes:
            normalized[key] = normalized_codes
    return normalized or None


def _infer_activity_name(items: list[dict[str, Any]]) -> str | None:
    activity_names: list[str] = []
    for item in items:
        corpus = str(item.get("corpus") or "")
        match = re.search(r"活动[:：]\s*([^。\n；;]+)", corpus)
        if match:
            activity_name = match.group(1).strip()
            if activity_name and activity_name not in activity_names:
                activity_names.append(activity_name)
    return activity_names[0] if len(activity_names) == 1 else None


def _infer_activity_name_from_display_name(display_name: str | None) -> str | None:
    text = str(display_name or "").strip()
    if not text:
        return None
    match = re.search(r"(\d{4}[^-_\s，,]*活动)", text)
    if match:
        return match.group(1).strip()
    if "旺玥" in text:
        return "0705旺玥活动"
    return text


def _infer_activity_name_from_asset_key(asset_key: str) -> str | None:
    if "wangyue" in str(asset_key or ""):
        return "0705旺玥活动"
    return None


def _infer_word_count(items: list[dict[str, Any]]) -> str | None:
    corpora = [str(item.get("corpus") or "") for item in items]
    has_mid = any("篇幅类型：中短文" in corpus for corpus in corpora)
    has_short = any("篇幅类型：短文" in corpus for corpus in corpora)
    mid_range = _first_match(
        corpora,
        r"(?:正文大约|可在)\s*(\d+)\s*-\s*(\d+)\s*字",
    )
    short_range = _first_match(corpora, r"篇幅类型：短文[\s\S]*?(?:正文大约|正文必须)?\s*(\d+)\s*-\s*(\d+)\s*字")
    if has_mid and has_short and mid_range and short_range:
        mid_start, mid_end = mid_range
        short_start, short_end = short_range
        return (
            f"逐条参考：中短文约{mid_start}-{mid_end}字，短一点但像真人不要硬扩写；"
            f"短文{short_start}-{short_end}字；标题不计；正文单段不换行"
        )
    return None


def _first_match(corpora: list[str], pattern: str) -> tuple[str, str] | None:
    for corpus in corpora:
        match = re.search(pattern, corpus)
        if match:
            return match.group(1), match.group(2)
    return None


def _read_rule_rows(file_content: bytes, *, source_name: str) -> list[dict[str, str]]:
    lower_name = source_name.lower()
    if lower_name.endswith(".xlsx"):
        return _read_xlsx_rows(file_content)
    if lower_name.endswith(".csv"):
        return _read_csv_rows(file_content)
    raise ValueError("only .csv and .xlsx files are supported")


def _read_csv_rows(file_content: bytes) -> list[dict[str, str]]:
    text = file_content.decode("utf-8-sig")
    content = "".join(line for line in text.splitlines(True) if not line.startswith("#"))
    return [
        {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
        for row in csv.DictReader(io.StringIO(content))
        if any(str(value or "").strip() for value in row.values())
    ]


def _read_xlsx_rows(file_content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows: list[dict[str, str]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {header: str(value or "").strip() for header, value in zip(headers, raw) if header}
        if any(row.values()):
            rows.append(row)
    return rows


def _row_to_rule_item(row: dict[str, str], index: int) -> dict[str, Any] | None:
    raw_rule = (
        row.get("业务规则名称")
        or row.get("业务规则")
        or row.get("规则名称")
        or row.get("business_rule")
        or ""
    ).strip()
    generation_instruction = (
        row.get("生文指令")
        or row.get("生成指令")
        or row.get("generation_instruction")
        or ""
    ).strip()
    content_direction = (
        row.get("内容方向")
        or row.get("这篇要写的事")
        or row.get("content_direction")
        or ""
    ).strip()
    corpus = (
        row.get("规则语料")
        or row.get("语料")
        or row.get("corpus")
        or content_direction
        or ""
    ).strip()
    if not raw_rule or not corpus:
        return None
    prompt_mode = (
        row.get("提示词模式")
        or row.get("prompt_mode")
        or row.get("generation_prompt_mode")
        or ""
    ).strip()
    activity_material = _cell_lines(row.get("活动素材") or row.get("activity_material"))
    prize_material = _cell_lines(
        row.get("奖品素材")
        or row.get("活动奖品素材")
        or row.get("prize_material")
    )
    batch_detection_material = _cell_lines(
        row.get("批批检素材")
        or row.get("批批检表达")
        or row.get("batch_detection_material")
    )
    selling_expression = (
        row.get("卖点表达")
        or row.get("selling_expression")
        or ""
    ).strip()
    selling_expression_note = (
        row.get("卖点表达说明")
        or row.get("卖点注意")
        or row.get("selling_expression_note")
        or ""
    ).strip()
    hard_boundaries = _cell_lines(
        row.get("事实与合规边界")
        or row.get("硬边界")
        or row.get("hard_boundaries")
    )
    writing_requirements = _cell_lines(
        row.get("成文要求")
        or row.get("写法")
        or row.get("writing_requirements")
    )
    layered_fields_present = any(
        (
            generation_instruction,
            content_direction,
            activity_material,
            prize_material,
            batch_detection_material,
            selling_expression,
            hard_boundaries,
            writing_requirements,
        )
    )
    if layered_fields_present and not prompt_mode:
        prompt_mode = "layered_article"
    variation_slots: list[dict[str, Any]] = []
    if prize_material:
        variation_slots.append(
            {
                "slot_code": "activity_prize",
                "slot_name": "活动奖品素材",
                "options": prize_material,
            }
        )
    if batch_detection_material:
        variation_slots.append(
            {
                "slot_code": "batch_detection",
                "slot_name": "批批检素材",
                "options": batch_detection_material,
            }
        )
    explicit_post_type = (
        row.get("帖子类型")
        or row.get("内容类型")
        or row.get("post_type")
        or ""
    ).strip()
    explicit_product_appearance_mode = (
        row.get("产品出现方式")
        or row.get("产品出现模式")
        or row.get("product_appearance_mode")
        or ""
    ).strip()
    inferred_post_type, inferred_product_mode = (
        _infer_product_permission_fields(raw_rule)
        if not explicit_post_type
        else (None, None)
    )
    post_type = (
        explicit_post_type
        or inferred_post_type
        or ""
    ).strip()
    product_appearance_mode = (
        explicit_product_appearance_mode
        or inferred_product_mode
        or ""
    ).strip()
    painpoint = (
        row.get("痛点")
        or row.get("核心痛点")
        or row.get("painpoint")
        or ""
    ).strip()
    selling_point = (
        row.get("卖点方向")
        or row.get("产品卖点")
        or row.get("产品依据")
        or row.get("selling_point")
        or row.get("product_basis")
        or ""
    ).strip()
    positive_evidence = (
        row.get("主正向证据")
        or row.get("正向证据")
        or row.get("效果证明")
        or row.get("positive_evidence")
        or ""
    ).strip()
    selling_point_surface = (
        row.get("卖点表达口吻")
        or row.get("卖点表述")
        or row.get("卖点表达")
        or row.get("selling_point_surface")
        or ""
    ).strip()
    ingredient_surface = (
        row.get("成分承接")
        or row.get("成分表达")
        or row.get("ingredient_surface")
        or ""
    ).strip()
    benefit_surface = (
        row.get("好处表达")
        or row.get("效果表达")
        or row.get("benefit_surface")
        or ""
    ).strip()
    selling_description = (
        row.get("卖点描述")
        or row.get("selling_description")
        or ""
    ).strip()
    selling_kernel = (
        row.get("种草内核")
        or row.get("selling_kernel")
        or _build_selling_kernel(
            painpoint=painpoint,
            selling_point=selling_point,
            positive_evidence=positive_evidence,
            selling_description=selling_description,
            selling_point_surface=selling_point_surface,
            ingredient_surface=ingredient_surface,
            benefit_surface=benefit_surface,
        )
        or ""
    ).strip()
    expression_mechanism = (
        row.get("表达机制")
        or row.get("写法机制")
        or row.get("expression_mechanism")
        or ""
    ).strip()
    ugc_post_type = (
        row.get("UGC类型")
        or row.get("帖子大类")
        or row.get("ugc_post_type")
        or ""
    ).strip()
    life_trigger = (
        row.get("生活动机")
        or row.get("发帖动机")
        or row.get("life_trigger")
        or ""
    ).strip()
    product_role = (
        row.get("产品角色")
        or row.get("product_role")
        or ""
    ).strip()
    product_relation = (
        row.get("产品关系")
        or row.get("product_relation")
        or _build_product_relation(
            product_appearance_mode=product_appearance_mode,
            product_role=product_role,
        )
        or ""
    ).strip()
    product_density = (
        row.get("产品浓度")
        or row.get("product_density")
        or ""
    ).strip()
    imperfection = (
        row.get("不完美感")
        or row.get("取舍")
        or row.get("imperfection")
        or ""
    ).strip()
    product_action_surface = (
        row.get("产品动作表面")
        or row.get("产品露出方式")
        or row.get("product_action_surface")
        or ""
    ).strip()
    title_shape_mode = (
        row.get("标题形态")
        or row.get("标题模式")
        or row.get("title_shape_mode")
        or ""
    ).strip()
    title_emoji_mode = (
        row.get("标题emoji")
        or row.get("标题 Emoji")
        or row.get("title_emoji_mode")
        or ""
    ).strip()
    scene_motive_bucket = (
        row.get("scene_motive_bucket")
        or row.get("正文场景")
        or row.get("场景动机")
        or row.get("场景桶")
        or ""
    ).strip()
    structure_slot = (
        row.get("结构槽位")
        or row.get("structure_slot")
        or ""
    ).strip()
    story_spine = (
        row.get("叙事主线")
        or row.get("内容主线")
        or row.get("story_spine")
        or ""
    ).strip()
    scene_constraint = (
        row.get("场景约束")
        or row.get("scene_constraint")
        or ""
    ).strip()
    product_position_mode = (
        row.get("产品出现位置")
        or row.get("产品位置")
        or row.get("product_position_mode")
        or ""
    ).strip()
    ending_mode = (
        row.get("收尾方式")
        or row.get("结尾方式")
        or row.get("ending_mode")
        or ""
    ).strip()
    examples = _line_examples(row.get("示例")) if "示例" in row else _examples_from_corpus(corpus)
    if "补充参考" in row:
        examples.extend(_line_examples(row.get("补充参考")))
    return {
        "rule_id": f"business_rule_{index:03d}",
        "business_rule": raw_rule,
        "corpus": corpus,
        "examples": examples,
        "supplements": [],
        "source_row_no": index,
        **({"prompt_mode": prompt_mode} if prompt_mode else {}),
        **({"generation_instruction": generation_instruction} if generation_instruction else {}),
        **({"content_direction": content_direction} if content_direction else {}),
        **({"activity_material": activity_material} if activity_material else {}),
        **({"selling_expression": selling_expression} if selling_expression else {}),
        **({"selling_expression_note": selling_expression_note} if selling_expression_note else {}),
        **({"hard_boundaries": hard_boundaries} if hard_boundaries else {}),
        **({"writing_requirements": writing_requirements} if writing_requirements else {}),
        **({"variation_slots": variation_slots} if variation_slots else {}),
        **({"post_type": post_type} if post_type else {}),
        **({"product_appearance_mode": product_appearance_mode} if product_appearance_mode else {}),
        **({"painpoint": painpoint} if painpoint else {}),
        **({"selling_point": selling_point} if selling_point else {}),
        **({"positive_evidence": positive_evidence} if positive_evidence else {}),
        **({"selling_point_surface": selling_point_surface} if selling_point_surface else {}),
        **({"ingredient_surface": ingredient_surface} if ingredient_surface else {}),
        **({"benefit_surface": benefit_surface} if benefit_surface else {}),
        **({"selling_description": selling_description} if selling_description else {}),
        **({"selling_kernel": selling_kernel} if selling_kernel else {}),
        **({"expression_mechanism": expression_mechanism} if expression_mechanism else {}),
        **({"ugc_post_type": ugc_post_type} if ugc_post_type else {}),
        **({"life_trigger": life_trigger} if life_trigger else {}),
        **({"product_role": product_role} if product_role else {}),
        **({"product_relation": product_relation} if product_relation else {}),
        **({"product_density": product_density} if product_density else {}),
        **({"imperfection": imperfection} if imperfection else {}),
        **({"product_action_surface": product_action_surface} if product_action_surface else {}),
        **({"title_shape_mode": title_shape_mode} if title_shape_mode else {}),
        **({"title_emoji_mode": title_emoji_mode} if title_emoji_mode else {}),
        **({"scene_motive_bucket": scene_motive_bucket} if scene_motive_bucket else {}),
        **({"structure_slot": structure_slot} if structure_slot else {}),
        **({"story_spine": story_spine} if story_spine else {}),
        **({"scene_constraint": scene_constraint} if scene_constraint else {}),
        **({"product_position_mode": product_position_mode} if product_position_mode else {}),
        **({"ending_mode": ending_mode} if ending_mode else {}),
    }


def _cell_lines(value: str | None) -> list[str]:
    lines: list[str] = []
    for raw in re.split(r"\r?\n|\s*\|\|\s*", str(value or "")):
        line = _clean_example_line(raw)
        if line and line not in lines:
            lines.append(line)
    return lines


def _build_selling_kernel(
    *,
    painpoint: str,
    selling_point: str,
    positive_evidence: str,
    selling_description: str,
    selling_point_surface: str,
    ingredient_surface: str,
    benefit_surface: str,
) -> str | None:
    parts: list[str] = []
    for label, value in [
        ("痛点", painpoint),
        ("卖点", selling_point),
        ("正向证据", positive_evidence),
        ("卖点描述", selling_description),
        ("卖点表达", selling_point_surface),
        ("成分承接", ingredient_surface),
        ("好处表达", benefit_surface),
    ]:
        text = str(value or "").strip().rstrip("。；;，, ")
        if text:
            parts.append(f"{label}：{text}")
    return "；".join(parts) if parts else None


def _build_product_relation(*, product_appearance_mode: str, product_role: str) -> str | None:
    parts: list[str] = []
    appearance = str(product_appearance_mode or "").strip().rstrip("。；;，, ")
    role = str(product_role or "").strip().rstrip("。；;，, ")
    if appearance:
        parts.append(f"出现方式：{appearance}")
    if role:
        parts.append(f"角色：{role}")
    return "；".join(parts) if parts else None


def _infer_product_permission_fields(raw_rule: str) -> tuple[str | None, str | None]:
    text = str(raw_rule or "").strip()
    if not text:
        return None, None
    for delimiter in ("｜", "|"):
        if delimiter not in text:
            continue
        post_type, product_mode = [part.strip() for part in text.split(delimiter, 1)]
        return post_type or None, product_mode or None
    return None, None


def _examples_from_corpus(corpus: str) -> list[str]:
    if "可参考素材" not in corpus:
        return []
    after = re.split(r"可参考素材[:：]", corpus, maxsplit=1)
    if len(after) < 2:
        return []
    body = re.split(r"\n\s*注意[:：]", after[1], maxsplit=1)[0]
    return [_clean_example_line(line) for line in body.splitlines() if _clean_example_line(line)]


def _line_examples(value: str | None) -> list[str]:
    examples: list[str] = []
    for raw in (value or "").splitlines():
        line = _clean_example_line(raw)
        if line:
            examples.append(line)
    return examples


def _clean_example_line(value: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^[\-\*•]\s*", "", text)
    text = re.sub(r"^\d+[、.．]\s*", "", text)
    return text.strip()


def _warnings_for_items(items: list[dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    missing_examples = [
        item["business_rule"]
        for item in items
        if not item.get("examples") and not item.get("supplements")
    ]
    if missing_examples:
        warnings.append(f"{len(missing_examples)} 条规则缺少示例")
    sparse_examples = [
        item["business_rule"]
        for item in items
        if 0 < len(item.get("examples") or []) + len(item.get("supplements") or []) < 3
    ]
    if sparse_examples:
        warnings.append(f"{len(sparse_examples)} 条规则示例少于3条")
    return warnings


async def _next_asset_version(db: AsyncSession, asset_types: tuple[str, ...], asset_key: str) -> int:
    result = await db.execute(
        select(AssetRegistry.version_no)
        .where(AssetRegistry.asset_type.in_(asset_types), AssetRegistry.asset_key == asset_key)
        .order_by(AssetRegistry.version_no.desc())
        .limit(1)
    )
    current = result.scalar_one_or_none()
    return int(current or 0) + 1
