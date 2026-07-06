"""Import operator-authored comment business-rule sets into MAGA assets."""
from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.maga_assets import AssetImportRun, AssetRegistry
from app.services.business_rule_asset_types import COMMENT_BUSINESS_RULE_ASSET_TYPES
from app.services.business_rule_asset_types import COMMENT_BUSINESS_RULE_ASSET_TYPE as COMMENT_BUSINESS_RULE_REGISTRY_TYPE

COMMENT_BUSINESS_RULE_ASSET_TYPE = COMMENT_BUSINESS_RULE_REGISTRY_TYPE
DEFAULT_COMMENT_BUSINESS_RULE_ASSET_KEY = "yuanyue_comment_activity"
DEFAULT_COMMENT_BATCH_TOPIC = "美素佳儿源悦活动评论"
DEFAULT_COMMENT_BATCH_LIMIT = 10
LEGACY_RULE_HEADER = "\u8bc4\u8bba" + "\u5207\u89d2"


@dataclass(slots=True)
class CommentBusinessRuleSetImportResult:
    import_run_id: int
    asset_id: int
    asset_key: str
    keyword_asset_key: str | None
    quality_guard_profile_key: str | None
    keyword_selection: dict[str, Any] | None
    source_hash: str
    rule_count: int
    example_count: int
    warnings: list[str]


async def import_comment_business_rule_set(
    db: AsyncSession,
    file_content: bytes,
    *,
    source_name: str,
    asset_key: str = DEFAULT_COMMENT_BUSINESS_RULE_ASSET_KEY,
    display_name: str | None = None,
    keyword_asset_key: str | None = None,
    quality_guard_profile_key: str | None = None,
    keyword_selection: dict[str, Any] | str | None = None,
    created_by: str = "maga-operator",
) -> CommentBusinessRuleSetImportResult:
    """Persist a comment business-rule CSV/XLSX as one versioned rule asset."""
    normalized_asset_key = (asset_key or DEFAULT_COMMENT_BUSINESS_RULE_ASSET_KEY).strip() or DEFAULT_COMMENT_BUSINESS_RULE_ASSET_KEY
    source_hash = hashlib.sha256(file_content).hexdigest()
    rows = _read_rule_rows(file_content, source_name=source_name)
    items = [_row_to_rule_item(row, index) for index, row in enumerate(rows, start=1)]
    items = [item for item in items if item is not None]
    if not items:
        raise ValueError("comment business rule set is empty")

    example_count = sum(len(item.get("examples") or []) + len(item.get("supplements") or []) for item in items)
    warnings = _warnings_for_items(items)
    activity_name = _activity_name_for_import(normalized_asset_key, display_name)
    content_json = {
        "rule_type": "business_rule",
        "rule_package_type": "business_rule",
        "business_rule_package": True,
        "activity_name": activity_name,
        "default_generation_count": DEFAULT_COMMENT_BATCH_LIMIT,
        "items": items,
    }
    normalized_keyword_asset_key = _normalize_keyword_asset_key(keyword_asset_key)
    if normalized_keyword_asset_key:
        content_json["keyword_asset_key"] = normalized_keyword_asset_key
    normalized_quality_guard_profile_key = _normalize_keyword_asset_key(quality_guard_profile_key)
    if normalized_quality_guard_profile_key:
        content_json["quality_guard_profile_key"] = normalized_quality_guard_profile_key
    normalized_keyword_selection = _normalize_keyword_selection(keyword_selection)
    if normalized_keyword_selection:
        content_json["keyword_selection"] = normalized_keyword_selection

    await db.execute(
        update(AssetRegistry)
        .where(
            AssetRegistry.asset_type.in_(COMMENT_BUSINESS_RULE_ASSET_TYPES),
            AssetRegistry.asset_key == normalized_asset_key,
            AssetRegistry.status == "active",
        )
        .values(status="archived")
    )
    asset = AssetRegistry(
        asset_type=COMMENT_BUSINESS_RULE_ASSET_TYPE,
        asset_key=normalized_asset_key,
        display_name=display_name or "源悦-评论业务规则",
        version_no=await _next_asset_version(db, COMMENT_BUSINESS_RULE_ASSET_TYPES, normalized_asset_key),
        status="active",
        asset_stage="production",
        source_name=source_name,
        source_uri=f"upload://{source_name}",
        source_hash=source_hash,
        content_json=content_json,
        metadata_json={
            "rule_type": "business_rule",
            "rule_package_type": "business_rule",
            "business_rule_package": True,
            "activity_name": activity_name,
            "default_generation_count": DEFAULT_COMMENT_BATCH_LIMIT,
            "rule_count": len(items),
            "example_count": example_count,
            "keyword_asset_key": normalized_keyword_asset_key,
            "quality_guard_profile_key": normalized_quality_guard_profile_key,
            "keyword_selection": normalized_keyword_selection,
            "warnings": warnings,
        },
        created_by=created_by,
    )
    db.add(asset)
    await db.flush()

    run = AssetImportRun(
        source_name=source_name,
        source_uri=f"upload://{source_name}",
        source_hash=source_hash,
        status="succeeded",
        imported_assets=1,
        summary_json={
            "asset_key": normalized_asset_key,
            "asset_type": COMMENT_BUSINESS_RULE_ASSET_TYPE,
            "rule_count": len(items),
            "example_count": example_count,
            "keyword_asset_key": normalized_keyword_asset_key,
            "quality_guard_profile_key": normalized_quality_guard_profile_key,
            "keyword_selection": normalized_keyword_selection,
            "warnings": warnings,
        },
        created_by=created_by,
    )
    db.add(run)
    await db.flush()
    return CommentBusinessRuleSetImportResult(
        import_run_id=run.id,
        asset_id=asset.id,
        asset_key=normalized_asset_key,
        keyword_asset_key=normalized_keyword_asset_key,
        quality_guard_profile_key=normalized_quality_guard_profile_key,
        keyword_selection=normalized_keyword_selection,
        source_hash=source_hash,
        rule_count=len(items),
        example_count=example_count,
        warnings=warnings,
    )


def business_rule_import_summary(result: CommentBusinessRuleSetImportResult) -> dict[str, Any]:
    return {
        "asset_type": COMMENT_BUSINESS_RULE_ASSET_TYPE,
        "asset_key": result.asset_key,
        "keyword_asset_key": result.keyword_asset_key,
        "quality_guard_profile_key": result.quality_guard_profile_key,
        "keyword_selection": result.keyword_selection,
        "rule_count": result.rule_count,
        "example_count": result.example_count,
        "warnings": result.warnings,
        "default_generation_count": DEFAULT_COMMENT_BATCH_LIMIT,
    }


def _normalize_keyword_asset_key(value: str | None) -> str | None:
    normalized = str(value or "").strip()
    return normalized or None


def _normalize_keyword_selection(value: dict[str, Any] | str | None) -> dict[str, Any] | None:
    if value is None:
        return None
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        try:
            value = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise ValueError("keyword_selection must be valid JSON") from exc
    if not isinstance(value, dict):
        raise ValueError("keyword_selection must be a JSON object")
    normalized: dict[str, list[str]] = {}
    for category_code, codes in value.items():
        key = str(category_code or "").strip()
        if not key:
            continue
        if isinstance(codes, dict):
            codes = codes.get("include") or codes.get("codes") or codes.get("keyword_codes")
        if isinstance(codes, str):
            codes = re.split(r"[,，\s]+", codes)
        if not isinstance(codes, list):
            raise ValueError("keyword_selection values must be arrays or comma-separated strings")
        normalized_codes = [str(code).strip() for code in codes if str(code).strip()]
        if normalized_codes:
            normalized[key] = normalized_codes
    return normalized or None


def _activity_name_for_import(asset_key: str, display_name: str | None) -> str:
    if asset_key == DEFAULT_COMMENT_BUSINESS_RULE_ASSET_KEY:
        return DEFAULT_COMMENT_BATCH_TOPIC
    normalized = str(display_name or "").strip()
    if normalized:
        normalized = re.sub(r"(?:评论)?业务规则(?:规则包)?$", "评论", normalized).strip()
        return normalized or str(display_name or "").strip()
    return asset_key


def _read_rule_rows(file_content: bytes, *, source_name: str) -> list[dict[str, str]]:
    lower_name = source_name.lower()
    if lower_name.endswith(".xlsx"):
        return _read_xlsx_rows(file_content)
    if lower_name.endswith(".csv"):
        return _read_csv_rows(file_content)
    raise ValueError("only .csv and .xlsx files are supported")


def _read_csv_rows(file_content: bytes) -> list[dict[str, str]]:
    text = file_content.decode("utf-8-sig")
    content = "".join(line for line in text.splitlines(True) if not line.startswith("#"))
    return [
        {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
        for row in csv.DictReader(io.StringIO(content))
        if any(str(value or "").strip() for value in row.values())
    ]


def _read_xlsx_rows(file_content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows: list[dict[str, str]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {
            header: str(value or "").strip()
            for header, value in zip(headers, raw)
            if header
        }
        if any(row.values()):
            rows.append(row)
    return rows


def _row_to_rule_item(row: dict[str, str], index: int) -> dict[str, Any] | None:
    business_rule, raw_corpus = _business_rule_and_corpus(row)
    if not business_rule or not raw_corpus:
        return None
    corpus, corpus_examples = _split_examples_from_corpus(raw_corpus)
    corpus = _clean_corpus_for_prompt(corpus, business_rule=business_rule)
    canonical_examples = _line_examples(row.get("示例") or row.get("examples"))
    examples = canonical_examples or _line_examples(row.get("评论示例")) or corpus_examples
    if not canonical_examples:
        examples.extend(_line_examples(row.get("评论补充")))
    item = {
        "rule_id": str(row.get("rule_id") or "").strip() or f"business_rule_{index:03d}",
        "business_rule": business_rule,
        "corpus": corpus,
        "examples": examples,
        "supplements": [],
        "source_row_no": index,
    }
    prompt_slots = _prompt_slots_from_row(row)
    if prompt_slots:
        item["prompt_slots"] = prompt_slots
    return item


def _business_rule_and_corpus(row: dict[str, str]) -> tuple[str, str]:
    """Read new business_rule columns while accepting old operator exports."""
    rule_value = (
        row.get("业务规则名称")
        or row.get("业务规则")
        or row.get("规则名称")
        or row.get("标题")
        or row.get("category")
        or row.get("business_rule")
        or row.get(LEGACY_RULE_HEADER)
        or ""
    )
    simple_corpus = _simple_operator_corpus(row)
    if simple_corpus:
        return str(rule_value).strip(), simple_corpus
    corpus_value = row.get("规则语料") or row.get("语料") or row.get("focus") or row.get("corpus") or ""
    if row.get("规则语料") or row.get("语料"):
        return str(rule_value).strip(), str(corpus_value).strip()
    if row.get("评论类型") and (row.get("业务规则") or row.get(LEGACY_RULE_HEADER)):
        return row.get("评论类型", "").strip(), str(rule_value).strip()
    return str(rule_value).strip(), str(corpus_value).strip()


def _simple_operator_corpus(row: dict[str, str]) -> str:
    write_what = str(row.get("写什么") or "").strip()
    how_to_say = str(row.get("怎么说") or "").strip()
    if not write_what and not how_to_say:
        return ""
    parts: list[str] = []
    if write_what:
        parts.append(f"写什么：{write_what}")
    if how_to_say:
        parts.append(f"怎么说：{how_to_say}")
    return "\n\n".join(parts).strip()


def _prompt_slots_from_row(row: dict[str, str]) -> dict[str, list[str]] | list[dict[str, Any]] | None:
    for field in ("prompt_slots", "comment_prompt_slots"):
        raw = str(row.get(field) or "").strip()
        if not raw:
            continue
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise ValueError(f"{field} must be valid JSON") from exc
        if isinstance(parsed, (dict, list)):
            return parsed
        raise ValueError(f"{field} must be a JSON object or array")

    slots: dict[str, list[str]] = {}
    for field in ("说话风格", "说话风格语料", "评论说话风格"):
        entries = _line_prompt_slot_entries(row.get(field), slot_name="说话风格")
        if entries:
            slots["说话风格"] = entries
            break
    return slots or None


def _line_prompt_slot_entries(value: str | None, *, slot_name: str) -> list[str]:
    entries: list[str] = []
    for raw in str(value or "").splitlines():
        line = _clean_example_line(raw)
        line = re.sub(rf"^{re.escape(slot_name)}\s*[:：]\s*", "", line).strip()
        if line:
            entries.append(line)
    return entries


def _line_examples(value: str | None) -> list[str]:
    examples: list[str] = []
    for raw in (value or "").splitlines():
        line = _clean_example_line(raw)
        if line:
            examples.append(line)
    return examples


def _examples_from_corpus(corpus: str) -> list[str]:
    return _split_examples_from_corpus(corpus)[1]


def _split_examples_from_corpus(corpus: str) -> tuple[str, list[str]]:
    if "示例" not in corpus:
        return corpus, []
    after = re.split(r"示例[:：]", corpus, maxsplit=1)
    if len(after) < 2:
        return corpus, []
    before = after[0].rstrip()
    remainder = after[1]
    note_match = re.search(r"\n\s*注意[:：]", remainder)
    if note_match:
        body = remainder[: note_match.start()]
        note = remainder[note_match.start() :].strip()
    else:
        body = remainder
        note = ""
    examples = [_clean_example_line(line) for line in body.splitlines() if _clean_example_line(line)]
    if not examples:
        return corpus, []
    # 重要逻辑：旧导出会把“示例”嵌在语料里；导入时先拆开，
    # 否则评论链路即使计划层只抽少量示例，也会通过 corpus 把全量示例塞进 prompt。
    cleaned_corpus = "\n\n".join(part for part in (before, note) if part).strip()
    return cleaned_corpus, examples


def _clean_corpus_for_prompt(corpus: str, *, business_rule: str | None = None) -> str:
    """Drop operator routing notes that should never be rendered into prompts."""
    lines: list[str] = []
    previous_blank = False
    stripped_rule = str(business_rule or "").strip().rstrip(":：")
    seen_rule_heading = False
    seen_headings: set[str] = set()
    for raw_line in str(corpus or "").splitlines():
        line = raw_line.rstrip()
        stripped = line.strip()
        if _is_operator_note_line(stripped):
            continue
        normalized_heading = stripped.rstrip(":：")
        if stripped.endswith((":", "：")) and normalized_heading:
            if normalized_heading in seen_headings:
                continue
            seen_headings.add(normalized_heading)
        if stripped_rule and stripped.rstrip(":：") == stripped_rule:
            if seen_rule_heading:
                continue
            seen_rule_heading = True
        if not stripped:
            if previous_blank:
                continue
            previous_blank = True
            lines.append("")
            continue
        previous_blank = False
        lines.append(line)
    return "\n".join(lines).strip()


def _is_operator_note_line(text: str) -> bool:
    if not text:
        return False
    return bool(re.match(r"^(?:关键词方向|关键词方向是)\s*[:：是]?.*", text))


def _clean_example_line(value: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^[\-\*•]\s*", "", text)
    text = re.sub(r"^\d+[、．]\s*", "", text)
    text = re.sub(r"^\d+\.\s+", "", text)
    return text.strip()


def _warnings_for_items(items: list[dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    missing_examples = [item["business_rule"] for item in items if not item.get("examples") and not item.get("supplements")]
    if missing_examples:
        warnings.append(f"{len(missing_examples)} 条规则缺少示例")
    return warnings


async def _next_asset_version(db: AsyncSession, asset_types: tuple[str, ...], asset_key: str) -> int:
    result = await db.execute(
        select(AssetRegistry.version_no)
        .where(AssetRegistry.asset_type.in_(asset_types), AssetRegistry.asset_key == asset_key)
        .order_by(AssetRegistry.version_no.desc())
        .limit(1)
    )
    current = result.scalar_one_or_none()
    return int(current or 0) + 1
