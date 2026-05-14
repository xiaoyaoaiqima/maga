"""Parse marketing training-rule workbooks for MAGA asset.import."""
from __future__ import annotations

import base64
import hashlib
from collections import OrderedDict
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def import_asset_package(input_payload: dict[str, Any]) -> dict[str, Any]:
    """Return structured MAGA assets from an uploaded workbook payload."""
    source_content_base64 = input_payload.get("source_content_base64")
    if not isinstance(source_content_base64, str) or not source_content_base64.strip():
        raise ValueError("source_content_base64 is required")
    workbook_bytes = base64.b64decode(source_content_base64)
    source_hash = input_payload.get("source_hash") or hashlib.sha256(workbook_bytes).hexdigest()
    asset_key = _clean(input_payload.get("asset_key")) or "yuanyue"
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True)

    assets = [
        {
            "asset_type": "brand_profile",
            "asset_key": asset_key,
            "display_name": "源悦品牌资料",
            "content_json": _parse_brand_profile(wb, asset_key),
        },
        {
            "asset_type": "product_selling_points",
            "asset_key": asset_key,
            "display_name": "源悦产品卖点",
            "content_json": {"items": _parse_product_selling_points(wb)},
        },
        {
            "asset_type": "painpoint_model",
            "asset_key": asset_key,
            "display_name": "源悦主题/痛点模型",
            "content_json": _parse_painpoint_model(wb),
        },
        {
            "asset_type": "ugc_expression_corpus",
            "asset_key": asset_key,
            "display_name": "源悦 UGC 卖点表述语料",
            "content_json": {
                "items": _parse_expression_sheet(
                    wb,
                    ["ugc卖点表述", "ugc常规-卖点表述"],
                    label_field="painpoint_or_selling_point",
                )
            },
        },
        {
            "asset_type": "compliance_rules",
            "asset_key": asset_key,
            "display_name": "源悦审核规则",
            "content_json": {"items": _parse_compliance_rules(wb)},
        },
    ]

    premium = _parse_expression_sheet(wb, ["精品-卖点表述", "ugc精品-卖点表述"], label_field="painpoint")
    if premium:
        assets.append(
            {
                "asset_type": "premium_expression_corpus",
                "asset_key": asset_key,
                "display_name": "源悦精品卖点表述语料",
                "content_json": {"items": premium},
            }
        )

    reference_examples = _parse_reference_examples(wb)
    if reference_examples:
        assets.append(
            {
                "asset_type": "reference_examples",
                "asset_key": asset_key,
                "display_name": "源悦参考例文",
                "content_json": {"items": reference_examples},
            }
        )

    return {
        "asset_key": asset_key,
        "source_hash": source_hash,
        "assets": assets,
        "warnings": _warnings_for_package(assets, wb.sheetnames),
    }


def _parse_brand_profile(wb, asset_key: str) -> dict[str, Any]:
    if "品牌资料整理" not in wb.sheetnames:
        return {"brand_key": asset_key}
    ws = wb["品牌资料整理"]
    return {
        "brand_key": asset_key,
        "brand_name": "源悦" if asset_key == "yuanyue" else asset_key,
        "content_focus": _clean(ws["C3"].value),
        "content_style": _clean(ws["C4"].value),
    }


def _parse_product_selling_points(wb) -> list[dict[str, Any]]:
    if "品牌资料整理" not in wb.sheetnames:
        return []
    ws = wb["品牌资料整理"]
    items_by_name: OrderedDict[str, dict[str, Any]] = OrderedDict()
    current_level = None
    current_selling_point = None
    for row in range(9, ws.max_row + 1):
        level = _clean(ws.cell(row, 2).value) or current_level
        selling_point = _clean(ws.cell(row, 3).value) or current_selling_point
        ingredient = _clean(ws.cell(row, 4).value)
        advantage = _clean(ws.cell(row, 5).value)
        vivid_explanation = _clean(ws.cell(row, 6).value)
        if _emptyish(selling_point) and _emptyish(ingredient) and _emptyish(advantage) and _emptyish(vivid_explanation):
            continue
        current_level = level
        current_selling_point = selling_point
        key = selling_point or f"selling_point_{len(items_by_name) + 1}"
        item = items_by_name.setdefault(
            key,
            {
                "level": level,
                "selling_point": selling_point,
                "ingredients": [],
                "advantages": [],
                "descriptions": [],
                "expressions": [],
            },
        )
        _append_unique(item["ingredients"], ingredient)
        _append_unique(item["advantages"], advantage)
        _append_unique(item["descriptions"], vivid_explanation)
        if not item.get("ingredient"):
            item["ingredient"] = ingredient
        if not item.get("advantage"):
            item["advantage"] = advantage
        if not item.get("vivid_explanation"):
            item["vivid_explanation"] = vivid_explanation
    return list(items_by_name.values())


def _parse_painpoint_model(wb) -> dict[str, Any]:
    if "内容模型" not in wb.sheetnames:
        return {"topics": [], "items": []}
    ws = wb["内容模型"]
    topics: OrderedDict[str, dict[str, Any]] = OrderedDict()
    current_stage = None
    current_topic = None
    for row in range(2, ws.max_row + 1):
        baby_stage = _clean(ws.cell(row, 2).value) or current_stage
        topic = _clean(ws.cell(row, 3).value) or current_topic
        symptom = _clean(ws.cell(row, 4).value)
        description = _clean(ws.cell(row, 5).value)
        selling_point = _clean(ws.cell(row, 6).value)
        extras = [_clean(ws.cell(row, col).value) for col in range(7, 11)]
        extras = [item for item in extras if not _emptyish(item)]
        if _emptyish(topic) and _emptyish(symptom) and _emptyish(description) and _emptyish(selling_point):
            continue
        current_stage = baby_stage
        current_topic = topic
        key = topic or f"topic_{len(topics) + 1}"
        topic_item = topics.setdefault(
            key,
            {
                "topic": topic,
                "painpoint": topic,
                "baby_stages": [],
                "descriptions": [],
                "selling_points": [],
            },
        )
        _append_unique(topic_item["baby_stages"], baby_stage)
        _append_unique(topic_item["descriptions"], symptom)
        _append_unique(topic_item["descriptions"], description)
        for extra in extras:
            _append_unique(topic_item["descriptions"], extra)
        if selling_point:
            _merge_topic_selling_point(topic_item, selling_point, [description, *extras])

    topic_values = list(topics.values())
    return {
        "topics": topic_values,
        # Legacy flat items keep current generation planner/runtime compatible
        # while the richer topic tree becomes the primary worker-owned output.
        "items": [_topic_to_legacy_item(topic) for topic in topic_values],
    }


def _merge_topic_selling_point(topic_item: dict[str, Any], selling_point: str, descriptions: list[str | None]) -> None:
    existing = next(
        (item for item in topic_item["selling_points"] if item.get("selling_point") == selling_point),
        None,
    )
    if existing is None:
        existing = {"selling_point": selling_point, "descriptions": [], "expressions": []}
        topic_item["selling_points"].append(existing)
    for description in descriptions:
        _append_unique(existing["descriptions"], description)


def _topic_to_legacy_item(topic: dict[str, Any]) -> dict[str, Any]:
    selling_points = [item.get("selling_point") for item in topic.get("selling_points") or [] if item.get("selling_point")]
    return {
        "baby_stage": (topic.get("baby_stages") or [None])[0],
        "painpoint": topic.get("painpoint") or topic.get("topic"),
        "description": "；".join(topic.get("descriptions") or []),
        "selling_point": selling_points[0] if selling_points else None,
        "selling_points": selling_points,
        "extra_descriptions": topic.get("descriptions") or [],
    }


def _parse_expression_sheet(wb, sheet_names: str | list[str], *, label_field: str) -> list[dict[str, Any]]:
    sheet_name = _first_existing_sheet(wb, sheet_names)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    items: list[dict[str, Any]] = []
    current_label = None
    for row in range(2, ws.max_row + 1):
        label = _clean(ws.cell(row, 2).value) or current_label
        expression = _clean(ws.cell(row, 3).value)
        owner = _clean(ws.cell(row, 4).value)
        if _emptyish(label) and _emptyish(expression):
            continue
        current_label = label
        if _emptyish(expression):
            continue
        items.append(
            {
                label_field: label,
                "expression": expression,
                "owner": None if _emptyish(owner) else owner,
            }
        )
    return items


def _parse_reference_examples(wb) -> list[dict[str, Any]]:
    if "例文收集" not in wb.sheetnames:
        return []
    ws = wb["例文收集"]
    header = [_clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    header_map = {name: idx + 1 for idx, name in enumerate(header) if name}
    real_workbook_shape = {"来源", "方向", "ID昵称", "发布链接", "发布形式", "笔记标题", "笔记正文"}.issubset(
        set(header_map)
    )

    items: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        if real_workbook_shape:
            title = _clean(ws.cell(row, header_map["笔记标题"]).value)
            body = _clean(ws.cell(row, header_map["笔记正文"]).value)
            if _emptyish(title) and _emptyish(body):
                continue
            source = _clean(ws.cell(row, header_map["来源"]).value)
            direction = _clean(ws.cell(row, header_map["方向"]).value)
            author = _clean(ws.cell(row, header_map["ID昵称"]).value)
            post_format = _clean(ws.cell(row, header_map["发布形式"]).value)
            items.append(
                {
                    "example_id": f"yuanyue_ref_{len(items) + 1:03d}",
                    "source": None if _emptyish(source) else source,
                    "direction": None if _emptyish(direction) else direction,
                    "author_name": None if _emptyish(author) else author,
                    "follower_count_w": None if "粉丝量" not in header_map else _clean(ws.cell(row, header_map["粉丝量"]).value),
                    "post_url": None if "发布链接" not in header_map else _clean(ws.cell(row, header_map["发布链接"]).value),
                    "post_format": None if _emptyish(post_format) else post_format,
                    "title": None if _emptyish(title) else title,
                    "body": None if _emptyish(body) else body,
                    "painpoint": None if "痛点" not in header_map else _clean(ws.cell(row, header_map["痛点"]).value),
                    "reference_type": None if _emptyish(post_format) else post_format,
                    "owner": None if _emptyish(author) else author,
                    "style_tags": [tag for tag in [post_format, direction] if not _emptyish(tag)],
                    "structure_tags": [],
                }
            )
            continue

        title = _clean(ws.cell(row, 2).value)
        body = _clean(ws.cell(row, 3).value)
        reference_type = _clean(ws.cell(row, 4).value)
        owner = _clean(ws.cell(row, 5).value)
        if _emptyish(title) and _emptyish(body):
            continue
        items.append(
            {
                "example_id": f"yuanyue_ref_{len(items) + 1:03d}",
                "title": None if _emptyish(title) else title,
                "body": None if _emptyish(body) else body,
                "reference_type": None if _emptyish(reference_type) else reference_type,
                "owner": None if _emptyish(owner) else owner,
                "style_tags": [] if _emptyish(reference_type) else [reference_type],
                "structure_tags": [],
            }
        )
    return items


def _parse_compliance_rules(wb) -> list[dict[str, Any]]:
    if "审核规则" not in wb.sheetnames:
        return []
    ws = wb["审核规则"]
    items: list[dict[str, Any]] = []
    current_content = None
    current_category = None
    for row in range(2, ws.max_row + 1):
        sequence = _clean(ws.cell(row, 1).value)
        audit_content = _clean(ws.cell(row, 2).value) or current_content
        category = _clean(ws.cell(row, 3).value) or current_category
        dimension = _clean(ws.cell(row, 4).value)
        feedback = _clean(ws.cell(row, 5).value)
        if _emptyish(dimension) and _emptyish(feedback):
            continue
        current_content = audit_content
        current_category = category
        items.append(
            {
                "sequence": sequence,
                "audit_content": audit_content,
                "category": category,
                "dimension": dimension,
                "feedback": feedback,
            }
        )
    return items


def _warnings_for_package(assets: list[dict[str, Any]], sheet_names: list[str]) -> list[str]:
    warnings: list[str] = []
    required = {"品牌资料整理", "内容模型", "审核规则"}
    missing = sorted(required - set(sheet_names))
    if missing:
        warnings.append(f"missing sheets: {', '.join(missing)}")
    for asset in assets:
        content = asset.get("content_json") or {}
        items = content.get("items") if isinstance(content, dict) else None
        topics = content.get("topics") if isinstance(content, dict) else None
        if asset["asset_type"] in {"painpoint_model", "product_selling_points", "compliance_rules"}:
            count = len(topics) if isinstance(topics, list) else len(items or [])
            if count == 0:
                warnings.append(f"{asset['asset_type']} parsed zero items")
    return warnings


def _first_existing_sheet(wb, sheet_names: str | list[str]) -> str:
    names = [sheet_names] if isinstance(sheet_names, str) else sheet_names
    return next((name for name in names if name in wb.sheetnames), names[0])


def _append_unique(items: list[Any], value: Any) -> None:
    if _emptyish(value):
        return
    text = str(value).strip()
    if text not in items:
        items.append(text)


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _emptyish(value: Any) -> bool:
    if value is None:
        return True
    return str(value).strip() in {"", "/", "\\"}
