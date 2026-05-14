"""
Extracts the EXPERT梳理.xlsx into structured corpus files.

Reads sheet '最新梳理完成版本' which has 14 cols:
  A 名称  B 类型(GE/AE)  C 能力  D 简介
  E 一级内容  F 内容说明  G 二级内容  H 理想作用语料
  I 语料解释  J 历史作用语料  K 输出支持
  L 模型  M 必选  N 评分类型

Forward-fills merged cells (Expert metadata stays the same across rows).
Emits:
  workspace/experts/_registry.yaml         - global expert index
  workspace/experts/<slug>/persona.md      - auto-stub from simplified 简介
  workspace/experts/<slug>/corpus.yaml     - L1 → L2 → ideal_corpus tree
"""
from __future__ import annotations

import argparse
from pathlib import Path
import openpyxl
import yaml

# Map Chinese expert names → English slugs
NAME_TO_SLUG = {
    "母婴行业-小红书-文章通用Expert": "ge_main",
    "法律Expert": "legal",
    "平台Expert": "platform",
    "品牌Expert": "brand",
    "产品痛点卖点Expert": "painpoint_selling",
    "活动Expert": "campaign",
    "拟人化Expert": "persona",
    "故事元素Expert": "story_elements",
    "内容结构Expert": "content_structure",
    "生活常识Expert": "common_sense",
    "文章优雅度Expert": "elegance",
    "AI味Expert": "ai_smell",
    "内容丰富度Expert": "content_richness",
    "人群多样性Expert": "persona_diversity",
}

# Normalize 输出支持 phrases
OUTPUT_MODE_MAP = {
    "固定输出": "fixed",
    "随机输出": "random",
    "根据比例、随机输出 指定痛卖点": "ratio_random",
    "根据比例、随机输出 指定人设": "ratio_random",
    "根据比例、随机输出 指定内容结构、写作手法、字数、emoji等": "ratio_random",
    "随机输出、固定输出": "mixed",
    "输出文章": "article_output",
    "-": None,
}

MUST_MAP = {"是": True, "否": False, None: None, "-": None}


def normalize_score(v):
    if v is None or v == "-":
        return None
    return str(v).strip()


def load_keyword_corpus(xlsx_path: Path) -> dict:
    """Read 大表模板 sheet for actual corpus content (keyword + 关键词语料 cols).

    Returns: {expert_name: {l1: {l2: [corpus_strings]}}}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["大表模板"]
    out: dict = {}
    sticky = {"expert": None, "l1": None, "l2": None}
    for r in range(2, ws.max_row + 1):
        expert = ws.cell(row=r, column=1).value
        l1     = ws.cell(row=r, column=5).value
        l2     = ws.cell(row=r, column=7).value
        kw     = ws.cell(row=r, column=8).value  # 关键词
        corpus = ws.cell(row=r, column=9).value  # 关键词语料
        if expert: sticky["expert"] = str(expert).strip()
        if l1: sticky["l1"] = str(l1).strip()
        if l2: sticky["l2"] = str(l2).strip()
        if not corpus or not str(corpus).strip():
            continue
        e = sticky["expert"]
        if not e:
            continue
        bucket = out.setdefault(e, {}).setdefault(sticky["l1"] or "_", {}).setdefault(sticky["l2"] or "_", [])
        bucket.append({"keyword": (str(kw).strip() if kw else None), "rule": str(corpus).strip()})
    return out


def load_rows(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["最新梳理完成版本"]
    rows = []
    last = {}
    for r in range(2, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        # Forward-fill: name, type, capability, desc, output_mode, model, must, score_type
        # Forward-fill L1 only when name persists
        name, type_, cap, desc, l1, l1_desc, l2, ideal, ideal_help, hist, out, model, must, score = cells

        if name:
            last = {
                "name": str(name).strip(),
                "type": (str(type_).strip() if type_ else None),
                "capability": (str(cap).strip() if cap else None),
                "desc": (str(desc).strip() if desc else None),
                "output_mode": (str(out).strip() if out else None),
                "model": (str(model).strip() if model else None),
                "must": (str(must).strip() if must else None),
                "score_type": (str(score).strip() if score else None),
                "l1": None,
                "l1_desc": None,
            }
        if l1:
            last["l1"] = str(l1).strip()
            last["l1_desc"] = (str(l1_desc).strip() if l1_desc else None)
        # update sticky output/model/must/score in case they appear later in same expert block
        if out:    last["output_mode"] = str(out).strip()
        if model:  last["model"] = str(model).strip()
        if must:   last["must"] = str(must).strip()
        if score:  last["score_type"] = str(score).strip()

        rows.append({
            **last,
            "l2": (str(l2).strip() if l2 else None),
            "ideal": (str(ideal).strip() if ideal else None),
            "ideal_help": (str(ideal_help).strip() if ideal_help else None),
            "hist": (str(hist).strip() if hist else None),
        })
    return rows


def build_experts(rows, kw_corpus: dict | None = None):
    experts: dict[str, dict] = {}
    kw_corpus = kw_corpus or {}
    for row in rows:
        name = row["name"]
        if not name:
            continue
        slug = NAME_TO_SLUG.get(name)
        if slug is None:
            print(f"  ⚠ unknown expert name: {name}, skipping")
            continue
        e = experts.setdefault(slug, {
            "slug": slug,
            "chinese_name": name,
            "type": row["type"],
            "capability": row["capability"],
            "desc": row["desc"],
            "output_mode": OUTPUT_MODE_MAP.get(row["output_mode"], row["output_mode"]),
            "model": row["model"],
            "must": MUST_MAP.get(row["must"]),
            "score_type": normalize_score(row["score_type"]),
            "l1_groups": {},  # l1_name → {desc, items: [{l2, ideal, hist, help}]}
        })
        l1 = row.get("l1")
        if not l1:
            continue
        g = e["l1_groups"].setdefault(l1, {"desc": row.get("l1_desc"), "items": []})
        if row["l2"] or row["ideal"] or row["hist"]:
            # Look up enrichment from 大表模板 by (chinese_name, l1, l2)
            rules = []
            try:
                bucket = kw_corpus.get(name, {}).get(l1, {}).get(row["l2"], [])
                for entry in bucket:
                    rules.append(entry.get("rule"))
            except Exception:
                pass
            g["items"].append({
                "l2": row["l2"],
                "ideal": row["ideal"],
                "ideal_help": row["ideal_help"],
                "hist_examples": (row["hist"] if row["hist"] and row["hist"] != "-" else None),
                "rules": rules or None,
            })
    return experts


def write_persona(out_dir: Path, e: dict):
    persona = f"""# {e['chinese_name']} ({e['slug']})

**类型**: {e['type']} | **能力**: {e['capability']}

## 角色简介

{e.get('desc') or '(待补充)'}

## 工作模式

- **生文前**: 接收 brief，从 `corpus.yaml` 按 `{e.get('output_mode')}` 模式抽样语料，输出**指令**给 GE
- **生文后**: 对草稿打分（评分类型: `{e.get('score_type')}`）

## 必选性

{'必选 — 在所有 brief_type 中强制启用' if e.get('must') else '可选 — 由 brief_type 决定是否启用'}

## 使用模型

`{e.get('model') or '(未指定)'}`
"""
    (out_dir / "persona.md").write_text(persona, encoding="utf-8")


def write_corpus(out_dir: Path, e: dict):
    corpus = {
        "expert": e["slug"],
        "chinese_name": e["chinese_name"],
        "output_mode": e["output_mode"],
        "score_type": e["score_type"],
        "groups": {},
    }
    for l1, g in e["l1_groups"].items():
        corpus["groups"][l1] = {
            "description": g.get("desc"),
            "items": [
                {
                    "l2": item["l2"],
                    "ideal": item["ideal"],
                    "ideal_help": item.get("ideal_help"),
                    "examples": item.get("hist_examples"),
                    "rules": item.get("rules"),
                }
                for item in g["items"]
                if item.get("ideal") or item.get("l2")
            ],
        }
    with (out_dir / "corpus.yaml").open("w", encoding="utf-8") as f:
        yaml.safe_dump(corpus, f, allow_unicode=True, sort_keys=False, width=120)


def write_registry(experts_dir: Path, experts: dict):
    registry = {"experts": {}}
    for slug, e in experts.items():
        registry["experts"][slug] = {
            "chinese_name": e["chinese_name"],
            "type": e["type"],
            "capability": e["capability"],
            "must": e["must"],
            "score_type": e["score_type"],
            "output_mode": e["output_mode"],
            "model": e["model"],
        }
    with (experts_dir / "_registry.yaml").open("w", encoding="utf-8") as f:
        yaml.safe_dump(registry, f, allow_unicode=True, sort_keys=False, width=120)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="/Users/luxifa/Downloads/EXPERT梳理 (5).xlsx")
    default_out = Path(__file__).resolve().parents[2] / "worker" / "profiles" / "maga-worker" / "experts"
    ap.add_argument("--out", default=str(default_out))
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    out_root = Path(args.out)
    out_root.mkdir(parents=True, exist_ok=True)

    rows = load_rows(xlsx_path)
    kw_corpus = load_keyword_corpus(xlsx_path)
    experts = build_experts(rows, kw_corpus)

    print(f"Parsed {len(experts)} experts.\n")
    for slug, e in experts.items():
        out_dir = out_root / slug
        out_dir.mkdir(parents=True, exist_ok=True)
        (out_dir / "lessons").mkdir(exist_ok=True)
        write_persona(out_dir, e)
        write_corpus(out_dir, e)
        n_groups = len(e["l1_groups"])
        n_items = sum(len(g["items"]) for g in e["l1_groups"].values())
        print(f"  ✓ {slug:25s} {n_groups} L1 groups, {n_items} items")

    write_registry(out_root, experts)
    print(f"\n  ✓ _registry.yaml written ({len(experts)} entries)")


if __name__ == "__main__":
    main()
