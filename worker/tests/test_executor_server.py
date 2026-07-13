from io import BytesIO
import base64

from fastapi.testclient import TestClient
from openpyxl import Workbook

from maga_worker.executor_server import app


def _client(monkeypatch):
    monkeypatch.setenv("MAGA_WORKER_EXECUTOR_TOKEN", "test-token")
    return TestClient(app)


def _headers(token="test-token"):
    return {
        "X-Maga-Protocol-Version": "0.1",
        "Authorization": f"Bearer {token}",
    }


def _envelope(capability, input_payload=None, stage_call_id="stage-001"):
    return {
        "protocol_version": "0.1",
        "run_id": 10,
        "task_id": 20,
        "stage_call_id": stage_call_id,
        "capability": capability,
        "executor_hints": {"timeout_seconds": 60},
        "input": input_payload or {},
    }


def test_invoke_rejects_missing_protocol_header(monkeypatch):
    client = _client(monkeypatch)

    response = client.post("/invoke", json=_envelope("content.generate"), headers={"Authorization": "Bearer test-token"})

    assert response.status_code == 400
    assert response.json()["detail"] == "unsupported protocol version"


def test_invoke_rejects_wrong_bearer_token(monkeypatch):
    client = _client(monkeypatch)

    response = client.post("/invoke", json=_envelope("content.generate"), headers=_headers("wrong-token"))

    assert response.status_code == 401


def test_asset_import_returns_structured_asset_package(monkeypatch):
    client = _client(monkeypatch)
    wb = Workbook()
    ws = wb.active
    ws.title = "品牌资料整理"
    ws["C3"] = "好消化易吸收，对应便便好，不上火"
    ws["C4"] = "高质量真实用户ugc"
    ws["B8"] = "分级"
    ws["C8"] = "卖点"
    ws["D8"] = "成分"
    ws["E8"] = "源悦优势"
    ws["F8"] = "生动理解"
    ws["B9"] = "核心卖点"
    ws["C9"] = "好消化易吸收"
    ws["D9"] = "软分子蛋白"
    ws["E9"] = "形成结构松散的软凝乳"
    ws["F9"] = "软软的米糊"

    ws2 = wb.create_sheet("内容模型")
    ws2.append(["序号", "宝宝阶段", "核心痛点", "具体表现", "痛点描述", "对应卖点"])
    ws2.append([None, None, "便便不规律", "羊屎蛋/干硬", "便便又干又硬", "好消化易吸收"])
    ws2.append([None, None, None, "拉臭费劲", "拉起来不轻松", "好消化易吸收"])

    ws3 = wb.create_sheet("ugc常规-卖点表述")
    ws3.append(["序号", "对应卖点", "卖点描述", "负责人"])
    ws3.append([None, "便便不规律", "便便基本一天一次，拉起来也不费劲", "东昕"])

    ws4 = wb.create_sheet("审核规则")
    ws4.append(["序号", "审核内容", "分类", "审核维度（问题分类）", "审核意见（返回给用户的）"])
    ws4.append([1, "文案审核", "草稿审核", "夸大产品效果或虚构使用经历", "文本不符合活动要求"])

    bio = BytesIO()
    wb.save(bio)

    response = client.post(
        "/invoke",
        json=_envelope(
            "asset.import",
            {
                "asset_key": "yuanyue",
                "source_name": "源悦种草活动-ai训练规则.xlsx",
                "source_hash": "hash-001",
                "source_content_base64": base64.b64encode(bio.getvalue()).decode("ascii"),
            },
            stage_call_id="stage-asset-import",
        ),
        headers=_headers(),
    )

    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "succeeded"
    assert data["stats"]["module"] == "asset-steward"
    output = data["output"]
    assert output["asset_key"] == "yuanyue"
    by_type = {asset["asset_type"]: asset for asset in output["assets"]}
    assert by_type["brand_profile"]["content_json"]["content_style"] == "高质量真实用户ugc"
    assert by_type["product_selling_points"]["content_json"]["items"][0]["selling_point"] == "好消化易吸收"
    topic = by_type["painpoint_model"]["content_json"]["topics"][0]
    assert topic["topic"] == "便便不规律"
    assert topic["descriptions"] == ["羊屎蛋/干硬", "便便又干又硬", "拉臭费劲", "拉起来不轻松"]
    assert topic["selling_points"][0]["selling_point"] == "好消化易吸收"


def test_content_generate_fake_mode_returns_comment_from_unified_input(monkeypatch):
    monkeypatch.setenv("MAGA_WORKER_RUNTIME_FAST_FAKE", "1")
    client = _client(monkeypatch)

    response = client.post(
        "/invoke",
        json=_envelope(
            "content.generate",
            {
                "content_type": "comment",
                "output_fields": ["comment"],
                "business_rule": {
                    "item_no": 1,
                    "business_rule": "整体适应",
                    "corpus": "整体适应：\n像妈妈在评论区聊刚开始喝源悦的观察。",
                    "examples": ["我家刚开始也在看源悦，想蹲蹲真实反馈"],
                },
                "selected_keywords": [
                    {"category_code": "persona", "keyword_name": "经验型妈妈", "corpus": ["自然交流"]}
                ],
                "expert": {"expert_config_code": "comment_generator_v1"},
                "rendered_prompt": "生成一条评论",
            },
            stage_call_id="stage-content-fake",
        ),
        headers=_headers(),
    )

    assert response.status_code == 200
    data = response.json()
    assert data["stage_call_id"] == "stage-content-fake"
    assert data["status"] == "succeeded"
    assert data["stats"]["module"] == "content-generator"
    assert data["output"]["comment"] == "我家刚开始也在看源悦，想蹲蹲真实反馈"
    assert data["output"]["runtime_result"]["mode"] == "content_fake"
    assert data["output"]["runtime_result"]["expert_config_code"] == "comment_generator_v1"


def test_content_generate_runtime_keeps_empty_comment_when_model_returns_empty(monkeypatch):
    monkeypatch.delenv("MAGA_WORKER_RUNTIME_FAST_FAKE", raising=False)

    def empty_model(*args, **kwargs):
        return ""

    monkeypatch.setattr("maga_worker.llm_runtime.call_model", empty_model)
    client = _client(monkeypatch)

    response = client.post(
        "/invoke",
        json=_envelope(
            "content.generate",
            {
                "content_type": "comment",
                "output_fields": ["comment"],
                "business_rule": {
                    "item_no": 1,
                    "business_rule": "奶量补充",
                    "examples": ["第一口愿意喝，我就放心点了"],
                },
                "model_config": {
                    "model_code": "test-model",
                    "provider_code": "test-provider",
                    "temperature": 0.85,
                    "max_tokens": 64,
                },
                "expert": {"expert_config_code": "comment_generator_v1"},
                "rendered_prompt": "生成一条评论",
            },
            stage_call_id="stage-content-runtime-empty",
        ),
        headers=_headers(),
    )

    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "succeeded"
    assert data["output"]["comment"] == ""
    runtime = data["output"]["runtime_result"]
    assert runtime["mode"] == "content_runtime"
    assert runtime["fallback"] is False
    assert runtime["empty_output"] is True
    assert runtime["empty_reason"] == "content.generate produced empty comment"
    assert runtime["model_attempts"] == 2


def test_content_generate_batch_retry_keeps_json_array_contract(monkeypatch):
    monkeypatch.delenv("MAGA_WORKER_RUNTIME_FAST_FAKE", raising=False)
    prompts = []

    def empty_model(*args, **kwargs):
        prompts.append(kwargs["user"])
        return ""

    monkeypatch.setattr("maga_worker.llm_runtime.call_model", empty_model)
    client = _client(monkeypatch)

    response = client.post(
        "/invoke",
        json=_envelope(
            "content.generate",
            {
                "content_type": "comment",
                "output_fields": ["comment"],
                "output_format_mode": "json_string_array",
                "expansion_count": 20,
                "model_config": {"model_code": "test-model"},
                "rendered_prompt": "生成 20 条评论",
            },
        ),
        headers=_headers(),
    )

    assert response.status_code == 200
    assert len(prompts) == 2
    assert "正好 20 条评论组成的 JSON 字符串数组" in prompts[1]
    assert "评论只输出一条评论正文" not in prompts[1]


def test_content_generate_runtime_uses_model_config_provider(monkeypatch):
    monkeypatch.delenv("MAGA_WORKER_RUNTIME_FAST_FAKE", raising=False)
    calls = []

    def model_with_provider(*args, **kwargs):
        calls.append(kwargs)
        return '{"title":"标题","body":"正文"}'

    monkeypatch.setattr("maga_worker.llm_runtime.call_model", model_with_provider)
    client = _client(monkeypatch)

    response = client.post(
        "/invoke",
        json=_envelope(
            "content.generate",
            {
                "content_type": "article",
                "output_fields": ["title", "body"],
                "business_rule": {"topic": "日常体验"},
                "model_config": {
                    "provider_code": "aihubmix",
                    "model_code": "deepseek-v4-flash",
                    "base_url": "https://aihubmix.example/v1",
                    "api_key": "db-key",
                },
                "expert": {"expert_config_code": "article_generator_v1"},
                "rendered_prompt": "生成一篇文章",
            },
            stage_call_id="stage-content-provider",
        ),
        headers=_headers(),
    )

    assert response.status_code == 200
    assert response.json()["status"] == "succeeded"
    assert calls[0]["base_url"] == "https://aihubmix.example/v1"
    assert calls[0]["api_key"] == "db-key"


def test_content_generate_runtime_accepts_items_json(monkeypatch):
    monkeypatch.delenv("MAGA_WORKER_RUNTIME_FAST_FAKE", raising=False)

    def model_with_items(*args, **kwargs):
        return (
            '{"items":['
            '{"title":"肚子响少了#儿童奶粉[话题]#","body":"旺玥喝了一周，肚子响少了。#旺玥#"},'
            '{"title":"精神好些","body":"娃晚上还在客厅玩。 #儿童成长"}'
            "]}"
        )

    monkeypatch.setattr("maga_worker.llm_runtime.call_model", model_with_items)
    client = _client(monkeypatch)

    response = client.post(
        "/invoke",
        json=_envelope(
            "content.generate",
            {
                "content_type": "article",
                "output_fields": ["title", "body"],
                "rendered_prompt": "一次生成 2 篇",
            },
            stage_call_id="stage-content-items",
        ),
        headers=_headers(),
    )

    assert response.status_code == 200
    output = response.json()["output"]
    assert output["title"] == "肚子响少了"
    assert output["body"] == "旺玥喝了一周，肚子响少了。"
    assert output["items"][1]["body"] == "娃晚上还在客厅玩。"


def test_content_rewrite_fake_mode_removes_forbidden_terms(monkeypatch):
    monkeypatch.setenv("MAGA_WORKER_RUNTIME_FAST_FAKE", "1")
    client = _client(monkeypatch)

    response = client.post(
        "/invoke",
        json=_envelope(
            "content.rewrite",
            {
                "content_type": "comment",
                "output_fields": ["comment"],
                "previous_content": {"comment": "我家刚开始也在看源悦，想蹲蹲真实反馈"},
                "forbidden_hits": ["源悦"],
                "rewrite_instructions": ["删除或自然替换命中的违禁词"],
            },
            stage_call_id="stage-content-rewrite-fake",
        ),
        headers=_headers(),
    )

    assert response.status_code == 200
    data = response.json()
    assert data["stage_call_id"] == "stage-content-rewrite-fake"
    assert data["status"] == "succeeded"
    assert data["stats"]["module"] == "content-generator"
    assert data["output"]["comment"] == "我家刚开始也在看，想蹲蹲真实反馈"
    assert "源悦" not in data["output"]["comment"]
    assert data["output"]["runtime_result"]["mode"] == "content_rewrite_fake"


def test_content_rewrite_fake_mode_uses_operator_feedback(monkeypatch):
    monkeypatch.setenv("MAGA_WORKER_RUNTIME_FAST_FAKE", "1")
    client = _client(monkeypatch)

    response = client.post(
        "/invoke",
        json=_envelope(
            "content.rewrite",
            {
                "content_type": "article",
                "output_fields": ["title", "body"],
                "previous_content": {"title": "原标题", "body": "原正文比较总结。"},
                "operator_feedback": "开头再具体一点，少一点总结腔。",
                "rewrite_instructions": ["根据运营修改意见调整内容"],
            },
            stage_call_id="stage-content-rewrite-feedback",
        ),
        headers=_headers(),
    )

    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "succeeded"
    assert data["output"]["title"] == "原标题"
    assert "按运营反馈调整：开头再具体一点" in data["output"]["body"]
    assert data["output"]["runtime_result"]["mode"] == "content_rewrite_fake"


def test_legacy_capabilities_return_protocol_failed_envelope(monkeypatch):
    client = _client(monkeypatch)

    for capability in ["xhs.generate_draft", "comment.generate"]:
        response = client.post(
            "/invoke",
            json=_envelope(capability, stage_call_id=f"stage-{capability}"),
            headers=_headers(),
        )

        assert response.status_code == 200
        assert response.json() == {
            "stage_call_id": f"stage-{capability}",
            "status": "failed",
            "error_code": "input_invalid",
            "error_message": f"unsupported capability: {capability}",
        }
