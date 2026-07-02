#!/usr/bin/env python3
"""Run A2 sentiment post smoke samples with BabyTree micro personas."""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
PLATFORM_SERVER = ROOT / "platform-server"
DEFAULT_RULE_DIR = ROOT / "prompts" / "a2_sentiment_post_activity"
DEFAULT_OUTPUT_DIR = ROOT / "outputs" / "a2_sentiment_post_new_persona_20260615"
PROFILE_KEY = "a2_sentiment_post_202606"

sys.path.insert(0, str(PLATFORM_SERVER))

from app.services.activity_quality_guard_service import ActivityQualityGuardService  # noqa: E402
from app.services.system_prompt_keyword_service import (  # noqa: E402
    fallback_system_prompt_keyword_content,
    normalize_system_prompt_keyword_content,
)
from app.services.unified_content_generation_service import (  # noqa: E402
    _fallback_expert_snapshot,
    _render_template,
    _select_keyword_bundle,
    _template_variables,
)


@dataclass(frozen=True)
class RuleSmokeConfig:
    rule_id: str
    persona_code: str
    writing_instruction_code: str
    perturbation_code: str
    writing_method_code: str
    format_code: str
    diversity_note: str


RULE_CONFIGS = [
    RuleSmokeConfig(
        "001",
        "babytree_consumption_volume_mom",
        "natural_article",
        "opening_shift",
        "scene_detail",
        "article_compact_clean",
        "把门店到货落到家里消耗量和续罐节奏。",
    ),
    RuleSmokeConfig(
        "002",
        "babytree_sale_stockup_mom",
        "light_article_experience",
        "stance_shift",
        "question_hook",
        "article_light_emoji",
        "线上有货写成先补一罐的囤货动作。",
    ),
    RuleSmokeConfig(
        "003",
        "babytree_after_feed_record_mom",
        "natural_article",
        "length_shift",
        "scene_detail",
        "article_compact_clean",
        "从刚喂完/家里口粮记录切入熟悉款续上。",
    ),
    RuleSmokeConfig(
        "004",
        "babytree_same_feeling_mom",
        "light_article_experience",
        "opening_shift",
        "plain_explain",
        "article_compact_clean",
        "试过别的后用同感补充的口吻写回到熟悉款。",
    ),
    RuleSmokeConfig(
        "005",
        "babytree_short_confirm_mom",
        "specific_expansion",
        "stance_shift",
        "question_hook",
        "article_light_emoji",
        "准备转 a2 时保留一点求确认，但不让别人做喂养决策。",
    ),
    RuleSmokeConfig(
        "006",
        "babytree_process_reminder_mom",
        "natural_article",
        "opening_shift",
        "scene_detail",
        "article_compact_clean",
        "到手扫码写成生活流程提醒，不变成攻略。",
    ),
    RuleSmokeConfig(
        "007",
        "babytree_ask_experienced_mom",
        "light_article_experience",
        "random_thinking_shift",
        "question_hook",
        "article_light_emoji",
        "质检和蜡样检测信息用找懂的妈妈的姿态承接。",
    ),
    RuleSmokeConfig(
        "008",
        "babytree_soft_reassure_mom",
        "specific_expansion",
        "stance_shift",
        "plain_explain",
        "article_clear_layout",
        "三方数据写得克制一点，只落到能看到、心里有底。",
    ),
    RuleSmokeConfig(
        "009",
        "babytree_dad_participation_mom",
        "natural_article",
        "opening_shift",
        "scene_detail",
        "article_compact_clean",
        "有货后核对报告加入家人协作的小动作。",
    ),
    RuleSmokeConfig(
        "010",
        "babytree_afraid_soldout_mom",
        "light_article_experience",
        "stance_shift",
        "plain_explain",
        "article_compact_clean",
        "转回 a2 的老用户轻带怕买不到，但不放大供应负面。",
    ),
]


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate local A2 sentiment post samples with new personas.")
    parser.add_argument("--rule-dir", type=Path, default=DEFAULT_RULE_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--limit", type=int, default=len(RULE_CONFIGS))
    parser.add_argument("--print-prompts-only", action="store_true")
    args = parser.parse_args()

    load_dotenv(ROOT / ".env")
    bridge_worker_env_to_direct_llm_env()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    results = asyncio.run(run_smoke(args.rule_dir, args.limit, print_prompts_only=args.print_prompts_only))
    json_path = args.output_dir / "a2_sentiment_post_new_persona_smoke.json"
    xlsx_path = args.output_dir / "a2_sentiment_post_new_persona_smoke.xlsx"
    prompt_path = args.output_dir / "rendered_prompts.md"

    json_path.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    write_workbook(results, xlsx_path)
    write_prompt_markdown(results, prompt_path)

    summary = {
        "count": len(results),
        "generated": sum(1 for item in results if item.get("status") == "generated"),
        "guard_pass": sum(1 for item in results if item.get("guard_pass") is True),
        "guard_fail": sum(1 for item in results if item.get("guard_pass") is False),
        "json_path": str(json_path),
        "xlsx_path": str(xlsx_path),
        "prompt_path": str(prompt_path),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


async def run_smoke(rule_dir: Path, limit: int, *, print_prompts_only: bool) -> list[dict[str, Any]]:
    keyword_content = normalize_system_prompt_keyword_content(fallback_system_prompt_keyword_content())
    guard = ActivityQualityGuardService()
    results: list[dict[str, Any]] = []
    for index, config in enumerate(RULE_CONFIGS[: max(0, limit)], start=1):
        rule_file = rule_dir / f"a2_sentiment_post_activity_post_rule_{config.rule_id}.txt"
        rule_text = rule_file.read_text(encoding="utf-8").strip()
        business_rule = build_business_rule(config, rule_file, rule_text)
        selected_keywords = _select_keyword_bundle(
            keyword_content,
            content_type="article",
            item_no=index,
            keyword_selection=business_rule["keyword_selection"],
        )
        expert = _fallback_expert_snapshot(
            "article_generator_v1",
            content_type="article",
            model_config=default_model_config(),
        )
        variables = _template_variables(
            content_type="article",
            output_fields=["title", "body"],
            business_rule=business_rule,
            selected_keywords=selected_keywords,
        )
        rendered_prompt = _render_template(expert["prompt_template"], variables)
        input_snapshot = {
            "schema_version": "1",
            "capability": "content.generate",
            "content_type": "article",
            "output_fields": ["title", "body"],
            "business_rule": business_rule,
            "selected_keywords": selected_keywords,
            "expert": expert,
            "model_config": expert["model_config"],
            "template_variables": variables,
            "rendered_prompt": rendered_prompt,
        }
        base_result = {
            "rule_id": config.rule_id,
            "rule_file": str(rule_file),
            "persona_code": config.persona_code,
            "persona_name": keyword_name(selected_keywords, "persona"),
            "diversity_note": config.diversity_note,
            "selected_keywords": selected_keywords,
            "selected_keywords_json": json.dumps(selected_keywords, ensure_ascii=False, indent=2),
            "rendered_prompt": rendered_prompt,
        }
        if print_prompts_only:
            results.append({**base_result, "status": "prompt_only", "title": "", "body": ""})
            continue
        try:
            output = await direct_content_generate(input_snapshot)
            item = review_with_guard(
                guard,
                title=str(output.get("title") or ""),
                body=str(output.get("body") or ""),
                business_rule=business_rule,
                selected_keywords=selected_keywords,
            )
            guard_payload = (item.quality_json or {}).get("activity_quality_guard") or {}
            issues = guard_payload.get("issues") or []
            results.append(
                {
                    **base_result,
                    "status": "generated",
                    "title": item.title,
                    "body": item.body,
                    "guard_pass": not issues,
                    "guard_issues": issues,
                    "guard_repairs": guard_payload.get("repairs") or [],
                    "runtime_result": output.get("runtime_result") or {},
                }
            )
        except Exception as exc:  # noqa: BLE001 - smoke output should preserve per-item failures.
            results.append(
                {
                    **base_result,
                    "status": "error",
                    "title": "",
                    "body": "",
                    "guard_pass": False,
                    "guard_issues": [{"code": type(exc).__name__, "message": str(exc)}],
                }
            )
    return results


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def bridge_worker_env_to_direct_llm_env() -> None:
    if not os.getenv("MAGA_DIRECT_CONTENT_MODEL"):
        os.environ["MAGA_DIRECT_CONTENT_MODEL"] = os.getenv("DEEPSEEK_MODEL") or "deepseek-v4-flash"
    if not os.getenv("MAGA_DIRECT_MODEL_BASE_URL") and os.getenv("MAGA_WORKER_MODEL_BASE_URL"):
        os.environ["MAGA_DIRECT_MODEL_BASE_URL"] = os.getenv("MAGA_WORKER_MODEL_BASE_URL", "")
    if not os.getenv("MAGA_DIRECT_MODEL_API_KEY") and os.getenv("MAGA_WORKER_MODEL_API_KEY"):
        os.environ["MAGA_DIRECT_MODEL_API_KEY"] = os.getenv("MAGA_WORKER_MODEL_API_KEY", "")


def default_model_config() -> dict[str, Any]:
    return {
        "provider_code": "aihubmix",
        "model_code": os.getenv("MAGA_DIRECT_CONTENT_MODEL") or os.getenv("DEEPSEEK_MODEL") or "deepseek-v4-flash",
        "temperature": 0.86,
        "max_tokens": 900,
        "timeout": 120,
        "system_prompt": "你是中文小红书母婴帖子生成器。严格按提示输出 JSON，不解释过程。",
    }


async def direct_content_generate(input_snapshot: dict[str, Any]) -> dict[str, Any]:
    model_config = dict(input_snapshot.get("model_config") or {})
    model = str(
        model_config.get("model_code")
        or os.getenv("MAGA_DIRECT_CONTENT_MODEL")
        or os.getenv("DEEPSEEK_MODEL")
        or "deepseek-v4-flash"
    )
    raw = await asyncio.to_thread(
        call_openai_compatible_model,
        model=model,
        system=str(model_config.get("system_prompt") or "你是中文小红书内容生成器，严格按用户提示输出，不解释过程。"),
        user=str(input_snapshot.get("rendered_prompt") or ""),
        temperature=float(model_config.get("temperature") or 0.86),
        max_tokens=int(model_config.get("max_tokens") or 900),
        timeout=float(model_config.get("timeout") or 120),
    )
    title, body = normalize_article_output(raw)
    if not body:
        raise ValueError("content.generate produced empty body")
    return {
        "title": title or "a2真实记录",
        "body": body,
        "runtime_result": {
            "mode": "local_openai_compatible_smoke",
            "fake": False,
            "model_code": model,
            "raw_output_length": len(raw),
        },
    }


def call_openai_compatible_model(
    *,
    model: str,
    system: str,
    user: str,
    temperature: float,
    max_tokens: int,
    timeout: float,
) -> str:
    endpoint = chat_completions_endpoint(
        os.getenv("MAGA_DIRECT_MODEL_BASE_URL")
        or os.getenv("OPENAI_BASE_URL")
        or os.getenv("AIHUBMIX_BASE_URL")
        or "https://aihubmix.com/v1"
    )
    api_key = (
        os.getenv("MAGA_DIRECT_MODEL_API_KEY")
        or os.getenv("AIHUBMIX_API_KEY")
        or os.getenv("OPENAI_API_KEY")
        or ""
    ).strip()
    if not api_key:
        raise RuntimeError("缺少 OpenAI-compatible API Key")
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    request = urllib.request.Request(
        endpoint,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        method="POST",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            data = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")[:300]
        raise RuntimeError(f"直连大模型 HTTP {exc.code}: {body}") from exc
    choices = data.get("choices") if isinstance(data, dict) else None
    if not isinstance(choices, list) or not choices:
        raise ValueError("直连大模型响应缺少 choices")
    first = choices[0] if isinstance(choices[0], dict) else {}
    message = first.get("message") if isinstance(first.get("message"), dict) else {}
    return str(message.get("content") or first.get("text") or "")


def chat_completions_endpoint(base_url: str) -> str:
    normalized = str(base_url or "").rstrip("/")
    if normalized.endswith("/chat/completions"):
        return normalized
    if normalized.endswith("/v1"):
        return f"{normalized}/chat/completions"
    return f"{normalized}/v1/chat/completions"


def normalize_article_output(raw: str) -> tuple[str, str]:
    parsed = parse_json_object(raw)
    title = str(parsed.get("title") or parsed.get("标题") or "").strip()
    body = str(parsed.get("body") or parsed.get("正文") or "").strip()
    if title and body:
        return title, body
    return loose_title_body_from_text(raw)


def parse_json_object(raw: str) -> dict[str, Any]:
    value = str(raw or "").strip()
    value = re.sub(r"^```(?:json)?\s*", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\s*```$", "", value).strip()
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", value, flags=re.DOTALL)
        if not match:
            return {}
        try:
            parsed = json.loads(match.group(0))
        except json.JSONDecodeError:
            return {}
    return parsed if isinstance(parsed, dict) else {}


def loose_title_body_from_text(raw: str) -> tuple[str, str]:
    lines = [line.strip() for line in str(raw or "").splitlines() if line.strip()]
    title = ""
    body_lines: list[str] = []
    for line in lines:
        normalized = re.sub(r"^(?:[-*•]\s*|\d+[、.．]\s*)", "", line).strip()
        title_match = re.match(r"^(?:标题|title)[:：]\s*(.+)$", normalized, flags=re.IGNORECASE)
        body_match = re.match(r"^(?:正文|body)[:：]\s*(.+)$", normalized, flags=re.IGNORECASE)
        if title_match:
            title = title_match.group(1).strip().strip('"“”')
            continue
        if body_match:
            body_lines.append(body_match.group(1).strip())
            continue
        body_lines.append(normalized)
    if not title and body_lines:
        first = body_lines[0]
        if len(first) <= 36 and len(body_lines) > 1:
            title = first.strip('"“”')
            body_lines = body_lines[1:]
    return title, "\n".join(body_lines).strip()


def build_business_rule(config: RuleSmokeConfig, rule_file: Path, rule_text: str) -> dict[str, Any]:
    heading = first_heading(rule_text)
    return {
        "rule_type": "article_business",
        "product_topic": "a2舆情改善帖子",
        "business_rule": heading,
        "corpus": rule_text,
        "asset_key": "a2_sentiment_post_activity",
        "rule_id": f"a2_sentiment_post_{config.rule_id}",
        "source_file": str(rule_file),
        "quality_guard_profile_key": PROFILE_KEY,
        "keyword_selection": {
            "persona": [config.persona_code],
            "writing_instruction": [config.writing_instruction_code],
            "perturbation_rule": [config.perturbation_code],
            "writing_method": [config.writing_method_code],
            "article_format_control": [config.format_code],
        },
        "brief_constraints": {
            "title": "8到18个字，像真实用户随手发帖标题，不写攻略/指南/建议口吻。",
            "body": "正文2到4段，120到260字，保留真实生活细节；不要照搬示例。",
        },
        "generation_requirements": (
            '只输出 JSON 对象，格式为 {"title": "...", "body": "..."}，不要 Markdown、编号或解释。\n'
            "标题和正文必须匹配同一个场景；不能让别人替自己做转奶、喂养或安全判断。\n"
            "报告信息只写能扫到、能看到、能核对、自己不一定看得专业；不要写成绝对安全或专业背书。\n"
            "报告段不要出现“安全、没问题、放心喝、保证”这类结论词，即使是否定句或引号里也不要出现。\n"
            "不要出现专项禁词：习惯性、断档、愿意喝着、愿意喝上。\n"
            "可以写个人松口气、先补一罐、慢慢来，但不要放大断货焦虑，不要踩竞品。"
        ),
    }


def first_heading(text: str) -> str:
    for line in text.splitlines():
        stripped = line.strip().strip("：:")
        if stripped:
            return stripped
    return "A2舆情改善帖子"


def keyword_name(selected_keywords: list[dict[str, Any]], category_code: str) -> str:
    for item in selected_keywords:
        if item.get("category_code") == category_code:
            return str(item.get("keyword_name") or item.get("keyword_code") or "")
    return ""


def review_with_guard(
    guard: ActivityQualityGuardService,
    *,
    title: str,
    body: str,
    business_rule: dict[str, Any],
    selected_keywords: list[dict[str, Any]],
) -> SimpleNamespace:
    # 重要逻辑：模拟 batch item 的 plan_json，让活动 guard 能读取人设/写法上下文并做同链路修复。
    item = SimpleNamespace(
        title=title.strip(),
        body=body.strip(),
        status="generated",
        plan_json={
            **business_rule,
            "unified_generation": {"selected_keywords": selected_keywords},
        },
        quality_json={},
    )
    guard.review_item(item, PROFILE_KEY)
    return item


def write_workbook(results: list[dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "samples"
    headers = [
        "rule_id",
        "persona_code",
        "persona_name",
        "diversity_note",
        "title",
        "body",
        "guard_pass",
        "guard_issues",
        "guard_repairs",
        "selected_keywords_json",
    ]
    ws.append(headers)
    for result in results:
        ws.append(
            [
                result.get("rule_id", ""),
                result.get("persona_code", ""),
                result.get("persona_name", ""),
                result.get("diversity_note", ""),
                result.get("title", ""),
                result.get("body", ""),
                "是" if result.get("guard_pass") is True else "否",
                json.dumps(result.get("guard_issues") or [], ensure_ascii=False),
                json.dumps(result.get("guard_repairs") or [], ensure_ascii=False),
                result.get("selected_keywords_json", ""),
            ]
        )
    style_sheet(ws)

    prompt_ws = wb.create_sheet("rendered_prompts")
    prompt_ws.append(["rule_id", "persona_code", "rendered_prompt"])
    for result in results:
        prompt_ws.append([result.get("rule_id", ""), result.get("persona_code", ""), result.get("rendered_prompt", "")])
    style_sheet(prompt_ws)

    wb.save(path)


def style_sheet(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        1: 10,
        2: 34,
        3: 18,
        4: 34,
        5: 24,
        6: 58,
        7: 12,
        8: 45,
        9: 35,
        10: 55,
    }
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_prompt_markdown(results: list[dict[str, Any]], path: Path) -> None:
    chunks = ["# A2 sentiment post new persona smoke prompts\n"]
    for result in results:
        chunks.append(f"## {result.get('rule_id')} / {result.get('persona_code')}\n")
        chunks.append("```text\n")
        chunks.append(str(result.get("rendered_prompt") or ""))
        chunks.append("\n```\n")
    path.write_text("\n".join(chunks), encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
