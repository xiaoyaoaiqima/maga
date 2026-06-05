#!/usr/bin/env python3
"""Check comment-angle example wording against approved reference comments.

The checker is intentionally stricter than a token diff. It reports:
- high risk: unsupported coined wording or AI-like summary patterns;
- medium risk: wording that may need copy review because it is unsupported
  and easy to overuse. Business anchors such as ml numbers, age stage,
  checkup/curve data, and pre-purchase question words are not risk findings.

Default inputs match the 源悦 comment corpus workflow used in this repo.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("openpyxl is required in the bundled Python runtime") from exc


REPO_ROOT = Path(__file__).resolve().parents[1]
LOCAL_DEPS = REPO_ROOT / ".codex_deps" / "python"
if LOCAL_DEPS.exists():
    sys.path.insert(0, str(LOCAL_DEPS))

try:
    import jieba
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit(
        "jieba is required. Install it with: "
        f"{sys.executable} -m pip install jieba --target {LOCAL_DEPS}"
    ) from exc


DEFAULT_CSV = REPO_ROOT / "批量生成" / "评论切角_子关键词导出.csv"
DEFAULT_BRAND = Path("/Users/luxifa/Downloads/品牌反馈-0602-源悦ai评论-25条.xlsx")
DEFAULT_REFERENCE = Path("/Users/luxifa/Downloads/评论示例-评论区【执行】Y26-1月-源悦-AO seeding.xlsx")

STOP_TOKENS = {
    "的",
    "了",
    "是",
    "在",
    "和",
    "或",
    "与",
    "就",
    "都",
    "也",
    "很",
    "还",
    "挺",
    "比较",
    "这个",
    "这款",
    "现在",
    "之前",
    "后面",
    "最近",
    "目前",
    "感觉",
    "觉得",
    "可以",
    "还是",
    "就是",
    "不是",
    "我们",
    "我家",
    "宝宝",
    "娃",
    "宝妈",
    "妈妈",
    "姐妹",
    "同款",
    "源悦",
    "奶粉",
    "转奶",
}

HIGH_RISK_TERMS = {
    "少操心",
    "纸尿裤",
    "尿不湿",
    "好收拾",
    "重敏",
    "盯得很紧",
    "功课做一堆",
    "楼上",
    "宝妈路过",
    "松口气",
    "不急着下结论",
    "下结论",
    "蹿个子",
    "救我",
    "支棱",
    "佛系",
    "期待太满",
    "袖口",
    "顺顺喝",
    "别乱就行",
    "活动价",
    "这单不亏",
    "买着比较踏实",
    "看得过去",
    "能干掉",
    "嗷嗷叫",
    "找奶嘴",
    "忍着没敢多给",
    "小鸡胃",
    "预算里",
    "奶瓶里不怎么剩",
    "我崽",
    "没那么严重",
    "没憋到脸红",
    "有没有憋着",
    "会不会用力",
    "第二顿直接",
}

NEGATIVE_FACTOR_TERMS = {
    "粑粑干",
    "便便变绿",
    "变绿",
    "有颗粒",
    "摇了好久",
    "摇了三分钟",
    "冲不开",
    "没啥反应",
    "没问题才",
    "过敏宝",
    "拉稀",
    "屁多",
    "颜色异常",
    "颜色深",
}

PRE_PURCHASE_QUESTION_MARKERS = (
    "能喝",
    "能试",
    "可以喝",
    "适合",
    "会不会",
    "容易不容易",
    "怕",
    "担心",
    "想问",
    "问问",
    "求问",
    "有没有",
    "正常不",
    "吗",
    "嘛",
    "么",
    "?",
    "？",
)

POST_USE_PATTERN = re.compile(
    r"(喝了|喝完|喝后|喝下来|喝着|换了|换这款|换源悦|转了|转源悦|"
    r"转奶后|试了|小罐试了|这罐|半罐|一罐|冲了|泡了|摇了好久|摇了三分钟|"
    r"之后|以后|还是有颗粒)"
)

PRODUCT_DIRECTION_PATTERNS = [
    (r"(刚转|转|试了|小罐试了|刚试)(能恩|水解)(?!转源悦)", "产品方向错误：源悦不是最终对象"),
    (r"喝水解", "产品方向错误：源悦不是最终对象"),
    (r"(能恩|水解)[^，。！？]{0,8}(刚试|试了|先试|继续看)", "产品方向错误：竞品/水解被写成正在试用对象"),
]

OPERATOR_0603_PATTERNS = [
    (r"价格[^，。！？]{0,6}看过|看过[^，。！？]{0,6}价格|配方和价格[^，。！？]{0,8}看", "运营0603：价格/配方“看过”表达不自然，改成性价比高/很划算/准备再囤"),
    (r"配这个价|这个价[^，。！？]{0,4}配这个配料|这个价位在预算里", "运营0603：性价比不要写成预算/配这个价，改成很划算/性价比高/准备再囤"),
    (r"每天(?:差不多|基本)?(?:都)?有(?![^，。！？]{0,8}(便便|大便|拉屎|拉粑粑|拉臭臭|拉💩))", "运营0603：便便指代不清，要写明基本一天一次/拉粑粑基本每天1次"),
    (r"每天差不多能看到一次", "运营0603：看到一次指代不清，改成拉粑粑基本每天1次"),
    (r"三顿\d+\s*ml", "运营0603：奶量数字有总量/单顿歧义，避免写成一顿不到常识量"),
    (r"总吐(?!奶)", "运营0603：未购口感痛点要写总吐奶，不要只写总吐"),
    (r"粑粑干", "运营0603：改成大便比较干燥"),
]

MEDIUM_REVIEW_TERMS = {
    "一点点",
}

SAFE_REFERENCE_TERMS = {
    "同款",
    "能恩转源悦",
    "不胀气",
    "便秘",
    "奶量",
    "喝奶秒没",
    "小罐",
    "奶源和配料",
}

HOMOGENEITY_WATCH_THRESHOLDS = {
    "先继续看": 2,
    "原来那款": 3,
    "之前那款": 3,
    "喝源悦不胀气": 2,
    "怎么观察": 2,
    "能接受": 5,
    "味道比较清淡": 2,
    "慢慢混着转": 2,
    "旧奶粉": 5,
    "分几顿": 5,
    "这个味道": 5,
    "称重": 3,
    "多了点": 3,
    "有点肉": 3,
    "自己那条线上": 3,
    "正常范围": 3,
    "儿保": 3,
    "稳": 4,
    "顺": 6,
    "拍完嗝能歇会儿": 2,
    "能接上": 5,
    "接上": 7,
    "问了一圈": 3,
    "有喝过的": 4,
    "想先问问": 3,
    "源悦味道": 3,
    "怎么安排": 3,
    "这阵子": 8,
    "最近": 10,
    "继续": 8,
    "咋样": 6,
    "有没有": 5,
    "怎么样": 12,
    "同款": 8,
    "顺手": 2,
}

HOMOGENEITY_OPENING_THRESHOLD = 5
HOMOGENEITY_ENDING_THRESHOLD = 4
TING_PHRASE_THRESHOLD = 2
TING_TOTAL_MIN_COUNT = 5
TING_TOTAL_RATIO_THRESHOLD = 0.05
SOFT_TEMPLATE_WATCHES = {
    "费不费劲": (3, "运营0603接受费不费劲，但同批要和不费劲了 / 基本一天一次 / 干不干轮换，且指代要清楚。"),
    "娃肯喝": (3, "拆成很喜欢 / 爱喝 / 还挺爱喝 / 准备再囤一点，别改成奶瓶里不怎么剩。"),
    "能喝完": (3, "拆成爱喝 / 还挺爱喝 / 奶量从X到Y / 准备再囤一点，注意 ml 数别有总量歧义。"),
    "喝得下": (3, "如果同批过敏/转奶都写喝得下，要换成接受度、奶量、是否剩奶等不同事实。"),
    "没怎么": (5, "改成更具体的小事实，例如少哼唧 / 没鼓起来 / 不用太使劲 / 没掉下来。"),
    "不太": (5, "避免把安全表达换成新的抽象模板，优先写具体行为或数据。"),
    "基本每天": (3, "要写清楚便便/拉粑粑基本每天1次，别只写每天基本有。"),
    "差不多每天": (3, "同批便便切角不要都用这个频次描述，可换成便便规律 / 基本一天一次 / 拉屎不费劲了。"),
    "能接受": (3, "拆成性价比高 / 很划算 / 准备再囤一点 / 我觉得很值。"),
    "没抗拒": (3, "拆成很喜欢 / 爱喝 / 接受度不错 / 准备再囤一点，不要都用没抗拒。"),
    "奶瓶里不怎么剩": (2, "运营0603判病句，改成奶粉没剩多少了 / 准备再囤一点 / 很喜欢。"),
    "爱喝": (6, "运营喜欢爱喝，但同批过多会模板化，要和很喜欢 / 喝着可以 / 喝得不错 / 接着喝轮换。"),
    "娃爱喝": (3, "不要把所有认可都写成娃爱喝，可换成家里少折腾、奶量数据、准备再囤、体检数据。"),
    "奶量没掉": (4, "生长发育可用，但同批过多要换成奶量正常 / 每顿量能接上 / 具体 ml / 体检身高数据。"),
    "少折腾": (3, "容易生病切角可用，但同批过多要换成换季状态、家里少闹、接着喝到三段等不同事实。"),
    "体检身高": (4, "生长发育要有数据感，但不要都用体检身高，可换成X个月X斤、身高数据在中上、衣服短了点。"),
}
SOFT_TEMPLATE_PATTERNS = [
    (
        "便便省事模板",
        re.compile(r"(拉屎|嗯嗯|便便|臭臭)[^，。！？]{0,12}(费劲|憋劲|憋着|脸红)"),
        5,
        "便便切角要打散：便便规律基本一天一次 / 拉屎不费劲了 / 干不干，避免憋着/脸红等怪表达。",
    ),
    (
        "短周期接受度模板",
        re.compile(r"(第一顿|前两顿|刚开|刚换|刚喝)[^，。！？]{0,14}(没抗拒|能喝完|喝得下|奶瓶里剩)"),
        2,
        "短周期只写观察，不要多条都用第一顿/前两顿 + 接受度模板。",
    ),
    (
        "奶后肚子模板",
        re.compile(r"(奶后|喝完|肚子|拍嗝)[^，。！？]{0,12}(没鼓|不胀|少哼唧|没那么费事|能歇会儿)"),
        4,
        "消化吸收切角要混用肚子、拍嗝、哼唧、夜里那顿等不同事实，不要都写奶后反应。",
    ),
    (
        "价格接受模板",
        re.compile(r"(能接受|在预算里|我觉得可以|配这个价|价格也能|价格也在)"),
        4,
        "性价比切角要加入奶源、配料表、渠道价、续罐动作，别都写接受/预算。",
    ),
]
HOMOGENEITY_SKIP_EDGE_TEXT = {
    "我家",
    "我们",
    "宝宝",
    "这个",
    "这款",
    "源悦",
    "奶粉",
    "喝完",
    "奶后",
}

SUGGESTIONS = {
    "便便问题": "优先换成：便便规律，基本一天一次 / 拉粑粑基本每天1次 / 拉屎不费劲了 / 大便比较干燥能喝吗",
    "奶量补充": "优先换成：6个月断奶后奶量接上了 / 一天4-5顿，每次200+ / 喝奶秒没选手",
    "过敏相关": "优先换成：敏感宝宝换奶按顿来 / 源悦小罐先试 / 严重的先问医生；不要提旧奶粉",
    "生长发育": "优先换成：体检身高在中上 / X个月X斤 / 奶量没掉还爱喝 / 对源悦有认可，别只写再看再称",
    "消化吸收": "优先换成：喝完奶不胀气 / 不哼哼唧唧了 / 没有不舒服 / 喝完奶肚子不鼓了",
    "性价比": "优先换成：性价比高 / 很划算 / 准备再囤一点 / 很喜欢；别写预算里、配这个价、奶瓶里不怎么剩",
}


@dataclass
class CorpusItem:
    row_no: int
    angle: str
    title: str
    text: str
    source: str
    line_no: int | None = None


@dataclass
class Finding:
    severity: str
    item: CorpusItem
    reasons: list[str] = field(default_factory=list)
    suggestion: str = ""

    def to_dict(self) -> dict[str, object]:
        return {
            "severity": self.severity,
            "line_no": self.item.line_no,
            "row_no": self.item.row_no,
            "source": self.item.source,
            "angle": self.item.angle,
            "title": self.item.title,
            "text": self.item.text,
            "reasons": self.reasons,
            "suggestion": self.suggestion,
        }


@dataclass
class HomogeneityObservation:
    kind: str
    phrase: str
    count: int
    samples: list[CorpusItem] = field(default_factory=list)
    suggestion: str = ""

    def to_dict(self) -> dict[str, object]:
        return {
            "kind": self.kind,
            "phrase": self.phrase,
            "count": self.count,
            "samples": [
                {
                    "line_no": item.line_no,
                    "row_no": item.row_no,
                    "angle": item.angle,
                    "text": item.text,
                }
                for item in self.samples
            ],
            "suggestion": self.suggestion,
        }


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", "", text).lower()


def normalize_example_for_duplicate(text: str) -> str:
    return re.sub(r"[\s~～。！!？?，,、；;：:（）()「」“”\"'`]+", "", text).lower()


def tokenize(text: str) -> list[str]:
    tokens: list[str] = []
    for token in jieba.cut(text):
        token = token.strip().lower()
        if not token:
            continue
        if re.fullmatch(r"[\W_]+", token):
            continue
        if re.fullmatch(r"\d+", token):
            continue
        if len(token) == 1:
            continue
        if token in STOP_TOKENS:
            continue
        tokens.append(token)
    return tokens


def chinese_ngrams(text: str, min_len: int = 2, max_len: int = 6) -> set[str]:
    phrases: set[str] = set()
    for chunk in re.findall(r"[\u4e00-\u9fff]+", text):
        for size in range(min_len, max_len + 1):
            for start in range(0, len(chunk) - size + 1):
                phrase = chunk[start : start + size]
                if phrase not in STOP_TOKENS:
                    phrases.add(phrase)
    return phrases


def read_reference_texts(brand_path: Path, reference_path: Path) -> list[str]:
    texts: list[str] = []

    brand_book = load_workbook(brand_path, read_only=True, data_only=True)
    brand_sheet = brand_book["Sheet1"]
    for row in range(3, brand_sheet.max_row + 1):
        for col in (3, 7):
            value = brand_sheet.cell(row, col).value
            if isinstance(value, str) and value.strip():
                texts.append(value.strip())

    reference_book = load_workbook(reference_path, read_only=True, data_only=True)
    reference_sheet = reference_book["评论区"]
    headers = [reference_sheet.cell(1, col).value for col in range(1, reference_sheet.max_column + 1)]
    header_index = {header: index + 1 for index, header in enumerate(headers) if header}
    text_col = header_index["文案范例"]
    for row in range(2, reference_sheet.max_row + 1):
        value = reference_sheet.cell(row, text_col).value
        if isinstance(value, str) and value.strip():
            texts.append(value.strip())

    return texts


def read_csv_rows(csv_path: Path) -> list[dict[str, str]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        content = "".join(line for line in handle if not line.startswith("#"))
    return list(csv.DictReader(io.StringIO(content)))


def extract_review_csv_items(csv_path: Path) -> list[CorpusItem]:
    rows = read_csv_rows(csv_path)
    items: list[CorpusItem] = []
    for row_no, row in enumerate(rows, start=2):
        text = (row.get("正文") or "").strip()
        if not text:
            continue
        reviewer = (row.get("审核人") or "").strip()
        title = f"审核人 {reviewer}" if reviewer else "审核CSV"
        items.append(
            CorpusItem(
                row_no=row_no,
                angle=(row.get("评论类型") or "").strip(),
                title=title,
                text=text,
                source="review-csv",
            )
        )
    return items


def line_number_map(csv_path: Path) -> dict[str, list[int]]:
    mapping: dict[str, list[int]] = defaultdict(list)
    for line_no, line in enumerate(csv_path.read_text(encoding="utf-8-sig").splitlines(), start=1):
        stripped = line.strip()
        if stripped.startswith("- "):
            mapping[stripped[2:].strip()].append(line_no)
    return mapping


def extract_article_pool_items(workbook_path: Path, sheet_name: str) -> list[CorpusItem]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    header_index = {header: index + 1 for index, header in enumerate(headers) if header}
    body_col = header_index["正文"]
    context_col = header_index.get("上下文变量(context_list)")
    id_col = header_index.get("ID")

    items: list[CorpusItem] = []
    for row_no in range(2, sheet.max_row + 1):
        text = sheet.cell(row_no, body_col).value
        if not isinstance(text, str) or not text.strip():
            continue
        angle = ""
        if context_col:
            raw_context = sheet.cell(row_no, context_col).value
            if isinstance(raw_context, str) and raw_context.strip():
                try:
                    angle = json.loads(raw_context).get("评论切角", "")
                except json.JSONDecodeError:
                    angle = ""
        content_id = sheet.cell(row_no, id_col).value if id_col else row_no
        items.append(
            CorpusItem(
                row_no=row_no,
                angle=angle,
                title=f"ID {content_id}",
                text=text.strip(),
                source="article-pool",
            )
        )
    return items


def extract_items(rows: list[dict[str, str]], line_map: dict[str, list[int]]) -> list[CorpusItem]:
    items: list[CorpusItem] = []
    for row_no, row in enumerate(rows, start=2):
        angle = row.get("评论切角", "")
        block = row.get("语料", "")
        title = block.split("\n", 1)[0].strip().rstrip("：:")
        in_examples = False
        for raw_line in block.splitlines():
            line = raw_line.strip()
            if line in {"示例：", "真人示例："}:
                in_examples = True
                continue
            if line.startswith("注意"):
                in_examples = False
            if in_examples and line.startswith("- "):
                text = line[2:].strip()
                item = CorpusItem(
                    row_no=row_no,
                    angle=angle,
                    title=title,
                    text=text,
                    source="example",
                    line_no=(line_map.get(text) or [None]).pop(0),
                )
                items.append(item)
    return items


def extract_rule_items(rows: list[dict[str, str]]) -> list[CorpusItem]:
    items: list[CorpusItem] = []
    for row_no, row in enumerate(rows, start=2):
        angle = row.get("评论切角", "")
        block = row.get("语料", "")
        title = block.split("\n", 1)[0].strip().rstrip("：:")
        in_examples = False
        for raw_line in block.splitlines()[1:]:
            line = raw_line.strip()
            if not line:
                continue
            if line in {"示例：", "真人示例："}:
                in_examples = True
                continue
            if line.startswith("注意"):
                in_examples = False
                continue
            if line.startswith("示例只作为语义素材"):
                continue
            if in_examples or line.startswith("- "):
                continue
            items.append(CorpusItem(row_no=row_no, angle=angle, title=title, text=line, source="rule"))
    return items


def build_reference_sets(texts: Iterable[str]) -> tuple[str, set[str], set[str]]:
    text_blob = "\n".join(texts)
    normalized = normalize_text(text_blob)
    token_set: set[str] = set()
    phrase_set: set[str] = set()
    for text in texts:
        token_set.update(tokenize(text))
        phrase_set.update(chinese_ngrams(text))
    return normalized, token_set, phrase_set


def is_supported(value: str, ref_text: str, ref_tokens: set[str], ref_phrases: set[str]) -> bool:
    return normalize_text(value) in ref_text or value.lower() in ref_tokens or value in ref_phrases


def sentence_pattern_reasons(text: str, ref_text: str) -> list[str]:
    normalized = normalize_text(text)
    reasons: list[str] = []
    if re.search(r"(?:还)?挺(稳|安稳|顺|适应|放心|认|对味)(?:的|了|吧|呀|啊|呢|啦|多了|些)?", text):
        reasons.append("句式风险：挺 + 抽象评价词")
    if normalized in ref_text:
        return reasons

    if re.search(r"[\u4e00-\u9fff]{1,8}这块[^，。！？]*?(还挺|挺|还算|比较)[^，。！？]*?(规律|稳|省心|顺|正常|可以|还行|舒服|轻松)", text):
        reasons.append("句式风险：X这块 + 挺/还挺 + 抽象总结词")
    if re.search(r"老母亲[^，。！？]*?(安心|省心|少操心)", text):
        reasons.append("句式风险：老母亲 + 安心/省心/少操心类安全总结")
    if re.search(r"(状态|节奏|规律)[^，。！？]{0,8}(挺|还挺|还算|比较|一直|慢慢|整体)?[^，。！？]{0,4}(稳|顺|正常|在线|还行|可以|好)", text):
        reasons.append("句式风险：状态/节奏/规律 + 总结口吻")
    if re.search(r"(肚肚|肚子|便便|臭臭|奶量|体质|吃喝|喂养|拍嗝|吐奶)[^，。！？]{0,8}(还算|还挺|挺|比较|整体|目前|一直|也)?[^，。！？]{0,4}(稳|稳定|省心|放心|安心|还行|还可以|在线)", text):
        reasons.append("句式风险：身体/喂养状态 + 稳/省心/还行类安全总结")
    if re.search(r"(拉屎|便便|嗯嗯|臭臭|拍嗝|喝奶|喝得|拉得|第一顿|这顿|奶后|喝完)[^，。！？]{0,8}(顺(?!序)|顺利|顺畅)", text):
        reasons.append("句式风险：身体/喂养状态 + 顺类安全总结")
    return reasons


def is_unpurchased_angle(angle: str) -> bool:
    return angle.strip().startswith("未购")


def is_negated_negative_factor(text: str, term: str) -> bool:
    return bool(re.search(rf"(没|没有|不|不会|未)[^，。！？?]{{0,4}}{re.escape(term)}", text))


def is_pre_purchase_negative_question(item: CorpusItem, text: str) -> bool:
    if not is_unpurchased_angle(item.angle):
        return False
    if POST_USE_PATTERN.search(text):
        return False
    return any(marker in text for marker in PRE_PURCHASE_QUESTION_MARKERS)


def hard_redline_reasons(item: CorpusItem) -> list[str]:
    text = item.text
    reasons: list[str] = []
    for pattern, reason in PRODUCT_DIRECTION_PATTERNS:
        if re.search(pattern, text):
            reasons.append(reason)
    for pattern, reason in OPERATOR_0603_PATTERNS:
        if re.search(pattern, text):
            reasons.append(reason)

    for term in sorted(NEGATIVE_FACTOR_TERMS, key=len, reverse=True):
        if term in text:
            if is_negated_negative_factor(text, term):
                continue
            if is_pre_purchase_negative_question(item, text):
                continue
            reasons.append(f"高风险负面因素：{term}")

    if "水解" in text and not re.search(r"(医生|严重|问医生|听医生)", text):
        reasons.append("产品方向风险：水解未作为医生边界出现")
    if "过敏相关" in item.angle and "旧奶粉" in text:
        reasons.append("运营0603：过敏相关不要提旧奶粉，改成源悦小罐/前几顿少量/按顿加")
    if "生长发育" in item.angle:
        if re.search(r"(奶量.*接上|再称|称重|再量|下次体检|再看|体重也没往下掉|奶量没掉.*再看)", text):
            reasons.append("运营0603：生长发育不能只写观察/没掉，要体现对源悦的认可或给出具体身高体重数据")
        if not re.search(r"(源悦|这款|奶粉|喝着|爱喝|奶量|体检|身高|体重|斤|cm|厘米|长个|长高|个子|肉)", text):
            reasons.append("运营0603：生长发育切角缺少奶粉认可或生长事实")
    if "容易生病" in item.angle and not re.search(r"(宝宝|娃|体质|感冒|生病|阿秋|闹|接受|爱喝|状态)", text):
        reasons.append("运营0603：容易生病切角不能只写跟风/喝到三段/没换，要带宝宝状态或体质相关事实")
    return reasons


def suggestion_for(angle: str) -> str:
    for keyword, suggestion in SUGGESTIONS.items():
        if keyword in angle:
            return suggestion
    return "优先换成两个参考文件里已出现的具体口径，避免自创安全总结句。"


def classify_item(
    item: CorpusItem,
    ref_text: str,
    ref_tokens: set[str],
    ref_phrases: set[str],
    verbose_diff: bool,
) -> Finding | None:
    high_reasons: list[str] = []
    medium_reasons: list[str] = []

    high_reasons.extend(hard_redline_reasons(item))

    for term in sorted(HIGH_RISK_TERMS, key=len, reverse=True):
        if term in item.text and not is_supported(term, ref_text, ref_tokens, ref_phrases):
            high_reasons.append(f"高风险词未见于参考池：{term}")

    high_reasons.extend(sentence_pattern_reasons(item.text, ref_text))

    for term in sorted(MEDIUM_REVIEW_TERMS, key=len, reverse=True):
        if term in item.text and not is_supported(term, ref_text, ref_tokens, ref_phrases):
            medium_reasons.append(f"中风险词需人工复核：{term}")

    if verbose_diff:
        token_diff = [
            token
            for token in sorted(set(tokenize(item.text)))
            if token not in ref_tokens and normalize_text(token) not in ref_text and token not in SAFE_REFERENCE_TERMS
        ]
        if token_diff:
            medium_reasons.append("分词差集：" + "、".join(token_diff[:8]))

        phrase_diff = [
            phrase
            for phrase in sorted(chinese_ngrams(item.text), key=lambda x: (-len(x), x))
            if len(phrase) >= 3 and phrase not in ref_phrases and normalize_text(phrase) not in ref_text
        ]
        phrase_diff = [
            phrase
            for phrase in phrase_diff
            if not any(phrase in reason for reason in high_reasons)
        ][:8]
        if phrase_diff:
            medium_reasons.append("短语差集：" + "、".join(phrase_diff))

    if high_reasons:
        return Finding("high", item, high_reasons, suggestion_for(item.angle))
    if medium_reasons:
        return Finding("medium", item, medium_reasons, suggestion_for(item.angle))
    return None


def duplicate_example_findings(items: list[CorpusItem]) -> list[Finding]:
    grouped: dict[str, list[CorpusItem]] = defaultdict(list)
    for item in items:
        normalized = normalize_example_for_duplicate(item.text)
        if normalized:
            grouped[normalized].append(item)

    findings: list[Finding] = []
    for duplicate_items in grouped.values():
        if len(duplicate_items) < 2:
            continue
        first = duplicate_items[0]
        first_line = f"L{first.line_no}" if first.line_no else f"row {first.row_no}"
        for item in duplicate_items[1:]:
            findings.append(
                Finding(
                    "high",
                    item,
                    [f"示例重复：去标点后与 {first_line} 相同"],
                    "同一份评论切角语料里保留一个入口，其余改成不同语义点，避免生文同质化。",
                )
            )
    return findings


def edge_key(text: str, side: str) -> str:
    normalized = re.sub(r"[\s~～。！!？?，,、；;：:（）()「」“”\"'`]+", "", text)
    return normalized[:4] if side == "opening" else normalized[-4:]


def should_skip_edge_key(key: str) -> bool:
    if not key or len(key) < 4:
        return True
    if re.search(r"\d|ml|ML", key):
        return True
    return key in HOMOGENEITY_SKIP_EDGE_TEXT


def homogeneity_observations(items: list[CorpusItem], sample_limit: int = 5) -> list[HomogeneityObservation]:
    observations: list[HomogeneityObservation] = []

    for phrase, threshold in HOMOGENEITY_WATCH_THRESHOLDS.items():
        matches = [item for item in items if phrase in item.text]
        if len(matches) >= threshold:
            observations.append(
                HomogeneityObservation("watch_phrase", phrase, len(matches), matches[:sample_limit])
            )

    for phrase, (threshold, suggestion) in SOFT_TEMPLATE_WATCHES.items():
        matches = [item for item in items if phrase in item.text]
        if len(matches) >= threshold:
            observations.append(
                HomogeneityObservation(
                    "soft_template",
                    phrase,
                    len(matches),
                    matches[:sample_limit],
                    suggestion,
                )
            )

    for phrase, pattern, threshold, suggestion in SOFT_TEMPLATE_PATTERNS:
        matches = [item for item in items if pattern.search(item.text)]
        if len(matches) >= threshold:
            observations.append(
                HomogeneityObservation(
                    "soft_template_pattern",
                    phrase,
                    len(matches),
                    matches[:sample_limit],
                    suggestion,
                )
            )

    ting_items = [item for item in items if "挺" in item.text]
    ting_total_threshold = max(TING_TOTAL_MIN_COUNT, int(len(items) * TING_TOTAL_RATIO_THRESHOLD))
    if ting_items and len(ting_items) >= ting_total_threshold:
        observations.append(
            HomogeneityObservation(
                "ting_total",
                f"含“挺”的示例/正文（{len(ting_items)}/{len(items)}）",
                len(ting_items),
                ting_items[:sample_limit],
            )
        )

    ting_phrase_groups: dict[str, list[CorpusItem]] = defaultdict(list)
    for item in items:
        for match in re.finditer(r"(?:还)?挺[^，。！？,.!?、；;：:\s]{1,6}", item.text):
            phrase = re.sub(r"(的|了|吧|呀|啊|呢|啦)$", "", match.group(0))
            if len(phrase) < 2:
                continue
            ting_phrase_groups[phrase].append(item)
    for phrase, matches in ting_phrase_groups.items():
        if len(matches) >= TING_PHRASE_THRESHOLD:
            observations.append(
                HomogeneityObservation("ting_phrase", phrase, len(matches), matches[:sample_limit])
            )

    for side, threshold in [
        ("opening", HOMOGENEITY_OPENING_THRESHOLD),
        ("ending", HOMOGENEITY_ENDING_THRESHOLD),
    ]:
        grouped: dict[str, list[CorpusItem]] = defaultdict(list)
        for item in items:
            key = edge_key(item.text, side)
            if should_skip_edge_key(key):
                continue
            grouped[key].append(item)
        for key, matches in grouped.items():
            if len(matches) >= threshold:
                observations.append(
                    HomogeneityObservation(side, key, len(matches), matches[:sample_limit])
                )

    observations.sort(key=lambda item: (-item.count, item.kind, item.phrase))
    return observations


def run_check(args: argparse.Namespace) -> tuple[list[Finding], dict[str, int]]:
    reference_texts = read_reference_texts(args.brand_feedback, args.reference_examples)
    ref_text, ref_tokens, ref_phrases = build_reference_sets(reference_texts)
    if args.article_pool_xlsx:
        rows = []
        example_items = extract_article_pool_items(args.article_pool_xlsx, args.article_pool_sheet)
        rule_items = []
    elif args.review_csv:
        rows = read_csv_rows(args.review_csv)
        example_items = extract_review_csv_items(args.review_csv)
        rule_items = []
    else:
        rows = read_csv_rows(args.csv)
        line_map = line_number_map(args.csv)
        example_items = extract_items(rows, line_map)
        rule_items = extract_rule_items(rows) if args.include_rules else []

    findings: list[Finding] = []
    if not args.article_pool_xlsx:
        findings.extend(duplicate_example_findings(example_items))
    for item in example_items:
        finding = classify_item(item, ref_text, ref_tokens, ref_phrases, args.verbose_diff)
        if finding:
            findings.append(finding)
    for item in rule_items:
        finding = classify_item(item, ref_text, ref_tokens, ref_phrases, False)
        if finding:
            finding.severity = "rule-" + finding.severity
            findings.append(finding)

    stats = {
        "reference_items": len(reference_texts),
        "csv_rows": len(rows),
        "example_items": len(example_items),
        "rule_items": len(rule_items),
        "high": sum(1 for item in findings if item.severity == "high"),
        "medium": sum(1 for item in findings if item.severity == "medium"),
        "rule_high": sum(1 for item in findings if item.severity == "rule-high"),
        "rule_medium": sum(1 for item in findings if item.severity == "rule-medium"),
    }
    return findings, stats


def collect_homogeneity_observations(args: argparse.Namespace) -> list[HomogeneityObservation]:
    if args.article_pool_xlsx:
        items = extract_article_pool_items(args.article_pool_xlsx, args.article_pool_sheet)
    elif args.review_csv:
        items = extract_review_csv_items(args.review_csv)
    else:
        rows = read_csv_rows(args.csv)
        items = extract_items(rows, line_number_map(args.csv))
    return homogeneity_observations(items)


def print_text_report(findings: list[Finding], stats: dict[str, int], max_items: int) -> None:
    print("评论切角参考语料一致性监测")
    print(
        "stats: "
        + ", ".join(f"{key}={value}" for key, value in stats.items())
    )

    for severity, title in [
        ("high", "高风险示例"),
        ("medium", "中风险示例"),
        ("rule-high", "高风险规则句"),
        ("rule-medium", "中风险规则句"),
    ]:
        grouped = [finding for finding in findings if finding.severity == severity]
        print(f"\n{title}: {len(grouped)}")
        for finding in grouped[:max_items]:
            line = f"L{finding.item.line_no}" if finding.item.line_no else f"row {finding.item.row_no}"
            print(f"- {line} | {finding.item.angle} | {finding.item.text}")
            print("  reason: " + "；".join(finding.reasons))
            print("  suggestion: " + finding.suggestion)


def print_homogeneity_report(observations: list[HomogeneityObservation], max_items: int) -> None:
    print(f"\n高频同质化观察: {len(observations)}")
    for observation in observations[:max_items]:
        print(f"- {observation.kind} | {observation.count}次 | {observation.phrase}")
        if observation.suggestion:
            print("  suggestion: " + observation.suggestion)
        for item in observation.samples[:3]:
            line = f"L{item.line_no}" if item.line_no else f"row {item.row_no}"
            print(f"  sample: {line} | {item.angle} | {item.text}")


def run_self_test(args: argparse.Namespace) -> None:
    reference_texts = read_reference_texts(args.brand_feedback, args.reference_examples)
    ref_text, ref_tokens, ref_phrases = build_reference_sets(reference_texts)

    cases = [
        ("测试", "粑粑这块还挺规律的，老母亲少操心", "high", ["少操心", "句式风险"]),
        ("测试", "我们家没拉肚子，肚肚还算稳", "high", ["句式风险"]),
        ("测试", "喝得挺顺的", "high", ["挺 + 抽象评价词"]),
        ("测试", "娃从半岁就喝这款，挺稳的", "high", ["挺 + 抽象评价词"]),
        ("测试", "源悦这个奶源我还挺认", "high", ["挺 + 抽象评价词"]),
        ("测试", "同款，从来没有便秘问题的路过~", "pass", []),
        ("测试", "转源悦快一个月，从2/3天一次到每天都拉", "pass", []),
        ("测试", "我们是能恩转源悦，适应得蛮好", "pass", []),
        ("测试", "源悦的奶源和配料都非常好，我很认", "pass", []),
        ("测试", "活动价入的，娃也愿意喝，这单不亏", "high", ["活动价", "这单不亏"]),
        ("测试", "刚转能恩，小罐试了三天还行", "high", ["产品方向错误"]),
        ("测试", "水解刚试了一顿，还没啥反应，继续看看", "high", ["产品方向", "没啥反应"]),
        ("未购，痛点类提问", "粑粑干能喝这款吗", "high", ["运营0603"]),
        ("已购，品牌新晋用户，便便问题", "喝了源悦之后粑粑干", "high", ["粑粑干"]),
        ("未购，冲泡类", "泡泡多吗，怕冲不开", "pass", []),
        ("未购，冲泡类", "摇了好久还是有颗粒，水温多少合适", "high", ["有颗粒"]),
        ("已购，品牌新晋用户，便便问题", "转源悦后便便变绿正常不？", "high", ["便便变绿"]),
        ("已购，品牌新晋用户，过敏相关", "从旧奶粉转过来的，前几顿先少量混着来", "high", ["不要提旧奶粉"]),
        ("已购，品牌新晋用户，性价比", "看了几款，奶源和价格都在预算里", "high", ["预算里"]),
        ("已购，品牌新晋用户，性价比", "刚拍下，奶瓶里不怎么剩，准备续罐了", "high", ["奶瓶里不怎么剩"]),
        ("已购，品牌新晋用户，生长发育", "这罐喝完再称一次，先看奶量有没有接上", "high", ["生长发育"]),
        ("已购，品牌新晋用户，便便问题", "转奶有一阵了，每天基本有，没憋到脸红", "high", ["没憋到脸红"]),
        ("未购，口感类", "源悦转奶好入口吗？娃之前喝别的总吐", "high", ["总吐奶"]),
    ]

    failures: list[str] = []
    for angle, text, expected, expected_reasons in cases:
        item = CorpusItem(0, angle, "测试", text, "example")
        finding = classify_item(item, ref_text, ref_tokens, ref_phrases, True)
        actual = "pass" if finding is None or finding.severity != "high" else "high"
        reason_blob = "；".join(finding.reasons if finding else [])
        if actual != expected:
            failures.append(f"{text}: expected {expected}, got {actual} ({reason_blob})")
            continue
        for reason in expected_reasons:
            if reason not in reason_blob:
                failures.append(f"{text}: missing reason {reason} ({reason_blob})")

    duplicate_findings = duplicate_example_findings(
        [
            CorpusItem(1, "测试", "测试", "同款，从来没有便秘问题的路过~", "example", 10),
            CorpusItem(2, "测试", "测试", "同款，从来没有便秘问题的路过", "example", 20),
        ]
    )
    if len(duplicate_findings) != 1 or duplicate_findings[0].severity != "high":
        failures.append("duplicate example check: expected one high finding")
    elif "示例重复" not in "；".join(duplicate_findings[0].reasons):
        failures.append("duplicate example check: missing duplicate reason")

    ting_observations = homogeneity_observations(
        [
            CorpusItem(1, "测试", "测试", "喝着挺稳的", "example", 10),
            CorpusItem(2, "测试", "测试", "奶量挺稳的", "example", 20),
            CorpusItem(3, "测试", "测试", "转奶挺顺的", "example", 30),
            CorpusItem(4, "测试", "测试", "喝得挺顺的", "example", 40),
        ]
    )
    if not any(item.kind == "ting_phrase" and item.phrase == "挺稳" for item in ting_observations):
        failures.append("ting phrase check: expected repeated 挺稳 observation")
    if not any(item.kind == "ting_phrase" and item.phrase == "挺顺" for item in ting_observations):
        failures.append("ting phrase check: expected repeated 挺顺 observation")

    soft_template_observations = homogeneity_observations(
        [
            CorpusItem(1, "测试", "测试", "娃肯喝，我先续一罐", "example", 10),
            CorpusItem(2, "测试", "测试", "源悦小罐娃肯喝", "example", 20),
            CorpusItem(3, "测试", "测试", "刚拍下，娃肯喝", "example", 30),
            CorpusItem(4, "测试", "测试", "第一顿奶瓶里剩一点", "example", 40),
            CorpusItem(5, "测试", "测试", "前两顿奶瓶里剩一点", "example", 50),
            CorpusItem(6, "测试", "测试", "拉屎没怎么费劲", "example", 60),
            CorpusItem(7, "测试", "测试", "嗯嗯没憋劲", "example", 70),
            CorpusItem(8, "测试", "测试", "便便没憋着", "example", 80),
            CorpusItem(9, "测试", "测试", "臭臭没憋到脸红", "example", 90),
            CorpusItem(10, "测试", "测试", "拉屎不用太费劲", "example", 100),
        ]
    )
    if not any(item.kind == "soft_template" and item.phrase == "娃肯喝" for item in soft_template_observations):
        failures.append("soft template check: expected repeated 娃肯喝 observation")
    if not any(item.kind == "soft_template_pattern" and item.phrase == "短周期接受度模板" for item in soft_template_observations):
        failures.append("soft template pattern check: expected short-cycle acceptance observation")
    if not any(item.kind == "soft_template_pattern" and item.phrase == "便便省事模板" for item in soft_template_observations):
        failures.append("soft template pattern check: expected便便省事模板 observation")

    if failures:
        print("self-test failed")
        for failure in failures:
            print("- " + failure)
        raise SystemExit(1)
    print("self-test passed")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    parser.add_argument("--brand-feedback", type=Path, default=DEFAULT_BRAND)
    parser.add_argument("--reference-examples", type=Path, default=DEFAULT_REFERENCE)
    parser.add_argument("--format", choices=("text", "json"), default="text")
    parser.add_argument("--max-items", type=int, default=50)
    parser.add_argument("--verbose-diff", action="store_true", help="also report raw token and 2-6 char phrase diffs")
    parser.add_argument("--include-rules", action="store_true")
    parser.add_argument("--homogeneity-report", action="store_true", help="also print report-only high-frequency wording observations")
    parser.add_argument("--article-pool-xlsx", type=Path)
    parser.add_argument("--article-pool-sheet", default="文章池数据")
    parser.add_argument("--review-csv", type=Path, help="review CSV with 正文/评论类型/审核人/运营反馈 columns")
    parser.add_argument("--fail-on-high", action="store_true")
    parser.add_argument("--fail-on-homogeneity", action="store_true", help="exit nonzero when homogeneity or soft-template observations exist")
    parser.add_argument("--self-test", action="store_true")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    if args.self_test:
        run_self_test(args)
        return 0

    findings, stats = run_check(args)
    needs_homogeneity = args.homogeneity_report or args.fail_on_homogeneity
    observations = collect_homogeneity_observations(args) if needs_homogeneity else []
    if args.format == "json":
        print(
            json.dumps(
                {
                    "stats": stats,
                    "findings": [f.to_dict() for f in findings],
                    "homogeneity_observations": [item.to_dict() for item in observations],
                },
                ensure_ascii=False,
                indent=2,
            )
        )
    else:
        print_text_report(findings, stats, args.max_items)
        if args.homogeneity_report:
            print_homogeneity_report(observations, args.max_items)

    if args.fail_on_high and (stats["high"] > 0 or stats["rule_high"] > 0):
        return 1
    if args.fail_on_homogeneity and observations:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
