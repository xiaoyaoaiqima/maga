#!/usr/bin/env python3
"""
High-concurrency LLM sentiment classifier for comment CSV/XLSX files.

Canonical script for the A2/XHS LLM sentiment pass:
  python3 llm_comment_sentiment.py INPUT.csv -o OUTPUT.csv

Input fields required:
  comment_type, parent_comment_id, comment_id, comment_text

The script calls an OpenAI-compatible chat completions API and writes a CSV with
the original columns plus sentiment_label, sentiment_confidence, sentiment_reason,
and llm_error.
"""

import argparse
import asyncio
import csv
import json
import os
import random
import re
import sys
import time
from pathlib import Path
from typing import Any

import aiohttp
from openpyxl import load_workbook


REQUIRED_FIELDS = [
    "comment_type",
    "parent_comment_id",
    "comment_id",
    "comment_text",
]

ADDED_FIELDS = [
    "row_index",
    "sentiment_label",
    "sentiment_confidence",
    "sentiment_reason",
    "llm_error",
]

SYSTEM_PROMPT = """你是中文社交媒体评论情感分析员。你的任务是判断评论文本对相关产品、品牌、购买体验或安全事件的态度。

判断视角：
- 目标产品/品牌：A2、a2、至初、A2至初、a2至初、紫白金、A2紫白金、a2紫白金、紫曜。
- 竞品/非目标产品示例：悦白、爱他美、领熠、澳爱、皇家、美素、优博瑞霂、蓝臻、山羊奶等。
- 情感必须按目标产品视角判断，不按评论表面情绪判断。

判定规则：
1. 负向：担忧、不信任、不敢买/喝/用、抱怨质量/安全/价格/缺货/售后、不适反应、召回恐慌、转奶换牌、退货投诉等。
2. 风险求证也判负向：询问安全、召回、批次、名单、还能不能喝/用、有没有问题，代表信任风险和潜在转牌风险。
3. 当前正在喝/刚买/囤货，同时担心召回、考虑换奶/转牌，判负向。
4. 转向竞品或非目标产品判负向：即使评论说“转了悦白/澳爱/爱他美后不吐奶、打嗝正常、挺好”，也代表对目标产品的流失或转牌风险，判负向。
5. 竞品偏好判负向：坚持喝/一直喝/继续喝/推荐爱他美、领熠、悦白、澳爱、皇家美素等竞品，或说竞品价格公道、品质有保证、没跟风换奶，均按目标产品视角判负向。
6. 转向目标产品判正向：转到/换成/喝上 A2、至初、紫白金后适应、便便正常、不吐奶、肚子舒服等，是目标产品正向，不是转牌流失。
7. 正向：认可目标产品/品牌、明确放心、继续使用目标产品、明确澄清目标产品无风险、推荐目标产品、回购、买到/有货、体验好等。
8. 如果开启三分类，中性用于普通知识咨询、无明确倾向、无意义文本、只@人、纯事实搬运或交易噪音；不要把安全/召回/批次求证判中性。
9. 注意否定词和反讽，例如“没有召回”“没问题”通常不是负向；“不放心”“不敢喝”是负向。
10. 只依据 comment_text 判断。comment_type、parent_comment_id、comment_id 只作为定位信息，不要臆测上下文。

必须输出严格 JSON，不要输出 Markdown，不要解释额外文字。"""


def load_env_file(path: Path = Path(".env")) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


def make_user_prompt(records: list[dict[str, Any]], label_mode: str, no_reason: bool) -> str:
    labels = ["正向", "负向"] if label_mode == "binary" else ["正向", "负向", "中性"]
    binary_note = ""
    if label_mode == "binary":
        binary_note = "- 二分类模式下，即使评论较弱，也必须在正向/负向中选择更接近的一类。\n"
    reason_note = "- reason 固定输出空字符串，减少输出 token。\n" if no_reason else "- reason 用中文，控制在 30 字以内。\n"
    reason_example = "" if no_reason else "简短原因"
    compact_records = [
        {
            "row_index": item["row_index"],
            "comment_text": item.get("comment_text", ""),
        }
        for item in records
    ]
    return (
        "请逐条判断下面评论的情感倾向。\n"
        f"允许的 sentiment_label 只能是：{', '.join(labels)}。\n"
        "输出 JSON 格式如下：\n"
        "{"
        f'"results":[{{"row_index":1,"sentiment_label":"正向","confidence":0.92,"reason":"{reason_example}"}}]'
        "}\n"
        "要求：\n"
        "- results 数量必须等于输入数量，不能漏行。\n"
        "- confidence 是 0 到 1 的数字。\n"
        f"{reason_note}"
        "- 纯询问安全、召回、批次、名单、能否继续使用，判负向。\n"
        "- A2/a2/至初/紫白金/a2紫白金/A2紫白金 是目标产品；转到这些产品且体验好，判正向。\n"
        "- 转到悦白/爱他美/领熠/澳爱/皇家/美素/蓝臻/山羊奶等非目标产品，判负向。\n"
        f"{binary_note}\n"
        "输入评论 JSON：\n"
        f"{json.dumps(compact_records, ensure_ascii=False)}"
    )


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_input(path: Path, sheet_name: str | None) -> tuple[list[dict[str, Any]], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames or []
            validate_fields(fieldnames, path)
            rows = []
            for i, row in enumerate(reader, 1):
                item = {key: normalize_cell(row.get(key, "")) for key in fieldnames}
                if item.get("comment_text"):
                    item["row_index"] = i
                    rows.append(item)
            return rows, fieldnames

    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [normalize_cell(v) for v in next(rows_iter)]
        except StopIteration:
            raise ValueError(f"{path} is empty") from None
        validate_fields(headers, path)
        rows = []
        for i, values in enumerate(rows_iter, 1):
            raw = {
                headers[col_i]: normalize_cell(values[col_i] if col_i < len(values) else "")
                for col_i in range(len(headers))
            }
            if raw.get("comment_text"):
                raw["row_index"] = i
                rows.append(raw)
        return rows, headers

    raise ValueError(f"unsupported input type: {path.suffix}; use .csv, .xlsx, or .xlsm")


def validate_fields(fieldnames: list[str], path: Path) -> None:
    missing = [name for name in REQUIRED_FIELDS if name not in fieldnames]
    if missing:
        raise ValueError(f"{path} missing required fields: {', '.join(missing)}")


def existing_row_indexes(path: Path) -> set[str]:
    if not path.exists() or path.stat().st_size == 0:
        return set()
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return {row.get("row_index", "") for row in reader if row.get("row_index")}


def chunked(items: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def extract_json_object(text: str) -> dict[str, Any]:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    match = re.search(r"\{.*\}", text, re.S)
    if not match:
        raise ValueError(f"model did not return JSON: {text[:300]}")
    return json.loads(match.group(0))


def normalize_model_results(
    payload: dict[str, Any],
    batch: list[dict[str, Any]],
    label_mode: str,
) -> list[dict[str, Any]]:
    allowed = {"正向", "负向"} if label_mode == "binary" else {"正向", "负向", "中性"}
    raw_results = payload.get("results")
    if not isinstance(raw_results, list):
        raise ValueError("JSON missing results array")

    by_index = {}
    for item in raw_results:
        if not isinstance(item, dict):
            continue
        row_index = str(item.get("row_index", ""))
        label = str(item.get("sentiment_label", "")).strip()
        if label not in allowed:
            raise ValueError(f"bad label for row {row_index}: {label}")
        try:
            confidence = float(item.get("confidence", 0))
        except (TypeError, ValueError):
            confidence = 0
        by_index[row_index] = {
            "sentiment_label": label,
            "sentiment_confidence": max(0, min(1, confidence)),
            "sentiment_reason": str(item.get("reason", "")).strip()[:80],
            "llm_error": "",
        }

    normalized = []
    for record in batch:
        row_index = str(record["row_index"])
        if row_index not in by_index:
            raise ValueError(f"model omitted row_index={row_index}")
        normalized.append({**record, **by_index[row_index]})
    return normalized


async def call_llm_batch(
    session: aiohttp.ClientSession,
    semaphore: asyncio.Semaphore,
    batch: list[dict[str, Any]],
    args: argparse.Namespace,
) -> list[dict[str, Any]]:
    async with semaphore:
        messages = [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": make_user_prompt(batch, args.label_mode, args.no_reason)},
        ]
        body: dict[str, Any] = {
            "model": args.model,
            "messages": messages,
            "temperature": 0,
        }
        if args.max_tokens:
            body["max_tokens"] = args.max_tokens
        if not args.no_response_format:
            body["response_format"] = {"type": "json_object"}

        last_error = ""
        for attempt in range(args.max_retries + 1):
            try:
                async with session.post(args.api_url, json=body, timeout=args.timeout) as resp:
                    response_text = await resp.text()
                    if resp.status in {429, 500, 502, 503, 504}:
                        raise RuntimeError(f"HTTP {resp.status}: {response_text[:300]}")
                    if resp.status >= 400:
                        raise RuntimeError(f"HTTP {resp.status}: {response_text[:600]}")

                response_json = json.loads(response_text)
                content = response_json["choices"][0]["message"]["content"]
                payload = extract_json_object(content)
                return normalize_model_results(payload, batch, args.label_mode)
            except Exception as exc:  # noqa: BLE001 - keep batch retry logic centralized.
                last_error = str(exc)
                if attempt >= args.max_retries:
                    break
                delay = min(args.retry_max_sleep, args.retry_base_sleep * (2**attempt))
                await asyncio.sleep(delay + random.uniform(0, 0.5))

        return [
            {
                **record,
                "sentiment_label": "",
                "sentiment_confidence": "",
                "sentiment_reason": "",
                "llm_error": last_error[:500],
            }
            for record in batch
        ]


async def run(args: argparse.Namespace) -> None:
    input_path = Path(args.input)
    output_path = Path(args.output)
    rows, input_fields = read_input(input_path, args.sheet)

    if args.resume:
        done = existing_row_indexes(output_path)
        rows = [row for row in rows if str(row["row_index"]) not in done]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = input_fields + [name for name in ADDED_FIELDS if name not in input_fields]
    write_header = not output_path.exists() or output_path.stat().st_size == 0 or not args.resume
    mode = "a" if args.resume else "w"

    headers = {
        "Authorization": f"Bearer {args.api_key}",
        "Content-Type": "application/json",
    }
    timeout = aiohttp.ClientTimeout(total=args.timeout)
    semaphore = asyncio.Semaphore(args.concurrency)
    batches = chunked(rows, args.batch_size)

    started = time.time()
    processed = 0
    failed = 0

    with output_path.open(mode, encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        if write_header:
            writer.writeheader()

        async with aiohttp.ClientSession(headers=headers, timeout=timeout) as session:
            tasks = [
                asyncio.create_task(call_llm_batch(session, semaphore, batch, args))
                for batch in batches
            ]
            for task in asyncio.as_completed(tasks):
                result_rows = await task
                writer.writerows(result_rows)
                f.flush()
                processed += len(result_rows)
                failed += sum(1 for row in result_rows if row.get("llm_error"))
                if processed % args.progress_every == 0 or processed == len(rows):
                    elapsed = max(0.001, time.time() - started)
                    print(
                        f"processed={processed}/{len(rows)} failed={failed} "
                        f"speed={processed / elapsed:.1f} rows/s",
                        file=sys.stderr,
                    )

    print(f"done input_rows={len(rows)} failed={failed} output={output_path}")


def build_api_url(args: argparse.Namespace) -> str:
    if args.api_url:
        return args.api_url.rstrip("/")
    base = args.api_base.rstrip("/")
    if base.endswith("/chat/completions"):
        return base
    return f"{base}/chat/completions"


def parse_args() -> argparse.Namespace:
    load_env_file()
    parser = argparse.ArgumentParser(
        description="Classify comment_text sentiment with a concurrent OpenAI-compatible LLM API.",
    )
    parser.add_argument("input", help="input .csv/.xlsx/.xlsm file")
    parser.add_argument("-o", "--output", required=True, help="output CSV path")
    parser.add_argument("--sheet", help="Excel sheet name; defaults to active sheet")
    parser.add_argument("--api-key", default=os.getenv("LLM_API_KEY") or os.getenv("OPENAI_API_KEY"))
    parser.add_argument("--api-base", default=os.getenv("LLM_API_BASE") or os.getenv("OPENAI_BASE_URL") or "https://api.openai.com/v1")
    parser.add_argument("--api-url", default=os.getenv("LLM_API_URL"), help="full chat completions URL; overrides --api-base")
    parser.add_argument("--model", default=os.getenv("LLM_MODEL", "deepseek-v4-flash"))
    parser.add_argument("--label-mode", choices=["binary", "ternary"], default="ternary")
    parser.add_argument("--batch-size", type=int, default=50)
    parser.add_argument("--concurrency", type=int, default=8)
    parser.add_argument("--max-tokens", type=int, default=0, help="optional max completion tokens per batch; 0 means provider default")
    parser.add_argument("--no-reason", action="store_true", help="write empty sentiment_reason to reduce output tokens and improve speed")
    parser.add_argument("--timeout", type=float, default=90)
    parser.add_argument("--max-retries", type=int, default=3)
    parser.add_argument("--retry-base-sleep", type=float, default=1.5)
    parser.add_argument("--retry-max-sleep", type=float, default=20)
    parser.add_argument("--progress-every", type=int, default=200)
    parser.add_argument("--resume", action="store_true", help="append and skip existing row_index values in output")
    parser.add_argument("--no-response-format", action="store_true", help="disable response_format=json_object for providers that do not support it")
    args = parser.parse_args()

    if not args.api_key:
        parser.error("missing API key; set LLM_API_KEY or OPENAI_API_KEY, or pass --api-key")
    if args.batch_size < 1:
        parser.error("--batch-size must be >= 1")
    if args.concurrency < 1:
        parser.error("--concurrency must be >= 1")

    args.api_url = build_api_url(args)
    return args


def main() -> None:
    args = parse_args()
    asyncio.run(run(args))


if __name__ == "__main__":
    main()
